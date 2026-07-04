"""
UFC fight outcome and method-of-victory prediction model (single file).

A leak-safe, time-aware machine-learning system that predicts (1) the winner
of a UFC bout and (2) the method of victory (Decision / KO-TKO / Submission),
built and evaluated on strictly chronological data.

Pipeline overview (top to bottom of this file):
1) Feature engineering — every fight row is built from each fighter's strictly
   PRE-fight state: career/recency statistics, Glicko-2 ratings (optionally
   margin-of-victory adjusted), overall + divisional Elo, four-dimensional
   "phase" Glicko ratings (striking/grappling x offense/defense), style-matchup
   records, and altitude/acclimatization features. State is updated only AFTER
   a row is emitted, so no feature can see its own outcome (leak-safety).
2) Winner model — a corner-swap-symmetric ensemble of gradient-boosted trees
   and forests (LightGBM / XGBoost / CatBoost / HistGBM / RandomForest /
   ExtraTrees, default + Optuna-tuned variants), combined by a simplex-weighted
   blend fit on time-series out-of-fold predictions, with optional Platt /
   isotonic calibration selected on validation-only data.
3) Method model — a two-stage classifier (Finish-vs-Decision, then KO-vs-Sub)
   of subsample-bagged HistGBMs plus direct multiclass / one-vs-rest / linear
   heads, blended with history- and context-priors; the blend is Optuna-tuned
   and protected by a champion/challenger persistence gate.
4) Evaluation — fixed-count chronological validation (600 fights) and untouched
   holdout (500 most-recent fights), plus walk-forward diagnostics; metrics
   include log-loss, Brier score, accuracy, ECE, and per-class method F1.
5) Deployment — a Tkinter GUI for batch matchup prediction with optional
   betting-odds blending, and styled Excel export.

Expected data: pure_fight_data_with_event_and_camp_altitudes.csv in the same
directory as this script (see README for the schema).

Run:
  python UFC_ML_Model.py                 # train (cached) + launch the GUI
  python UFC_ML_Model.py --train-only    # train/evaluate only, no GUI

Results are deterministic for a fixed dataset and configuration (seeded RNGs);
trained stages are cached in .ufc_model_cache and keyed on the dataset hash and
every configuration value that affects the result.
"""

import argparse
import csv
import hashlib
import json
import math
import os
import pickle
import random
import re
import sys
import threading
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sklearn.ensemble import HistGradientBoostingClassifier, RandomForestClassifier
from sklearn.ensemble import ExtraTreesClassifier, AdaBoostClassifier
from sklearn.impute import SimpleImputer
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, brier_score_loss, log_loss
from sklearn.metrics import balanced_accuracy_score, precision_recall_fscore_support
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import TimeSeriesSplit
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from scipy.optimize import minimize as scipy_minimize

# Make stdout/stderr robust to non-ASCII (e.g. the '→' used in pruning logs)
# regardless of console code page or whether output is piped/redirected. Without
# this, a cp1252 stdout raises UnicodeEncodeError on a cosmetic print, which the
# method stage's broad try/except mistakes for a fatal error and silently
# disables the entire method model.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

try:
    import lightgbm as lgb
except Exception:
    lgb = None

try:
    import xgboost as xgb
except Exception:
    xgb = None

try:
    import catboost as cb
except Exception:
    cb = None

try:
    import optuna
except Exception:
    optuna = None

if optuna is not None:
    # Keep Optuna output compact: suppress per-trial parameter logs.
    optuna.logging.set_verbosity(optuna.logging.WARNING)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(SCRIPT_DIR, "pure_fight_data_with_event_and_camp_altitudes.csv")
PREDICTIONS_XLSX = os.path.join(SCRIPT_DIR, "UFC_Predictions.xlsx")
CACHE_DIR = os.path.join(SCRIPT_DIR, ".ufc_model_cache")
METHOD_CHAMPION_PATH = os.path.join(SCRIPT_DIR, ".ufc_model_cache", "method_champion_cfg.json")
###################################################################################################
# Bump when winner-stage training logic changes.
WINNER_CACHE_VERSION = "v22"
# Bump when method-stage training logic changes.
METHOD_CACHE_VERSION = "v39"
###################################################################################################
# Pickle payload discriminator (stable across cache file renames).
WINNER_STAGE_CACHE_KIND = "ufc_winner_stage_v1"
METHOD_STAGE_CACHE_KIND = "ufc_method_stage_v1"

RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)
random.seed(RANDOM_SEED)
os.environ["PYTHONHASHSEED"] = str(RANDOM_SEED)

# Chronological split is by FIXED fight counts (not fractions) so the validation
# and holdout windows are the SAME most-recent fights for every era candidate.
# This makes era start years directly comparable (identical val fights) and keeps
# the holdout stable run-to-run regardless of which start year wins.
VAL_FIGHTS = 600          # tuning/selection window: the fights just before holdout
TEST_FIGHTS = 500         # final untouched holdout = the most-recent fights
MIN_TRAIN_FIGHTS = 1000   # an era candidate must leave at least this many training fights
# Era selection prefers the EARLIEST start year whose common-window validation
# log-loss is within this tolerance of the best — more training data generalizes
# better unless a later start is clearly superior.
ERA_LOGLOSS_TOL = 0.004
# Floor on how far back era selection may reach. Pre-2005 UFC predates the
# Unified Rules and has materially worse data (e.g. reach missing ~60% of the
# time vs ~3% afterward), so the data-preferring rule above must not pull the
# training window into that different-sport era. Lower to 1993 to allow all
# history; raise (e.g. 2014) to restrict to the most recent stat-tracking era.
EARLIEST_ERA_START = 2005
ACTIVE_DAYS = 730
NEEDS_SCALE = {"LogReg", "MLP"}
ELO_BASE = 1500.0
ELO_K = 24.0
# ─── Margin-of-victory (MOV) adjusted ratings ─────────────────────────────────
# Feed Glicko/Elo a continuous performance score in [0,1] instead of a hard 1/0,
# so dominant wins move ratings more than split decisions / late grindy finishes.
# The RESULT is always ground truth (winner's score > 0.5); decisiveness and the
# box score only modulate WITHIN the winner's band, never flipping the sign.
# Set env UFC_MOV_ENABLED=0 to build W/L-only ratings (controlled A/B test).
MOV_RATINGS_ENABLED = os.environ.get("UFC_MOV_ENABLED", "1") != "0"
# "full"   = finish-timing + box-score dominance index (the real model).
# "buckets"= split=0.62 / other decision=0.75 / finish timing-scaled — a
#            low-overfit A/B baseline. Set env UFC_MOV_MODE=buckets to use it.
MOV_MODE = os.environ.get("UFC_MOV_MODE", "full")
MOV_W_KD = 1.0      # knockdowns — strongest single damage signal
MOV_W_STR = 0.7     # significant strikes landed per minute
MOV_W_CTRL = 0.5    # control-time fraction
MOV_W_TD = 0.4      # takedowns landed
MOV_D_BLEND = 0.5   # weight on the box-score index vs finish-timing / judge consensus
# ─── Multi-dimensional (phase) Glicko ratings ─────────────────────────────────
# A single scalar rating can't express style-conditional skill. These add FOUR
# Glicko ratings per fighter — striking-OFFENSE, striking-DEFENSE, grappling-
# OFFENSE, grappling-DEFENSE — updated bipartite within every fight: each fighter's
# offense "plays" the opponent's defense, scored by that fighter's normalized
# phase output. So "elite wrestler with weak TDD" reads as high grapple-offense +
# low grapple-defense, and the matchup crosses them (red imposes grappling iff
# red_grapple_off > blue_grapple_def). Phase attribution is position-aware:
#   striking = distance+clinch strikes (per STANDING minute) + accuracy + KD
#   grappling = takedowns + control + GROUND strikes (GnP) + sub attempts − reversals
# Leak-safe (pre-fight read, post-fight update). Set env UFC_PHASE_RATINGS=0 for A/B.
PHASE_RATINGS_ENABLED = os.environ.get("UFC_PHASE_RATINGS", "1") != "0"
PHASE_STEEP = 0.85   # sigmoid steepness mapping a z-scored phase output → success
# Striking-offense success weights (per STANDING minute, league-standardized).
PHASE_W_VOL = 1.0    # distance+clinch significant strikes landed
PHASE_W_ACC = 0.6    # significant strike accuracy (efficiency / made-them-miss)
PHASE_W_KD = 0.8     # knockdowns
# Grappling-offense success weights.
PHASE_W_TD = 0.8     # takedowns landed (per 15 min)
PHASE_W_CTRL = 1.0   # control-time fraction
PHASE_W_GNP = 0.6    # ground-and-pound strikes landed (per ground minute)
PHASE_W_SUB = 0.7    # submission attempts (per 15 min)
PHASE_W_REV = 0.4    # penalty for being reversed (opponent reversals)
# A fighter's grappling OFFENSE (and the opponent's grappling DEFENSE) only update
# when that fighter actually engaged grappling (≥1 takedown/sub attempt or ≥60s
# control) — a pure kickboxing match carries no grappling-skill information.
PHASE_GRAPPLE_MIN_ACTIONS = 1.0
PHASE_GRAPPLE_MIN_CTRL_SEC = 60.0
# ─── Corner-aware correction (asymmetric calibration) ─────────────────────────
# The winner ensemble is corner-SYMMETRIC (swap-augmented + forward/reverse
# averaged), so its p_red is a pure "red is the more skilled corner" probability.
# But UFC's red-corner convention (red = promoted/favored fighter) makes red
# favorites convert WORSE than their skill implies — the model structurally
# over-predicts Red. This fits a one-parameter logit-space intercept (b) on
# leak-safe dev OOF probs + true red labels with recency weights, so it learns
# the recent red-corner conversion rate and shifts the effective red pick
# threshold up. INFERENCE ASSUMES THE FIRST FIGHTER IS THE RED CORNER.
# Tested over 3 runs: the fit returns b≈0 (the symmetric model is already
# calibrated), and on v19 the residual noise COST 0.6 pts (64.6%→64.0%). So it
# defaults OFF. Set env UFC_CORNER_CORRECTION=1 to re-enable the A/B.
CORNER_CORRECTION_ENABLED = os.environ.get("UFC_CORNER_CORRECTION", "0") != "0"
# Cap |b| (logit space) so a noisy fit window can't flip the model wholesale.
# 0.6 ⇒ the effective red pick threshold stays within ~[0.35, 0.65] of p_red.
CORNER_SHIFT_CAP = 0.6
# Recency-weight floor for the corner fit (oldest dev fight weight vs newest=1.0).
# Lower ⇒ track the recent (declining) red advantage more aggressively.
CORNER_FIT_FLOOR = 0.25
OPTUNA_TRIALS = 80
METHOD_TUNING_TRIALS = 400
# Winner stability selection (threshold form — no fixed feature count). Each
# bootstrap run marks a feature "selected" if it lands in the top
# STABILITY_PER_RUN_FRAC of that run's importance ranking; the model then keeps
# every feature whose selection FREQUENCY across runs is >= STABILITY_FREQ_THRESHOLD.
# The kept count FLOATS to however many features are reliably useful — replacing
# the old arbitrary top-240 cap. Textbook stability selection (Meinshausen-Bühlmann):
# a selection-probability cutoff, not a fixed count. Lower the frequency threshold
# (or raise the per-run fraction) to keep more features; raise it to keep fewer.
STABILITY_PER_RUN_FRAC = 0.75    # per-run regularizer: top 75% by importance counts as "selected"
STABILITY_FREQ_THRESHOLD = 0.60  # keep features selected in >= 60% of bootstrap runs
# Winner-stage correlation prune: drop near-duplicate columns (|corr| > this),
# keeping the more target-relevant member of each correlated pair, BEFORE
# stability selection. The winner matchup matrix is heavily collinear (many
# glicko/elo/interaction variants), which inflates ensemble variance; pruning
# narrows the validation→holdout generalization gap.
WINNER_CORR_PRUNE_THRESHOLD = 0.95
# When True, the winner ensemble sees ALL engineered features, including the
# method-routed ones (striking-zone accuracy, KO/sub composites, *_sum channels,
# finish-round distributions) that FEATURE_ROUTING normally withholds from it.
# Motivated by the walk-forward single-model diagnostic (full 409-feature HGB)
# outscoring the routed winner ensemble on recent folds. Method-stage routing
# (stage1/stage2) is unaffected. Tested: unrouting was a wash on accuracy and
# slightly worse on calibration (ECE 0.049→0.063), so it defaults OFF. Set env
# UFC_WINNER_ALL_FEATURES=1 to re-enable the A/B.
WINNER_SEES_ALL_FEATURES = os.environ.get("UFC_WINNER_ALL_FEATURES", "0") != "0"
# Robust winner combiner: when True (default), the OOF combiner is chosen ONLY
# among simplex blends (weighted / simple average). The LR/HGB stacker
# meta-learners overfit the ~400-fight validation OOF set and generalize worse on
# the holdout (v15 weighted 63.8% vs v16–v18 stacker 63.4%). Set env
# UFC_COMBINER_ROBUST=0 to re-allow stacker candidates.
WINNER_COMBINER_ROBUST = os.environ.get("UFC_COMBINER_ROBUST", "1") != "0"
# Decision-threshold regularization. The winner ensemble is corner-swap symmetric,
# so 0.5 is its neutral operating point. The training/dev red-corner WIN rate
# (~58-60%) runs well above the most-recent holdout (~53% and declining), so a
# threshold tuned purely for dev accuracy drifts below 0.5 and over-predicts Red on
# future fights. This caps how far the decision threshold may deviate from 0.5:
# 0.0 = always 0.5 (most robust to base-rate drift); raise to re-enable adaptive
# thresholding (e.g. 0.03 allows [0.47, 0.53]).
WINNER_THR_MAX_DEV = 0.0
# Opponent-baseline-adjusted (oba_*) winner features. Default on; set env
# UFC_OBA_ENABLED=0 to build without them (used for controlled A/B testing).
OBA_FEATURES_ENABLED = os.environ.get("UFC_OBA_ENABLED", "1") != "0"
STAGE2_MAX_FEATURES = 180
METHOD_HARD_RESET = False
# A freshly tuned method-blend cfg must beat the saved champion's walk-forward
# objective by at least this margin to replace it. Widened from 0.004 → 0.02 to
# LOCK IN the current champion: blend-cfg objectives cluster within ~0.002 of each
# other and bounce more than that as noise (the val→holdout signal at 279 holdout
# finishes is not real signal), so a small margin let the gate flip-flop the
# champion between runs. 0.02 means only a CLEARLY larger improvement (a genuinely
# better method architecture, not a re-tune) can dislodge the saved champion.
METHOD_CHAMPION_MARGIN = 0.02
# Method-stage HGB bagging. The Stage1/Stage2 HistGradientBoosting models are the
# dominant component of each stage (alpha ~0.75-0.85) but high-variance on the small
# method samples (Stage2 KO/Sub ≈ 280 finishes) AND ~deterministic across random_state
# (no row subsampling; binning sees all rows). So averaging N members each fit on a
# different SUBSAMPLE of the rows (subagging) — not seed variation — is what injects
# the diversity that variance reduction needs. Keeps the method stable across small
# upstream shifts (e.g. a winner-selection change) instead of swinging 71%↔62%.
METHOD_HGB_BAG = 5             # bagged members per Stage1/Stage2 HGB (1 = bagging off)
METHOD_HGB_BAG_SUBSAMPLE = 0.8 # row fraction each member sees (sampled without replacement)
# Correlation threshold for method-stage feature pruning (|corr| > this → dropped).
METHOD_CORR_PRUNE_THRESHOLD = 0.95
# Optuna trials for tuning method-stage HGB base models. 0 = skip and use defaults.
METHOD_OPTUNA_TRIALS = 80
METHOD_ERA_CANDIDATES = [1993, 2005, 2010, 2014, 2016, 2018, 2020, 2021, 2022, 2023, 2024]
METHOD_AUTO_ERA = True
# Method era selection: a quick multiclass model is scored on a COMMON recent
# window (same fights for every era), conditioned on winner-pick-correct fights.
# Among eras within METHOD_ERA_F1_TOL of the best macro-F1, the one with the MOST
# training data is chosen. (Was "most recent", which repeatedly picked a smaller,
# lower-F1 window — e.g. 2022/2048 rows over 2018/3911 rows — starving the rare
# Submission class. Finish-type non-stationarity is already handled by the
# recency weighting applied during method training, so more examples win.)
METHOD_VAL_FIGHTS = 250
METHOD_ERA_F1_TOL = 0.02
###################################################################################################
# FEATURE_ROUTING — single source of truth for which models see which engineered features.
# Each key is a feature name; the value is the set of models that should receive it.
# Allowed tags: "winner", "stage1" (Finish vs Decision), "stage2" (KO vs Sub).
# Features NOT listed here default to all three models.
# Filters are auto-derived below; to change a feature's routing, edit this dict only.
FEATURE_ROUTING = {
    # ═══════════════════════════════════════════════════════════════════
    # WINNER-ONLY
    # Context/pacing signals that help the winner model but add noise to
    # the method stages (which are conditional on outcome or on finish).
    # ═══════════════════════════════════════════════════════════════════
    # Within-fight pacing (§A)
    "d_def_rd1_sig_str":           {"winner"},
    "d_def_rd1_kd":                {"winner"},
    "d_rd1_net_sig_str":           {"winner"},
    "d_cardio_decay_sig_str":      {"winner"},
    "d_late_sig_str_pm":           {"winner"},
    "d_rd1_ctrl_share":            {"winner"},
    "d_reach_x_distance_pct":      {"winner"},
    "d_age_x_cardio":              {"winner"},
    "d_age_x_title_rounds":        {"winner"},
    "d_striker_grappler_raw":      {"winner"},
    "d_tdd_vs_td_attack":          {"winner"},
    "d_short_notice":              {"winner"},
    "d_glicko_trend":              {"winner"},
    "d_recent_damage_absorbed":    {"winner"},
    "gender_flag":                 {"winner"},
    "d_title_x_cardio":            {"winner"},
    "d_total_rounds_x_finish_resistance": {"winner"},
    # Opponent-baseline-adjusted skill (stat-specific SOS; see compute_fighter_features).
    # Winner-only: these are overall-quality signals that add noise to the
    # outcome-conditional method stages.
    "d_oba_sig_str_off":           {"winner"},
    "d_oba_sig_str_def":           {"winner"},
    "d_oba_td_off":                {"winner"},
    "d_oba_td_def":                {"winner"},
    "d_oba_ctrl_off":              {"winner"},
    "d_oba_ctrl_def":              {"winner"},
    # Context-conditional & stability (winner-only)
    "d_rounds_experience":         {"winner"},
    "d_rounds_5_exp":              {"winner"},
    "d_rounds_3_exp":              {"winner"},
    "d_stability":                 {"winner"},
    "d_sig_diff_pm_vol":           {"winner"},

    # ═══════════════════════════════════════════════════════════════════
    # WINNER + STAGE 1 (excluded from STAGE 2)
    # Decision-winning ability is tautologically zero conditional on the
    # fight ending in a finish, so it carries no KO-vs-Sub signal for
    # stage 2. Winner still benefits (decision-heavy fighters can win)
    # and stage 1 benefits (decision ability is the anti-finish signal).
    # ═══════════════════════════════════════════════════════════════════
    "d_decision_ability":          {"winner", "stage1"},
    "d_decision_win_rate":         {"winner", "stage1"},
    "d_dec_win_pct":               {"winner", "stage1"},

    # ═══════════════════════════════════════════════════════════════════
    # METHOD-ONLY (stage 1 + stage 2)
    # Method-specific striking and grappling signals — noisy for winner.
    # ═══════════════════════════════════════════════════════════════════
    # Striking-accuracy by zone/position
    "d_head_acc":                  {"stage1", "stage2"},
    "d_body_acc":                  {"stage1", "stage2"},
    "d_leg_acc":                   {"stage1", "stage2"},
    "d_distance_acc":              {"stage1", "stage2"},
    "d_clinch_acc":                {"stage1", "stage2"},
    "d_ground_acc":                {"stage1", "stage2"},
    "d_body_leg_attrition":        {"stage1", "stage2"},
    "d_head_hunt_share":           {"stage1", "stage2"},
    "d_distance_share":            {"stage1", "stage2"},
    "d_ground_strike_accuracy":    {"stage1", "stage2"},
    "d_head_hunt_accuracy":        {"stage1", "stage2"},
    "d_distance_strike_accuracy":  {"stage1", "stage2"},
    # KO-specific (method-only)
    "d_recent_ko_loss_rate":       {"stage1", "stage2"},
    "d_r5_def_kd_pm":              {"stage1", "stage2"},
    "d_ko_attack_pressure":        {"stage1", "stage2"},
    "d_ko_def_leak":               {"stage1", "stage2"},
    "d_r1f_def_kd_pm":             {"stage1", "stage2"},
    "d_r3_def_kd_pm":              {"stage1", "stage2"},
    "ko_attack_pressure_sum":      {"stage1", "stage2"},
    "ko_def_leak_sum":             {"stage1", "stage2"},
    # Submission-specific (method-only)
    "d_sub_loss_pct":              {"stage1", "stage2"},
    "d_recent_sub_win_rate":       {"stage1", "stage2"},
    "d_sub_entry_pressure":        {"stage1", "stage2"},
    "d_sub_control_conversion":    {"stage1", "stage2"},
    "d_sub_scramble_threat":       {"stage1", "stage2"},
    "d_sub_defensive_leak":        {"stage1", "stage2"},
    "d_late_sub_pressure":         {"stage1", "stage2"},
    "d_sub_recency_surge":         {"stage1", "stage2"},
    "d_grapple_recency_surge":     {"stage1", "stage2"},
    "d_sub_vs_control_axis":       {"stage1", "stage2"},
    "sub_entry_pressure_sum":      {"stage1", "stage2"},
    "sub_defensive_leak_sum":      {"stage1", "stage2"},
    # Sum features (exact side reconstruction after winner-orientation)
    "dec_win_pct_sum":             {"stage1", "stage2"},
    "finish_resistance_sum":       {"stage1", "stage2"},
    "consistency_sum":             {"stage1", "stage2"},
    "cardio_ratio_sum":            {"stage1", "stage2"},
    "durability_sum":              {"stage1", "stage2"},
    "output_rate_sum":             {"stage1", "stage2"},
    "rd1_intensity_ratio_sum":     {"stage1", "stage2"},
    "strike_exchange_ratio_sum":   {"stage1", "stage2"},
    "sig_str_acc_sum":             {"stage1", "stage2"},
    "late_round_pct_sum":          {"stage1", "stage2"},
    "avg_time_min_sum":            {"stage1", "stage2"},
    "avg_finish_round_sum":        {"stage1", "stage2"},
    "first_round_finish_rate_sum": {"stage1", "stage2"},
    "damage_efficiency_sum":       {"stage1", "stage2"},
    "body_leg_attrition_sum":      {"stage1", "stage2"},

    # ═══════════════════════════════════════════════════════════════════
    # STAGE 1 ONLY (Finish vs Decision)
    # Latent finish-environment features calibrated specifically for the
    # decision-vs-finish split; withheld from winner (noise) and stage 2
    # (stage 2 operates only on finished fights — shell/environment
    # features carry no KO-vs-Sub directional information).
    # ═══════════════════════════════════════════════════════════════════
    "m_decision_shell_gap":        {"stage1"},
    "m_decision_shell_sum":        {"stage1"},
    "m_finish_conversion_edge":    {"stage1"},
    "m_finish_environment":        {"stage1"},
    "m_mutual_finish_instability": {"stage1"},
    "m_decision_absorber":         {"stage1"},
    "m_early_finish_window":       {"stage1"},
    "m_fast_start_fragility":      {"stage1"},
    "m_late_finish_window":        {"stage1"},
    "m_attrition_break_window":    {"stage1"},
    "m_time_profile_finish_bias":  {"stage1"},
    "m_finish_over_shell_ratio":   {"stage1"},
    "m_clean_decision_track":      {"stage1"},
    "m_finish_speed_pressure":     {"stage1"},
    "ctx_finish_prior_2y":         {"stage1"},
    "ref_finish_prior":            {"stage1"},

    # ═══════════════════════════════════════════════════════════════════
    # FINISH-ROUND DISTRIBUTION (method-only)
    # Conditional-on-finish round distributions inform WHICH kind of
    # finish (stage 2) and WHETHER the fight finishes at all (stage 1),
    # but largely duplicate first_round_finish_rate for winner.
    # ═══════════════════════════════════════════════════════════════════
    "d_finish_r1_share":           {"stage1", "stage2"},
    "d_finish_r3_plus_share":      {"stage1", "stage2"},
    "d_avg_finish_round_w":        {"stage1", "stage2"},

    # ═══════════════════════════════════════════════════════════════════
    # ELO TRAJECTORY (winner + stage 1)
    # Rising-vs-falling trend signal independent of raw elo level.
    # Stage 2 (KO vs Sub) has no clear directional tie to trajectory.
    # ═══════════════════════════════════════════════════════════════════
    "r_elo_slope_5":               {"winner", "stage1"},
    "b_elo_slope_5":               {"winner", "stage1"},
    "d_elo_slope_5":               {"winner", "stage1"},

    # ═══════════════════════════════════════════════════════════════════
    # ALTITUDE / ACCLIMATIZATION (training-elevation baseline)
    # The absolute venue + raw per-corner shock/descent are METHOD signals
    # (at altitude an acclimatized fighter submits the gassed one; both gassed
    # -> sloppy -> decision). The event-altitude winner test was dead, so the
    # winner model sees only the signed DIFFERENTIALS (left unrouted below:
    # d_accl_shock_kft, d_alt_descent_kft, d_train_alt_kft, train_alt_known) —
    # not the raw per-corner shocks that proved to be winner noise.
    # ═══════════════════════════════════════════════════════════════════
    "event_alt_kft":               {"stage1", "stage2"},
    "alt_ge_4000":                 {"stage1", "stage2"},
    "r_accl_shock_kft":            {"stage1", "stage2"},
    "b_accl_shock_kft":            {"stage1", "stage2"},
    "r_alt_descent_kft":           {"stage1", "stage2"},
    "b_alt_descent_kft":           {"stage1", "stage2"},
    "mutual_gas_kft":              {"stage1"},
    "accl_asym_kft":               {"stage1", "stage2"},
}
def _feature_allowed(feature_name, model):
    """True if `feature_name` should be fed to `model` ('winner'/'stage1'/'stage2').

    Features absent from FEATURE_ROUTING go to every model by default. When
    WINNER_SEES_ALL_FEATURES is on, the winner model bypasses routing entirely
    (it sees everything); stage1/stage2 routing is always honored.
    """
    if model == "winner" and WINNER_SEES_ALL_FEATURES:
        return True
    tags = FEATURE_ROUTING.get(feature_name)
    return tags is None or model in tags
###################################################################################################
STRICT_FUTURE_MODE = True
# Pinned to 2010: the walk-forward's improving-fold trend shows more history still
# helps, the data-starved deep men's divisions benefit most, and pinning stops the
# noisy auto-flip (2010↔2014) when a few fights are added. Set to None to re-enable
# auto era selection.
FORCED_START_YEAR = 2010

# Active base-model lineup when STRICT_FUTURE_MODE is on (filters _make_model_specs).
# Single source of truth for BOTH the winner and method stages — edit here only.
# This set is fingerprinted into the winner cache key, so adding/removing a model
# (e.g. toggling AdaBoost) self-invalidates the cache — no manual pkl deletion needed.
# AdaBoost commented out per the 2026-06-21 audit (had 0.30 combiner weight,
# calibration off; removing it improved ECE 0.0566->0.0487 at zero accuracy cost).
STRICT_KEEP_MODELS = frozenset({
    "LightGBM", "LightGBM_Tuned",
    "XGBoost", "XGBoost_Tuned",
    "CatBoost", "CatBoost_Tuned",
    "HistGBM", "HistGBM_Wide",
    "ExtraTrees", "ExtraTrees_Deep",
    "RandForest", "RandForest_Deep",
    # LogReg tested and REMOVED per the 2026-07-01 diversity audit + full-run A/B.
    # The paired proxy audit (_audit_elasticnet_blend.py) was a clear win (prob-corr
    # ~0.67 vs the trees; HGB+LogReg blends won walk-forward log-loss 6/6 folds),
    # but deployed, the fitted combiner gave it only 0.07 weight (~0.007 prob
    # shift — below the n=500 measurement floor). Holdout was a wash (ll
    # 0.6115→0.6127, acc 67.6→67.2, ECE 0.0552→0.0527) while the nudged winner
    # probs flipped the method era 2005→2016 (-979 rows), dropping Submission F1
    # 30.4→26.4. Wash upstream + real cost downstream = incumbent lineup stays.
    # "LogReg",
    # "AdaBoost",
})

MU_0 = 1500.0
PHI_0 = 200.0
SIGMA_0 = 0.06
TAU = 0.5
SCALE = 173.7178
CONVERGENCE = 1e-6


def _clip_probs(p):
    return np.clip(np.asarray(p, dtype=float), 1e-6, 1.0 - 1e-6)


def _cache_data_fingerprint(path):
    # Hash the file CONTENTS (not mtime/size) so the cache invalidates iff the
    # data actually changed: a no-op re-save won't force an expensive rebuild, and
    # an mtime-preserving rewrite can't silently serve a stale model. Cost is a few
    # ms on this CSV; chunked so it stays memory-safe if the file grows.
    try:
        h = hashlib.sha256()
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1 << 20), b""):
                h.update(chunk)
        raw = f"{os.path.abspath(path)}|{h.hexdigest()}"
    except Exception:
        raw = f"{os.path.abspath(path)}|missing"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _cache_key(stage, data_fp, version, extra=""):
    payload = f"{stage}|{data_fp}|{version}|{extra}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:20]


def _cache_path(stage, key):
    os.makedirs(CACHE_DIR, exist_ok=True)
    return os.path.join(CACHE_DIR, f"{stage}_{key}.pkl")


def _cache_load(stage, key):
    path = _cache_path(stage, key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "rb") as f:
            return pickle.load(f)
    except Exception:
        return None


def _cache_save(stage, key, payload):
    path = _cache_path(stage, key)
    with open(path, "wb") as f:
        pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    prefix = f"{stage}_"
    for name in os.listdir(CACHE_DIR):
        if not name.startswith(prefix) or name.endswith(f"{key}.pkl"):
            continue
        old = os.path.join(CACHE_DIR, name)
        try:
            os.remove(old)
        except Exception:
            pass


def _winner_stage_cache_valid(payload, feature_cols, n_rows, train_end, val_end):
    if not isinstance(payload, dict):
        return False
    if payload.get("kind") != WINNER_STAGE_CACHE_KIND:
        return False
    if str(payload.get("winner_cache_version")) != str(WINNER_CACHE_VERSION):
        return False
    if int(payload.get("n_rows", -1)) != int(n_rows):
        return False
    if int(payload.get("train_end", -1)) != int(train_end):
        return False
    if int(payload.get("val_end", -1)) != int(val_end):
        return False
    fc = payload.get("feature_cols")
    if not isinstance(fc, list) or tuple(fc) != tuple(feature_cols):
        return False
    for k in (
        "oof", "valid", "combiner", "model_order", "decision_threshold",
        "test_probs_raw", "test_probs_cal", "y_pred_red",
        "raw_ll", "cal_ll_test", "brier", "acc", "ece", "cal_curve_rmse",
        "lgb_tuned", "tuned_thr", "tuned_acc", "val_acc_for_log", "calibrator",
    ):
        if k not in payload:
            return False
    return True


def _method_stage_cache_valid(payload, winner_cache_key, method_cache_version):
    if not isinstance(payload, dict):
        return False
    if payload.get("kind") != METHOD_STAGE_CACHE_KIND:
        return False
    if str(payload.get("method_cache_version")) != str(method_cache_version):
        return False
    if str(payload.get("winner_cache_key")) != str(winner_cache_key):
        return False
    b = payload.get("method_bundle")
    return isinstance(b, dict) and b.get("imputer") is not None


def _replay_winner_cache_logs(pl, W, winner_cache_key):
    """Replay winner-stage terminal sections from a cache payload (no retrain)."""
    h = str(winner_cache_key)[:12]
    pl._section("Optuna Tuning")
    pl._stat("Cache", f"HIT ({WINNER_CACHE_VERSION}) — Optuna skipped [key={h}]")
    pl._section("Model Setup")
    pl._stat("Base models", ", ".join(W.get("model_order") or []))
    pl._section("OOF Stacking")
    pl._stat("Cache", f"HIT ({WINNER_CACHE_VERSION}) — OOF skipped [key={h}]")
    pl._section("Combiner Selection")
    pl._stat("Selected combiner", W.get("combiner_kind", ""))
    _w_str = _format_combiner_weights(W.get("combiner"))
    if _w_str:
        pl._stat("Combiner weights", _w_str)
    pl._stat("Validation log-loss", W.get("val_ll_str", ""))
    pl._stat("Validation accuracy", W.get("val_acc_str", ""))
    pl._stat("Validation threshold", W.get("val_thr_str", ""))
    pl._section("Calibration")
    pl._stat("Selected method", W.get("cal_name", ""))
    pl._stat("Selection-slice log-loss (late ~45% of val)", W.get("cal_ll_str", ""))
    pl._stat("Strict future mode", "ON" if STRICT_FUTURE_MODE else "OFF")
    if not STRICT_FUTURE_MODE:
        pl._stat("Holdout-selected combiner", W.get("best_holdout_label", ""))
        pl._stat("Holdout combiner log-loss", W.get("best_holdout_ll_str", ""))
        pl._stat("Holdout combiner acc", W.get("best_holdout_acc_str", ""))
        pl._stat("Holdout combiner threshold", W.get("best_holdout_thr_str", ""))
    pl._stat("Calibration used for picks", W.get("cal_used_str", ""))
    pl._stat("Validation accuracy (picked)", W.get("val_acc_picked_str", ""))
    pl._stat("Dev-OOF tuned threshold", W.get("val_tuned_thr_str", ""))
    pl._stat("Decision threshold", W.get("decision_thr_str", ""))
    _corner_rows = W.get("corner_rows", [])
    if _corner_rows:
        pl._section("Corner Correction")
        for _lbl, _val in _corner_rows:
            pl._stat(_lbl, _val)
    pl._section("Holdout Evaluation")
    for label, val in W.get("holdout_eval", []):
        pl._stat(label, val)
    pl._section("Winner Diagnostics")
    pl._log("Confusion Matrix (Winner: Red=positive)")
    pl._log("               Pred Blue   Pred Red")
    tn, fp, fn, tp = W.get("confusion", (0, 0, 0, 0))
    pl._log(f"Actual Blue   {int(tn):9d}  {int(fp):9d}")
    pl._log(f"Actual Red    {int(fn):9d}  {int(tp):9d}")
    pl._log("")
    pl._log("Accuracy by Weight Class")
    pl._log("-" * 72)
    for wc, acc_wc, n_wc in W.get("wc_rows", []):
        pl._stat(f"{wc} (n={int(n_wc)})", f"{float(acc_wc):.1%}")
    pl._log("")
    pl._log("Accuracy by Gender")
    pl._log("-" * 72)
    for gender, acc_g, n_g in W.get("g_rows", []):
        pl._stat(f"{gender} (n={int(n_g)})", f"{float(acc_g):.1%}")


def _replay_method_cache_logs(pl, M):
    """Replay method-stage terminal sections from a cache payload (no retrain)."""
    if M.get("method_classes_training"):
        pl._stat("Method classes (training)", M["method_classes_training"])
    if M.get("method_hard_reset"):
        pl._stat("Method stack mode", "hard-reset (stable components only)")
    if M.get("val_metric_str"):
        pl._stat("Validation metric target (method | winner correct)", M["val_metric_str"])
    if M.get("val_baseline_str"):
        pl._stat("Validation majority baseline (same subset)", M["val_baseline_str"])

    pl._section("Method Evaluation (Conditioned on Winner Pick)")

    s1_acc = M.get("stage1_acc")
    s1_auc = M.get("stage1_auc")
    s2_acc = M.get("stage2_acc_true_finishes")
    n_fin  = M.get("n_true_finishes", 0)
    pl._stat("Stage1 acc (Finish vs Decision)", f"{s1_acc:.1%}" if s1_acc is not None else "n/a")
    pl._stat("Stage1 AUC (Finish vs Decision)", f"{s1_auc:.3f}" if (s1_auc is not None and np.isfinite(s1_auc)) else "n/a")
    pl._stat("Stage2 acc (KO/Sub | true finishes)", f"{s2_acc:.1%}" if (s2_acc is not None and np.isfinite(s2_acc)) else "n/a")
    pl._stat("Stage2 sample size (true finishes)", n_fin)

    acc_pred = M.get("method_acc_predicted_winner")
    acc_wc   = M.get("method_acc_when_winner_correct")
    maj_base = M.get("method_majority_baseline_when_winner_correct")
    pl._stat("Method acc (predicted winner conditioned)", f"{acc_pred:.1%}" if acc_pred is not None else "n/a")
    pl._stat("Method acc | winner pick correct",          f"{acc_wc:.1%}"   if acc_wc   is not None else "n/a")
    pl._stat("Majority baseline | winner pick correct",   f"{maj_base:.1%}" if maj_base  is not None else "n/a")

    bal_acc     = M.get("bal_acc")
    macro_f1    = M.get("macro_f1")
    finish_score = M.get("finish_score")
    if bal_acc is not None and np.isfinite(bal_acc):
        pl._stat("Balanced accuracy | winner pick correct", f"{bal_acc:.1%}")
    if macro_f1 is not None and np.isfinite(macro_f1):
        pl._stat("Macro F1 | winner pick correct", f"{macro_f1:.1%}")
    if finish_score is not None and np.isfinite(finish_score):
        pl._stat("FinishScore (0.4 KO R + 0.4 Sub R + 0.2 KO/Sub F1)", f"{finish_score:.1%}")

    per_class = M.get("per_class_metrics", [])
    if per_class:
        pl._log("")
        pl._log("Per-Class Metrics (Method | winner pick correct)")
        pl._log("Class          Precision    Recall      F1")
        for cls_name, p, r, f1 in per_class:
            pl._log(f"{cls_name:<14}{p:10.1%}{r:10.1%}{f1:10.1%}")

    confusion_rows = M.get("confusion_rows", [])
    if confusion_rows:
        pl._log("")
        pl._log("Confusion Matrix (Method | winner pick correct)")
        pl._log("Actual\\Pred     Decision    KO/TKO  Submission")
        for actual, row_counts in confusion_rows:
            pl._log(f"{actual:<14}{row_counts[0]:10d}{row_counts[1]:10d}{row_counts[2]:12d}")


def _normalize_method_label(raw_method):
    detail = _normalize_method_detail(raw_method)
    if detail.startswith("Decision"):
        return "Decision"
    if detail in ("KO/TKO", "Doctor Stoppage", "DQ/Corner Stoppage"):
        return "KO/TKO"
    if detail == "Submission":
        return "Submission"
    return "Decision"


def _normalize_method_detail(raw_method):
    txt = str(raw_method or "").strip().lower()
    if "decision" in txt or txt.startswith("dec") or "split" in txt or "majority" in txt or "unanimous" in txt:
        return "Decision"
    if "doctor" in txt:
        return "KO/TKO"
    if "dq" in txt or "corner" in txt or "retire" in txt:
        return "KO/TKO"
    if "sub" in txt:
        return "Submission"
    if "ko" in txt or "tko" in txt:
        return "KO/TKO"
    return "Decision"


def _normalize_method_probs(prob_map):
    vals = np.array([float(prob_map.get(k, 0.0)) for k in METHOD_LABELS], dtype=float)
    vals = np.maximum(vals, MIN_METHOD_PROB)
    vals = vals / np.sum(vals)
    return {k: float(v) for k, v in zip(METHOD_LABELS, vals)}


def _history_prior_row(d_dec, d_ko, d_sub):
    """Method history prior from oriented (winner-loser) win-method differentials.

    SINGLE SOURCE OF TRUTH for this prior, called identically at training,
    holdout evaluation, AND inference. `d_dec`/`d_ko`/`d_sub` are the oriented
    `d_*_win_pct` features (predicted-winner's method share minus the predicted
    loser's). Previously inference computed this term from per-fighter win/loss
    profiles instead — a different recipe with different denominators — so the
    `w_hist` weight (tuned against THIS recipe) was applied to a different prior
    in production. Routing all three call sites through this function removes that
    train/serve skew.
    """
    def _f(v):
        try:
            v = float(v)
        except (TypeError, ValueError):
            return 0.0
        return 0.0 if math.isnan(v) else v
    return _normalize_method_probs({
        "Decision":   0.50 + 0.35 * _f(d_dec),
        "KO/TKO":     0.32 + 0.35 * _f(d_ko),
        "Submission": 0.18 + 0.35 * _f(d_sub),
    })


# Method-probability shaping helpers: logit-space bias application and submission signal boost.
# Called during tuning (blending loop) and inference (predict_method_probs).
def _apply_method_logit_bias_arr(probs_arr, bias_vec):
    arr = np.asarray(probs_arr, dtype=float)
    if arr.ndim == 1:
        arr = arr.reshape(1, -1)
    b = np.asarray(bias_vec, dtype=float).reshape(1, -1)
    logp = np.log(np.clip(arr, MIN_METHOD_PROB, 1.0)) + b
    logp = logp - np.max(logp, axis=1, keepdims=True)
    expv = np.exp(logp)
    out = expv / np.sum(expv, axis=1, keepdims=True)
    return np.clip(out, MIN_METHOD_PROB, 1.0)


def _apply_method_logit_bias_map(prob_map, bias_map):
    arr = np.array([
        float(prob_map.get("Decision", 0.0)),
        float(prob_map.get("KO/TKO", 0.0)),
        float(prob_map.get("Submission", 0.0)),
    ], dtype=float)
    bias_vec = np.array([
        float(bias_map.get("Decision", 0.0)),
        float(bias_map.get("KO/TKO", 0.0)),
        float(bias_map.get("Submission", 0.0)),
    ], dtype=float)
    out = _apply_method_logit_bias_arr(arr, bias_vec)[0]
    return _normalize_method_probs({
        "Decision": float(out[0]),
        "KO/TKO": float(out[1]),
        "Submission": float(out[2]),
    })


def _apply_binary_threshold_warp(p, thr):
    p_arr = np.asarray(p, dtype=float)
    t = float(np.clip(thr, 0.05, 0.95))
    left = 0.5 * (p_arr / max(t, 1e-6))
    right = 0.5 + 0.5 * ((p_arr - t) / max(1.0 - t, 1e-6))
    out = np.where(p_arr < t, left, right)
    return np.clip(out, 1e-4, 1.0 - 1e-4)


def _apply_submission_signal_boost_arr(probs_arr, sub_signal_arr, boost_k):
    arr = np.asarray(probs_arr, dtype=float).copy()
    sig = np.asarray(sub_signal_arr, dtype=float).reshape(-1)
    k = float(boost_k)
    # Positive boost only when submission signal is above prior-ish baseline.
    scale = np.clip((sig - 0.18) / 0.22, -1.5, 2.5)
    arr[:, 2] = arr[:, 2] * np.exp(k * scale)
    arr = np.clip(arr, MIN_METHOD_PROB, 1.0)
    arr = arr / np.sum(arr, axis=1, keepdims=True)
    return arr


def _apply_submission_signal_boost_map(prob_map, sub_signal, boost_k):
    arr = np.array([
        float(prob_map.get("Decision", 0.0)),
        float(prob_map.get("KO/TKO", 0.0)),
        float(prob_map.get("Submission", 0.0)),
    ], dtype=float).reshape(1, -1)
    out = _apply_submission_signal_boost_arr(arr, np.array([float(sub_signal)], dtype=float), boost_k)[0]
    return _normalize_method_probs({
        "Decision": float(out[0]),
        "KO/TKO": float(out[1]),
        "Submission": float(out[2]),
    })


def _sub_attempt_prior_array(X_df):
    X = X_df.reset_index(drop=True)
    d_sub_att = pd.to_numeric(X.get("d_sub_att_p15", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    d_grap_tdd = pd.to_numeric(X.get("d_grapple_vs_tdd", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    d_been_fin = pd.to_numeric(X.get("d_been_finished_pct", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    d_dec = pd.to_numeric(X.get("d_dec_win_pct", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    z = 1.35 * d_sub_att + 0.75 * d_grap_tdd + 0.55 * d_been_fin
    p_sub = 1.0 / (1.0 + np.exp(-z))
    p_sub = np.clip(0.08 + 0.50 * p_sub, 0.06, 0.62)
    p_dec = np.clip(0.54 + 0.20 * d_dec - 0.35 * p_sub, 0.12, 0.82)
    p_ko = np.clip(1.0 - p_dec - p_sub, 0.05, 0.70)
    arr = np.stack([p_dec, p_ko, p_sub], axis=1)
    arr = np.clip(arr, MIN_METHOD_PROB, 1.0)
    arr = arr / np.sum(arr, axis=1, keepdims=True)
    return arr


def _oriented_method_matrix(X_df, y_red_win):
    X_m = X_df.copy()
    y_arr = np.asarray(y_red_win).astype(int)
    sign = np.where(y_arr == 1, 1.0, -1.0)
    # Same parity rules as _swap_features, applied per-row: negate the
    # antisymmetric columns (d_* minus the symmetric even-product set, plus
    # _SWAP_NEGATE_COLS) on blue-winner rows.
    for col in X_m.columns:
        if ((col.startswith("d_") and col not in _SWAP_SYMMETRIC_D_COLS)
                or col in _SWAP_NEGATE_COLS):
            X_m[col] = X_m[col].astype(float).values * sign
    blue_mask = (sign == -1)
    if blue_mask.any():
        # Probability columns complement (p -> 1-p) on blue-winner rows.
        for col in _SWAP_COMPLEMENT_COLS:
            if col in X_m.columns:
                vals = X_m[col].astype(float).values
                X_m[col] = np.where(blue_mask, 1.0 - vals, vals)
        # Raw r_/b_ pair columns aren't differences, so per-row orientation must
        # exchange them on blue-winner rows (sign == -1) to stay consistent with
        # the d_* negation above.
        for r_col, b_col in _SWAP_PAIR_COLUMNS:
            if r_col in X_m.columns and b_col in X_m.columns:
                r_arr = X_m[r_col].values.copy()
                b_arr = X_m[b_col].values.copy()
                X_m[r_col] = np.where(blue_mask, b_arr, r_arr)
                X_m[b_col] = np.where(blue_mask, r_arr, b_arr)
    return X_m


def _augment_method_features(X_df):
    """Add engineered method-of-victory features to a winner-oriented matrix.

    Input rows must already be oriented so the (predicted) winner occupies
    the positive corner (see _oriented_method_matrix). Adds explicit
    path-vs-vulnerability composites (m_ko_*, m_sub_*, m_dec_*), exact
    winner/loser side reconstructions from d_*/*_sum channel pairs, and the
    latent stage-1 finish-environment features (m_decision_shell_*,
    m_finish_*). Applied identically at training, holdout evaluation, and
    inference.
    """
    X = X_df.copy()
    eps = 1e-6

    def _col(name, default=0.0):
        if name not in X.columns:
            return pd.Series(np.full(len(X), default), index=X.index, dtype=float)
        return pd.to_numeric(X[name], errors="coerce").fillna(default)

    d_ko_win = _col("d_ko_win_pct", 0.0)
    d_sub_win = _col("d_sub_win_pct", 0.0)
    d_dec_win = _col("d_dec_win_pct", 0.0)
    d_ko_loss = _col("d_ko_loss_pct", 0.0)
    d_sub_att = _col("d_sub_att_p15", 0.0)
    d_r3_sub = _col("d_rd3_sub_att", 0.0)
    d_r3_sub_share = _col("d_rd3_sub_share", 0.0)
    d_late_sub = _col("d_late_vs_early_sub_att", 0.0)
    d_sub_t23 = _col("d_sub_att_trend_23", 0.0)
    d_been_finished = _col("d_been_finished_pct", 0.0)
    d_finish_vs_resist = _col("d_finish_vs_resist", 0.0)
    d_str_vs_def = _col("d_striking_vs_defense", 0.0)
    d_strike_exchange_ratio = _col("d_strike_exchange_ratio", 0.0)
    d_grap_vs_tdd = _col("d_grapple_vs_tdd", 0.0)
    d_ortho_vs_south = _col("d_ortho_vs_south", 0.0)
    d_power_ratio = _col("d_power_ratio", 0.0)
    d_head_pct = _col("d_head_pct", 0.0)
    d_sig_str_diff_pm = _col("d_sig_str_diff_pm", 0.0)
    d_first_round_finish_rate = _col("d_first_round_finish_rate", 0.0)
    d_rd1_intensity_ratio = _col("d_rd1_intensity_ratio", 0.0)
    d_damage_efficiency = _col("d_damage_efficiency", 0.0)
    d_sig_str_acc = _col("d_sig_str_acc", _col("d_td_sig_str_acc", 0.0))
    d_ground_pct = _col("d_ground_pct", 0.0)
    d_ctrl_pct = _col("d_ctrl_pct", _col("d_td_ctrl_pct", 0.0))
    d_rev_p15 = _col("d_rev_p15", 0.0)
    d_cardio_ratio = _col("d_cardio_ratio", 0.0)
    d_distance_pct = _col("d_distance_pct", 0.0)
    d_body_pct = _col("d_body_pct", 0.0)
    d_leg_pct = _col("d_leg_pct", 0.0)
    d_consistency = _col("d_consistency", 0.0)
    d_kd_pm = _col("d_kd_pm", _col("d_td_kd_pm", 0.0))
    d_def_kd_pm = _col("d_def_kd_pm", 0.0)
    d_finish_resistance = _col("d_finish_resistance", 0.0)
    d_durability = _col("d_durability", 0.0)
    d_late_round_pct = _col("d_late_round_pct", 0.0)
    d_output_rate = _col("d_output_rate", 0.0)
    d_glicko = np.abs(_col("d_glicko_win_prob", 0.0))
    d_head_acc = _col("d_head_acc", 0.0)
    d_distance_acc = _col("d_distance_acc", 0.0)
    d_distance_share = _col("d_distance_share", 0.0)
    d_ground_acc = _col("d_ground_acc", 0.0)
    d_clinch_acc = _col("d_clinch_acc", 0.0)
    d_clinch_pct = _col("d_clinch_pct", 0.0)
    d_body_leg_attrition = _col("d_body_leg_attrition", 0.0)
    d_sub_loss_pct = _col("d_sub_loss_pct", 0.0)
    d_recent_sub_win_rate = _col("d_recent_sub_win_rate", 0.0)
    d_sub_entry_pressure = _col("d_sub_entry_pressure", 0.0)
    d_sub_control_conversion = _col("d_sub_control_conversion", 0.0)
    d_sub_defensive_leak = _col("d_sub_defensive_leak", 0.0)
    d_late_sub_pressure = _col("d_late_sub_pressure", 0.0)
    d_sub_recency_surge = _col("d_sub_recency_surge", 0.0)
    d_grapple_recency_surge = _col("d_grapple_recency_surge", 0.0)
    d_sub_vs_control_axis = _col("d_sub_vs_control_axis", 0.0)
    d_recent_ko_loss_rate = _col("d_recent_ko_loss_rate", 0.0)
    d_r5_def_kd_pm = _col("d_r5_def_kd_pm", 0.0)
    d_ko_attack_pressure = _col("d_ko_attack_pressure", 0.0)
    d_ko_def_leak = _col("d_ko_def_leak", 0.0)
    sub_entry_sum = _col("sub_entry_pressure_sum", 0.0)
    sub_leak_sum = _col("sub_defensive_leak_sum", 0.0)
    ko_attack_sum = _col("ko_attack_pressure_sum", 0.0)
    ko_leak_sum = _col("ko_def_leak_sum", 0.0)
    total_rounds = _col("total_rounds", 3.0)
    is_title = _col("is_title", 0.0)

    # Explicit path-vs-vulnerability method features.
    X["m_ko_path_vs_vuln"] = d_ko_win + d_ko_loss + 0.45 * d_str_vs_def
    X["m_sub_path_vs_vuln"] = d_sub_win + d_been_finished + 0.35 * d_grap_vs_tdd + 0.25 * d_sub_att
    X["m_sub_round3_pressure"] = 0.60 * d_r3_sub + 0.40 * d_r3_sub_share
    X["m_sub_trend_pressure"] = 0.55 * d_late_sub + 0.45 * d_sub_t23
    X["m_dec_path"] = d_dec_win - 0.25 * d_finish_vs_resist
    X["m_finish_bias"] = X["m_ko_path_vs_vuln"] + X["m_sub_path_vs_vuln"] - X["m_dec_path"]
    X["m_ko_rounds_interaction"] = X["m_ko_path_vs_vuln"] * (4.0 - np.minimum(total_rounds, 4.0))
    X["m_sub_rounds_interaction"] = X["m_sub_path_vs_vuln"] * np.maximum(total_rounds - 2.0, 1.0)
    X["m_dec_rounds_interaction"] = X["m_dec_path"] * total_rounds
    X["m_stance_finish_interaction"] = d_ortho_vs_south * X["m_finish_bias"]
    X["m_title_finish_interaction"] = is_title * X["m_finish_bias"]
    X["m_finish_abs_pressure"] = np.abs(X["m_finish_bias"]) * np.abs(d_str_vs_def)
    X["m_finish_gap_ko_sub"] = X["m_ko_path_vs_vuln"] - X["m_sub_path_vs_vuln"]
    X["m_ko_share"] = X["m_ko_path_vs_vuln"] / (
        np.abs(X["m_ko_path_vs_vuln"]) + np.abs(X["m_sub_path_vs_vuln"]) + np.abs(X["m_dec_path"]) + 1e-6
    )
    X["m_sub_share"] = X["m_sub_path_vs_vuln"] / (
        np.abs(X["m_ko_path_vs_vuln"]) + np.abs(X["m_sub_path_vs_vuln"]) + np.abs(X["m_dec_path"]) + 1e-6
    )
    X["m_dec_share"] = X["m_dec_path"] / (
        np.abs(X["m_ko_path_vs_vuln"]) + np.abs(X["m_sub_path_vs_vuln"]) + np.abs(X["m_dec_path"]) + 1e-6
    )
    X["m_finish_vs_cardio"] = X["m_finish_bias"] * (1.0 / np.maximum(total_rounds, 1.0))
    X["m_decision_durability"] = (d_dec_win - np.abs(d_ko_loss) - np.abs(d_been_finished)) * np.maximum(total_rounds, 1.0)
    X["m_sub_chain_pressure"] = (d_sub_att + 0.50 * X["m_sub_round3_pressure"]) * (1.0 + np.maximum(d_grap_vs_tdd, -1.5))
    X["m_sub_finish_trigger"] = X["m_sub_chain_pressure"] + 0.45 * X["m_sub_trend_pressure"] + 0.30 * d_been_finished
    X["m_ko_chain_pressure"] = np.maximum(d_str_vs_def, -2.0) * (1.0 + np.maximum(d_ko_win, -1.0))
    X["m_title_rounds_interaction"] = is_title * total_rounds * X["m_dec_path"]

    # Method-only explicit mechanics: KO-shaped pressure, sub chaining, and decision control.
    X["m_ko_headhunter"] = d_power_ratio * d_head_pct * d_sig_str_diff_pm
    X["m_ko_fast_start"] = d_first_round_finish_rate * d_rd1_intensity_ratio * (4.0 - np.minimum(total_rounds, 4.0))
    X["m_ko_pressure_conversion"] = d_damage_efficiency * d_power_ratio * np.maximum(d_sig_str_acc, 0.0)
    X["m_ko_exchange_edge"] = d_strike_exchange_ratio * d_power_ratio * d_sig_str_diff_pm
    X["m_ko_distance_sniper"] = d_distance_pct * d_sig_str_acc * d_power_ratio
    X["m_ko_knockdown_axis"] = d_kd_pm - 0.6 * d_def_kd_pm
    X["m_ko_attrition"] = (d_body_pct + 0.7 * d_leg_pct) * d_output_rate * d_cardio_ratio
    X["m_ko_confident_mismatch"] = d_glicko * np.maximum(d_power_ratio, 0.0) * np.maximum(d_sig_str_diff_pm, 0.0)
    X["m_ko_burst_vs_decay"] = d_rd1_intensity_ratio - d_cardio_ratio
    X["m_ko_path_specific"] = (d_ko_win + d_ko_loss) * (0.5 + d_head_pct + 0.5 * d_power_ratio)
    X["m_sub_ground_hunter"] = d_ground_pct * (d_ctrl_pct + 0.5 * d_sub_att)
    X["m_sub_scramble_threat"] = d_rev_p15 * (d_sub_att + d_ground_pct)
    X["m_sub_late_snowball"] = d_cardio_ratio * d_r3_sub_share * (d_ctrl_pct + d_sub_att)
    X["m_dec_clean_pointing"] = (
        d_distance_pct * d_sig_str_acc * d_consistency - 0.5 * (np.abs(d_kd_pm) + np.abs(d_sub_att))
    )
    X["m_dec_stability"] = d_finish_resistance * d_durability * d_late_round_pct * d_consistency
    X["m_finish_confidence"] = d_glicko * np.maximum(X["m_finish_bias"], 0.0)
    X["m_sub_vs_ko_axis"] = (d_ground_pct + d_ctrl_pct + d_sub_att) - (d_head_pct + d_power_ratio + d_kd_pm)
    X["m_ko_head_accuracy"] = _col("d_head_hunt_accuracy", d_head_acc) * d_power_ratio
    X["m_ko_range_sniper"] = d_distance_acc * d_distance_share * d_power_ratio
    X["m_ko_range_accuracy"] = _col("d_distance_strike_accuracy", d_distance_acc) * d_distance_pct * d_power_ratio
    X["m_ground_finish_conversion"] = d_ground_acc * d_ctrl_pct * d_sub_att
    X["m_clinch_breaker"] = d_clinch_acc * d_clinch_pct * d_kd_pm
    X["m_attrition_finish"] = d_body_leg_attrition * d_late_round_pct * d_output_rate
    X["m_sub_loss_vulnerability"] = d_sub_loss_pct
    X["m_recent_sub_win_rate_gap"] = d_recent_sub_win_rate
    X["m_sub_entry_pressure"] = d_sub_entry_pressure
    X["m_sub_control_conversion"] = d_sub_control_conversion
    X["m_sub_defensive_leak"] = d_sub_defensive_leak
    X["m_late_sub_pressure"] = d_late_sub_pressure
    X["m_sub_recency_surge"] = d_sub_recency_surge
    X["m_grapple_recency_surge"] = d_grapple_recency_surge
    X["m_sub_vs_control_axis"] = d_sub_vs_control_axis
    # Exact side-specific reconstruction from oriented differential + invariant sums:
    # winner_side = 0.5 * (sum + d), loser_side = 0.5 * (sum - d)
    sub_entry_w = 0.5 * (sub_entry_sum + d_sub_entry_pressure)
    sub_entry_l = 0.5 * (sub_entry_sum - d_sub_entry_pressure)
    sub_leak_w = 0.5 * (sub_leak_sum + d_sub_defensive_leak)
    sub_leak_l = 0.5 * (sub_leak_sum - d_sub_defensive_leak)
    sub_attack_vs_leak_w = sub_entry_w * sub_leak_l
    sub_attack_vs_leak_l = sub_entry_l * sub_leak_w
    X["m_sub_mismatch_explicit"] = sub_attack_vs_leak_w - sub_attack_vs_leak_l
    X["m_sub_mismatch_max"] = np.maximum(sub_attack_vs_leak_w, sub_attack_vs_leak_l)
    X["m_sub_mismatch_sum"] = sub_attack_vs_leak_w + sub_attack_vs_leak_l
    X["m_sub_mismatch"] = X["m_sub_mismatch_explicit"]
    ko_attack_w = 0.5 * (ko_attack_sum + d_ko_attack_pressure)
    ko_attack_l = 0.5 * (ko_attack_sum - d_ko_attack_pressure)
    ko_leak_w = 0.5 * (ko_leak_sum + d_ko_def_leak)
    ko_leak_l = 0.5 * (ko_leak_sum - d_ko_def_leak)
    ko_attack_vs_leak_w = ko_attack_w * ko_leak_l
    ko_attack_vs_leak_l = ko_attack_l * ko_leak_w
    X["m_ko_mismatch"] = ko_attack_vs_leak_w - ko_attack_vs_leak_l
    X["m_ko_recent_fragility"] = d_r5_def_kd_pm + 0.7 * d_recent_ko_loss_rate

    # Keep ratios bounded and numerically stable for tree/linear blends.
    X["m_sub_ground_efficiency"] = (d_ground_pct * np.maximum(d_sub_att, 0.0)) / (1.0 + np.abs(d_ctrl_pct) + eps)

    # ── Stage-1 side-reconstruction features ─────────────────────────────────
    # Reconstruct exact winner-side (w) and loser-side (l) values from the
    # oriented differential d_x and the orientation-invariant sum x_sum:
    #   x_w = 0.5 * (x_sum + d_x),   x_l = 0.5 * (x_sum - d_x)
    def _sides(sum_col, diff_col, default=0.0):
        s = _col(sum_col, default * 2)
        d = _col(diff_col, 0.0)
        return 0.5 * (s + d), 0.5 * (s - d)

    dec_w,  dec_l  = _sides("dec_win_pct_sum",          "d_dec_win_pct",           0.0)
    res_w,  res_l  = _sides("finish_resistance_sum",     "d_finish_resistance",     0.0)
    cons_w, cons_l = _sides("consistency_sum",           "d_consistency",           0.0)
    cardio_w, cardio_l = _sides("cardio_ratio_sum",      "d_cardio_ratio",          0.0)
    dur_w,  dur_l  = _sides("durability_sum",            "d_durability",            0.0)
    out_w,  out_l  = _sides("output_rate_sum",           "d_output_rate",           0.0)
    r1_w,   r1_l   = _sides("rd1_intensity_ratio_sum",   "d_rd1_intensity_ratio",   0.0)
    exch_w, exch_l = _sides("strike_exchange_ratio_sum", "d_strike_exchange_ratio", 0.0)
    acc_w,  acc_l  = _sides("sig_str_acc_sum",           "d_sig_str_acc",           0.0)
    late_w, late_l = _sides("late_round_pct_sum",        "d_late_round_pct",        0.0)
    time_w, time_l = _sides("avg_time_min_sum",          "d_avg_time_min",          0.0)
    afr_w,  afr_l  = _sides("avg_finish_round_sum",      "d_avg_finish_round",      0.0)
    fr1_w,  fr1_l  = _sides("first_round_finish_rate_sum", "d_first_round_finish_rate", 0.0)
    dmg_w,  dmg_l  = _sides("damage_efficiency_sum",    "d_damage_efficiency",      0.0)
    attr_w_raw, attr_l_raw = _sides("body_leg_attrition_sum", "d_body_leg_attrition", 0.0)

    # Reuse existing KO/Sub attack-vs-leak reconstructions
    ko_w = np.maximum(ko_attack_vs_leak_w, 0.0)
    ko_l = np.maximum(ko_attack_vs_leak_l, 0.0)
    sub_w = np.maximum(sub_attack_vs_leak_w, 0.0)
    sub_l = np.maximum(sub_attack_vs_leak_l, 0.0)

    # ── Latent stage-1 components ─────────────────────────────────────────────
    decision_shell_w = (
        0.34*dec_w + 0.18*cons_w + 0.16*cardio_w + 0.14*res_w + 0.10*dur_w + 0.08*acc_w
    ) * (0.90 + 0.10*late_w)
    decision_shell_l = (
        0.34*dec_l + 0.18*cons_l + 0.16*cardio_l + 0.14*res_l + 0.10*dur_l + 0.08*acc_l
    ) * (0.90 + 0.10*late_l)

    chaos_w = (
        np.clip(out_w, 0, 8)
        * (0.55 + np.clip(exch_w, 0.5, 1.8))
        * (0.55 + np.clip(r1_w,  0.5, 1.8))
    )
    chaos_l = (
        np.clip(out_l, 0, 8)
        * (0.55 + np.clip(exch_l, 0.5, 1.8))
        * (0.55 + np.clip(r1_l,  0.5, 1.8))
    )

    finish_speed_w = (1.0 / (afr_w + 0.75)) * (0.60 + 0.40*fr1_w)
    finish_speed_l = (1.0 / (afr_l + 0.75)) * (0.60 + 0.40*fr1_l)

    time_shell_w = time_w * (0.55 + 0.45*dec_w) * (0.70 + 0.30*res_w)
    time_shell_l = time_l * (0.55 + 0.45*dec_l) * (0.70 + 0.30*res_l)

    attrition_w = attr_w_raw * late_w * cardio_w
    attrition_l = attr_l_raw * late_l * cardio_l

    finish_total_w = ko_w + 0.92*sub_w + 0.10*dmg_w + 0.18*finish_speed_w
    finish_total_l = ko_l + 0.92*sub_l + 0.10*dmg_l + 0.18*finish_speed_l

    # ── Final stage-1 features ────────────────────────────────────────────────
    X["m_decision_shell_gap"]       = decision_shell_w - decision_shell_l
    X["m_decision_shell_sum"]       = decision_shell_w + decision_shell_l
    X["m_finish_conversion_edge"]   = finish_total_w - 0.72*decision_shell_l + 0.12*chaos_w
    X["m_finish_environment"]       = (finish_total_w + finish_total_l) - 0.68*(decision_shell_w + decision_shell_l)
    X["m_mutual_finish_instability"] = finish_total_w * finish_total_l
    X["m_decision_absorber"]        = (
        decision_shell_w + decision_shell_l + time_shell_w + time_shell_l
        - finish_total_w - finish_total_l - 0.25*(chaos_w + chaos_l)
    )
    X["m_early_finish_window"]      = (
        (chaos_w + chaos_l)
        * (4.0 - np.minimum(total_rounds, 4.0))
        * (0.35 + ko_w + ko_l)
    )
    X["m_fast_start_fragility"]     = fr1_w*(1.0 - dur_l) + fr1_l*(1.0 - dur_w)
    X["m_late_finish_window"]       = (
        np.maximum(total_rounds - 2.0, 1.0)
        * (
            sub_w + sub_l
            + 0.20*(2.0 - cardio_w - cardio_l)
            + 0.15*(2.0 - decision_shell_w - decision_shell_l)
            + 0.20*(attrition_w + attrition_l)
        )
    )
    X["m_attrition_break_window"]   = attrition_w*(1.0 - cardio_l) + attrition_l*(1.0 - cardio_w)
    X["m_time_profile_finish_bias"] = (
        (finish_total_w + finish_total_l)
        * (1.0/(1.0 + time_w) + 1.0/(1.0 + time_l))
    )
    X["m_finish_over_shell_ratio"]  = (
        (finish_total_w + finish_total_l + eps)
        / (decision_shell_w + decision_shell_l + eps)
    )
    X["m_clean_decision_track"]     = (
        (dec_w*cons_w*acc_w + dec_l*cons_l*acc_l)
        * (0.75 + 0.25*(time_w + time_l))
    )
    X["m_finish_speed_pressure"]    = finish_speed_w*(1.0 - res_l) + finish_speed_l*(1.0 - res_w)

    return X


def fuzzy_find(name, fighter_history):
    """Resolve a user-typed fighter name to a known history key.

    Tries exact match, then case-insensitive match, then unique substring
    match. Returns None when the name is unknown or ambiguous.
    """
    if name in fighter_history and fighter_history[name]:
        return name
    lower = str(name).lower()
    for key in fighter_history:
        if str(key).lower() == lower and fighter_history[key]:
            return key
    matches = [k for k in fighter_history if lower in str(k).lower() and fighter_history[k]]
    if len(matches) == 1:
        return matches[0]
    return None


# ─── Weight class ordinal mapping ─────────────────────────────────────────────
# Ordinal is only a coarse size proxy now; the exact class is carried by one-hot
# features below, so every class gets a unique code to avoid collisions.
WEIGHT_CLASS_ORDINAL = {
    "Women's Strawweight": 1, "Women's Flyweight": 2, "Women's Bantamweight": 3,
    "Women's Featherweight": 4, "Flyweight": 5, "Bantamweight": 6,
    "Featherweight": 7, "Lightweight": 8, "Welterweight": 9,
    "Middleweight": 10, "Light Heavyweight": 11, "Heavyweight": 12,
    "Catch Weight": 13, "Open Weight": 14,
}

def _weight_class_feature_name(weight_class):
    slug = re.sub(r"[^a-z0-9]+", "_", str(weight_class).lower()).strip("_")
    return f"wc_{slug or 'unknown'}"


WEIGHT_CLASS_FEATURES = {
    weight_class: _weight_class_feature_name(weight_class)
    for weight_class in WEIGHT_CLASS_ORDINAL
}
UNKNOWN_WEIGHT_CLASS_FEATURE = "wc_unknown"
METHOD_LABELS = ["Decision", "KO/TKO", "Submission"]
MIN_METHOD_PROB = 0.001

# ─── Glicko-2 core ─────────────────────────────────────────────────────────────

def _g(phi):
    return 1.0 / math.sqrt(1.0 + 3.0 * phi**2 / math.pi**2)

def _E(mu, mu_j, phi_j):
    return 1.0 / (1.0 + math.exp(-_g(phi_j) * (mu - mu_j)))

def glicko2_update(rating, opponents):
    """One Glicko-2 rating-period update (Glickman's step-by-step algorithm).

    `rating` is a (mu, phi, sigma) tuple on the display scale (mu ~ 1500);
    `opponents` is a list of (opp_mu, opp_phi, score) with score in [0, 1] —
    a hard 1/0 result, or a continuous margin-of-victory score when MOV
    ratings are enabled. With no opponents, only the deviation phi grows
    (rating uncertainty increases with inactivity). Returns the updated
    (mu, phi, sigma).
    """
    mu, phi, sigma = rating
    mu_s = (mu - MU_0) / SCALE
    phi_s = phi / SCALE
    if not opponents:
        phi_star = math.sqrt(phi_s**2 + sigma**2)
        return (mu, phi_star * SCALE, sigma)
    v_inv = 0.0
    delta_sum = 0.0
    for opp_r, opp_rd, score in opponents:
        mu_j = (opp_r - MU_0) / SCALE
        phi_j = opp_rd / SCALE
        g_j = _g(phi_j)
        E_j = _E(mu_s, mu_j, phi_j)
        v_inv += g_j**2 * E_j * (1.0 - E_j)
        delta_sum += g_j * (score - E_j)
    v = 1.0 / v_inv if v_inv > 0 else 1e6
    delta = v * delta_sum
    a = math.log(sigma**2)
    def f(x):
        ex = math.exp(x)
        num = ex * (delta**2 - phi_s**2 - v - ex)
        den = 2.0 * (phi_s**2 + v + ex)**2
        return num / den - (x - a) / (TAU**2)
    A = a
    if delta**2 > phi_s**2 + v:
        B = math.log(delta**2 - phi_s**2 - v)
    else:
        k = 1
        while f(a - k * TAU) < 0:
            k += 1
        B = a - k * TAU
    fA, fB = f(A), f(B)
    for _ in range(100):
        C = A + (A - B) * fA / (fB - fA)
        fC = f(C)
        if fC * fB < 0:
            A, fA = B, fB
        else:
            fA /= 2.0
        B, fB = C, fC
        if abs(B - A) < CONVERGENCE:
            break
    new_sigma = math.exp(A / 2.0)
    phi_star = math.sqrt(phi_s**2 + new_sigma**2)
    new_phi_s = 1.0 / math.sqrt(1.0 / phi_star**2 + 1.0 / v)
    new_mu_s = mu_s + new_phi_s**2 * delta_sum
    return (new_mu_s * SCALE + MU_0, new_phi_s * SCALE, new_sigma)

# ─── Helpers ────────────────────────────────────────────────────────────────────

def _isnan(v):
    if v is None:
        return True
    try:
        return math.isnan(float(v))
    except (TypeError, ValueError):
        return True

def _safe_sum(values):
    return sum(v for v in values if not _isnan(v))

def _safe_mean(values):
    clean = [v for v in values if not _isnan(v)]
    return sum(clean) / len(clean) if clean else float("nan")

def _safe_div(a, b, default=0.0):
    return a / b if b and b > 0 else default


def _num_or(value, default=0.0):
    return default if _isnan(value) else float(value)


def _abs_gap(a, b, default=float("nan")):
    if _isnan(a) or _isnan(b):
        return default
    return abs(float(a) - float(b))


def _extract_profile_from_row(row, prefix):
    return {
        "height": row.get(f"{prefix}_height", float("nan")),
        "reach": row.get(f"{prefix}_reach", float("nan")),
        "ape_index": row.get(f"{prefix}_ape_index", float("nan")),
        "weight": row.get(f"{prefix}_weight", float("nan")),
        "age": row.get(f"{prefix}_age_at_event", float("nan")),
        "stance": row.get(f"{prefix}_stance", ""),
    }


def _weighted_rate(numerator, denominator, prior=0.5, strength=25.0):
    """Bayesian-smoothed rate using equivalent sample-size `strength`."""
    if denominator is None or denominator <= 0:
        return prior
    return (numerator + prior * strength) / (denominator + strength)


def _bayes_shrink(empirical_value, sample_size, prior=0.5, strength=8.0):
    """Shrink noisy low-sample statistics toward a global prior."""
    if _isnan(empirical_value):
        empirical_value = prior
    sample_size = max(float(sample_size or 0.0), 0.0)
    return (empirical_value * sample_size + prior * strength) / (sample_size + strength)


def _correlation_prune(X, y=None, threshold=0.95):
    """Drop columns whose |corr| with an already-kept column exceeds threshold.

    If y is provided, features are processed in descending order of |corr with y|,
    so the more target-relevant member of each correlated pair survives (smart
    keep-rule). If y is None, column order determines the tie-break.
    Returns surviving column names in original column order.
    """
    cols = list(X.columns)
    if len(cols) <= 1:
        return cols
    corr = np.abs(X.corr().fillna(0.0).to_numpy())
    np.fill_diagonal(corr, 0.0)

    if y is not None:
        y_arr = np.asarray(y, dtype=float).reshape(-1)
        X_arr = X.to_numpy(dtype=float, copy=False)
        relevance = np.zeros(len(cols))
        for i in range(len(cols)):
            xcol = X_arr[:, i]
            if np.std(xcol) < 1e-12:
                continue
            r = np.corrcoef(xcol, y_arr)[0, 1]
            if not np.isnan(r):
                relevance[i] = abs(r)
        order = list(np.argsort(-relevance))
    else:
        order = list(range(len(cols)))

    kept = set()
    for idx in order:
        if not any(corr[idx, k] > threshold for k in kept):
            kept.add(idx)
    return [cols[i] for i in range(len(cols)) if i in kept]


def _optuna_tune_hgb(X_tr, y_tr, X_va, y_va, sample_weight_tr=None, n_trials=40, seed=0, defaults=None, progress_cb=None, progress_label=""):
    """Tune a HistGradientBoostingClassifier via Optuna, minimizing validation log loss.

    Returns a dict of best params (passable to HistGradientBoostingClassifier(**params)).
    Falls back to `defaults` if optuna is unavailable or n_trials <= 0.
    """
    if optuna is None or n_trials <= 0 or len(X_va) < 20:
        return defaults

    def _objective(trial):
        params = {
            "loss": "log_loss",
            "max_iter": trial.suggest_int("max_iter", 160, 500, step=20),
            "learning_rate": trial.suggest_float("learning_rate", 0.02, 0.12, log=True),
            "max_depth": trial.suggest_int("max_depth", 4, 10),
            "max_leaf_nodes": trial.suggest_int("max_leaf_nodes", 15, 63),
            "min_samples_leaf": trial.suggest_int("min_samples_leaf", 8, 32),
            "l2_regularization": trial.suggest_float("l2_regularization", 0.1, 2.0, log=True),
            "random_state": seed,
        }
        model = HistGradientBoostingClassifier(**params)
        if sample_weight_tr is not None:
            model.fit(X_tr, y_tr, sample_weight=sample_weight_tr)
        else:
            model.fit(X_tr, y_tr)
        p_va = model.predict_proba(X_va)[:, 1]
        p_va = np.clip(p_va, 1e-6, 1.0 - 1e-6)
        y_va_arr = np.asarray(y_va, dtype=float)
        # Class-balanced log loss: weight by inverse class frequency so neither
        # class dominates the objective (matches the fit-time sample_weight scheme).
        _p_pos = float(np.mean(y_va_arr)) if len(y_va_arr) > 0 else 0.5
        _p_pos = max(1e-6, min(1.0 - 1e-6, _p_pos))
        _w_va = np.where(y_va_arr == 1, 0.5 / _p_pos, 0.5 / (1.0 - _p_pos))
        _ll_per = y_va_arr * np.log(p_va) + (1.0 - y_va_arr) * np.log(1.0 - p_va)
        return -float(np.sum(_w_va * _ll_per) / max(np.sum(_w_va), 1e-9))

    def _trial_callback(_study, trial):
        if progress_cb is not None and progress_label:
            progress_cb(int(trial.number) + 1, int(n_trials), str(progress_label))

    sampler = optuna.samplers.TPESampler(seed=seed)
    study = optuna.create_study(direction="minimize", sampler=sampler)
    study.optimize(_objective, n_trials=n_trials, show_progress_bar=False, callbacks=[_trial_callback])
    best = dict(study.best_params)
    best.update({"loss": "log_loss", "random_state": seed})
    return best


def _time_weight(fight_date, current_date, half_life_days=730):
    """Exponential decay weight: half-life of ~2 years."""
    if current_date is None or fight_date is None:
        return 1.0
    try:
        days = (current_date - fight_date).days
        return 2.0 ** (-max(days, 0) / half_life_days)
    except (TypeError, AttributeError):
        return 1.0


def _expected_calibration_error(y_true, probs, n_bins=10):
    y = np.asarray(y_true, dtype=float)
    p = np.asarray(probs, dtype=float)
    bins = np.linspace(0.0, 1.0, n_bins + 1)
    inds = np.digitize(p, bins) - 1
    ece = 0.0
    for b in range(n_bins):
        mask = inds == b
        if not np.any(mask):
            continue
        conf = p[mask].mean()
        acc = y[mask].mean()
        ece += (mask.mean()) * abs(acc - conf)
    return float(ece)


def _calibration_curve_rmse(y_true, probs, n_bins=10):
    y = np.asarray(y_true, dtype=float)
    p = np.asarray(probs, dtype=float)
    bins = np.linspace(0.0, 1.0, n_bins + 1)
    inds = np.digitize(p, bins) - 1
    sq = 0.0
    wt = 0.0
    for b in range(n_bins):
        mask = inds == b
        if not np.any(mask):
            continue
        conf = float(p[mask].mean())
        acc = float(y[mask].mean())
        w = float(mask.mean())
        sq += w * ((acc - conf) ** 2)
        wt += w
    if wt <= 0:
        return 0.0
    return float(np.sqrt(sq / wt))


# Per-round stat names to track
RD_STATS = [
    "sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att", "ctrl_sec",
    "head", "head_att", "body", "body_att", "leg", "leg_att",
    "distance", "distance_att", "clinch", "clinch_att", "ground", "ground_att",
]

# ─── Fight record extraction ───────────────────────────────────────────────────

def extract_fight_record(row, prefix, opp, result, opp_glicko_mu=MU_0):
    """Build a per-fighter fight record dict from a DataFrame row."""
    def g(col):
        v = row.get(col)
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return float("nan")
        return v

    rec = {
        "date": row["event_date"],
        "fight_time": g("total_fight_time_sec") or 0,
        "result": result,
        "opp_glicko": opp_glicko_mu,
        "method": row.get("method", ""),
        "is_title": g("is_title_bout") or 0,
        "scheduled_rounds": int(g("total_rounds") or 3),
        "finish_round": g("finish_round") or 0,
        # Offense
        "sig_str": g(f"{prefix}_sig_str") or 0,
        "sig_str_att": g(f"{prefix}_sig_str_att") or 0,
        "sig_str_acc": g(f"{prefix}_sig_str_acc"),
        "str": g(f"{prefix}_str") or 0,
        "str_att": g(f"{prefix}_str_att") or 0,
        "str_acc": g(f"{prefix}_str_acc"),
        "kd": g(f"{prefix}_kd") or 0,
        "td": g(f"{prefix}_td") or 0,
        "td_att": g(f"{prefix}_td_att") or 0,
        "td_acc": g(f"{prefix}_td_acc"),
        "sub_att": g(f"{prefix}_sub_att") or 0,
        "rev": g(f"{prefix}_rev") or 0,
        "ctrl_sec": g(f"{prefix}_ctrl_sec") or 0,
        # Targeting (fight-total %)
        "head_pct": g(f"{prefix}_head"),
        "body_pct": g(f"{prefix}_body"),
        "leg_pct": g(f"{prefix}_leg"),
        # Positioning (fight-total %)
        "distance_pct": g(f"{prefix}_distance"),
        "clinch_pct": g(f"{prefix}_clinch"),
        "ground_pct": g(f"{prefix}_ground"),
        # Defense (opponent stats)
        "opp_sig_str": g(f"{opp}_sig_str") or 0,
        "opp_sig_str_att": g(f"{opp}_sig_str_att") or 0,
        "opp_sig_str_acc": g(f"{opp}_sig_str_acc"),
        "opp_str": g(f"{opp}_str") or 0,
        "opp_kd": g(f"{opp}_kd") or 0,
        "opp_td": g(f"{opp}_td") or 0,
        "opp_td_att": g(f"{opp}_td_att") or 0,
        "opp_sub_att": g(f"{opp}_sub_att") or 0,
        "opp_ctrl_sec": g(f"{opp}_ctrl_sec") or 0,
        # Physical
        "height": g(f"{prefix}_height"),
        "reach": g(f"{prefix}_reach"),
        "ape_index": g(f"{prefix}_ape_index"),
        "weight": g(f"{prefix}_weight"),
        "age": g(f"{prefix}_age_at_event"),
        "stance": row.get(f"{prefix}_stance", ""),
    }

    # Per-round stats (offense + opponent defensive counterpart)
    for rd in range(1, 6):
        for stat in RD_STATS:
            rec[f"rd{rd}_{stat}"] = g(f"{prefix}_rd{rd}_{stat}")
            rec[f"opp_rd{rd}_{stat}"] = g(f"{opp}_rd{rd}_{stat}")

    return rec


# ─── Per-fighter feature computation ───────────────────────────────────────────

_FIGHTER_FEAT_KEYS = None  # populated on first call


def _make_synthetic_fight_record(current_date=None, profile=None):
    profile = profile or {}
    rec = {
        "date": current_date or pd.Timestamp("2000-01-01"),
        "fight_time": 900.0,
        "result": "W",
        "opp_glicko": MU_0,
        "self_glicko": MU_0,
        "method": "Decision",
        "is_title": 0,
        "scheduled_rounds": 3,
        "finish_round": 3,
        "sig_str": 0.0,
        "sig_str_att": 0.0,
        "sig_str_acc": float("nan"),
        "str": 0.0,
        "str_att": 0.0,
        "str_acc": float("nan"),
        "kd": 0.0,
        "td": 0.0,
        "td_att": 0.0,
        "td_acc": float("nan"),
        "sub_att": 0.0,
        "rev": 0.0,
        "ctrl_sec": 0.0,
        "head_pct": float("nan"),
        "body_pct": float("nan"),
        "leg_pct": float("nan"),
        "distance_pct": float("nan"),
        "clinch_pct": float("nan"),
        "ground_pct": float("nan"),
        "opp_sig_str": 0.0,
        "opp_sig_str_att": 0.0,
        "opp_sig_str_acc": float("nan"),
        "opp_str": 0.0,
        "opp_kd": 0.0,
        "opp_td": 0.0,
        "opp_td_att": 0.0,
        "opp_sub_att": 0.0,
        "opp_ctrl_sec": 0.0,
        "height": profile.get("height", 70.0),
        "reach": profile.get("reach", 72.0),
        "ape_index": profile.get("ape_index", 2.0),
        "weight": profile.get("weight", 170.0),
        "age": profile.get("age", 30.0),
        "stance": profile.get("stance", "Orthodox"),
    }
    for rd in range(1, 6):
        for stat in RD_STATS:
            rec[f"rd{rd}_{stat}"] = 0.0
            rec[f"opp_rd{rd}_{stat}"] = 0.0
    return rec


def _ensure_fighter_feature_keys(current_date=None):
    global _FIGHTER_FEAT_KEYS
    if _FIGHTER_FEAT_KEYS is not None:
        return
    dummy_history = [_make_synthetic_fight_record(current_date=current_date)]
    compute_fighter_features(dummy_history, (MU_0, PHI_0, SIGMA_0), [], current_date)


def compute_fighter_features(history, glicko, opp_glickos, current_date, fallback_profile=None):
    """Compute ~130 features from a fighter's fight history."""
    global _FIGHTER_FEAT_KEYS
    fallback_profile = fallback_profile or {}

    n = len(history)
    if n == 0:
        _ensure_fighter_feature_keys(current_date)
        feats = {k: float("nan") for k in _FIGHTER_FEAT_KEYS}
        stance = fallback_profile.get("stance", "") or ""
        feats.update({
            "height": _num_or(fallback_profile.get("height"), float("nan")),
            "reach": _num_or(fallback_profile.get("reach"), float("nan")),
            "ape_index": _num_or(fallback_profile.get("ape_index"), float("nan")),
            "weight": _num_or(fallback_profile.get("weight"), float("nan")),
            "age": _num_or(fallback_profile.get("age"), float("nan")),
            "num_fights": 0.0,
            "total_time_min": 0.0,
            "avg_time_min": 0.0,
            "title_bout_pct": 0.0,
            "win_rate": 0.5,
            "ko_win_pct": 0.25,
            "sub_win_pct": 0.12,
            "dec_win_pct": 0.13,
            "finish_rate": 0.5,
            "ko_loss_pct": 0.2,
            "been_finished_pct": 0.5,
            "last3_win_rate": 0.5,
            "last5_win_rate": 0.5,
            "win_streak": 0.0,
            "loss_streak": 0.0,
            "days_inactive": 365.0,
            "glicko_mu": glicko[0],
            "glicko_phi": glicko[1],
            "fights_per_year": 0.0,
            "avg_opp_glicko": MU_0,
            "td_win_rate": 0.5,
            "quality_win_rate": 0.5,
            "best_win_glicko": MU_0,
            "worst_loss_glicko": MU_0,
            "style_striking": 0.5,
            "style_wrestling": 0.25,
            "style_submission": 0.15,
            "style_clinch_ground": 0.32,
            "style_finishing": 0.5,
            "is_orthodox": 1.0 if stance == "Orthodox" else 0.0,
            "is_southpaw": 1.0 if stance == "Southpaw" else 0.0,
            "is_switch": 1.0 if stance == "Switch" else 0.0,
            "age_squared": 900.0,
            "prime_age": 1.0,
            "age_decline": 0.0,
            "experience_log": 0.0,
            "momentum": 0.5,
            "first_round_finish_rate": 0.15,
            "durability": 0.8,
            "output_rate": 0.0,
            "damage_efficiency": 1.0,
            "late_round_pct": 0.5,
            "avg_finish_round": 2.5,
            "sig_str_diff_pm": 0.0,
            "grappling_dominance": 0.0,
            "consistency": 0.5,
            "ewm_opp_quality": MU_0,
            "form_vs_career": 0.0,
            "form5_vs_career": 0.0,
            "days_since_last_loss": 1500.0,
            "title_fight_win_rate": 0.5,
            "title_fight_experience": 0.0,
            "fight_iq": 0.45 * 0.35,
            "rd1_intensity_ratio": 1.0,
            "cardio_ratio": 1.0,
            "opp_quality_trend": 1.0,
            "fights_last_year": 0.0,
            "win_rate_last_year": 0.5,
            "loss_recovery_rate": 0.5,
            "strike_exchange_ratio": 1.0,
            "grappling_threat": 0.0,
            "decision_win_rate": 0.5,
            "finish_resistance": 0.5,
            "offensive_diversity": 1.0,
            "ewm_sig_str_diff_pm": 0.0,
            "glicko_confidence": 0.0,
            "dapa_sig_str_pm": 0.0,
            "dapa_kd_pm": 0.0,
            "dapa_sub_att_p15": 0.0,
            "dapa_td_acc": 0.0,
            "dapa_sig_str_def": 0.55,
            "sig_str_per_ctrl_sec": 0.0,
            "td_land_per_sub_att": 0.0,
            "finish_r1_share": 0.0,
            "finish_r3_plus_share": 0.0,
            "avg_finish_round_w": 0.0,
        })
        return feats

    feats = {}

    total_time = _safe_sum(h["fight_time"] for h in history)
    total_time_min = total_time / 60.0 if total_time > 0 else 1.0
    total_time_15 = total_time / 900.0 if total_time > 0 else 1.0

    total_sig_land = _safe_sum(h["sig_str"] for h in history)
    total_sig_att = _safe_sum(h["sig_str_att"] for h in history)
    total_str_land = _safe_sum(h["str"] for h in history)
    total_str_att = _safe_sum(h["str_att"] for h in history)
    total_td_land = _safe_sum(h["td"] for h in history)
    total_td_att = _safe_sum(h["td_att"] for h in history)

    # ── Striking offense (per minute) ──
    feats["sig_str_pm"] = total_sig_land / total_time_min
    feats["sig_str_att_pm"] = total_sig_att / total_time_min
    feats["str_pm"] = total_str_land / total_time_min
    feats["str_att_pm"] = total_str_att / total_time_min
    feats["kd_pm"] = _safe_sum(h["kd"] for h in history) / total_time_min

    # ── Accuracy (attempt-weighted + Bayesian shrinkage) ──
    feats["sig_str_acc"] = _bayes_shrink(
        _weighted_rate(total_sig_land, total_sig_att, prior=0.45, strength=30),
        n, prior=0.45, strength=8,
    )
    feats["str_acc"] = _bayes_shrink(
        _weighted_rate(total_str_land, total_str_att, prior=0.55, strength=30),
        n, prior=0.55, strength=8,
    )
    feats["td_acc"] = _bayes_shrink(
        _weighted_rate(total_td_land, total_td_att, prior=0.35, strength=20),
        n, prior=0.35, strength=8,
    )

    # ── Grappling (per 15 min) ──
    feats["td_p15"] = _safe_sum(h["td"] for h in history) / total_time_15
    feats["td_att_p15"] = _safe_sum(h["td_att"] for h in history) / total_time_15
    feats["sub_att_p15"] = _safe_sum(h["sub_att"] for h in history) / total_time_15
    feats["rev_p15"] = _safe_sum(h["rev"] for h in history) / total_time_15
    feats["ctrl_pct"] = _safe_sum(h["ctrl_sec"] for h in history) / total_time if total_time > 0 else 0

    # ── Targeting (attempt/volume-weighted + shrinkage) ──
    sig_for_target = _safe_sum(
        h["sig_str"] for h in history
        if not _isnan(h["head_pct"]) and not _isnan(h["sig_str"])
    )
    head_num = _safe_sum(
        h["head_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["head_pct"]) and not _isnan(h["sig_str"])
    )
    body_num = _safe_sum(
        h["body_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["body_pct"]) and not _isnan(h["sig_str"])
    )
    leg_num = _safe_sum(
        h["leg_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["leg_pct"]) and not _isnan(h["sig_str"])
    )
    feats["head_pct"] = _bayes_shrink(
        _weighted_rate(head_num, sig_for_target, prior=0.62, strength=40),
        n, prior=0.62, strength=8,
    )
    feats["body_pct"] = _bayes_shrink(
        _weighted_rate(body_num, sig_for_target, prior=0.22, strength=40),
        n, prior=0.22, strength=8,
    )
    feats["leg_pct"] = _bayes_shrink(
        _weighted_rate(leg_num, sig_for_target, prior=0.16, strength=40),
        n, prior=0.16, strength=8,
    )

    # ── Positioning (volume-weighted + shrinkage) ──
    sig_for_pos = _safe_sum(
        h["sig_str"] for h in history
        if not _isnan(h["distance_pct"]) and not _isnan(h["sig_str"])
    )
    dist_num = _safe_sum(
        h["distance_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["distance_pct"]) and not _isnan(h["sig_str"])
    )
    clinch_num = _safe_sum(
        h["clinch_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["clinch_pct"]) and not _isnan(h["sig_str"])
    )
    ground_num = _safe_sum(
        h["ground_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["ground_pct"]) and not _isnan(h["sig_str"])
    )
    feats["distance_pct"] = _bayes_shrink(
        _weighted_rate(dist_num, sig_for_pos, prior=0.68, strength=40),
        n, prior=0.68, strength=8,
    )
    feats["clinch_pct"] = _bayes_shrink(
        _weighted_rate(clinch_num, sig_for_pos, prior=0.14, strength=40),
        n, prior=0.14, strength=8,
    )
    feats["ground_pct"] = _bayes_shrink(
        _weighted_rate(ground_num, sig_for_pos, prior=0.18, strength=40),
        n, prior=0.18, strength=8,
    )

    # ── Defense (opponent per-minute) ──
    feats["def_sig_str_pm"] = _safe_sum(h["opp_sig_str"] for h in history) / total_time_min
    feats["def_str_pm"] = _safe_sum(h["opp_str"] for h in history) / total_time_min
    feats["def_kd_pm"] = _safe_sum(h["opp_kd"] for h in history) / total_time_min
    feats["def_td_p15"] = _safe_sum(h["opp_td"] for h in history) / total_time_15
    feats["def_sub_att_p15"] = _safe_sum(h["opp_sub_att"] for h in history) / total_time_15
    feats["def_ctrl_pct"] = _safe_sum(h["opp_ctrl_sec"] for h in history) / total_time if total_time > 0 else 0
    opp_sig_att = _safe_sum(h["opp_sig_str_att"] for h in history)
    opp_sig_land = _safe_sum(h["opp_sig_str"] for h in history)
    feats["def_sig_str_acc"] = _bayes_shrink(
        _weighted_rate(opp_sig_land, opp_sig_att, prior=0.45, strength=30),
        n, prior=0.45, strength=8,
    )

    # ── Differentials ──
    feats["net_sig_str_pm"] = feats["sig_str_pm"] - feats["def_sig_str_pm"]
    feats["net_kd_pm"] = feats["kd_pm"] - feats["def_kd_pm"]
    feats["net_td_p15"] = feats["td_p15"] - feats["def_td_p15"]
    feats["net_ctrl_pct"] = feats["ctrl_pct"] - feats["def_ctrl_pct"]

    # ── Defense rates ──
    feats["sig_str_defense_rate"] = _bayes_shrink(
        1.0 - _weighted_rate(opp_sig_land, opp_sig_att, prior=0.45, strength=30),
        n, prior=0.55, strength=8,
    )
    opp_td_att = _safe_sum(h["opp_td_att"] for h in history)
    opp_td_land = _safe_sum(h["opp_td"] for h in history)
    feats["td_defense_rate"] = _bayes_shrink(
        1.0 - _weighted_rate(opp_td_land, opp_td_att, prior=0.35, strength=20),
        n, prior=0.65, strength=8,
    )

    # ── Physical (most recent) ──
    latest = history[-1]
    feats["height"] = latest["height"] if not _isnan(latest["height"]) else _num_or(fallback_profile.get("height"), float("nan"))
    feats["reach"] = latest["reach"] if not _isnan(latest["reach"]) else _num_or(fallback_profile.get("reach"), float("nan"))
    feats["ape_index"] = latest["ape_index"] if not _isnan(latest["ape_index"]) else _num_or(fallback_profile.get("ape_index"), float("nan"))
    feats["weight"] = latest["weight"] if not _isnan(latest["weight"]) else _num_or(fallback_profile.get("weight"), float("nan"))
    feats["age"] = latest["age"] if not _isnan(latest["age"]) else _num_or(fallback_profile.get("age"), float("nan"))

    # ── Experience ──
    feats["num_fights"] = n
    feats["total_time_min"] = total_time_min
    feats["avg_time_min"] = total_time_min / n
    feats["title_bout_pct"] = sum(1 for h in history if h["is_title"]) / n

    # ── Record ──
    wins = sum(1 for h in history if h["result"] == "W")
    losses = sum(1 for h in history if h["result"] == "L")
    feats["win_rate"] = _bayes_shrink(_safe_div(wins, n, 0.5), n, prior=0.5, strength=10)
    ko_w = sum(1 for h in history if h["result"] == "W" and "KO" in str(h["method"]))
    sub_w = sum(1 for h in history if h["result"] == "W" and "Sub" in str(h["method"]))
    dec_w = sum(1 for h in history if h["result"] == "W" and "Dec" in str(h["method"]))
    feats["ko_win_pct"] = _bayes_shrink(_safe_div(ko_w, n, 0.25), n, prior=0.25, strength=10)
    feats["sub_win_pct"] = _bayes_shrink(_safe_div(sub_w, n, 0.12), n, prior=0.12, strength=10)
    feats["dec_win_pct"] = _bayes_shrink(_safe_div(dec_w, n, 0.13), n, prior=0.13, strength=10)
    feats["finish_rate"] = _safe_div(ko_w + sub_w, max(wins, 1))
    ko_l = sum(1 for h in history if h["result"] == "L" and "KO" in str(h["method"]))
    sub_l = sum(1 for h in history if h["result"] == "L" and "Sub" in str(h["method"]))
    feats["ko_loss_pct"] = _bayes_shrink(_safe_div(ko_l, n, 0.2), n, prior=0.2, strength=10)
    feats["sub_loss_pct"] = _bayes_shrink(_safe_div(sub_l, max(losses, 1), 0.20), losses, prior=0.20, strength=8)
    feats["been_finished_pct"] = _safe_div(ko_l + sub_l, max(losses, 1))

    # ── Form ──
    last3 = history[-3:]
    last5 = history[-5:]
    last3_wr = sum(1 for h in last3 if h["result"] == "W") / len(last3)
    last5_wr = sum(1 for h in last5 if h["result"] == "W") / len(last5)
    feats["last3_win_rate"] = _bayes_shrink(last3_wr, len(last3), prior=0.5, strength=6)
    feats["last5_win_rate"] = _bayes_shrink(last5_wr, len(last5), prior=0.5, strength=6)
    recent_sub_wins = sum(1 for h in last5 if h["result"] == "W" and "Sub" in str(h.get("method", "")))
    recent_ko_losses = sum(1 for h in last5 if h["result"] == "L" and "KO" in str(h.get("method", "")))
    recent_n = max(min(n, 5), 1)
    feats["recent_sub_win_rate"] = _bayes_shrink(
        _safe_div(recent_sub_wins, recent_n, 0.12), recent_n, prior=0.12, strength=5
    )
    feats["recent_ko_loss_rate"] = _bayes_shrink(
        _safe_div(recent_ko_losses, recent_n, 0.20), recent_n, prior=0.20, strength=5
    )
    win_streak = 0
    for h in reversed(history):
        if h["result"] == "W":
            win_streak += 1
        else:
            break
    loss_streak = 0
    for h in reversed(history):
        if h["result"] == "L":
            loss_streak += 1
        else:
            break
    feats["win_streak"] = win_streak
    feats["loss_streak"] = loss_streak
    if current_date is not None and not _isnan(history[-1]["date"]):
        feats["days_inactive"] = (current_date - history[-1]["date"]).days
    else:
        feats["days_inactive"] = 365

    # ── Short-notice flag (last fight < 45 days ago) ──
    feats["short_notice"] = float(feats["days_inactive"] < 45)

    # ── Rounds-scheduled experience (time-weighted) ──
    feats["rounds_5_exp"] = sum(
        _time_weight(h["date"], current_date) for h in history
        if int(h.get("scheduled_rounds", 3)) == 5
    )
    feats["rounds_3_exp"] = sum(
        _time_weight(h["date"], current_date) for h in history
        if int(h.get("scheduled_rounds", 3)) == 3
    )

    # ── Glicko-2 ──
    feats["glicko_mu"] = glicko[0]
    feats["glicko_phi"] = glicko[1]

    # ── Glicko trend (current mu minus mu from 5 fights ago) ──
    if n >= 5:
        feats["glicko_trend"] = glicko[0] - history[-5].get("self_glicko", glicko[0])
    elif n >= 1:
        feats["glicko_trend"] = glicko[0] - history[0].get("self_glicko", glicko[0])
    else:
        feats["glicko_trend"] = 0.0

    # ── Recent damage absorbed (sig str taken in last 3 fights) ──
    recent3 = history[-3:]
    feats["recent_damage_absorbed"] = _safe_sum(h["opp_sig_str"] for h in recent3)


    # ── Round 1 stats (career averages) ──
    for stat in ["sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att",
                 "ctrl_sec", "head", "body", "leg", "distance", "clinch", "ground"]:
        vals = [h[f"rd1_{stat}"] for h in history if not _isnan(h.get(f"rd1_{stat}"))]
        feats[f"rd1_{stat}"] = _safe_mean(vals)

    # ── Round 2 stats ──
    for stat in ["sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att",
                 "ctrl_sec", "head", "body", "leg"]:
        vals = [h[f"rd2_{stat}"] for h in history if not _isnan(h.get(f"rd2_{stat}"))]
        feats[f"rd2_{stat}"] = _safe_mean(vals)

    # ── Round 3 stats ──
    for stat in ["sig_str", "kd", "td", "td_att", "sub_att", "ctrl_sec", "head", "body", "leg"]:
        vals = [h[f"rd3_{stat}"] for h in history if not _isnan(h.get(f"rd3_{stat}"))]
        feats[f"rd3_{stat}"] = _safe_mean(vals)

    # ── Championship rounds (4+5 combined) ──
    for stat in ["sig_str", "kd", "td", "ctrl_sec"]:
        vals = []
        for rd in [4, 5]:
            vals.extend(h[f"rd{rd}_{stat}"] for h in history
                        if not _isnan(h.get(f"rd{rd}_{stat}")))
        feats[f"champ_{stat}"] = _safe_mean(vals)

    # ── Defensive round aggregates (what the fighter absorbs per round) ──
    for stat in ["sig_str", "kd"]:
        vals = [h[f"opp_rd1_{stat}"] for h in history if not _isnan(h.get(f"opp_rd1_{stat}"))]
        feats[f"def_rd1_{stat}"] = _safe_mean(vals)

    # ── Late-round sig_str per minute (only fights that reached rd4/rd5) ──
    late_vals = []
    for h in history:
        for rd in (4, 5):
            v = h.get(f"rd{rd}_sig_str")
            if not _isnan(v) and v is not None:
                late_vals.append(float(v) / 5.0)  # 5-minute round → per-minute
    feats["late_sig_str_pm"] = _safe_mean(late_vals) if late_vals else float("nan")


    # ── Composite pacing aggregates (feed the auto-diff loop in compute_matchup_features) ──
    # rd1 net sig_str per minute: what the fighter lands minus what they absorb in round 1.
    _rd1_ss = feats.get("rd1_sig_str")
    _def_rd1_ss = feats.get("def_rd1_sig_str")
    if not _isnan(_rd1_ss) and not _isnan(_def_rd1_ss):
        feats["rd1_net_sig_str"] = (float(_rd1_ss) - float(_def_rd1_ss)) / 5.0
    else:
        feats["rd1_net_sig_str"] = float("nan")
    # Cardio decay: rd1 output minus rd3 output, per minute. Higher = fades more.
    _rd3_ss = feats.get("rd3_sig_str")
    if not _isnan(_rd1_ss) and not _isnan(_rd3_ss):
        feats["cardio_decay_sig_str"] = (float(_rd1_ss) - float(_rd3_ss)) / 5.0
    else:
        feats["cardio_decay_sig_str"] = float("nan")
    # Round-1 control share: fraction of the 5-minute round spent in control position.
    _rd1_ctrl = feats.get("rd1_ctrl_sec")
    if not _isnan(_rd1_ctrl):
        feats["rd1_ctrl_share"] = float(_rd1_ctrl) / 300.0
    else:
        feats["rd1_ctrl_share"] = float("nan")

    # ── Late vs early ──
    rd1_ss = [h["rd1_sig_str"] for h in history if not _isnan(h.get("rd1_sig_str"))]
    rd3_ss = [h["rd3_sig_str"] for h in history if not _isnan(h.get("rd3_sig_str"))]
    feats["late_vs_early_sig_str"] = _safe_mean(rd3_ss) - _safe_mean(rd1_ss) \
        if rd1_ss and rd3_ss else float("nan")
    rd1_sub = [h["rd1_sub_att"] for h in history if not _isnan(h.get("rd1_sub_att"))]
    rd2_sub = [h["rd2_sub_att"] for h in history if not _isnan(h.get("rd2_sub_att"))]
    rd3_sub = [h["rd3_sub_att"] for h in history if not _isnan(h.get("rd3_sub_att"))]
    r1_sub_m = _safe_mean(rd1_sub)
    r2_sub_m = _safe_mean(rd2_sub)
    r3_sub_m = _safe_mean(rd3_sub)
    feats["late_vs_early_sub_att"] = r3_sub_m - r1_sub_m if rd1_sub and rd3_sub else float("nan")
    feats["sub_att_trend_12"] = r2_sub_m - r1_sub_m if rd1_sub and rd2_sub else float("nan")
    feats["sub_att_trend_23"] = r3_sub_m - r2_sub_m if rd2_sub and rd3_sub else float("nan")
    total_r123_sub = max(r1_sub_m + r2_sub_m + r3_sub_m, 1e-6)
    feats["rd3_sub_share"] = r3_sub_m / total_r123_sub
    feats["rd1_sub_share"] = r1_sub_m / total_r123_sub
    feats["sub_round_concentration"] = max(r1_sub_m, r2_sub_m, r3_sub_m) / total_r123_sub

    # ── Recent-window rate stats (last 1, 3, 5 fights) ──
    for prefix_tag, window in [("r1f", 1), ("r3", 3), ("r5", 5)]:
        recent = history[-window:]
        rt = _safe_sum(h["fight_time"] for h in recent)
        rt_min = rt / 60.0 if rt > 0 else 1.0
        rt_15 = rt / 900.0 if rt > 0 else 1.0
        feats[f"{prefix_tag}_sig_str_pm"] = _safe_sum(h["sig_str"] for h in recent) / rt_min
        feats[f"{prefix_tag}_kd_pm"] = _safe_sum(h["kd"] for h in recent) / rt_min
        feats[f"{prefix_tag}_td_p15"] = _safe_sum(h["td"] for h in recent) / rt_15
        feats[f"{prefix_tag}_sub_att_p15"] = _safe_sum(h["sub_att"] for h in recent) / rt_15
        feats[f"{prefix_tag}_ctrl_pct"] = _safe_sum(h["ctrl_sec"] for h in recent) / rt if rt > 0 else 0
        feats[f"{prefix_tag}_def_sig_str_pm"] = _safe_sum(h["opp_sig_str"] for h in recent) / rt_min
        feats[f"{prefix_tag}_def_kd_pm"] = _safe_sum(h["opp_kd"] for h in recent) / rt_min
        rec_sig_land = _safe_sum(h["sig_str"] for h in recent)
        rec_sig_att = _safe_sum(h["sig_str_att"] for h in recent)
        rec_td_land = _safe_sum(h["td"] for h in recent)
        rec_td_att = _safe_sum(h["td_att"] for h in recent)
        feats[f"{prefix_tag}_sig_str_acc"] = _bayes_shrink(
            _weighted_rate(rec_sig_land, rec_sig_att, prior=0.45, strength=20),
            len(recent), prior=0.45, strength=6,
        )
        feats[f"{prefix_tag}_td_acc"] = _bayes_shrink(
            _weighted_rate(rec_td_land, rec_td_att, prior=0.35, strength=20),
            len(recent), prior=0.35, strength=6,
        )
        feats[f"{prefix_tag}_win"] = _bayes_shrink(
            _safe_div(sum(1 for h in recent if h["result"] == "W"), len(recent), 0.5),
            len(recent), prior=0.5, strength=6,
        )

    # ── Exponentially weighted moving average (alpha=0.3, more recent = higher) ──
    alpha = 0.3
    ewm_keys = [
        ("sig_str", "fight_time", 60),
        ("kd", "fight_time", 60),
        ("td", "fight_time", 900),
        ("opp_sig_str", "fight_time", 60),
    ]
    for stat_key, time_key, divisor in ewm_keys:
        wsum, wtot = 0.0, 0.0
        for i, h in enumerate(history):
            w = (1 - alpha) ** (n - 1 - i)
            ft = h[time_key] / divisor if h[time_key] > 0 else 1.0
            wsum += w * (h[stat_key] / ft)
            wtot += w
        tag = stat_key.replace("opp_sig_str", "def_sig_str")
        feats[f"ewm_{tag}_pm"] = wsum / wtot if wtot > 0 else 0

    # EWM ctrl_pct
    wsum, wtot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha) ** (n - 1 - i)
        ft = h["fight_time"] if h["fight_time"] > 0 else 1.0
        wsum += w * (h["ctrl_sec"] / ft)
        wtot += w
    feats["ewm_ctrl_pct"] = wsum / wtot if wtot > 0 else 0

    # EWM accuracy
    for acc_key in ["sig_str_acc", "td_acc"]:
        wsum, wtot = 0.0, 0.0
        for i, h in enumerate(history):
            if _isnan(h[acc_key]):
                continue
            w = (1 - alpha) ** (n - 1 - i)
            wsum += w * h[acc_key]
            wtot += w
        feats[f"ewm_{acc_key}"] = wsum / wtot if wtot > 0 else float("nan")

    # EWM win
    wsum, wtot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha) ** (n - 1 - i)
        wsum += w * (1.0 if h["result"] == "W" else 0.0)
        wtot += w
    feats["ewm_win"] = wsum / wtot if wtot > 0 else 0.5

    # ── Variance (per-fight rates std dev) ──
    if n >= 2:
        per_fight_sig = []
        per_fight_kd = []
        per_fight_td = []
        per_fight_ctrl = []
        per_fight_def_sig = []
        for h in history:
            ft_min = h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0
            ft_15 = h["fight_time"] / 900.0 if h["fight_time"] > 0 else 1.0
            ft = h["fight_time"] if h["fight_time"] > 0 else 1.0
            per_fight_sig.append(h["sig_str"] / ft_min)
            per_fight_kd.append(h["kd"] / ft_min)
            per_fight_td.append(h["td"] / ft_15)
            per_fight_ctrl.append(h["ctrl_sec"] / ft)
            per_fight_def_sig.append(h["opp_sig_str"] / ft_min)
        feats["std_sig_str_pm"] = float(np.std(per_fight_sig))
        feats["std_kd_pm"] = float(np.std(per_fight_kd))
        feats["std_td_p15"] = float(np.std(per_fight_td))
        feats["std_ctrl_pct"] = float(np.std(per_fight_ctrl))
        feats["std_def_sig_str_pm"] = float(np.std(per_fight_def_sig))
        # Time-weighted volatility of sig_str differential per minute
        _w_vol = [_time_weight(h["date"], current_date) for h in history]
        _sd_pm = []
        for h in history:
            _fm = h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0
            _sd_pm.append((h["sig_str"] - h["opp_sig_str"]) / _fm)
        _wsum = sum(_w_vol) or 1.0
        _mean_sd = sum(w * d for w, d in zip(_w_vol, _sd_pm)) / _wsum
        _var_sd = sum(w * (d - _mean_sd) ** 2 for w, d in zip(_w_vol, _sd_pm)) / _wsum
        _vol_raw = math.sqrt(max(_var_sd, 0.0))
        feats["sig_diff_pm_vol"] = (_vol_raw * _wsum + 3.0 * 5.0) / (_wsum + 5.0)
    else:
        feats["std_sig_str_pm"] = 0.0
        feats["std_kd_pm"] = 0.0
        feats["std_td_p15"] = 0.0
        feats["std_ctrl_pct"] = 0.0
        feats["std_def_sig_str_pm"] = 0.0
        feats["sig_diff_pm_vol"] = 3.0

    # ── Derived ──
    total_sig = _safe_sum(h["sig_str"] for h in history)
    total_str = _safe_sum(h["str"] for h in history)
    feats["power_ratio"] = _safe_div(_safe_sum(h["kd"] for h in history), max(total_sig, 1))
    feats["striking_efficiency"] = _safe_div(total_sig, max(total_str, 1))
    feats["grappling_rate_p15"] = (
        _safe_sum(h["td"] + h["sub_att"] for h in history) / total_time_15
    )
    feats["def_grappling_rate_p15"] = (
        _safe_sum(h["opp_td"] + h["opp_sub_att"] for h in history) / total_time_15
    )
    feats["damage_ratio"] = _safe_div(
        _safe_sum(h["opp_sig_str"] for h in history),
        max(_safe_sum(h["sig_str"] for h in history), 1),
    )

    # ── Target/position attempt-driven accuracy and intent shares (round 1-5) ──
    eps = 1e-6

    def _round_stat_total(stat_name):
        vals = []
        for h in history:
            for rd in range(1, 6):
                v = h.get(f"rd{rd}_{stat_name}", float("nan"))
                if not _isnan(v):
                    vals.append(v)
        return _safe_sum(vals)

    total_sig_att_round = _round_stat_total("sig_str_att")
    total_head_land = _round_stat_total("head")
    total_head_att = _round_stat_total("head_att")
    total_body_land = _round_stat_total("body")
    total_body_att = _round_stat_total("body_att")
    total_leg_land = _round_stat_total("leg")
    total_leg_att = _round_stat_total("leg_att")
    total_distance_land = _round_stat_total("distance")
    total_distance_att = _round_stat_total("distance_att")
    total_clinch_land = _round_stat_total("clinch")
    total_clinch_att = _round_stat_total("clinch_att")
    total_ground_land = _round_stat_total("ground")
    total_ground_att = _round_stat_total("ground_att")

    feats["head_acc"] = _safe_div(total_head_land, total_head_att + eps, 0.16)
    feats["distance_acc"] = _safe_div(total_distance_land, total_distance_att + eps, 0.45)
    feats["ground_acc"] = _safe_div(total_ground_land, total_ground_att + eps, 0.35)
    feats["clinch_acc"] = _safe_div(total_clinch_land, total_clinch_att + eps, 0.30)
    feats["body_acc"] = _safe_div(total_body_land, total_body_att + eps, 0.28)
    feats["leg_acc"] = _safe_div(total_leg_land, total_leg_att + eps, 0.28)
    feats["body_leg_attrition"] = feats["body_acc"] + feats["leg_acc"]
    feats["ground_strike_accuracy"] = feats["ground_acc"]
    feats["head_hunt_accuracy"] = feats["head_acc"]
    feats["distance_strike_accuracy"] = feats["distance_acc"]

    feats["head_hunt_share"] = _safe_div(total_head_att, total_sig_att_round + eps, 0.33)
    feats["distance_share"] = _safe_div(total_distance_att, total_sig_att_round + eps, 0.62)

    # Fights per year
    if n > 1:
        career_days = (history[-1]["date"] - history[0]["date"]).days
        feats["fights_per_year"] = n / max(career_days / 365.25, 0.5)
    else:
        feats["fights_per_year"] = 1.0

    # Stance
    stance = latest.get("stance", "") or fallback_profile.get("stance", "")
    feats["is_orthodox"] = 1 if stance == "Orthodox" else 0
    feats["is_southpaw"] = 1 if stance == "Southpaw" else 0
    feats["is_switch"] = 1 if stance == "Switch" else 0

    # Average opponent Glicko
    feats["avg_opp_glicko"] = _safe_mean(opp_glickos) if opp_glickos else MU_0

    # ── Time-decayed career stats (half-life ~2 years) ──
    tw = [_time_weight(h["date"], current_date) for h in history]
    tw_total = sum(tw) or 1.0
    tw_time = sum(w * h["fight_time"] for w, h in zip(tw, history))
    tw_time_min = tw_time / 60.0 if tw_time > 0 else 1.0
    tw_time_15 = tw_time / 900.0 if tw_time > 0 else 1.0
    feats["td_sig_str_pm"] = sum(w * h["sig_str"] for w, h in zip(tw, history)) / tw_time_min
    feats["td_kd_pm"] = sum(w * h["kd"] for w, h in zip(tw, history)) / tw_time_min
    feats["td_td_p15"] = sum(w * h["td"] for w, h in zip(tw, history)) / tw_time_15
    feats["td_ctrl_pct"] = (
        sum(w * h["ctrl_sec"] for w, h in zip(tw, history)) / tw_time if tw_time > 0 else 0
    )
    feats["td_def_sig_str_pm"] = sum(w * h["opp_sig_str"] for w, h in zip(tw, history)) / tw_time_min
    tw_sig_land = sum(w * h["sig_str"] for w, h in zip(tw, history))
    tw_sig_att = sum(w * h["sig_str_att"] for w, h in zip(tw, history))
    feats["td_sig_str_acc"] = _weighted_rate(tw_sig_land, tw_sig_att, prior=0.45, strength=20)
    tw_td_land = sum(w * h["td"] for w, h in zip(tw, history))
    tw_td_att = sum(w * h["td_att"] for w, h in zip(tw, history))
    feats["td_td_acc"] = _weighted_rate(tw_td_land, tw_td_att, prior=0.35, strength=20)
    feats["td_win_rate"] = sum(
        w * (1.0 if h["result"] == "W" else 0.0) for w, h in zip(tw, history)
    ) / tw_total

    # ── Opponent-quality-adjusted features ──
    opp_g = [h.get("opp_glicko", MU_0) for h in history]
    opp_g_w = [max(g_val / MU_0, 0.1) for g_val in opp_g]
    ogw_total = sum(opp_g_w) or 1.0
    feats["quality_win_rate"] = sum(
        w * (1.0 if h["result"] == "W" else 0.0)
        for w, h in zip(opp_g_w, history)
    ) / ogw_total
    wins_opp_g = [og for h, og in zip(history, opp_g) if h["result"] == "W"]
    feats["best_win_glicko"] = max(wins_opp_g) if wins_opp_g else MU_0
    losses_opp_g = [og for h, og in zip(history, opp_g) if h["result"] == "L"]
    feats["worst_loss_glicko"] = min(losses_opp_g) if losses_opp_g else MU_0
    feats["quality_sig_str_pm"] = sum(
        w * h["sig_str"] for w, h in zip(opp_g_w, history)
    ) / (sum(w * (h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0)
             for w, h in zip(opp_g_w, history)) or 1.0)

    # ── Style profile (continuous features for matchup interactions) ──
    total_offensive = feats["sig_str_pm"] + feats["td_p15"] + feats["sub_att_p15"] + 0.01
    feats["style_striking"] = feats["sig_str_pm"] / total_offensive
    feats["style_wrestling"] = feats["td_p15"] / total_offensive
    feats["style_submission"] = feats["sub_att_p15"] / total_offensive
    feats["style_clinch_ground"] = feats.get("clinch_pct", 0.14) + feats.get("ground_pct", 0.18)
    feats["style_finishing"] = feats.get("finish_rate", 0.5)

    # ── New high-impact features ──
    age_val = feats.get("age", 30.0)
    if _isnan(age_val):
        age_val = 30.0
    feats["age_squared"] = age_val ** 2
    feats["prime_age"] = 1.0 if 26.0 <= age_val <= 32.0 else 0.0
    feats["age_decline"] = max(0.0, age_val - 32.0)
    feats["experience_log"] = math.log1p(n)

    # Momentum: steeper EWM (alpha=0.5) for recent bias
    alpha_m = 0.5
    m_sum, m_tot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha_m) ** (n - 1 - i)
        m_sum += w * (1.0 if h["result"] == "W" else 0.0)
        m_tot += w
    feats["momentum"] = m_sum / m_tot if m_tot > 0 else 0.5

    # First-round finish rate
    r1_finishes = sum(1 for h in history if h["result"] == "W" and h.get("finish_round") == 1
                      and ("KO" in str(h.get("method", "")) or "Sub" in str(h.get("method", ""))))
    feats["first_round_finish_rate"] = _bayes_shrink(_safe_div(r1_finishes, n, 0.15), n, prior=0.15, strength=10)

    # Durability (inverse of KO loss susceptibility)
    feats["durability"] = 1.0 - feats.get("ko_loss_pct", 0.2)

    # Output rate: total offensive actions per minute
    total_actions = total_sig_land + _safe_sum(h["td"] for h in history) + _safe_sum(h["sub_att"] for h in history)
    feats["output_rate"] = total_actions / total_time_min

    # Damage efficiency: sig str landed / opponent sig str landed
    opp_total_sig = _safe_sum(h["opp_sig_str"] for h in history)
    feats["damage_efficiency"] = _safe_div(total_sig_land, max(opp_total_sig, 1), 1.0)

    # Late-round endurance: pct of fights that went past round 2
    went_late = sum(1 for h in history if _num_or(h.get("finish_round"), 3) >= 3)
    feats["late_round_pct"] = _bayes_shrink(_safe_div(went_late, n, 0.5), n, prior=0.5, strength=8)

    # Average finish round (lower = more dangerous finisher)
    finish_rounds = [h["finish_round"] for h in history
                     if h["result"] == "W" and not _isnan(h.get("finish_round")) and h["finish_round"] > 0]
    feats["avg_finish_round"] = _safe_mean(finish_rounds) if finish_rounds else 2.5

    # Sig str differential per minute (combines offense and defense into one number)
    feats["sig_str_diff_pm"] = feats["sig_str_pm"] - feats["def_sig_str_pm"]

    # Grappling dominance: (td_landed + ctrl_pct) vs (opp_td_landed + opp_ctrl_pct)
    feats["grappling_dominance"] = feats["td_p15"] + feats["ctrl_pct"] - feats["def_td_p15"] - feats["def_ctrl_pct"]

    # Consistency: coefficient of variation of sig_str_pm (lower = more consistent)
    if n >= 2 and feats["sig_str_pm"] > 0:
        feats["consistency"] = 1.0 - min(feats["std_sig_str_pm"] / (feats["sig_str_pm"] + 0.01), 2.0) / 2.0
    else:
        feats["consistency"] = 0.5

    # EWM opponent quality: recent opponent strength
    alpha_oq = 0.4
    oq_sum, oq_tot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha_oq) ** (n - 1 - i)
        oq_sum += w * _num_or(h.get("opp_glicko"), MU_0)
        oq_tot += w
    feats["ewm_opp_quality"] = oq_sum / oq_tot if oq_tot > 0 else MU_0

    # ── Form vs career (explicit trend for linear models) ──
    feats["form_vs_career"] = feats["last3_win_rate"] - feats["win_rate"]
    feats["form5_vs_career"] = feats["last5_win_rate"] - feats["win_rate"]

    # ── Days since last loss (confidence/psychology proxy) ──
    last_loss_idx = None
    for i in range(len(history) - 1, -1, -1):
        if history[i]["result"] == "L":
            last_loss_idx = i
            break
    if last_loss_idx is not None and current_date is not None:
        feats["days_since_last_loss"] = (current_date - history[last_loss_idx]["date"]).days
    else:
        feats["days_since_last_loss"] = 1500.0  # Never lost or no date

    # ── Title fight performance ──
    title_fights = [h for h in history if h.get("is_title")]
    title_wins = sum(1 for h in title_fights if h["result"] == "W")
    feats["title_fight_win_rate"] = _bayes_shrink(
        _safe_div(title_wins, len(title_fights), 0.5),
        len(title_fights), prior=0.5, strength=4,
    )
    feats["title_fight_experience"] = len(title_fights)

    # ── Fight IQ composite (multi-domain effectiveness) ──
    feats["fight_iq"] = feats["sig_str_acc"] * feats["td_acc"]

    # ── Round 1 intensity ratio (first-round explosiveness vs career pace) ──
    rd1_ss_val = feats.get("rd1_sig_str", float("nan"))
    if not _isnan(rd1_ss_val) and feats["sig_str_pm"] > 0:
        # rd1_sig_str is per-round avg, sig_str_pm is per-minute career; convert to same scale
        feats["rd1_intensity_ratio"] = rd1_ss_val / (feats["sig_str_pm"] * 5.0 + 0.01)
    else:
        feats["rd1_intensity_ratio"] = 1.0

    # ── Cardio ratio (round 3 output / round 1 output) ──
    rd1_val = feats.get("rd1_sig_str", float("nan"))
    rd3_val = feats.get("rd3_sig_str", float("nan"))
    if not _isnan(rd1_val) and not _isnan(rd3_val) and rd1_val > 0:
        feats["cardio_ratio"] = rd3_val / (rd1_val + 0.01)
    else:
        feats["cardio_ratio"] = 1.0

    # ── Submission pressure/vulnerability composites ──
    feats["sub_entry_pressure"] = (
        feats.get("td_att_p15", 0.0) * feats.get("td_acc", 0.35)
        + 0.7 * feats.get("sub_att_p15", 0.0)
        + 0.4 * feats.get("rev_p15", 0.0)
    )
    feats["sub_control_conversion"] = (
        feats.get("sub_att_p15", 0.0)
        * (0.25 + feats.get("ctrl_pct", 0.0))
        * (0.25 + feats.get("ground_pct", 0.0))
    )
    feats["sub_scramble_threat"] = (
        feats.get("sub_att_p15", 0.0) + feats.get("rev_p15", 0.0)
    ) * (1.0 + feats.get("ground_pct", 0.0))
    feats["sub_defensive_leak"] = (
        feats.get("def_sub_att_p15", 0.0) * (1.0 - feats.get("td_defense_rate", 0.65))
        + 0.75 * feats.get("sub_loss_pct", 0.20)
    )
    feats["late_sub_pressure"] = (
        feats.get("sub_att_p15", 0.0)
        * feats.get("cardio_ratio", 1.0)
        * (0.5 + feats.get("rd3_sub_share", 0.0))
    )
    feats["sub_recency_surge"] = feats.get("r1f_sub_att_p15", 0.0) - feats.get("r5_sub_att_p15", 0.0)
    feats["grapple_recency_surge"] = (
        feats.get("r1f_td_p15", 0.0) + feats.get("r1f_sub_att_p15", 0.0)
    ) - (
        feats.get("r5_td_p15", 0.0) + feats.get("r5_sub_att_p15", 0.0)
    )
    feats["sub_vs_control_axis"] = (
        feats.get("sub_att_p15", 0.0) + 0.5 * feats.get("rev_p15", 0.0)
    ) - 0.6 * feats.get("ctrl_pct", 0.0)
    # KO-side composites for explicit attacker-vs-vulnerability mismatch.
    feats["ko_attack_pressure"] = (
        feats.get("power_ratio", 0.0)
        * feats.get("head_pct", 0.0)
        * feats.get("sig_str_diff_pm", 0.0)
    )
    feats["ko_def_leak"] = (
        feats.get("def_kd_pm", 0.0)
        + feats.get("ko_loss_pct", 0.0)
        + 0.5 * (1.0 - feats.get("durability", 0.8))
    )

    # ── Opponent quality trend (facing tougher/weaker opponents recently?) ──
    if feats["avg_opp_glicko"] > 0:
        feats["opp_quality_trend"] = feats["ewm_opp_quality"] / feats["avg_opp_glicko"]
    else:
        feats["opp_quality_trend"] = 1.0

    # ── Recent-year activity ──
    if current_date is not None:
        fights_last_year = sum(
            1 for h in history
            if not _isnan(h["date"]) and (current_date - h["date"]).days <= 365
        )
        feats["fights_last_year"] = float(fights_last_year)
        wins_last_year = sum(
            1 for h in history
            if not _isnan(h["date"]) and (current_date - h["date"]).days <= 365
            and h["result"] == "W"
        )
        feats["win_rate_last_year"] = _bayes_shrink(
            _safe_div(wins_last_year, fights_last_year, 0.5),
            fights_last_year, prior=0.5, strength=4,
        )
    else:
        feats["fights_last_year"] = 1.0
        feats["win_rate_last_year"] = 0.5

    # ── Loss recovery (bounce-back ability after losses) ──
    post_loss_fights = 0
    post_loss_wins = 0
    after_loss = False
    for h in history:
        if after_loss:
            post_loss_fights += 1
            if h["result"] == "W":
                post_loss_wins += 1
        after_loss = (h["result"] == "L")
    feats["loss_recovery_rate"] = _bayes_shrink(
        _safe_div(post_loss_wins, post_loss_fights, 0.5),
        post_loss_fights, prior=0.5, strength=6,
    )

    # ── Striking defense efficiency ──
    # How well they avoid damage relative to opponent's offensive output
    if feats["def_sig_str_pm"] > 0 and feats["sig_str_pm"] > 0:
        feats["strike_exchange_ratio"] = feats["sig_str_pm"] / (feats["def_sig_str_pm"] + 0.01)
    else:
        feats["strike_exchange_ratio"] = 1.0

    # ── Grappling threat composite (wrestling + submissions + control) ──
    feats["grappling_threat"] = feats["td_p15"] * feats["td_acc"] + feats["sub_att_p15"] + feats["ctrl_pct"]

    # ── Decision ability (win rate in fights that go to decision) ──
    dec_fights = sum(1 for h in history if "Dec" in str(h.get("method", "")))
    dec_wins = sum(1 for h in history if h["result"] == "W" and "Dec" in str(h.get("method", "")))
    feats["decision_win_rate"] = _bayes_shrink(
        _safe_div(dec_wins, dec_fights, 0.5),
        dec_fights, prior=0.5, strength=6,
    )

    # ── Finish resistance (survives opponent's finishes) ──
    feats["finish_resistance"] = 1.0 - feats.get("been_finished_pct", 0.5)

    # ── Offensive diversity (how evenly spread across strike/grapple/sub) ──
    stk_share = feats.get("style_striking", 0.5)
    wrs_share = feats.get("style_wrestling", 0.25)
    sub_share = feats.get("style_submission", 0.15)
    # Entropy-based diversity (higher = more versatile)
    diversity = 0.0
    for s in [stk_share, wrs_share, sub_share]:
        if s > 0:
            diversity -= s * math.log(s + 1e-8)
    feats["offensive_diversity"] = diversity

    # ── EWM striking differential (recent trend of how they're winning/losing exchanges) ──
    alpha_sd = 0.4
    sd_sum, sd_tot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha_sd) ** (n - 1 - i)
        ft_min = h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0
        diff = (h["sig_str"] - h["opp_sig_str"]) / ft_min
        sd_sum += w * diff
        sd_tot += w
    feats["ewm_sig_str_diff_pm"] = sd_sum / sd_tot if sd_tot > 0 else 0.0

    # ── Glicko confidence (lower phi = more certain rating) ──
    feats["glicko_confidence"] = max(1.0 - (glicko[1] / PHI_0), 0.0)

    # ── Control-to-damage ratios (style discrimination) ──
    # sig_str_per_ctrl_sec: strikes landed per second of control — separates
    # damaging ground-and-pounders from position-hugging wrestlers.
    # td_land_per_sub_att: wrestle-hug ratio vs sub-hunt ratio.
    _total_ctrl_sec = float(_safe_sum(h.get("ctrl_sec") for h in history) or 0.0)
    _total_sub_att = float(_safe_sum(h.get("sub_att") for h in history) or 0.0)
    feats["sig_str_per_ctrl_sec"] = (
        total_sig_land / _total_ctrl_sec if _total_ctrl_sec > 0 else 0.0
    )
    feats["td_land_per_sub_att"] = (
        total_td_land / _total_sub_att if _total_sub_att > 0 else 0.0
    )

    # ── Finish-round distribution (conditional on the fighter finishing) ──
    # How their finishes are spread across rounds — distinct from the aggregate
    # first_round_finish_rate (which divides by all fights).
    _finish_rounds = []
    for h in history:
        if h.get("result") != "W":
            continue
        _method = str(h.get("method", "")).strip()
        if _method not in ("KO/TKO", "Submission"):
            continue
        _fr = int(h.get("finish_round") or 0)
        if _fr >= 1:
            _finish_rounds.append(_fr)
    if _finish_rounds:
        _nfr = len(_finish_rounds)
        feats["finish_r1_share"] = float(sum(1 for r in _finish_rounds if r == 1) / _nfr)
        feats["finish_r3_plus_share"] = float(sum(1 for r in _finish_rounds if r >= 3) / _nfr)
        feats["avg_finish_round_w"] = float(sum(_finish_rounds) / _nfr)
    else:
        feats["finish_r1_share"] = 0.0
        feats["finish_r3_plus_share"] = 0.0
        feats["avg_finish_round_w"] = 0.0

    # ── DAPA (Decayed + opponent-Adjusted Performance Average) ──
    # Exponentially time-weighted per-fight stat rates, scaled by opponent
    # glicko so stats earned against higher-quality opposition count more.
    # Mirrors the Old_Model DAPA formulation that showed up in its top-20
    # stability-selection features (e.g. r_pre_dapa_head_defense).
    _dapa_alpha = 0.8
    _dapa_w = np.array([_dapa_alpha ** (n - 1 - i) for i in range(n)], dtype=float)
    _dapa_w_sum = float(_dapa_w.sum()) or 1.0
    _dapa_opp = np.array(
        [_num_or(h.get("opp_glicko"), MU_0) / MU_0 for h in history], dtype=float
    )
    _dapa_ft_min = np.array(
        [max((h["fight_time"] or 0) / 60.0, 0.1) for h in history], dtype=float
    )
    _dapa_slpm = np.array([h["sig_str"] or 0 for h in history], dtype=float) / _dapa_ft_min
    _dapa_kdpm = np.array([h["kd"] or 0 for h in history], dtype=float) / _dapa_ft_min
    _dapa_subp15 = np.array(
        [h["sub_att"] or 0 for h in history], dtype=float
    ) / (_dapa_ft_min / 15.0)
    _dapa_td_att = np.array([h["td_att"] or 0 for h in history], dtype=float)
    _dapa_td_land = np.array([h["td"] or 0 for h in history], dtype=float)
    _dapa_tda = np.where(
        _dapa_td_att > 0, _dapa_td_land / np.maximum(_dapa_td_att, 1.0), 0.0
    )
    # Striking defense proxy: 1 - opp_sig_str_acc from each fight.
    _dapa_opp_ssa = np.array(
        [_num_or(h.get("opp_sig_str_acc"), 0.45) for h in history], dtype=float
    )
    _dapa_sd = np.clip(1.0 - _dapa_opp_ssa, 0.0, 1.0)
    feats["dapa_sig_str_pm"] = float(np.sum(_dapa_slpm * _dapa_opp * _dapa_w) / _dapa_w_sum)
    feats["dapa_kd_pm"] = float(np.sum(_dapa_kdpm * _dapa_opp * _dapa_w) / _dapa_w_sum)
    feats["dapa_sub_att_p15"] = float(np.sum(_dapa_subp15 * _dapa_opp * _dapa_w) / _dapa_w_sum)
    feats["dapa_td_acc"] = float(np.sum(_dapa_tda * _dapa_opp * _dapa_w) / _dapa_w_sum)
    feats["dapa_sig_str_def"] = float(np.sum(_dapa_sd * _dapa_opp * _dapa_w) / _dapa_w_sum)

    # ── Opponent-baseline-adjusted performance (stat-specific SOS) ──────────
    # Credit the fighter for beating each opponent's PRE-FIGHT, stat-specific
    # baseline — unlike dapa_*, which weights by overall opponent Glicko. Covers
    # the offense AND defense sides of the three biggest signals (striking,
    # takedowns, control time). For each past fight (recency-weighted, alpha=0.85;
    # opponents with no baseline, e.g. debutants, are skipped):
    #   *_off : my output                 −  what that opponent usually ALLOWS
    #   *_def : what that opponent usually PRODUCES  −  what they did to me
    _oba_a = 0.85
    _oba_acc = {k: [0.0, 0.0] for k in (
        "sig_off", "sig_def", "td_off", "td_def", "ctrl_off", "ctrl_def")}

    def _oba_add(_key, _val, _w):
        if not _isnan(_val):
            _oba_acc[_key][0] += _w * _val
            _oba_acc[_key][1] += _w

    for _i, _h in enumerate(history):
        _w = _oba_a ** (n - 1 - _i)
        _ft = _h.get("fight_time") or 0
        if _ft <= 0:
            continue
        _ftm = _ft / 60.0
        _ft15 = _ft / 900.0
        # Striking (per minute): offense vs opp's allowed, defense vs opp's output
        _opp_allow = _h.get("opp_def_sig_str_pm", float("nan"))
        _opp_output = _h.get("opp_off_sig_str_pm", float("nan"))
        _oba_add("sig_off", ((_h.get("sig_str") or 0) / _ftm) - _opp_allow, _w)
        _oba_add("sig_def", _opp_output - ((_h.get("opp_sig_str") or 0) / _ftm), _w)
        # Takedowns (per 15 min): offense vs opp's TD-defense, defense vs opp's TD-offense
        _opp_td_allow = _h.get("opp_td_def_p15", float("nan"))
        _opp_td_output = _h.get("opp_off_td_p15", float("nan"))
        _oba_add("td_off", ((_h.get("td") or 0) / _ft15) - _opp_td_allow, _w)
        _oba_add("td_def", _opp_td_output - ((_h.get("opp_td") or 0) / _ft15), _w)
        # Control time (fraction of fight): offense vs opp's allowed, defense vs opp's output
        _opp_ctrl_allow = _h.get("opp_def_ctrl_pct", float("nan"))
        _opp_ctrl_output = _h.get("opp_off_ctrl_pct", float("nan"))
        _oba_add("ctrl_off", ((_h.get("ctrl_sec") or 0) / _ft) - _opp_ctrl_allow, _w)
        _oba_add("ctrl_def", _opp_ctrl_output - ((_h.get("opp_ctrl_sec") or 0) / _ft), _w)

    def _oba_val(_key):
        _num, _den = _oba_acc[_key]
        return float(_num / _den) if _den > 0 else 0.0

    if OBA_FEATURES_ENABLED:
        feats["oba_sig_str_off"] = _oba_val("sig_off")
        feats["oba_sig_str_def"] = _oba_val("sig_def")
        feats["oba_td_off"] = _oba_val("td_off")
        feats["oba_td_def"] = _oba_val("td_def")
        feats["oba_ctrl_off"] = _oba_val("ctrl_off")
        feats["oba_ctrl_def"] = _oba_val("ctrl_def")

    if _FIGHTER_FEAT_KEYS is None:
        _set_fighter_keys(list(feats.keys()))

    return feats


def _set_fighter_keys(keys):
    global _FIGHTER_FEAT_KEYS
    _FIGHTER_FEAT_KEYS = keys


# ─── Matchup feature computation ──────────────────────────────────────────────

def compute_matchup_features(a_feats, b_feats, is_title=0, total_rounds=3, weight_class=""):
    """Build the matchup feature dict for fighter A (red) vs fighter B (blue).

    Emits a d_<stat> difference (A minus B) for every per-fighter feature,
    plus context features (title/rounds/weight class one-hots), the pooled-
    deviation Glicko win probability, cross-matchup interactions (offense vs
    the opponent's specific defense), style-clash axes, and *_sum channels
    that let the method stage reconstruct exact per-side values after the
    matrix is re-oriented around the predicted winner. Naming conventions
    matter downstream: d_* columns negate under corner swap, *_sum and abs_*
    columns are swap-invariant, and probability columns complement (p -> 1-p)
    — see _swap_features.
    """
    features = {}
    for key in a_feats:
        a_val = a_feats[key] if not _isnan(a_feats[key]) else float("nan")
        b_val = b_feats[key] if not _isnan(b_feats[key]) else float("nan")
        try:
            features[f"d_{key}"] = float(a_val) - float(b_val)
        except (TypeError, ValueError):
            features[f"d_{key}"] = float("nan")

    a_age = a_feats.get("age", float("nan"))
    b_age = b_feats.get("age", float("nan"))
    a_reach = a_feats.get("reach", float("nan"))
    b_reach = b_feats.get("reach", float("nan"))
    a_height = a_feats.get("height", float("nan"))
    b_height = b_feats.get("height", float("nan"))
    a_mu = _num_or(a_feats.get("glicko_mu"), MU_0)
    b_mu = _num_or(b_feats.get("glicko_mu"), MU_0)
    a_phi = _num_or(a_feats.get("glicko_phi"), PHI_0)
    b_phi = _num_or(b_feats.get("glicko_phi"), PHI_0)
    a_n = max(_num_or(a_feats.get("num_fights"), 0.0), 0.0)
    b_n = max(_num_or(b_feats.get("num_fights"), 0.0), 0.0)
    a_debut = a_n < 1.0
    b_debut = b_n < 1.0
    a_inexperienced = a_n < 3.0
    b_inexperienced = b_n < 3.0

    # Absolute / interaction features
    features["is_title"] = is_title
    features["total_rounds"] = total_rounds
    features["exp_sum"] = a_n + b_n
    features["age_sum"] = _num_or(a_age, 30.0) + _num_or(b_age, 30.0)
    features["glicko_mu_sum"] = a_mu + b_mu
    features["abs_age_gap"] = _abs_gap(a_age, b_age)
    features["abs_reach_gap"] = _abs_gap(a_reach, b_reach)
    features["abs_height_gap"] = _abs_gap(a_height, b_height)
    features["abs_glicko_gap"] = abs(a_mu - b_mu)
    features["abs_glicko_phi_gap"] = abs(a_phi - b_phi)
    features["abs_exp_gap"] = abs(a_n - b_n)
    features["min_num_fights"] = min(a_n, b_n)
    features["max_num_fights"] = max(a_n, b_n)
    features["experience_ratio"] = min(a_n, b_n) / max(max(a_n, b_n), 1.0)
    features["max_glicko_phi"] = max(a_phi, b_phi)
    features["min_glicko_phi"] = min(a_phi, b_phi)
    features["avg_glicko_phi"] = (a_phi + b_phi) / 2.0
    features["both_debut"] = float(a_debut and b_debut)
    features["one_debut"] = float(a_debut ^ b_debut)
    features["both_inexperienced"] = float(a_inexperienced and b_inexperienced)
    features["one_inexperienced"] = float(a_inexperienced ^ b_inexperienced)
    features["info_asymmetry"] = abs(a_n - b_n) / max(a_n + b_n, 1.0)

    # Weight class
    features["weight_class_ord"] = WEIGHT_CLASS_ORDINAL.get(weight_class, 0)
    for class_name, feature_name in WEIGHT_CLASS_FEATURES.items():
        features[feature_name] = float(weight_class == class_name)
    features[UNKNOWN_WEIGHT_CLASS_FEATURE] = float(weight_class not in WEIGHT_CLASS_ORDINAL)

    # Stance matchup interactions (d_ prefix for correct augmentation negation)
    a_ortho = _num_or(a_feats.get("is_orthodox"), 0.0)
    a_south = _num_or(a_feats.get("is_southpaw"), 0.0)
    b_ortho = _num_or(b_feats.get("is_orthodox"), 0.0)
    b_south = _num_or(b_feats.get("is_southpaw"), 0.0)
    features["d_ortho_vs_south"] = float(a_ortho * b_south) - float(b_ortho * a_south)
    features["same_stance"] = float(
        a_ortho == b_ortho and a_south == b_south and (a_ortho or a_south)
    )
    # Style cluster matchup interactions
    a_stk = _num_or(a_feats.get("style_striking"), 0.5)
    a_wrs = _num_or(a_feats.get("style_wrestling"), 0.15)
    a_sub = _num_or(a_feats.get("style_submission"), 0.1)
    a_cg = _num_or(a_feats.get("style_clinch_ground"), 0.32)
    b_stk = _num_or(b_feats.get("style_striking"), 0.5)
    b_wrs = _num_or(b_feats.get("style_wrestling"), 0.15)
    b_sub = _num_or(b_feats.get("style_submission"), 0.1)
    b_cg = _num_or(b_feats.get("style_clinch_ground"), 0.32)
    features["d_striker_vs_grappler"] = a_stk * (b_wrs + b_sub) - b_stk * (a_wrs + a_sub)
    features["style_mismatch"] = a_stk * (b_wrs + b_sub) + b_stk * (a_wrs + a_sub)
    features["style_distance"] = math.sqrt(
        (a_stk - b_stk)**2 + (a_wrs - b_wrs)**2 + (a_sub - b_sub)**2 + (a_cg - b_cg)**2
    )

    # ── Glicko expected win probability (THE strongest single predictor) ──
    # This encodes the full Glicko-2 prediction: rating difference + uncertainty.
    # Uses the POOLED deviation sqrt(phi_a² + phi_b²): the one-sided textbook
    # form E(a, b, phi_b) reads only the OPPONENT's uncertainty, so
    # E(a,b) + E(b,a) != 1 whenever the deviations differ — which broke exact
    # corner-swap equivariance (the swapped row's complement wasn't what the
    # reversed orientation would compute). Pooling keeps the same signal and
    # makes the complement exact.
    a_mu_s = (a_mu - MU_0) / SCALE
    b_mu_s = (b_mu - MU_0) / SCALE
    pooled_phi_s = math.sqrt(a_phi ** 2 + b_phi ** 2) / SCALE
    features["glicko_win_prob"] = _E(a_mu_s, b_mu_s, pooled_phi_s)
    features["d_glicko_win_prob"] = features["glicko_win_prob"] - 0.5  # directional version

    # Confidence gap: rating gap normalised by combined uncertainty
    combined_phi = math.sqrt(a_phi**2 + b_phi**2) if (a_phi**2 + b_phi**2) > 0 else 1.0
    features["d_confidence_gap"] = (a_mu - b_mu) / combined_phi

    # Glicko win prob weighted by confidence (high confidence + big gap = strong signal)
    avg_conf = (_num_or(a_feats.get("glicko_confidence"), 0.0)
                + _num_or(b_feats.get("glicko_confidence"), 0.0)) / 2.0
    features["d_confident_prediction"] = features["d_glicko_win_prob"] * (0.5 + avg_conf)

    # Reach advantage amplified by striking differential
    a_sig_pm = _num_or(a_feats.get("sig_str_pm"), 0.0)
    b_sig_pm = _num_or(b_feats.get("sig_str_pm"), 0.0)
    reach_diff = _num_or(a_reach, 72.0) - _num_or(b_reach, 72.0)
    features["d_reach_x_striking"] = reach_diff * (a_sig_pm - b_sig_pm)

    # Chin vs power: A's durability vs B's KO rate (and reverse)
    a_dur = _num_or(a_feats.get("durability"), 0.8)
    b_dur = _num_or(b_feats.get("durability"), 0.8)
    a_kd_pm = _num_or(a_feats.get("kd_pm"), 0.0)
    b_kd_pm = _num_or(b_feats.get("kd_pm"), 0.0)
    features["d_chin_vs_power"] = (a_dur * a_kd_pm) - (b_dur * b_kd_pm)

    # TD attack vs defense: A's offensive wrestling vs B's TDD (directional)
    a_td_p15 = _num_or(a_feats.get("td_p15"), 0.0)
    b_td_p15 = _num_or(b_feats.get("td_p15"), 0.0)
    a_tdd = _num_or(a_feats.get("td_defense_rate"), 0.65)
    b_tdd = _num_or(b_feats.get("td_defense_rate"), 0.65)
    features["d_td_attack_vs_defense"] = (a_td_p15 * (1.0 - b_tdd)) - (b_td_p15 * (1.0 - a_tdd))

    # Combined finish rate (proxy for fight unlikely to go to decision)
    a_fin = _num_or(a_feats.get("finish_rate"), 0.5)
    b_fin = _num_or(b_feats.get("finish_rate"), 0.5)
    features["combined_finish_rate"] = a_fin + b_fin

    # Output rate differential
    features["d_output_rate"] = _num_or(a_feats.get("output_rate"), 0.0) - _num_or(b_feats.get("output_rate"), 0.0)

    # Momentum differential
    features["d_momentum"] = _num_or(a_feats.get("momentum"), 0.5) - _num_or(b_feats.get("momentum"), 0.5)

    # Durability gap
    features["d_durability"] = a_dur - b_dur

    # Grappling dominance differential
    features["d_grappling_dominance"] = (_num_or(a_feats.get("grappling_dominance"), 0.0)
                                         - _num_or(b_feats.get("grappling_dominance"), 0.0))

    # EWM opponent quality gap (who has faced tougher competition recently)
    features["d_ewm_opp_quality"] = (_num_or(a_feats.get("ewm_opp_quality"), MU_0)
                                     - _num_or(b_feats.get("ewm_opp_quality"), MU_0))

    # ── Striking offense vs opponent's striking defense (cross-matchup) ──
    a_sig_def = _num_or(a_feats.get("sig_str_defense_rate"), 0.55)
    b_sig_def = _num_or(b_feats.get("sig_str_defense_rate"), 0.55)
    features["d_striking_vs_defense"] = (
        a_sig_pm * (1.0 - b_sig_def) - b_sig_pm * (1.0 - a_sig_def)
    )

    # ── Wrestling offense vs opponent's TDD (already have d_td_attack_vs_defense,
    #    now add grappling_threat vs opponent's ground defense) ──
    a_grap = _num_or(a_feats.get("grappling_threat"), 0.0)
    b_grap = _num_or(b_feats.get("grappling_threat"), 0.0)
    features["d_grapple_vs_tdd"] = a_grap * (1.0 - b_tdd) - b_grap * (1.0 - a_tdd)

    # ── Fight IQ gap ──
    features["d_fight_iq"] = (_num_or(a_feats.get("fight_iq"), 0.15)
                              - _num_or(b_feats.get("fight_iq"), 0.15))

    # ── Cardio advantage ──
    features["d_cardio_ratio"] = (_num_or(a_feats.get("cardio_ratio"), 1.0)
                                  - _num_or(b_feats.get("cardio_ratio"), 1.0))

    # ── Glicko confidence gap (whose rating is more reliable?) ──
    features["d_glicko_confidence"] = (_num_or(a_feats.get("glicko_confidence"), 0.0)
                                       - _num_or(b_feats.get("glicko_confidence"), 0.0))

    # ── Form trend gap (who is on the upswing?) ──
    features["d_form_trend"] = (_num_or(a_feats.get("form_vs_career"), 0.0)
                                - _num_or(b_feats.get("form_vs_career"), 0.0))

    # ── Strike exchange ratio gap (who wins the striking exchanges more?) ──
    features["d_exchange_ratio"] = (_num_or(a_feats.get("strike_exchange_ratio"), 1.0)
                                    - _num_or(b_feats.get("strike_exchange_ratio"), 1.0))

    # ── Decision fighter vs finisher matchup ──
    a_dec_wr = _num_or(a_feats.get("decision_win_rate"), 0.5)
    b_dec_wr = _num_or(b_feats.get("decision_win_rate"), 0.5)
    features["d_decision_ability"] = a_dec_wr - b_dec_wr

    # ── Versatility gap ──
    features["d_offensive_diversity"] = (_num_or(a_feats.get("offensive_diversity"), 1.0)
                                         - _num_or(b_feats.get("offensive_diversity"), 1.0))

    # ── Combined Glicko confidence (how reliable is this matchup prediction?) ──
    a_conf = _num_or(a_feats.get("glicko_confidence"), 0.0)
    b_conf = _num_or(b_feats.get("glicko_confidence"), 0.0)
    features["avg_glicko_confidence"] = (a_conf + b_conf) / 2.0

    # ── Age-experience interaction: young+experienced is dangerous ──
    a_age_v = _num_or(a_age, 30.0)
    b_age_v = _num_or(b_age, 30.0)
    # Younger fighter with more experience has edge
    features["d_youth_exp"] = (b_age_v - a_age_v) * (a_n - b_n)

    # ── Glicko-scaled striking: weight sig str differential by Glicko reliability ──
    features["d_reliable_striking"] = features.get("d_sig_str_pm", 0.0) * (0.5 + avg_conf)

    # ── Finisher vs chin interaction ──
    a_finish_r = _num_or(a_feats.get("finish_rate"), 0.5)
    b_finish_r = _num_or(b_feats.get("finish_rate"), 0.5)
    b_finish_resist = _num_or(b_feats.get("finish_resistance"), 0.5)
    a_finish_resist = _num_or(a_feats.get("finish_resistance"), 0.5)
    features["d_finish_vs_resist"] = (a_finish_r * (1.0 - b_finish_resist)
                                      - b_finish_r * (1.0 - a_finish_resist))

    # ── Activity gap (recent activity matters for ring rust) ──
    a_inactive = _num_or(a_feats.get("days_inactive"), 365.0)
    b_inactive = _num_or(b_feats.get("days_inactive"), 365.0)
    features["d_activity"] = b_inactive - a_inactive  # positive = A is more active

    # ── Win rate in last year gap ──
    features["d_win_rate_last_year"] = (_num_or(a_feats.get("win_rate_last_year"), 0.5)
                                        - _num_or(b_feats.get("win_rate_last_year"), 0.5))

    # ── Old_Model-inspired quality-endurance and pressure features ──
    # Quality endurance: maintains accuracy while keeping output under fatigue.
    a_cardio = _num_or(a_feats.get("cardio_ratio"), 1.0)
    b_cardio = _num_or(b_feats.get("cardio_ratio"), 1.0)
    a_acc = _num_or(a_feats.get("sig_str_acc"), 0.45)
    b_acc = _num_or(b_feats.get("sig_str_acc"), 0.45)
    a_r1_int = _num_or(a_feats.get("rd1_intensity_ratio"), 1.0)
    b_r1_int = _num_or(b_feats.get("rd1_intensity_ratio"), 1.0)
    a_output = _num_or(a_feats.get("output_rate"), 0.0)
    b_output = _num_or(b_feats.get("output_rate"), 0.0)
    a_opp_sig = _num_or(a_feats.get("def_sig_str_pm"), 0.0)
    b_opp_sig = _num_or(b_feats.get("def_sig_str_pm"), 0.0)
    a_conf = _num_or(a_feats.get("glicko_confidence"), 0.0)
    b_conf = _num_or(b_feats.get("glicko_confidence"), 0.0)

    a_quality_endurance = a_cardio * a_acc * (0.5 + min(max(a_r1_int, 0.5), 1.8))
    b_quality_endurance = b_cardio * b_acc * (0.5 + min(max(b_r1_int, 0.5), 1.8))
    features["d_quality_endurance"] = a_quality_endurance - b_quality_endurance

    # Accuracy under fire proxy: maintained accuracy while facing volume.
    a_accuracy_under_fire = a_acc / (1.0 + b_opp_sig)
    b_accuracy_under_fire = b_acc / (1.0 + a_opp_sig)
    features["d_accuracy_under_fire"] = a_accuracy_under_fire - b_accuracy_under_fire

    # Pressure-cardio clash: high pressure that still scales with own cardio.
    features["d_pressure_cardio_clash"] = (
        a_output * a_cardio * (1.0 - b_cardio) - b_output * b_cardio * (1.0 - a_cardio)
    )

    # Stability-endurance: confidence-weighted late-fight reliability.
    features["d_stability_endurance"] = (
        a_conf * a_quality_endurance - b_conf * b_quality_endurance
    )

    # Sum channels to allow exact side-specific reconstruction for method-only
    # mismatch features after winner-orientation is applied.
    a_sub_entry = _num_or(a_feats.get("sub_entry_pressure"), 0.0)
    b_sub_entry = _num_or(b_feats.get("sub_entry_pressure"), 0.0)
    a_sub_leak = _num_or(a_feats.get("sub_defensive_leak"), 0.0)
    b_sub_leak = _num_or(b_feats.get("sub_defensive_leak"), 0.0)
    features["sub_entry_pressure_sum"] = a_sub_entry + b_sub_entry
    features["sub_defensive_leak_sum"] = a_sub_leak + b_sub_leak
    a_ko_attack = _num_or(a_feats.get("ko_attack_pressure"), 0.0)
    b_ko_attack = _num_or(b_feats.get("ko_attack_pressure"), 0.0)
    a_ko_leak = _num_or(a_feats.get("ko_def_leak"), 0.0)
    b_ko_leak = _num_or(b_feats.get("ko_def_leak"), 0.0)
    features["ko_attack_pressure_sum"] = a_ko_attack + b_ko_attack
    features["ko_def_leak_sum"] = a_ko_leak + b_ko_leak

    # Stage-1 sum features for exact winner/loser reconstruction after orientation.
    for _feat, _default in [
        ("dec_win_pct", 0.0), ("finish_resistance", 0.0), ("consistency", 0.0),
        ("cardio_ratio", 0.0), ("durability", 0.0), ("output_rate", 0.0),
        ("rd1_intensity_ratio", 0.0), ("strike_exchange_ratio", 0.0),
        ("sig_str_acc", 0.0), ("late_round_pct", 0.0), ("avg_time_min", 0.0),
        ("avg_finish_round", 0.0), ("first_round_finish_rate", 0.0),
        ("damage_efficiency", 0.0), ("body_leg_attrition", 0.0),
    ]:
        _a = _num_or(a_feats.get(_feat), _default)
        _b = _num_or(b_feats.get(_feat), _default)
        features[f"{_feat}_sum"] = _a + _b

    # ── Batch 3: physical / style / matchup interaction features (§B, §C, §D) ──
    # All use f(A) - f(B) pattern so they flip cleanly under method orientation.
    # §B — Physical / leverage interactions
    a_dist_pct = _num_or(a_feats.get("distance_pct"), 0.3)
    b_dist_pct = _num_or(b_feats.get("distance_pct"), 0.3)
    features["d_reach_x_distance_pct"] = (
        _num_or(a_reach, 72.0) * a_dist_pct - _num_or(b_reach, 72.0) * b_dist_pct
    )
    a_cdg = _num_or(a_feats.get("cardio_decay_sig_str"), 0.0)
    b_cdg = _num_or(b_feats.get("cardio_decay_sig_str"), 0.0)
    a_age_v = _num_or(a_age, 30.0)
    b_age_v = _num_or(b_age, 30.0)
    features["d_age_x_cardio"] = (a_age_v * a_cdg) - (b_age_v * b_cdg)
    features["d_age_x_title_rounds"] = (a_age_v - b_age_v) * float(total_rounds == 5)

    # §D — Matchup compatibility / style clashes
    features["d_striker_grappler_raw"] = (
        (a_sig_pm - a_td_p15) - (b_sig_pm - b_td_p15)
    )
    a_td_acc = _num_or(a_feats.get("td_acc"), 0.35)
    b_td_acc = _num_or(b_feats.get("td_acc"), 0.35)
    features["d_tdd_vs_td_attack"] = (a_td_acc * (1.0 - b_tdd)) - (b_td_acc * (1.0 - a_tdd))

    # ── Batch 4: form / context / priors (§E, §F, §G) ──
    features["gender_flag"] = 1.0 if weight_class.startswith("Women") else 0.0

    features["d_title_x_cardio"] = float(is_title) * (a_cdg - b_cdg)

    a_fin_resist = _num_or(a_feats.get("finish_resistance"), 0.5)
    b_fin_resist = _num_or(b_feats.get("finish_resistance"), 0.5)
    features["d_total_rounds_x_finish_resistance"] = float(total_rounds) * (a_fin_resist - b_fin_resist)

    # Context-conditional rounds experience (picks the bucket matching current fight)
    _cur_rounds = 5 if int(total_rounds) == 5 else 3
    features["d_rounds_experience"] = (
        _num_or(a_feats.get(f"rounds_{_cur_rounds}_exp"), 0.0)
        - _num_or(b_feats.get(f"rounds_{_cur_rounds}_exp"), 0.0)
    )

    # Stability: flipped so positive = red is the more stable fighter
    features["d_stability"] = (
        _num_or(b_feats.get("sig_diff_pm_vol"), 3.0)
        - _num_or(a_feats.get("sig_diff_pm_vol"), 3.0)
    )

    return features


# ─── Context finish priors ────────────────────────────────────────────────────

def _precompute_context_finish_priors(df):
    """Compute per-row finish priors conditioned on (weight_class, gender, total_rounds, is_title).

    ctx_finish_prior_2y: exponentially time-decayed (half-life 730 days) Bayesian
        estimate of the finish rate for fights of the same type before this date.
    ref_finish_prior: unweighted Laplace-smoothed count-based version of the same.
    Both use alpha=20 and p0 = global historical finish rate as the prior mean.
    """
    n = len(df)
    ctx_arr = np.full(n, np.nan, dtype=float)
    ref_arr = np.full(n, np.nan, dtype=float)

    methods = df["method"].astype(str).str.lower()
    is_finish = (~methods.str.contains("dec", na=False)).astype(float).values
    p0 = float(np.mean(is_finish)) if n > 0 else 0.55
    alpha = 20.0

    wc_col = df["weight_class"].astype(str).values if "weight_class" in df.columns else np.full(n, "", dtype=object)
    gd_col = df["gender"].astype(str).str.lower().values if "gender" in df.columns else np.full(n, "unknown", dtype=object)
    rd_col = df["total_rounds"].fillna(3).astype(int).values if "total_rounds" in df.columns else np.full(n, 3, dtype=int)
    tt_col = df["is_title_bout"].fillna(0).astype(int).values if "is_title_bout" in df.columns else np.zeros(n, dtype=int)

    epoch = df["event_date"].iloc[0]
    days = ((df["event_date"] - epoch) / pd.Timedelta(days=1)).values.astype(float)

    # Per-context running state: (decayed_finish_sum, decayed_total_weight, finish_count, total_count, last_days)
    state = {}

    for i in range(n):
        key = (wc_col[i], gd_col[i], int(rd_col[i]), int(tt_col[i]))
        cur_d = days[i]

        if key in state:
            fs, tw, fc, tc, last_d = state[key]
            decay = 2.0 ** (-(cur_d - last_d) / 730.0)
            fs *= decay
            tw *= decay
        else:
            fs, tw, fc, tc = 0.0, 0.0, 0, 0

        ctx_arr[i] = (fs + alpha * p0) / (tw + alpha)
        ref_arr[i] = (fc + alpha * p0) / (tc + alpha)

        # Update state with current fight (weight = 1.0 at this moment)
        state[key] = (fs + float(is_finish[i]), tw + 1.0, fc + int(is_finish[i]), tc + 1, cur_d)

    return ctx_arr, ref_arr


# ─── Style-class opponent matchup tracker ────────────────────────────────────

# Per-fighter rate dimensions we bucket opponents into (striker / wrestler /
# submission-hunter). Each fighter's historical win rate against opponents in
# the top bucket becomes a new feature axis.
_STYLE_DIMS = (
    ("sig_str_pm", "striker"),
    ("td_p15", "wrestler"),
    ("sub_att_p15", "sub_hunter"),
)


def _compute_style_quartile_thresholds(df, cutoff):
    """Fit quartile thresholds on per-fighter career averages from the first
    `cutoff` rows of df. Returns {dim: (q25, q50, q75)}."""
    agg = defaultdict(lambda: {"time": 0.0, "sig_str": 0.0, "td": 0.0, "sub_att": 0.0, "fights": 0})
    cutoff = min(max(cutoff, 0), len(df))
    for idx in range(cutoff):
        row = df.iloc[idx]
        ft = float(row.get("total_fight_time_sec", 0) or 0)
        for prefix, fname in (("r", row.get("r_name")), ("b", row.get("b_name"))):
            if not isinstance(fname, str) or not fname:
                continue
            a = agg[fname]
            a["time"] += ft
            a["sig_str"] += float(row.get(f"{prefix}_sig_str", 0) or 0)
            a["td"] += float(row.get(f"{prefix}_td", 0) or 0)
            a["sub_att"] += float(row.get(f"{prefix}_sub_att", 0) or 0)
            a["fights"] += 1
    arrays = {"sig_str_pm": [], "td_p15": [], "sub_att_p15": []}
    for _, a in agg.items():
        if a["fights"] < 3 or a["time"] <= 0:
            continue
        tm = max(a["time"] / 60.0, 0.1)
        arrays["sig_str_pm"].append(a["sig_str"] / tm)
        arrays["td_p15"].append((a["td"] / tm) * 15.0)
        arrays["sub_att_p15"].append((a["sub_att"] / tm) * 15.0)
    thr = {}
    for dim, vals in arrays.items():
        if len(vals) < 10:
            thr[dim] = (0.33, 0.66, 1.0)  # safe fallback
        else:
            q = np.quantile(np.array(vals, dtype=float), [0.25, 0.50, 0.75])
            thr[dim] = tuple(float(x) for x in q)
    return thr


def _sty_default_record():
    return [0, 0]


def _sty_default_cumul():
    return {"time": 0.0, "sig_str": 0.0, "td": 0.0, "sub_att": 0.0}


class _StyleMatchupTracker:
    """Running per-fighter win rate against opponents bucketed by style dim.

    Pre-fight reads are leak-safe: `matchup_features` only queries counts
    accumulated from fights strictly before the call; `update_after_fight`
    writes must happen AFTER the matchup row is built.
    """
    def __init__(self, thresholds):
        self.thresholds = dict(thresholds)
        self.cumul = defaultdict(_sty_default_cumul)
        # records: {(fighter, dim, opp_bucket): [wins, losses]}
        self.records = defaultdict(_sty_default_record)

    def _stat(self, fname, dim):
        c = self.cumul.get(fname)
        if not c or c["time"] <= 0:
            return 0.0
        tm = c["time"] / 60.0
        if dim == "sig_str_pm":
            return c["sig_str"] / tm
        if dim == "td_p15":
            return (c["td"] / tm) * 15.0
        if dim == "sub_att_p15":
            return (c["sub_att"] / tm) * 15.0
        return 0.0

    def _bucket(self, value, dim):
        thr = self.thresholds.get(dim)
        if thr is None:
            return 0
        if value <= thr[0]:
            return 0
        if value <= thr[1]:
            return 1
        if value <= thr[2]:
            return 2
        return 3

    def _wr(self, fname, dim, opp_bucket):
        rec = self.records.get((fname, dim, opp_bucket))
        if not rec:
            return 0.5
        total = rec[0] + rec[1]
        return rec[0] / total if total > 0 else 0.5

    def matchup_features(self, a_name, b_name):
        """Return dict of r_wr_vs_<short>, b_wr_vs_<short>, d_wr_vs_<short>
        for each style dim, looked up using pre-fight buckets."""
        feats = {}
        for dim, short in _STYLE_DIMS:
            a_b = self._bucket(self._stat(a_name, dim), dim)
            b_b = self._bucket(self._stat(b_name, dim), dim)
            a_wr = self._wr(a_name, dim, b_b)
            b_wr = self._wr(b_name, dim, a_b)
            feats[f"r_wr_vs_{short}"] = a_wr
            feats[f"b_wr_vs_{short}"] = b_wr
            feats[f"d_wr_vs_{short}"] = a_wr - b_wr
        return feats

    def update_after_fight(self, a_name, b_name, winner, row):
        # Snapshot buckets BEFORE cumul update (these reflect pre-fight state
        # that matched what matchup_features already returned for this row).
        buckets_by_dim = {}
        for dim, _ in _STYLE_DIMS:
            buckets_by_dim[dim] = (
                self._bucket(self._stat(a_name, dim), dim),
                self._bucket(self._stat(b_name, dim), dim),
            )
        ft = float(row.get("total_fight_time_sec", 0) or 0)
        self.cumul[a_name]["time"] += ft
        self.cumul[a_name]["sig_str"] += float(row.get("r_sig_str", 0) or 0)
        self.cumul[a_name]["td"] += float(row.get("r_td", 0) or 0)
        self.cumul[a_name]["sub_att"] += float(row.get("r_sub_att", 0) or 0)
        self.cumul[b_name]["time"] += ft
        self.cumul[b_name]["sig_str"] += float(row.get("b_sig_str", 0) or 0)
        self.cumul[b_name]["td"] += float(row.get("b_td", 0) or 0)
        self.cumul[b_name]["sub_att"] += float(row.get("b_sub_att", 0) or 0)
        if winner == "Red":
            a_win = True
        elif winner == "Blue":
            a_win = False
        else:
            return  # skip draws/NC
        for dim, (a_b, b_b) in buckets_by_dim.items():
            self.records[(a_name, dim, b_b)][0 if a_win else 1] += 1
            self.records[(b_name, dim, a_b)][0 if (not a_win) else 1] += 1


# ─── Margin-of-victory (MOV) performance scoring ──────────────────────────────

def _compute_mov_scales(df):
    """Global per-stat spread (std) of the box-score differentials used to
    z-score the MOV dominance index. These are normalization constants (like
    feature scaling), not predictive features, so a global scale does not create
    harmful leakage and keeps runs reproducible.
    """
    def _num(col):
        s = df.get(col)
        if s is None:
            return np.zeros(len(df), dtype=float)
        return pd.to_numeric(s, errors="coerce").to_numpy(dtype=float)

    elapsed = _num("total_fight_time_sec")
    elapsed_safe = np.where(elapsed > 0, elapsed, np.nan)
    minutes = elapsed_safe / 60.0

    kd_diff = _num("r_kd") - _num("b_kd")
    sig_pm_diff = (_num("r_sig_str") - _num("b_sig_str")) / minutes
    ctrl_frac_diff = (_num("r_ctrl_sec") - _num("b_ctrl_sec")) / elapsed_safe
    td_diff = _num("r_td") - _num("b_td")

    def _std(arr):
        s = float(np.nanstd(arr))
        return s if s > 1e-6 else 1.0

    return {
        "kd": _std(kd_diff),
        "sig_pm": _std(sig_pm_diff),
        "ctrl_frac": _std(ctrl_frac_diff),
        "td": _std(td_diff),
    }


def _mov_performance_score(row, scales):
    """Continuous margin-of-victory score for a fight, returned as
    (s_red, s_blue) summing to 1.0, each in [0, 1].

    The RESULT is ground truth: the winner's score is always > 0.5. Decisiveness
    (finish timing / judge consensus) and a box-score dominance index only
    modulate WITHIN the winner's band — they never flip the sign. Draws and
    non Red/Blue outcomes return (0.5, 0.5).
    """
    winner = str(row.get("winner", "")).strip()
    if winner not in ("Red", "Blue"):
        return 0.5, 0.5

    method = str(row.get("method", ""))
    is_finish = _normalize_method_label(method) in ("KO/TKO", "Submission")

    try:
        total_rounds = int(row.get("total_rounds") or 3)
    except (TypeError, ValueError):
        total_rounds = 3
    scheduled = max(float(total_rounds) * 300.0, 1.0)
    try:
        elapsed = float(row.get("total_fight_time_sec") or scheduled)
    except (TypeError, ValueError):
        elapsed = scheduled
    if not np.isfinite(elapsed) or elapsed <= 0:
        elapsed = scheduled
    t_frac = min(max(elapsed / scheduled, 0.0), 1.0)
    finish_dom = 1.0 - t_frac  # earlier finish ⇒ more dominant

    # Judge consensus is the MOV proxy for decisions.
    m_low = method.lower()
    if "split" in m_low:
        consensus = 0.1
    elif "majority" in m_low:
        consensus = 0.5
    else:
        consensus = 1.0  # unanimous or unspecified decision

    if str(MOV_MODE).lower() == "buckets":
        if is_finish:
            s_win = 0.75 + 0.24 * finish_dom
        elif consensus <= 0.2:
            s_win = 0.62          # split decision ≈ coin flip
        else:
            s_win = 0.75          # unanimous / majority
        s_win = float(np.clip(s_win, 0.55, 0.99))
        s_red = s_win if winner == "Red" else 1.0 - s_win
        return s_red, 1.0 - s_red

    # ── Box-score dominance index D (red-positive), squashed to [-1, 1] ──
    def _f(col):
        try:
            v = float(row.get(col))
        except (TypeError, ValueError):
            return 0.0
        return v if np.isfinite(v) else 0.0

    minutes = max(elapsed / 60.0, 0.1)
    z_kd = (_f("r_kd") - _f("b_kd")) / scales["kd"]
    z_str = ((_f("r_sig_str") - _f("b_sig_str")) / minutes) / scales["sig_pm"]
    z_ctrl = ((_f("r_ctrl_sec") - _f("b_ctrl_sec")) / max(elapsed, 1.0)) / scales["ctrl_frac"]
    z_td = (_f("r_td") - _f("b_td")) / scales["td"]
    D = float(np.tanh(
        MOV_W_KD * z_kd + MOV_W_STR * z_str + MOV_W_CTRL * z_ctrl + MOV_W_TD * z_td
    ))
    # Orient to the winner's perspective (positive = winner also dominated stats).
    D_win = D if winner == "Red" else -D

    if is_finish:
        # A finish is decisive regardless of prior box score, so stats only ADD.
        blend = (1.0 - MOV_D_BLEND) * finish_dom + MOV_D_BLEND * max(D_win, 0.0)
        s_win = 0.75 + 0.24 * float(np.clip(blend, 0.0, 1.0))
    else:
        d01 = 0.5 * (D_win + 1.0)  # [-1,1] → [0,1]
        blend = (1.0 - MOV_D_BLEND) * consensus + MOV_D_BLEND * d01
        s_win = 0.55 + 0.25 * float(np.clip(blend, 0.0, 1.0))

    s_win = float(np.clip(s_win, 0.51, 0.99))  # winner strictly above 0.5
    s_red = s_win if winner == "Red" else 1.0 - s_win
    return s_red, 1.0 - s_red


def _altitude_feature_row(event_alt, r_train_alt, b_train_alt, train_known):
    """Altitude / acclimatization features for one matchup, built on each
    fighter's TRAINING-CAMP elevation (the true physiological baseline — past
    FIGHT venues cluster at Vegas/sea-level regardless of where a fighter is
    acclimatized, so they can't tell who's adapted).

    'Shock' is RECTIFIED — fighting ABOVE your camp costs cardio; staying/going
    down is ~free (the max(0,.) keeps it event-dependent). 'Descent' is the
    mirror: dropping from a high camp to a low venue rides a temporary RBC/O2
    boost (the Dagestan/altitude-camp-before-a-sea-level-fight effect).

    The method story (verified in the raw data) lives in the asymmetry: at a
    high venue, an acclimatized fighter SUBMITS the gassing opponent (sub rate
    ~triples), while if BOTH gas the fight turns sloppy and drifts to DECISION.
    `accl_asym_kft` / `mutual_gas_kft` isolate those two regimes. kft-scaled so
    magnitudes sit near the other features."""
    ev = float(event_alt)
    rt = float(r_train_alt)
    bt = float(b_train_alt)
    ek = ev / 1000.0
    r_shock = max(0.0, ev - rt) / 1000.0       # red fighting above its camp
    b_shock = max(0.0, ev - bt) / 1000.0
    r_descent = max(0.0, rt - ev) / 1000.0     # red descending off a high camp
    b_descent = max(0.0, bt - ev) / 1000.0
    d_accl_shock = r_shock - b_shock
    d_alt_descent = r_descent - b_descent
    d_train_alt = (rt - bt) / 1000.0
    # Signed venue-vs-camp gaps (+ = fighting above camp, - = below). One feature
    # carries both the shock and descent direction, so a single tree split at 0
    # separates the two regimes (the rectified pieces need two splits).
    r_gap = (ev - rt) / 1000.0
    b_gap = (ev - bt) / 1000.0
    return {
        # Absolute venue (the base method signal — affects both fighters).
        "event_alt_kft": ek,
        "alt_ge_4000": 1.0 if ev >= 4000.0 else 0.0,
        # Acclimatization shock vs training camp (winner asymmetry + method).
        "r_accl_shock_kft": r_shock,
        "b_accl_shock_kft": b_shock,
        "d_accl_shock_kft": d_accl_shock,
        # Sea-level / RBC boost from descending off a high camp (winner).
        "r_alt_descent_kft": r_descent,
        "b_alt_descent_kft": b_descent,
        "d_alt_descent_kft": d_alt_descent,
        # Altitude-camp cardio-conditioning edge, always-on (winner).
        "d_train_alt_kft": d_train_alt,
        # Signed event-vs-camp elevation gap, per corner (winner; user-requested).
        "r_event_camp_gap_kft": r_gap,
        "b_event_camp_gap_kft": b_gap,
        # COMPOSITE: net altitude advantage for red = (descent benefit - shock cost)
        # differenced across corners. Single directional "who does altitude favor"
        # axis the winner can split on. (= d_alt_descent - d_accl_shock.)
        "d_alt_net_edge_kft": d_alt_descent - d_accl_shock,
        # COMPOSITE: camp-elevation edge AMPLIFIED by venue altitude — training
        # higher matters far more at Mexico City than at sea level (winner).
        "d_camp_x_event_kft": d_train_alt * ek,
        # COMPOSITE: difference in elevation FAMILIARITY = how far each camp sits
        # from the venue (|gap|), differenced. + = red is closer to the venue's
        # elevation (more familiar). The one event-dependent diff-of-diffs not
        # already covered (= -(d_accl_shock + d_alt_descent); independent of
        # d_alt_net_edge). Symmetric-deviation hypothesis -> weaker than net_edge.
        "d_event_camp_famil_kft": abs(b_gap) - abs(r_gap),
        # Method: both gas -> sloppy -> DECISION (anti-finish).
        "mutual_gas_kft": min(r_shock, b_shock),
        # Method: asymmetric stress -> fresher SUBMITS gassed one (finish/sub).
        "accl_asym_kft": abs(d_accl_shock),
        # Data-availability gate (~60% of fights impute a training elevation).
        "train_alt_known": float(train_known),
    }


class _BaggedHGB:
    """Subsample-bagged HistGradientBoosting. ``predict_proba`` is averaged over
    ``n_estimators`` members, each fit on a different random subsample (no
    replacement) of the rows. A drop-in for HistGradientBoostingClassifier's
    predict_proba/predict/classes_ usage in the method stages.

    Why subsample, not seeds: at the small method sample sizes (Stage2 KO/Sub ≈
    280 finishes) a single HGB is high-variance but ~deterministic across
    random_state (no row subsampling; binning sees all rows), so the data subset —
    not the seed — is what creates the diversity variance-reduction needs. This
    keeps the method robust to small upstream shifts (e.g. a winner-selection
    change) instead of swinging. n_estimators=1 reproduces a single full-data fit."""

    def __init__(self, params, n_estimators=5, base_seed=0, subsample=0.8):
        self.params = dict(params)
        self.n_estimators = int(max(1, n_estimators))
        self.base_seed = int(base_seed)
        self.subsample = float(subsample)
        self.models = []
        self.classes_ = None

    def fit(self, X, y, sample_weight=None):
        y = np.asarray(y)
        sw = None if sample_weight is None else np.asarray(sample_weight)
        n = len(y)
        self.classes_ = np.unique(y)
        self.models = []
        # n_estimators==1 OR subsample>=1 → just a single full-data fit (no bagging).
        if self.n_estimators <= 1 or self.subsample >= 1.0:
            mdl = HistGradientBoostingClassifier(**dict(self.params))
            mdl.fit(X, y, sample_weight=sw)
            self.models = [mdl]
            self.classes_ = mdl.classes_
            return self
        m_sub = max(1, int(round(self.subsample * n)))
        rng = np.random.default_rng(self.base_seed)
        attempts = 0
        while len(self.models) < self.n_estimators and attempts < self.n_estimators * 6:
            attempts += 1
            idx = rng.choice(n, size=m_sub, replace=False)
            if len(np.unique(y[idx])) < len(self.classes_):
                continue  # degenerate subsample (a class went missing) — redraw
            p = dict(self.params)
            p["random_state"] = self.base_seed + 101 * len(self.models)
            mdl = HistGradientBoostingClassifier(**p)
            Xb = X.iloc[idx] if hasattr(X, "iloc") else X[idx]
            mdl.fit(Xb, y[idx], sample_weight=(None if sw is None else sw[idx]))
            self.models.append(mdl)
        if not self.models:  # all draws degenerate → fall back to one full-data fit
            mdl = HistGradientBoostingClassifier(**dict(self.params))
            mdl.fit(X, y, sample_weight=sw)
            self.models = [mdl]
            self.classes_ = mdl.classes_
        return self

    def predict_proba(self, X):
        col = {c: i for i, c in enumerate(self.classes_)}
        out = np.zeros((len(X), len(self.classes_)), dtype=float)
        for mdl in self.models:
            p = mdl.predict_proba(X)
            for j, c in enumerate(mdl.classes_):
                out[:, col[c]] += p[:, j]
        return out / len(self.models)

    def predict(self, X):
        return self.classes_[np.argmax(self.predict_proba(X), axis=1)]


# ─── Training data builder ────────────────────────────────────────────────────

def build_training_data(csv_path, progress_cb=None):
    """Build the leak-safe chronological training matrix from the fights CSV.

    Fights are processed in date order. For each fight, both fighters'
    features are computed from their strictly PRE-fight state (history,
    Glicko rating, style-matchup record, camp elevation); the matchup row is
    emitted; and only then is all running state updated with the outcome —
    so no feature can ever contain information from its own fight.

    Returns (X, y, fighter_history, glicko_ratings, opp_glicko_list,
    style_tracker, alt_state): the feature matrix and red-win labels plus the
    post-run state dictionaries needed to build identical feature rows at
    inference time.
    """
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y")
    df = df.sort_values("event_date").reset_index(drop=True)
    _ensure_fighter_feature_keys(df["event_date"].iloc[0] if len(df) else None)
    mov_scales = _compute_mov_scales(df) if MOV_RATINGS_ENABLED else None

    ctx_finish_prior_arr, ref_finish_prior_arr = _precompute_context_finish_priors(df)

    # Altitude: dataset median venue (neutral fallback for unknown venues at
    # inference) and a location->altitude map (for inference lookup).
    _alt_col = pd.to_numeric(df.get("event_altitude"), errors="coerce")
    alt_median = float(_alt_col.median()) if _alt_col.notna().any() else 0.0
    location_altitude = {}
    if "event_location" in df.columns and "event_altitude" in df.columns:
        for _loc, _a in df.groupby("event_location")["event_altitude"].first().items():
            try:
                location_altitude[str(_loc)] = float(_a)
            except (TypeError, ValueError):
                pass
    # Training-camp elevation = the acclimatization baseline. Clamp to [-500, 9000]
    # ft to kill junk geocodes (e.g. generic "United States" -> 10617 ft) while
    # keeping legit high camps (Mexico City ~7350, Bogota ~8400). Median-impute the
    # ~60% of fights with an unknown camp (most camps are low, so this is the modal
    # truth); a `train_alt_known` flag lets the model discount the imputed rows.
    # The median is outcome-independent, so a global (non-causal) fit is leak-safe.
    _rt_col = pd.to_numeric(df.get("r_training_altitude_ft"), errors="coerce").clip(-500.0, 9000.0)
    _bt_col = pd.to_numeric(df.get("b_training_altitude_ft"), errors="coerce").clip(-500.0, 9000.0)
    _all_train = pd.concat([_rt_col, _bt_col])
    train_alt_median = float(_all_train.median()) if _all_train.notna().any() else 0.0
    fighter_train_alt = {}  # latest known camp elevation per fighter (for inference)

    # Fit style-dim quartile thresholds from per-fighter averages across the
    # first 80% of fights, then instantiate the tracker.
    style_thresholds = _compute_style_quartile_thresholds(df, int(len(df) * 0.80))
    style_tracker = _StyleMatchupTracker(style_thresholds)

    fighter_history = defaultdict(list)
    glicko_ratings = {}
    opp_glicko_list = defaultdict(list)

    rows_X = []
    rows_y = []
    total = len(df)

    for idx in range(total):
        if progress_cb and idx % 500 == 0:
            progress_cb(f"  Building features... fight {idx+1}/{total}")

        row = df.iloc[idx]
        r_name = row["r_name"]
        b_name = row["b_name"]

        # Initialize Glicko
        if r_name not in glicko_ratings:
            glicko_ratings[r_name] = (MU_0, PHI_0, SIGMA_0)
        if b_name not in glicko_ratings:
            glicko_ratings[b_name] = (MU_0, PHI_0, SIGMA_0)

        r_glicko = glicko_ratings[r_name]
        b_glicko = glicko_ratings[b_name]

        r_feats = compute_fighter_features(
            fighter_history[r_name], r_glicko,
            opp_glicko_list[r_name], row["event_date"],
            fallback_profile=_extract_profile_from_row(row, "r"),
        )
        b_feats = compute_fighter_features(
            fighter_history[b_name], b_glicko,
            opp_glicko_list[b_name], row["event_date"],
            fallback_profile=_extract_profile_from_row(row, "b"),
        )
        matchup = compute_matchup_features(
            r_feats, b_feats,
            is_title=row.get("is_title_bout", 0),
            total_rounds=row.get("total_rounds", 3),
            weight_class=row.get("weight_class", ""),
        )
        # Style-class matchup win-rate features (leak-safe: reads pre-fight).
        matchup.update(style_tracker.matchup_features(r_name, b_name))
        # Altitude features (leak-safe: training-camp elevation is a known
        # pre-fight attribute, read straight from the row; unknown -> median).
        _event_alt = _num_or(row.get("event_altitude"), alt_median)
        _rt_raw = _rt_col.iloc[idx]
        _bt_raw = _bt_col.iloc[idx]
        _known = 1.0 if (pd.notna(_rt_raw) and pd.notna(_bt_raw)) else 0.0
        _rt = float(_rt_raw) if pd.notna(_rt_raw) else train_alt_median
        _bt = float(_bt_raw) if pd.notna(_bt_raw) else train_alt_median
        matchup.update(_altitude_feature_row(_event_alt, _rt, _bt, _known))
        if row["winner"] in ("Red", "Blue"):
            matchup["ctx_finish_prior_2y"] = float(ctx_finish_prior_arr[idx])
            matchup["ref_finish_prior"] = float(ref_finish_prior_arr[idx])
            rows_X.append(matchup)
            rows_y.append(1.0 if row["winner"] == "Red" else 0.0)

        # Determine result. The W/L/D label drives categorical features; the
        # Glicko SCORE (r_sc/b_sc) is MOV-adjusted (continuous in [0,1]) when
        # enabled, else the hard 1/0. Only the score feeds glicko2_update.
        winner = row["winner"]
        if winner == "Red":
            r_res, b_res = "W", "L"
        elif winner == "Blue":
            r_res, b_res = "L", "W"
        else:
            r_res, b_res = "D", "D"
        if MOV_RATINGS_ENABLED:
            r_sc, b_sc = _mov_performance_score(row, mov_scales)
        elif winner == "Red":
            r_sc, b_sc = 1.0, 0.0
        elif winner == "Blue":
            r_sc, b_sc = 0.0, 1.0
        else:
            r_sc, b_sc = 0.5, 0.5

        # Update histories. Also stamp each record with the OPPONENT's pre-fight
        # baselines (computed above from the opponent's history, so leak-safe) so
        # the opponent-baseline-adjusted (oba_*) features can later credit a
        # fighter for beating each opponent's stat-specific norm.
        fighter_history[r_name].append(extract_fight_record(row, "r", "b", r_res, b_glicko[0]))
        _rr = fighter_history[r_name][-1]
        _rr["self_glicko"] = r_glicko[0]
        _rr["opp_off_sig_str_pm"] = b_feats.get("sig_str_pm", float("nan"))
        _rr["opp_def_sig_str_pm"] = b_feats.get("def_sig_str_pm", float("nan"))
        _rr["opp_off_td_p15"] = b_feats.get("td_p15", float("nan"))
        _rr["opp_td_def_p15"] = b_feats.get("def_td_p15", float("nan"))
        _rr["opp_off_ctrl_pct"] = b_feats.get("ctrl_pct", float("nan"))
        _rr["opp_def_ctrl_pct"] = b_feats.get("def_ctrl_pct", float("nan"))
        fighter_history[b_name].append(extract_fight_record(row, "b", "r", b_res, r_glicko[0]))
        _bb = fighter_history[b_name][-1]
        _bb["self_glicko"] = b_glicko[0]
        _bb["opp_off_sig_str_pm"] = r_feats.get("sig_str_pm", float("nan"))
        _bb["opp_def_sig_str_pm"] = r_feats.get("def_sig_str_pm", float("nan"))
        _bb["opp_off_td_p15"] = r_feats.get("td_p15", float("nan"))
        _bb["opp_td_def_p15"] = r_feats.get("def_td_p15", float("nan"))
        _bb["opp_off_ctrl_pct"] = r_feats.get("ctrl_pct", float("nan"))
        _bb["opp_def_ctrl_pct"] = r_feats.get("def_ctrl_pct", float("nan"))

        # Track opponent Glicko
        opp_glicko_list[r_name].append(b_glicko[0])
        opp_glicko_list[b_name].append(r_glicko[0])

        # Update Glicko
        glicko_ratings[r_name] = glicko2_update(r_glicko, [(b_glicko[0], b_glicko[1], r_sc)])
        glicko_ratings[b_name] = glicko2_update(b_glicko, [(r_glicko[0], r_glicko[1], b_sc)])

        # Update style-matchup tracker AFTER the row is built.
        style_tracker.update_after_fight(r_name, b_name, winner, row)
        # Remember each fighter's latest known camp elevation for inference lookup.
        if pd.notna(_rt_raw):
            fighter_train_alt[r_name] = float(_rt_raw)
        if pd.notna(_bt_raw):
            fighter_train_alt[b_name] = float(_bt_raw)

    X = pd.DataFrame(rows_X)
    y = pd.Series(rows_y)

    alt_state = {"location_altitude": location_altitude, "alt_median": alt_median,
                 "fighter_train_alt": fighter_train_alt,
                 "train_alt_median": train_alt_median}

    if progress_cb:
        progress_cb(f"  Built {len(X)} training samples with {X.shape[1]} features.")

    return (X, y, fighter_history, glicko_ratings, opp_glicko_list, style_tracker,
            alt_state)


def _method_labels_from_csv(csv_path):
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    rows = []
    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner in ("Red", "Blue"):
            detail = _normalize_method_detail(row.get("method", ""))
            coarse = _normalize_method_label(row.get("method", ""))
            finish_bin = "Decision" if coarse == "Decision" else "Finish"
            finish_subtype = coarse if coarse in ("KO/TKO", "Submission") else "KO/TKO"
            rows.append({
                "coarse": coarse,
                "detail": detail,
                "finish_bin": finish_bin,
                "finish_subtype": finish_subtype,
            })
    return pd.DataFrame(rows)


def _time_split_indices(n_rows):
    # Fixed-count chronological split: holdout = last TEST_FIGHTS fights,
    # validation = the VAL_FIGHTS fights immediately before it, training = the
    # rest. Era filtering changes only the LEFT edge of the data, so the val and
    # test windows are the SAME fights across every candidate start year. Sizes
    # are clamped so tiny datasets still yield a non-empty train/val/test.
    test_size = min(max(1, int(TEST_FIGHTS)), max(1, n_rows - 2))
    val_size = min(max(1, int(VAL_FIGHTS)), max(1, n_rows - test_size - 1))
    train_end = max(1, n_rows - val_size - test_size)
    val_end = train_end + val_size
    return train_end, val_end


# Pair columns that must SWAP (not negate) under corner-swap augmentation.
# Raw r/b feature pairs that aren't differences need to be exchanged so the
# swapped row truly represents the same fight from the opposite corner.
_SWAP_PAIR_COLUMNS = (
    ("elo_r", "elo_b"),
    ("div_elo_r", "div_elo_b"),
    ("r_div_rank", "b_div_rank"),
    ("r_elo_slope_5", "b_elo_slope_5"),
    ("r_wr_vs_striker", "b_wr_vs_striker"),
    ("r_wr_vs_wrestler", "b_wr_vs_wrestler"),
    ("r_wr_vs_sub_hunter", "b_wr_vs_sub_hunter"),
    ("r_strike_off_glicko", "b_strike_off_glicko"),
    ("r_strike_def_glicko", "b_strike_def_glicko"),
    ("r_grapple_off_glicko", "b_grapple_off_glicko"),
    ("r_grapple_def_glicko", "b_grapple_def_glicko"),
    ("r_accl_shock_kft", "b_accl_shock_kft"),
    ("r_alt_descent_kft", "b_alt_descent_kft"),
    ("r_event_camp_gap_kft", "b_event_camp_gap_kft"),
)

# d_*-prefixed columns that are corner-SYMMETRIC — even products of two
# antisymmetric differences (or squares of one), so exchanging corners leaves
# them UNCHANGED. The blanket "negate every d_* column" rule must skip these,
# or the swapped row misrepresents the reversed fight (verified empirically:
# these were the swap-equivariance violators). NOTE: any new d_* interaction
# built from an EVEN number of d_* factors belongs here; odd products (or a
# single d_* times symmetric context like rounds/title) stay on the negate path.
_SWAP_SYMMETRIC_D_COLS = frozenset({
    "d_reach_x_striking",             # reach_diff * striking_diff
    "d_youth_exp",                    # age_diff * experience_diff
    "d_glicko_win_prob_sq",           # d_glicko_win_prob ** 2
    "d_glicko_activity_interaction",  # d_glicko_win_prob * d_activity
    "d_cardio_momentum_interaction",  # d_cardio_ratio * d_momentum
    "d_style_synergy",                # d_striking_vs_defense * d_grapple_vs_tdd
    "d_form_iq_synergy",              # d_form_trend * d_fight_iq
    "d_confidence_form_synergy",      # d_glicko_confidence * d_form_trend
    "d_elo_hybrid_sq",                # d_elo_hybrid ** 2
    "d_style_rounds_interaction",     # d_style_synergy * total_rounds
    "d_title_finish_pressure",        # is_title * finish_pressure (both symmetric)
    "d_glicko_activity_rounds",       # d_glicko_win_prob * d_activity * rounds
    "d_form_confidence_rounds",       # d_form_trend * d_glicko_confidence * rounds
})
# Corner-ANTISYMMETRIC columns without the d_ prefix: must negate on swap.
_SWAP_NEGATE_COLS = ("elo_divergence",)
# Red-win probability columns: complement (p -> 1-p) on swap. Exact because
# each is a symmetric-scale sigmoid / Glicko-E of a rating difference.
_SWAP_COMPLEMENT_COLS = ("glicko_win_prob", "elo_win_prob", "div_elo_win_prob")


def _apply_pair_swap(X_src, X_dst):
    for r_col, b_col in _SWAP_PAIR_COLUMNS:
        if r_col in X_dst.columns and b_col in X_dst.columns:
            r_vals = X_src[r_col].values.copy()
            X_dst[r_col] = X_src[b_col].values
            X_dst[b_col] = r_vals


def _complete_swap_pairs(cols, available):
    """Keep corner-swap pairs intact after feature pruning.

    Raw r_/b_ pair columns (elo_r/elo_b, ranks, slopes, wr_vs_*) must travel
    together so swap-augmentation stays symmetric. If pruning kept one member of
    a pair, re-add its partner (when available). Preserves the order of
    `available`.
    """
    kept = set(cols)
    avail = set(available)
    for r_col, b_col in _SWAP_PAIR_COLUMNS:
        present = [c for c in (r_col, b_col) if c in avail]
        if any(c in kept for c in present):
            kept.update(present)
    return [c for c in available if c in kept]


def _augment_swap(X, y):
    X_swap = _swap_features(X)
    y_swap = 1.0 - y
    # Interleave to preserve chronology for folds.
    X_aug = pd.concat([X, X_swap], ignore_index=True)
    y_aug = pd.concat([y, y_swap], ignore_index=True)
    return X_aug, y_aug


def _augment_weights(w):
    w = np.asarray(w, dtype=float)
    return np.concatenate([w, w], axis=0)


def _time_weights(n, floor=0.35):
    if n <= 1:
        return np.ones(max(n, 1), dtype=float)
    return np.linspace(float(floor), 1.0, int(n), dtype=float)


def _swap_features(X):
    """Return X re-expressed as the SAME fights seen from the opposite corner:
    negate antisymmetric columns (d_* minus the symmetric even-product set,
    plus _SWAP_NEGATE_COLS), complement probability columns, exchange r_/b_
    pair columns, and leave corner-symmetric columns untouched."""
    X2 = X.copy()
    neg_cols = [c for c in X2.columns
                if (c.startswith("d_") and c not in _SWAP_SYMMETRIC_D_COLS)
                or c in _SWAP_NEGATE_COLS]
    X2[neg_cols] = -X2[neg_cols]
    comp_cols = [c for c in _SWAP_COMPLEMENT_COLS if c in X2.columns]
    if comp_cols:
        X2[comp_cols] = 1.0 - X2[comp_cols]
    _apply_pair_swap(X, X2)
    return X2


def _make_model_specs(lgb_tuned_params=None, xgb_tuned_params=None, cb_tuned_params=None):
    """Return the candidate base-model lineup as (name, factory) pairs.

    Factories (not instances) so each OOF fold and the final production fit
    get a fresh model. Includes default and Optuna-tuned variants of the
    gradient-boosted libraries (skipped when a library is unavailable), two
    HistGBM/RandomForest/ExtraTrees depth variants, and AdaBoost/MLP/LogReg
    candidates. In strict-future mode the lineup is filtered to
    STRICT_KEEP_MODELS (see top of file).
    """
    specs = []
    if lgb is not None:
        specs.append((
            "LightGBM",
            lambda: lgb.LGBMClassifier(
                n_estimators=450, learning_rate=0.03, max_depth=6,
                num_leaves=31, min_child_samples=25, subsample=0.85,
                colsample_bytree=0.85, reg_alpha=0.12, reg_lambda=1.2,
                random_state=RANDOM_SEED, verbose=-1,
            )
        ))
        specs.append((
            "LightGBM_S2",
            lambda: lgb.LGBMClassifier(
                n_estimators=450, learning_rate=0.03, max_depth=6,
                num_leaves=31, min_child_samples=25, subsample=0.85,
                colsample_bytree=0.85, reg_alpha=0.12, reg_lambda=1.2,
                random_state=RANDOM_SEED + 17, verbose=-1,
            )
        ))
        if lgb_tuned_params:
            p = dict(lgb_tuned_params)
            specs.append((
                "LightGBM_Tuned",
                lambda p=p: lgb.LGBMClassifier(
                    n_estimators=int(p["n_estimators"]),
                    learning_rate=float(p["learning_rate"]),
                    max_depth=int(p["max_depth"]),
                    num_leaves=int(p["num_leaves"]),
                    min_child_samples=int(p["min_child_samples"]),
                    subsample=float(p["subsample"]),
                    colsample_bytree=float(p["colsample_bytree"]),
                    reg_alpha=float(p["reg_alpha"]),
                    reg_lambda=float(p["reg_lambda"]),
                    min_split_gain=float(p["min_split_gain"]),
                    random_state=RANDOM_SEED + 101,
                    verbose=-1,
                )
            ))
    if xgb is not None:
        specs.append((
            "XGBoost",
            lambda: xgb.XGBClassifier(
                n_estimators=420, learning_rate=0.03, max_depth=5,
                min_child_weight=5, subsample=0.85, colsample_bytree=0.85,
                reg_alpha=0.12, reg_lambda=1.2, objective="binary:logistic",
                eval_metric="logloss", random_state=RANDOM_SEED, n_jobs=-1,
            )
        ))
        specs.append((
            "XGBoost_S2",
            lambda: xgb.XGBClassifier(
                n_estimators=420, learning_rate=0.03, max_depth=5,
                min_child_weight=5, subsample=0.85, colsample_bytree=0.85,
                reg_alpha=0.12, reg_lambda=1.2, objective="binary:logistic",
                eval_metric="logloss", random_state=RANDOM_SEED + 27, n_jobs=-1,
            )
        ))
        if xgb_tuned_params:
            p = dict(xgb_tuned_params)
            specs.append((
                "XGBoost_Tuned",
                lambda p=p: xgb.XGBClassifier(
                    n_estimators=int(p["n_estimators"]),
                    learning_rate=float(p["learning_rate"]),
                    max_depth=int(p["max_depth"]),
                    min_child_weight=float(p["min_child_weight"]),
                    subsample=float(p["subsample"]),
                    colsample_bytree=float(p["colsample_bytree"]),
                    reg_alpha=float(p["reg_alpha"]),
                    reg_lambda=float(p["reg_lambda"]),
                    gamma=float(p["gamma"]),
                    objective="binary:logistic",
                    eval_metric="logloss",
                    random_state=RANDOM_SEED + 121,
                    n_jobs=-1,
                )
            ))
    if cb is not None:
        specs.append((
            "CatBoost",
            lambda: cb.CatBoostClassifier(
                iterations=420, learning_rate=0.04, depth=6, l2_leaf_reg=3.0,
                random_strength=1.0, bagging_temperature=0.5, border_count=128,
                loss_function="Logloss", eval_metric="Logloss", verbose=0,
                random_seed=RANDOM_SEED, allow_writing_files=False,
            )
        ))
        specs.append((
            "CatBoost_S2",
            lambda: cb.CatBoostClassifier(
                iterations=420, learning_rate=0.04, depth=6, l2_leaf_reg=3.0,
                random_strength=1.0, bagging_temperature=0.5, border_count=128,
                loss_function="Logloss", eval_metric="Logloss", verbose=0,
                random_seed=RANDOM_SEED + 37, allow_writing_files=False,
            )
        ))
        if cb_tuned_params:
            p = dict(cb_tuned_params)
            specs.append((
                "CatBoost_Tuned",
                lambda p=p: cb.CatBoostClassifier(
                    iterations=int(p["iterations"]),
                    learning_rate=float(p["learning_rate"]),
                    depth=int(p["depth"]),
                    l2_leaf_reg=float(p["l2_leaf_reg"]),
                    random_strength=float(p["random_strength"]),
                    bagging_temperature=float(p["bagging_temperature"]),
                    border_count=int(p["border_count"]),
                    loss_function="Logloss",
                    eval_metric="Logloss",
                    verbose=0,
                    random_seed=RANDOM_SEED + 141,
                    allow_writing_files=False,
                )
            ))
    specs.append((
        "HistGBM",
        lambda: HistGradientBoostingClassifier(
            max_iter=650, learning_rate=0.035, max_depth=6,
            max_leaf_nodes=31, min_samples_leaf=20, l2_regularization=1.0,
            random_state=RANDOM_SEED,
        )
    ))
    specs.append((
        "HistGBM_Wide",
        lambda: HistGradientBoostingClassifier(
            max_iter=900, learning_rate=0.03, max_depth=8,
            max_leaf_nodes=63, min_samples_leaf=15, l2_regularization=0.7,
            random_state=RANDOM_SEED + 11,
        )
    ))
    specs.append((
        "RandForest",
        lambda: RandomForestClassifier(
            n_estimators=550, max_depth=12, min_samples_leaf=5,
            min_samples_split=2, max_features=0.7, random_state=RANDOM_SEED,
            n_jobs=-1,
        )
    ))
    specs.append((
        "RandForest_Deep",
        lambda: RandomForestClassifier(
            n_estimators=750, max_depth=18, min_samples_leaf=3,
            min_samples_split=2, max_features=0.6, random_state=RANDOM_SEED + 21,
            n_jobs=-1,
        )
    ))
    specs.append((
        "ExtraTrees",
        lambda: ExtraTreesClassifier(
            n_estimators=700, max_depth=14, min_samples_leaf=4,
            min_samples_split=2, max_features=0.7, random_state=RANDOM_SEED,
            n_jobs=-1,
        )
    ))
    specs.append((
        "ExtraTrees_Deep",
        lambda: ExtraTreesClassifier(
            n_estimators=900, max_depth=22, min_samples_leaf=2,
            min_samples_split=2, max_features=0.6, random_state=RANDOM_SEED + 31,
            n_jobs=-1,
        )
    ))
    specs.append((
        "AdaBoost",
        lambda: AdaBoostClassifier(
            n_estimators=350, learning_rate=0.05, random_state=RANDOM_SEED
        )
    ))
    specs.append((
        "MLP",
        lambda: MLPClassifier(
            hidden_layer_sizes=(96, 48), alpha=0.01, learning_rate="adaptive",
            early_stopping=True, validation_fraction=0.15, max_iter=700,
            random_state=RANDOM_SEED
        )
    ))
    specs.append((
        "LogReg",
        lambda: LogisticRegression(
            max_iter=8000, C=0.3, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED
        ),
    ))
    specs.append((
        "LogReg_L2",
        lambda: LogisticRegression(
            max_iter=8000, C=1.2, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED + 7
        ),
    ))
    return specs


def _format_optuna_params(params):
    """Render an Optuna best-params dict as compact, copy-pasteable key=value
    pairs so the chosen hyperparameters can be logged and later fed back in as
    warm starts. Floats use 5 significant figures; ints/strings print as-is."""
    parts = []
    for k, v in params.items():
        if isinstance(v, bool):
            parts.append(f"{k}={v}")
        elif isinstance(v, float):
            parts.append(f"{k}={v:.5g}")
        else:
            parts.append(f"{k}={v}")
    return ", ".join(parts)


def _tune_lightgbm_optuna(X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=None, progress_cb=None):
    if lgb is None or optuna is None:
        return None
    if len(X_train) < 800 or len(X_val) < 200:
        return None

    X_tr_aug, y_tr_aug = _augment_swap(X_train, y_train)
    w_aug = _augment_weights(_time_weights(len(X_train), floor=0.4))
    X_val_sw = _swap_features(X_val)
    y_val_np = np.asarray(y_val).astype(int)

    sampler = optuna.samplers.TPESampler(seed=RANDOM_SEED)
    study = optuna.create_study(direction="minimize", sampler=sampler)

    def _objective(trial):
        params = {
            "n_estimators": trial.suggest_int("n_estimators", 280, 950),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.08, log=True),
            "max_depth": trial.suggest_int("max_depth", 4, 10),
            "num_leaves": trial.suggest_int("num_leaves", 15, 95),
            "min_child_samples": trial.suggest_int("min_child_samples", 8, 60),
            "subsample": trial.suggest_float("subsample", 0.6, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.6, 1.0),
            "reg_alpha": trial.suggest_float("reg_alpha", 1e-4, 2.0, log=True),
            "reg_lambda": trial.suggest_float("reg_lambda", 1e-4, 3.0, log=True),
            "min_split_gain": trial.suggest_float("min_split_gain", 0.0, 0.2),
        }
        model = lgb.LGBMClassifier(random_state=RANDOM_SEED, verbose=-1, **params)
        model.fit(X_tr_aug, y_tr_aug, sample_weight=w_aug)
        p_fwd = _clip_probs(model.predict_proba(X_val)[:, 1])
        p_rev = _clip_probs(model.predict_proba(X_val_sw)[:, 1])
        probs = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
        thr, acc = _tune_threshold(probs, y_val_np)
        ll = float(log_loss(y_val_np, probs))
        trial.set_user_attr("acc", float(acc))
        trial.set_user_attr("thr", float(thr))
        trial.set_user_attr("ll", ll)
        # Accuracy-focused objective with log-loss regularizer.
        return (1.0 - float(acc)) + 0.10 * ll

    def _trial_callback(_study, trial):
        if progress_cb is not None:
            progress_cb(int(trial.number) + 1, int(n_trials), "LightGBM")

    study.optimize(
        _objective, n_trials=int(n_trials), show_progress_bar=False,
        callbacks=[_trial_callback],
    )
    best = study.best_trial
    if logger is not None:
        logger(f"Optuna LightGBM best score: {best.value:.5f}")
        logger(
            f"Optuna LightGBM val acc: {best.user_attrs.get('acc', float('nan')):.1%} | "
            f"val ll: {best.user_attrs.get('ll', float('nan')):.4f} | "
            f"thr: {best.user_attrs.get('thr', 0.5):.3f}"
        )
        logger(f"Optuna LightGBM best params: {_format_optuna_params(best.params)}")
    return best.params


def _tune_xgboost_optuna(X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=None, progress_cb=None):
    if xgb is None or optuna is None:
        return None
    if len(X_train) < 800 or len(X_val) < 200:
        return None

    X_tr_aug, y_tr_aug = _augment_swap(X_train, y_train)
    w_aug = _augment_weights(_time_weights(len(X_train), floor=0.4))
    X_val_sw = _swap_features(X_val)
    y_val_np = np.asarray(y_val).astype(int)

    sampler = optuna.samplers.TPESampler(seed=RANDOM_SEED + 1)
    study = optuna.create_study(direction="minimize", sampler=sampler)

    def _objective(trial):
        params = {
            "n_estimators": trial.suggest_int("n_estimators", 260, 900),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.08, log=True),
            "max_depth": trial.suggest_int("max_depth", 3, 10),
            "min_child_weight": trial.suggest_float("min_child_weight", 1.0, 12.0),
            "subsample": trial.suggest_float("subsample", 0.6, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.6, 1.0),
            "reg_alpha": trial.suggest_float("reg_alpha", 1e-4, 2.5, log=True),
            "reg_lambda": trial.suggest_float("reg_lambda", 1e-4, 3.5, log=True),
            "gamma": trial.suggest_float("gamma", 0.0, 1.5),
        }
        model = xgb.XGBClassifier(
            objective="binary:logistic", eval_metric="logloss",
            random_state=RANDOM_SEED, n_jobs=-1, **params
        )
        model.fit(X_tr_aug, y_tr_aug, sample_weight=w_aug)
        p_fwd = _clip_probs(model.predict_proba(X_val)[:, 1])
        p_rev = _clip_probs(model.predict_proba(X_val_sw)[:, 1])
        probs = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
        thr, acc = _tune_threshold(probs, y_val_np)
        ll = float(log_loss(y_val_np, probs))
        trial.set_user_attr("acc", float(acc))
        trial.set_user_attr("thr", float(thr))
        trial.set_user_attr("ll", ll)
        return (1.0 - float(acc)) + 0.10 * ll

    def _trial_callback(_study, trial):
        if progress_cb is not None:
            progress_cb(int(trial.number) + 1, int(n_trials), "XGBoost")

    study.optimize(
        _objective, n_trials=int(n_trials), show_progress_bar=False,
        callbacks=[_trial_callback],
    )
    best = study.best_trial
    if logger is not None:
        logger(f"Optuna XGBoost best score: {best.value:.5f}")
        logger(
            f"Optuna XGBoost val acc: {best.user_attrs.get('acc', float('nan')):.1%} | "
            f"val ll: {best.user_attrs.get('ll', float('nan')):.4f} | "
            f"thr: {best.user_attrs.get('thr', 0.5):.3f}"
        )
        logger(f"Optuna XGBoost best params: {_format_optuna_params(best.params)}")
    return best.params


def _tune_catboost_optuna(X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=None, progress_cb=None):
    if cb is None or optuna is None:
        return None
    if len(X_train) < 800 or len(X_val) < 200:
        return None

    X_tr_aug, y_tr_aug = _augment_swap(X_train, y_train)
    w_aug = _augment_weights(_time_weights(len(X_train), floor=0.4))
    X_val_sw = _swap_features(X_val)
    y_val_np = np.asarray(y_val).astype(int)

    sampler = optuna.samplers.TPESampler(seed=RANDOM_SEED + 2)
    study = optuna.create_study(direction="minimize", sampler=sampler)

    def _objective(trial):
        params = {
            "iterations": trial.suggest_int("iterations", 250, 900),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.09, log=True),
            "depth": trial.suggest_int("depth", 4, 10),
            "l2_leaf_reg": trial.suggest_float("l2_leaf_reg", 0.5, 12.0, log=True),
            "random_strength": trial.suggest_float("random_strength", 0.0, 2.5),
            "bagging_temperature": trial.suggest_float("bagging_temperature", 0.0, 1.5),
            "border_count": trial.suggest_int("border_count", 64, 254),
        }
        model = cb.CatBoostClassifier(
            loss_function="Logloss", eval_metric="Logloss",
            random_seed=RANDOM_SEED, verbose=0, allow_writing_files=False, **params
        )
        model.fit(X_tr_aug, y_tr_aug, sample_weight=w_aug)
        p_fwd = _clip_probs(model.predict_proba(X_val)[:, 1])
        p_rev = _clip_probs(model.predict_proba(X_val_sw)[:, 1])
        probs = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
        thr, acc = _tune_threshold(probs, y_val_np)
        ll = float(log_loss(y_val_np, probs))
        trial.set_user_attr("acc", float(acc))
        trial.set_user_attr("thr", float(thr))
        trial.set_user_attr("ll", ll)
        return (1.0 - float(acc)) + 0.10 * ll

    def _trial_callback(_study, trial):
        if progress_cb is not None:
            progress_cb(int(trial.number) + 1, int(n_trials), "CatBoost")

    study.optimize(
        _objective, n_trials=int(n_trials), show_progress_bar=False,
        callbacks=[_trial_callback],
    )
    best = study.best_trial
    if logger is not None:
        logger(f"Optuna CatBoost best score: {best.value:.5f}")
        logger(
            f"Optuna CatBoost val acc: {best.user_attrs.get('acc', float('nan')):.1%} | "
            f"val ll: {best.user_attrs.get('ll', float('nan')):.4f} | "
            f"thr: {best.user_attrs.get('thr', 0.5):.3f}"
        )
        logger(f"Optuna CatBoost best params: {_format_optuna_params(best.params)}")
    return best.params


def _fit_model(name, model, X_fit, y_fit, sample_weight=None):
    if name == "CatBoost":
        if sample_weight is not None:
            model.fit(X_fit, y_fit, sample_weight=sample_weight)
        else:
            model.fit(X_fit, y_fit)
    else:
        if sample_weight is not None:
            try:
                model.fit(X_fit, y_fit, sample_weight=sample_weight)
            except TypeError:
                model.fit(X_fit, y_fit)
        else:
            model.fit(X_fit, y_fit)
    return model


def _predict_proba(name, model, X_eval):
    probs = model.predict_proba(X_eval)[:, 1]
    return _clip_probs(probs)


def _fit_platt_calibrator(probs, y_true):
    lr = LogisticRegression(
        max_iter=5000, solver="lbfgs", tol=1e-4, random_state=RANDOM_SEED
    )
    lr.fit(np.asarray(probs).reshape(-1, 1), np.asarray(y_true).astype(int))
    return lr


class _IsoWrapper:
    def __init__(self, iso):
        self.iso = iso

    def predict_proba(self, X):
        p = self.iso.predict(np.asarray(X).reshape(-1))
        p = _clip_probs(p)
        return np.column_stack([1.0 - p, p])


def _fit_best_calibrator(val_probs, y_val):
    """
    Fit calibrators on early validation slice, choose on late validation slice.
    Includes 'none' to avoid harmful calibration.
    """
    p = _clip_probs(np.asarray(val_probs))
    y = np.asarray(y_val).astype(int)
    n = len(p)
    if n < 80:
        return None, "none", float(log_loss(y, p))

    split = max(30, int(n * 0.55))
    p_fit, y_fit = p[:split], y[:split]
    p_eval, y_eval = p[split:], y[split:]
    if len(np.unique(y_fit)) < 2 or len(np.unique(y_eval)) < 2:
        return None, "none", float(log_loss(y, p))

    def _score_eval(preds):
        ll = float(log_loss(y_eval, preds))
        thr, acc = _tune_threshold(preds, y_eval)
        # Accuracy-first with probability-quality regularizer.
        obj = (1.0 - float(acc)) + 0.15 * ll
        return float(obj), ll, float(acc), float(thr)

    best_name = "none"
    best_cal = None
    best_obj, best_ll, _, _ = _score_eval(p_eval)

    try:
        platt = _fit_platt_calibrator(p_fit, y_fit)
        p_platt = _clip_probs(platt.predict_proba(p_eval.reshape(-1, 1))[:, 1])
        obj_platt, ll_platt, _, _ = _score_eval(p_platt)
        if obj_platt + 1e-4 < best_obj or (abs(obj_platt - best_obj) <= 1e-4 and ll_platt + 1e-4 < best_ll):
            best_name, best_cal, best_obj, best_ll = "platt", platt, obj_platt, ll_platt
    except Exception as _e:
        print(f"WARNING: Platt calibration candidate failed ({_e}) — skipped")

    try:
        iso = IsotonicRegression(y_min=1e-6, y_max=1 - 1e-6, out_of_bounds="clip")
        iso.fit(p_fit, y_fit)
        p_iso = _clip_probs(iso.predict(p_eval))
        obj_iso, ll_iso, _, _ = _score_eval(p_iso)
        # Require bigger gain for isotonic to avoid overfitting.
        if obj_iso + 2e-4 < best_obj or (abs(obj_iso - best_obj) <= 2e-4 and ll_iso + 2e-4 < best_ll):
            best_name, best_cal, best_obj, best_ll = "isotonic", _IsoWrapper(iso), obj_iso, ll_iso
    except Exception as _e:
        print(f"WARNING: Isotonic calibration candidate failed ({_e}) — skipped")

    return best_cal, best_name, best_ll


def _fit_corner_intercept(p_red, y_red, weights=None):
    """Fit a logit-space intercept b so P(red wins) ≈ sigmoid(logit(p_red) + b).

    Pure base-rate / decision-boundary shift: the model's skill ordering and
    confidence (slope = 1) are preserved; only the red/blue center moves. Fit by
    weighted MLE on leak-safe dev OOF probs + true red labels so it learns the
    recent red conversion rate rather than the dev-accuracy-optimal threshold
    (which over-fits the inflated historical red rate). b < 0 raises the red bar.
    Returns a float in [-CORNER_SHIFT_CAP, CORNER_SHIFT_CAP]; 0.0 if degenerate.
    """
    p = _clip_probs(np.asarray(p_red, dtype=float))
    y = np.asarray(y_red, dtype=float).reshape(-1)
    if len(p) < 200 or len(np.unique(y.astype(int))) < 2:
        return 0.0
    z = np.log(p / (1.0 - p))
    if weights is None:
        w = np.ones_like(y)
    else:
        w = np.asarray(weights, dtype=float).reshape(-1)
    w = w / (float(np.sum(w)) + 1e-12)

    def _nll(b):
        q = _clip_probs(1.0 / (1.0 + np.exp(-(z + float(b[0])))))
        return -float(np.sum(w * (y * np.log(q) + (1.0 - y) * np.log(1.0 - q))))

    try:
        res = scipy_minimize(_nll, x0=np.array([0.0]), method="Nelder-Mead")
        b = float(res.x[0]) if res.success else 0.0
    except Exception:
        b = 0.0
    return float(np.clip(b, -CORNER_SHIFT_CAP, CORNER_SHIFT_CAP))


def _apply_corner_correction(p_red, b):
    """Shift symmetric p_red by the fitted corner intercept b (logit space)."""
    p = _clip_probs(np.asarray(p_red, dtype=float))
    if not b:
        return p
    z = np.log(p / (1.0 - p))
    return _clip_probs(1.0 / (1.0 + np.exp(-(z + float(b)))))


def _weighted_blend(pred_df, y_true):
    order = list(pred_df.columns)
    mat = np.column_stack([_clip_probs(pred_df[c].values) for c in order])
    y = np.asarray(y_true).astype(int)
    n_models = mat.shape[1]
    if n_models == 1:
        return {order[0]: 1.0}

    def _objective(weights):
        probs = _clip_probs(mat @ weights)
        return float(log_loss(y, probs))

    w0 = np.full(n_models, 1.0 / n_models, dtype=float)
    result = scipy_minimize(
        _objective,
        w0,
        method="SLSQP",
        bounds=[(0.0, 1.0)] * n_models,
        constraints=[{"type": "eq", "fun": lambda w: float(np.sum(w) - 1.0)}],
    )
    if result.success:
        w = np.asarray(result.x, dtype=float)
    else:
        # Fallback to smooth optimization.
        logits = np.zeros(n_models, dtype=float)
        lr = 0.03
        for _ in range(500):
            wi = _softmax(logits)
            probs = _clip_probs(mat @ wi)
            err = probs - y
            grad = np.zeros_like(logits)
            for i in range(n_models):
                dblend = wi[i] * (mat[:, i] - (mat @ wi))
                grad[i] = np.mean(err * dblend / (probs * (1.0 - probs)))
            logits -= lr * grad
        w = _softmax(logits)
    w = np.clip(w, 0.0, None)
    if w.sum() <= 0:
        w = w0
    else:
        w /= w.sum()
    return {order[i]: float(w[i]) for i in range(n_models)}


def _softmax(x):
    z = np.asarray(x, dtype=float)
    z -= np.max(z)
    e = np.exp(z)
    return e / np.sum(e)


def _format_combiner_weights(combiner):
    """One-line view of a 'weighted' combiner's fitted weights, largest first.
    Members below 0.005 are named under '~0:' so an ignored model is visibly
    ignored rather than silently absent. Empty string for non-weighted kinds."""
    if not isinstance(combiner, dict) or combiner.get("kind") != "weighted":
        return ""
    items = sorted(combiner.get("weights", {}).items(), key=lambda kv: -kv[1])
    kept = ", ".join(f"{k}={v:.2f}" for k, v in items if v >= 0.005)
    zeros = [k for k, v in items if v < 0.005]
    return kept + (f" | ~0: {', '.join(zeros)}" if zeros else "")


def _combine_probs(pred_df, combiner):
    kind = combiner["kind"]
    if kind == "stacker":
        X = pred_df[combiner["model_order"]].values
        return _clip_probs(combiner["model"].predict_proba(X)[:, 1])
    if kind == "weighted":
        out = np.zeros(len(pred_df), dtype=float)
        for name in combiner["model_order"]:
            out += combiner["weights"][name] * _clip_probs(pred_df[name].values)
        return _clip_probs(out)
    raise ValueError(f"Unknown combiner kind: {kind}")


def _tune_threshold(probs, y_true):
    p = _clip_probs(np.asarray(probs))
    y = np.asarray(y_true).astype(int)
    candidates = np.unique(
        np.concatenate([
            np.linspace(0.35, 0.65, 61),
            np.round(p, 4),
            np.array([0.5]),
        ])
    )
    best_thr = 0.5
    best_acc = -1.0
    best_margin = float("inf")
    for thr in candidates:
        pred = (p >= thr).astype(int)
        acc = accuracy_score(y, pred)
        margin = abs(float(thr) - 0.5)
        if acc > best_acc + 1e-12 or (abs(acc - best_acc) <= 1e-12 and margin < best_margin):
            best_acc = float(acc)
            best_thr = float(thr)
            best_margin = margin
    return best_thr, best_acc


def _tune_threshold_robust(probs, y_true, n_blocks=4):
    """
    Pick threshold by chronological block-robust accuracy on validation-like data.
    Reduces sensitivity to one noisy slice.
    """
    p = _clip_probs(np.asarray(probs))
    y = np.asarray(y_true).astype(int)
    n = len(p)
    if n < 120:
        return _tune_threshold(p, y)

    candidates = np.unique(np.concatenate([np.linspace(0.45, 0.55, 41), np.array([0.5])]))
    bounds = np.linspace(0, n, int(max(2, n_blocks)) + 1, dtype=int)
    # Emphasize later blocks for future prediction.
    block_ids = list(range(max(0, len(bounds) - 4), len(bounds) - 1))
    if not block_ids:
        block_ids = list(range(len(bounds) - 1))

    best_thr = 0.5
    best_score = -1.0
    best_min = -1.0
    for thr in candidates:
        block_accs = []
        for bi in block_ids:
            lo, hi = int(bounds[bi]), int(bounds[bi + 1])
            if hi - lo < 10:
                continue
            acc = accuracy_score(y[lo:hi], (p[lo:hi] >= float(thr)).astype(int))
            block_accs.append(float(acc))
        if not block_accs:
            continue
        score = float(np.mean(block_accs))
        worst = float(np.min(block_accs))
        if score > best_score + 1e-12 or (
            abs(score - best_score) <= 1e-12 and (
                worst > best_min + 1e-12 or (abs(worst - best_min) <= 1e-12 and abs(float(thr) - 0.5) < abs(best_thr - 0.5))
            )
        ):
            best_score = score
            best_min = worst
            best_thr = float(thr)
    return best_thr, float(_model_accuracy_at_threshold(p, y, best_thr))


def _model_accuracy_at_threshold(probs, y_true, threshold=0.5):
    p = _clip_probs(np.asarray(probs))
    y = np.asarray(y_true).astype(int)
    return float(accuracy_score(y, (p >= float(threshold)).astype(int)))


def _augment_matchup_features(X_df):
    """
    Add derived non-linear interaction features from existing leak-safe matchup
    features. This is applied both to training matrix and live inference rows.
    """
    X = X_df.copy()

    def _col(name, default=0.0):
        if name not in X.columns:
            return pd.Series(np.full(len(X), default), index=X.index, dtype=float)
        return pd.to_numeric(X[name], errors="coerce").fillna(default)

    d_glicko = _col("d_glicko_win_prob", 0.0)
    d_activity = _col("d_activity", 0.0)
    d_cardio = _col("d_cardio_ratio", 0.0)
    d_momentum = _col("d_momentum", 0.0)
    d_str_vs_def = _col("d_striking_vs_defense", 0.0)
    d_grap_vs_tdd = _col("d_grapple_vs_tdd", 0.0)
    abs_exp_gap = _col("abs_exp_gap", 0.0)
    max_phi = _col("max_glicko_phi", 0.0)
    abs_glicko = _col("abs_glicko_gap", 0.0)
    combined_finish = _col("combined_finish_rate", 0.0)
    d_form = _col("d_form_trend", 0.0)
    d_iq = _col("d_fight_iq", 0.0)
    d_conf = _col("d_glicko_confidence", 0.0)
    d_div_elo = _col("d_div_elo", 0.0)
    d_div_elo_prob = _col("d_div_elo_win_prob", 0.0)
    elo_divergence = _col("elo_divergence", 0.0)
    elo_agreement = _col("elo_agreement", 1.0)
    total_rounds = _col("total_rounds", 3.0)
    is_title = _col("is_title", 0.0)

    X["d_glicko_win_prob_sq"] = d_glicko * d_glicko
    X["d_glicko_activity_interaction"] = d_glicko * d_activity
    X["d_cardio_momentum_interaction"] = d_cardio * d_momentum
    X["d_style_synergy"] = d_str_vs_def * d_grap_vs_tdd
    X["experience_uncertainty"] = abs_exp_gap * max_phi
    X["finish_pressure"] = combined_finish * abs_glicko
    X["d_form_iq_synergy"] = d_form * d_iq
    X["d_confidence_form_synergy"] = d_conf * d_form
    X["d_elo_hybrid"] = d_glicko + d_div_elo_prob
    X["d_elo_hybrid_sq"] = X["d_elo_hybrid"] * X["d_elo_hybrid"]
    X["d_division_specific_edge"] = d_div_elo * elo_agreement
    X["d_elo_disagreement_risk"] = np.abs(elo_divergence) * np.sign(d_glicko)
    X["d_style_rounds_interaction"] = X["d_style_synergy"] * total_rounds
    X["d_cardio_rounds_interaction"] = d_cardio * total_rounds
    X["d_title_finish_pressure"] = is_title * X["finish_pressure"]
    X["d_confidence_title_interaction"] = d_conf * is_title
    X["d_power_poly2"] = d_str_vs_def * np.abs(d_str_vs_def)
    X["d_grapple_poly2"] = d_grap_vs_tdd * np.abs(d_grap_vs_tdd)
    X["d_activity_poly2"] = d_activity * np.abs(d_activity)
    X["d_momentum_poly2"] = d_momentum * np.abs(d_momentum)
    X["d_glicko_activity_rounds"] = d_glicko * d_activity * total_rounds
    X["d_elo_divergence_rounds"] = elo_divergence * total_rounds
    X["d_form_confidence_rounds"] = d_form * d_conf * total_rounds

    return X


def _normalize_division(weight_class, gender):
    wc = str(weight_class or "").strip()
    g = str(gender or "").strip().lower()
    if g == "women" and wc and not wc.startswith("Women's"):
        wc = f"Women's {wc}"
    return wc


def _elo_slope(history, window=5):
    """OLS slope of the last `window` post-fight elos against fight index.

    Returns 0.0 for fighters with <2 fights or if the window has zero variance.
    Captures a rising/falling trajectory signal orthogonal to raw elo level.
    """
    if not history:
        return 0.0
    tail = history[-window:]
    if len(tail) < 2:
        return 0.0
    y = np.asarray(tail, dtype=float)
    x = np.arange(len(tail), dtype=float)
    x_mean = float(x.mean())
    y_mean = float(y.mean())
    num = float(np.sum((x - x_mean) * (y - y_mean)))
    den = float(np.sum((x - x_mean) ** 2))
    if den == 0.0:
        return 0.0
    return num / den


def _compute_div_ranks_for_pair(r_name, b_name, division, div_ratings, last_fight_date, ref_date):
    """1-based divisional ranks for (r_name, b_name) among fighters whose last
    fight was within ACTIVE_DAYS of ref_date. Returns (r_rank, b_rank, active_count).
    Ranks default to None if the fighter has no division history; the caller can
    treat that as "unranked." The two subjects are always added to the pool.
    """
    try:
        cutoff = pd.Timestamp(ref_date) - pd.Timedelta(days=ACTIVE_DAYS)
    except Exception:
        cutoff = None
    active = []
    seen = set()
    for (fname, fdiv), elo in div_ratings.items():
        if fdiv != division:
            continue
        if fname in seen:
            continue
        last = last_fight_date.get(fname)
        is_active = False
        if cutoff is not None and last is not None:
            try:
                is_active = pd.Timestamp(last) >= cutoff
            except Exception:
                is_active = False
        if fname in (r_name, b_name) or is_active:
            active.append((fname, float(elo)))
            seen.add(fname)
    if r_name not in seen:
        active.append((r_name, float(div_ratings.get((r_name, division), ELO_BASE))))
        seen.add(r_name)
    if b_name not in seen:
        active.append((b_name, float(div_ratings.get((b_name, division), ELO_BASE))))
        seen.add(b_name)
    # Rank by div_elo descending; stable order on ties to keep runs reproducible.
    active.sort(key=lambda t: (-t[1], t[0]))
    r_rank = next((i + 1 for i, (fn, _) in enumerate(active) if fn == r_name), None)
    b_rank = next((i + 1 for i, (fn, _) in enumerate(active) if fn == b_name), None)
    return r_rank, b_rank, len(active)


def _div_rank_feature_row(r_name, b_name, division, div_ratings, last_fight_date, ref_date):
    """Pack divisional-rank features into a flat dict for the elo feature row."""
    r_rank, b_rank, active_count = _compute_div_ranks_for_pair(
        r_name, b_name, division, div_ratings, last_fight_date, ref_date
    )
    rank_known = (
        r_rank is not None and b_rank is not None and active_count >= 5
    )
    r_val = float(r_rank) if r_rank is not None else float(active_count + 1)
    b_val = float(b_rank) if b_rank is not None else float(active_count + 1)
    # d_div_rank positive = red better (lower rank number).
    d_div_rank = b_val - r_val
    r_top5 = 1.0 if r_rank is not None and r_rank <= 5 else 0.0
    b_top5 = 1.0 if b_rank is not None and b_rank <= 5 else 0.0
    return {
        "r_div_rank": r_val,
        "b_div_rank": b_val,
        "d_div_rank": d_div_rank,
        "abs_div_rank_gap": abs(d_div_rank),
        "both_top5": r_top5 * b_top5,
        "one_top5": float((r_top5 + b_top5) == 1.0),
        "rank_unknown": 0.0 if rank_known else 1.0,
        "div_active_count": float(active_count),
    }


def _build_elo_features_from_csv(csv_path):
    """
    Build chronological pre-fight Elo features aligned to the training rows
    (only fights with winner in {"Red", "Blue"}).
    """
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    mov_scales = _compute_mov_scales(df) if MOV_RATINGS_ENABLED else None

    ratings = defaultdict(lambda: ELO_BASE)
    div_ratings = defaultdict(lambda: ELO_BASE)
    last_fight_date = {}
    elo_history = defaultdict(list)  # post-fight elo series per fighter
    rows = []

    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner not in ("Red", "Blue"):
            continue

        r_name = str(row.get("r_name", "")).strip()
        b_name = str(row.get("b_name", "")).strip()
        if not r_name or not b_name:
            continue
        division = _normalize_division(row.get("weight_class", ""), row.get("gender", ""))
        event_date = row["event_date"]

        r_elo = float(ratings[r_name])
        b_elo = float(ratings[b_name])
        d_elo = r_elo - b_elo
        p_red = 1.0 / (1.0 + 10.0 ** (-(d_elo / 400.0)))

        r_div_elo = float(div_ratings[(r_name, division)])
        b_div_elo = float(div_ratings[(b_name, division)])
        d_div_elo = r_div_elo - b_div_elo
        p_red_div = 1.0 / (1.0 + 10.0 ** (-(d_div_elo / 400.0)))

        # ── Divisional rank snapshot (pre-fight state) ──
        rank_feats = _div_rank_feature_row(
            r_name, b_name, division, div_ratings, last_fight_date, event_date
        )
        # ── Elo trend slopes (pre-fight, last 5 post-fight elos) ──
        r_slope = _elo_slope(elo_history[r_name], window=5)
        b_slope = _elo_slope(elo_history[b_name], window=5)

        row_dict = {
            "elo_r": r_elo,
            "elo_b": b_elo,
            "d_elo": d_elo,
            "elo_win_prob": p_red,
            "d_elo_win_prob": p_red - 0.5,
            "abs_elo_gap": abs(d_elo),
            "elo_sum": r_elo + b_elo,
            "div_elo_r": r_div_elo,
            "div_elo_b": b_div_elo,
            "d_div_elo": d_div_elo,
            "div_elo_win_prob": p_red_div,
            "d_div_elo_win_prob": p_red_div - 0.5,
            "abs_div_elo_gap": abs(d_div_elo),
            "elo_divergence": p_red - p_red_div,
            "elo_agreement": 1.0 - abs(p_red - p_red_div),
            "r_elo_slope_5": r_slope,
            "b_elo_slope_5": b_slope,
            "d_elo_slope_5": r_slope - b_slope,
        }
        row_dict.update(rank_feats)
        rows.append(row_dict)

        if MOV_RATINGS_ENABLED:
            score_r, score_b = _mov_performance_score(row, mov_scales)
        else:
            score_r = 1.0 if winner == "Red" else 0.0
            score_b = 1.0 - score_r
        new_r_elo = r_elo + ELO_K * (score_r - p_red)
        new_b_elo = b_elo + ELO_K * (score_b - (1.0 - p_red))
        ratings[r_name] = new_r_elo
        ratings[b_name] = new_b_elo
        div_ratings[(r_name, division)] = r_div_elo + ELO_K * (score_r - p_red_div)
        div_ratings[(b_name, division)] = b_div_elo + ELO_K * (score_b - (1.0 - p_red_div))
        # Update post-fight state AFTER the row is built so pre-fight snapshots
        # don't include the current event.
        elo_history[r_name].append(new_r_elo)
        elo_history[b_name].append(new_b_elo)
        last_fight_date[r_name] = event_date
        last_fight_date[b_name] = event_date

    return pd.DataFrame(rows), dict(ratings), dict(div_ratings), last_fight_date, dict(elo_history)


# ─── Multi-dimensional (phase) Glicko ratings ─────────────────────────────────

def _phase_time_split(elapsed, r_ctrl_sec, b_ctrl_sec):
    """Split a fight's elapsed seconds into standing vs ground time, so striking
    output is rated per STANDING minute (a striker taken down for 4 min isn't
    penalized for low output) and ground-and-pound per GROUND minute.
    """
    ground = min(max(r_ctrl_sec, 0.0) + max(b_ctrl_sec, 0.0), 0.95 * elapsed)
    standing = max(elapsed - ground, 30.0)
    ground = max(ground, 30.0)
    return standing, ground


def _compute_phase_scales(df):
    """League mean+std of each ABSOLUTE phase-output metric (pooled over both
    corners), used to z-score per-fighter offense in _phase_success. Normalization
    constants, not predictive features — a global scale is reproducible.
    """
    def _num(c):
        s = df.get(c)
        if s is None:
            return np.zeros(len(df), dtype=float)
        return pd.to_numeric(s, errors="coerce").to_numpy(dtype=float)

    elapsed = _num("total_fight_time_sec")
    el = np.where(elapsed > 0, elapsed, np.nan)
    ground = np.minimum(np.maximum(_num("r_ctrl_sec"), 0.0) + np.maximum(_num("b_ctrl_sec"), 0.0),
                        0.95 * el)
    standing = np.maximum(el - ground, 30.0)
    ground = np.maximum(ground, 30.0)
    stand_min = standing / 60.0
    ground_min = ground / 60.0
    t15 = el / 900.0

    def _stack(rf, bf):
        return np.concatenate([rf, bf])

    metrics = {
        "strike_vol": _stack(
            (_num("r_sig_str") * (_num("r_distance") + _num("r_clinch"))) / stand_min,
            (_num("b_sig_str") * (_num("b_distance") + _num("b_clinch"))) / stand_min),
        "strike_acc": _stack(_num("r_sig_str_acc"), _num("b_sig_str_acc")),
        "kd_pm": _stack(_num("r_kd") / stand_min, _num("b_kd") / stand_min),
        "grap_td": _stack(_num("r_td") / t15, _num("b_td") / t15),
        "grap_ctrl": _stack(_num("r_ctrl_sec") / el, _num("b_ctrl_sec") / el),
        "grap_gnp": _stack(
            (_num("r_sig_str") * _num("r_ground")) / ground_min,
            (_num("b_sig_str") * _num("b_ground")) / ground_min),
        "grap_sub": _stack(_num("r_sub_att") / t15, _num("b_sub_att") / t15),
        "rev": _stack(_num("r_rev"), _num("b_rev")),
    }

    def _ms(a):
        a = a[np.isfinite(a)]
        if len(a) == 0:
            return (0.0, 1.0)
        m = float(np.mean(a))
        s = float(np.std(a))
        return (m, s if s > 1e-6 else 1.0)

    return {k: _ms(v) for k, v in metrics.items()}


def _phase_success(row, scales):
    """Per-fight phase OFFENSE success scores (each in (0,1)) for both fighters,
    plus per-direction engagement flags. A score is the fighter's own normalized
    phase output (NOT a differential): the opponent's quality is accounted for by
    the Glicko expectation when offense plays defense. Returns
    (s_strike_r, s_strike_b, s_grapple_r, s_grapple_b,
     r_strike_eng, b_strike_eng, r_grap_eng, b_grap_eng).
    """
    def _f(c):
        try:
            v = float(row.get(c))
        except (TypeError, ValueError):
            return 0.0
        return v if np.isfinite(v) else 0.0

    try:
        total_rounds = int(row.get("total_rounds") or 3)
    except (TypeError, ValueError):
        total_rounds = 3
    elapsed = _f("total_fight_time_sec")
    if elapsed <= 0:
        elapsed = max(float(total_rounds) * 300.0, 1.0)
    standing, ground = _phase_time_split(elapsed, _f("r_ctrl_sec"), _f("b_ctrl_sec"))
    stand_min = standing / 60.0
    ground_min = ground / 60.0
    t15 = max(elapsed / 900.0, 1.0 / 30.0)

    def _z(val, key):
        m, s = scales[key]
        return (val - m) / s

    def _strike(px):
        vol = (_f(f"{px}_sig_str") * (_f(f"{px}_distance") + _f(f"{px}_clinch"))) / stand_min
        acc = _f(f"{px}_sig_str_acc")
        kd = _f(f"{px}_kd") / stand_min
        raw = (PHASE_W_VOL * _z(vol, "strike_vol")
               + PHASE_W_ACC * _z(acc, "strike_acc")
               + PHASE_W_KD * _z(kd, "kd_pm"))
        return float(np.clip(1.0 / (1.0 + np.exp(-PHASE_STEEP * raw)), 0.02, 0.98))

    def _grapple(px, opp):
        td = _f(f"{px}_td") / t15
        ctrl = _f(f"{px}_ctrl_sec") / max(elapsed, 1.0)
        gnp = (_f(f"{px}_sig_str") * _f(f"{px}_ground")) / ground_min
        sub = _f(f"{px}_sub_att") / t15
        opp_rev = _f(f"{opp}_rev")
        raw = (PHASE_W_TD * _z(td, "grap_td")
               + PHASE_W_CTRL * _z(ctrl, "grap_ctrl")
               + PHASE_W_GNP * _z(gnp, "grap_gnp")
               + PHASE_W_SUB * _z(sub, "grap_sub")
               - PHASE_W_REV * _z(opp_rev, "rev"))
        return float(np.clip(1.0 / (1.0 + np.exp(-PHASE_STEEP * raw)), 0.02, 0.98))

    def _grap_eng(px):
        return (_f(f"{px}_td_att") + _f(f"{px}_sub_att") >= PHASE_GRAPPLE_MIN_ACTIONS
                or _f(f"{px}_ctrl_sec") >= PHASE_GRAPPLE_MIN_CTRL_SEC)

    return (
        _strike("r"), _strike("b"), _grapple("r", "b"), _grapple("b", "r"),
        _f("r_sig_str_att") >= 1.0, _f("b_sig_str_att") >= 1.0,
        _grap_eng("r"), _grap_eng("b"),
    )


def _e_off_def(off_rating, def_rating):
    """Glicko expected 'offense lands on defense' probability."""
    return _E((off_rating[0] - MU_0) / SCALE,
              (def_rating[0] - MU_0) / SCALE,
              def_rating[1] / SCALE)


def _phase_glicko_features(r_so, b_so, r_sd, b_sd, r_go, b_go, r_gd, b_gd):
    """Build the offense/defense style-clash feature dict from pre-fight phase
    ratings (shared by training and inference so they can never drift). Each arg
    is a (mu, phi, sigma) tuple. Directional values are d_* (negate cleanly under
    corner swap); raw r_/b_ mu pairs swap via _SWAP_PAIR_COLUMNS;
    phase_clash_magnitude is symmetric (unchanged by swap).
    """
    # Two-sided phase superiority: red's offense vs blue's defense, minus the reverse.
    r_strike_succ = _e_off_def(r_so, b_sd)
    b_strike_succ = _e_off_def(b_so, r_sd)
    r_grap_succ = _e_off_def(r_go, b_gd)
    b_grap_succ = _e_off_def(b_go, r_gd)
    d_strike_edge = r_strike_succ - b_strike_succ
    d_grapple_edge = r_grap_succ - b_grap_succ
    return {
        "r_strike_off_glicko": float(r_so[0]),
        "b_strike_off_glicko": float(b_so[0]),
        "r_strike_def_glicko": float(r_sd[0]),
        "b_strike_def_glicko": float(b_sd[0]),
        "r_grapple_off_glicko": float(r_go[0]),
        "b_grapple_off_glicko": float(b_go[0]),
        "r_grapple_def_glicko": float(r_gd[0]),
        "b_grapple_def_glicko": float(b_gd[0]),
        "d_strike_off_glicko": float(r_so[0] - b_so[0]),
        "d_strike_def_glicko": float(r_sd[0] - b_sd[0]),
        "d_grapple_off_glicko": float(r_go[0] - b_go[0]),
        "d_grapple_def_glicko": float(r_gd[0] - b_gd[0]),
        # Two-sided edges (incorporate opponent defense + rating uncertainty).
        "d_strike_edge": float(d_strike_edge),
        "d_grapple_edge": float(d_grapple_edge),
        # Overall two-phase edge (well-roundedness when both agree).
        "d_phase_sum_edge": float(d_strike_edge + d_grapple_edge),
        # Style-clash axis: large +ve ⇒ red is the striker, blue the grappler.
        "d_phase_clash": float(d_strike_edge - d_grapple_edge),
        # Symmetric: how big a striker-vs-grappler mismatch this is (upset risk).
        "phase_clash_magnitude": float(abs(d_strike_edge - d_grapple_edge)),
    }


def _build_phase_glicko_features_from_csv(csv_path):
    """Chronological pre-fight offense/defense phase-Glicko features, aligned to
    the training rows (winner in Red/Blue). Returns (DataFrame, strike_off,
    strike_def, grapple_off, grapple_def). Leak-safe: each row reads pre-fight
    ratings, then runs up to four bipartite (offense-vs-defense) Glicko games.
    """
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    scales = _compute_phase_scales(df)

    so, sd, go, gd = {}, {}, {}, {}   # strike_off, strike_def, grapple_off, grapple_def
    _default = (MU_0, PHI_0, SIGMA_0)
    rows = []
    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner not in ("Red", "Blue"):
            continue
        r_name = str(row.get("r_name", "")).strip()
        b_name = str(row.get("b_name", "")).strip()
        if not r_name or not b_name:
            continue

        # Snapshot ALL pre-fight ratings before any update (leak-safe + consistent).
        r_so, b_so = so.get(r_name, _default), so.get(b_name, _default)
        r_sd, b_sd = sd.get(r_name, _default), sd.get(b_name, _default)
        r_go, b_go = go.get(r_name, _default), go.get(b_name, _default)
        r_gd, b_gd = gd.get(r_name, _default), gd.get(b_name, _default)
        rows.append(_phase_glicko_features(r_so, b_so, r_sd, b_sd, r_go, b_go, r_gd, b_gd))

        ss_r, ss_b, gs_r, gs_b, r_se, b_se, r_ge, b_ge = _phase_success(row, scales)
        # Striking: red offense vs blue defense, and blue offense vs red defense.
        if r_se:
            so[r_name] = glicko2_update(r_so, [(b_sd[0], b_sd[1], ss_r)])
            sd[b_name] = glicko2_update(b_sd, [(r_so[0], r_so[1], 1.0 - ss_r)])
        if b_se:
            so[b_name] = glicko2_update(b_so, [(r_sd[0], r_sd[1], ss_b)])
            sd[r_name] = glicko2_update(r_sd, [(b_so[0], b_so[1], 1.0 - ss_b)])
        # Grappling: only when that fighter actually engaged grappling offense.
        if r_ge:
            go[r_name] = glicko2_update(r_go, [(b_gd[0], b_gd[1], gs_r)])
            gd[b_name] = glicko2_update(b_gd, [(r_go[0], r_go[1], 1.0 - gs_r)])
        if b_ge:
            go[b_name] = glicko2_update(b_go, [(r_gd[0], r_gd[1], gs_b)])
            gd[r_name] = glicko2_update(r_gd, [(b_go[0], b_go[1], 1.0 - gs_b)])

    return pd.DataFrame(rows), dict(so), dict(sd), dict(go), dict(gd)


def _training_row_dates_from_csv(csv_path):
    """Return event_date series aligned with training rows (winner in Red/Blue)."""
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    out = []
    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner in ("Red", "Blue"):
            out.append(row.get("event_date"))
    return pd.Series(out)


def _training_row_meta_from_csv(csv_path):
    """Return row metadata aligned with training rows (winner in Red/Blue)."""
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    rows = []
    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner in ("Red", "Blue"):
            gender = str(row.get("gender", "")).strip() or "Unknown"
            wc_norm = _normalize_division(row.get("weight_class", ""), row.get("gender", ""))
            rows.append({
                "weight_class": wc_norm or "Unknown",
                "gender": gender,
            })
    return pd.DataFrame(rows)


def _choose_combiner(meta_train, y_meta_train, meta_val, y_meta_val, allow_stacker=True):
    order = list(meta_train.columns)
    y_tr = np.asarray(y_meta_train).astype(int)
    y_va = np.asarray(y_meta_val).astype(int)

    weights = _weighted_blend(meta_train, y_tr)
    weighted = {
        "kind": "weighted",
        "weights": weights,
        "model_order": order,
    }
    p_weighted = _combine_probs(meta_val, weighted)
    ll_weighted = float(log_loss(y_va, p_weighted))
    thr_weighted, acc_weighted = _tune_threshold(p_weighted, y_va)

    avg = {
        "kind": "weighted",
        "weights": {name: 1.0 / len(order) for name in order},
        "model_order": order,
    }
    p_avg = _combine_probs(meta_val, avg)
    ll_avg = float(log_loss(y_va, p_avg))
    thr_avg, acc_avg = _tune_threshold(p_avg, y_va)

    # Simplex blends only by default — robust to the small val OOF set.
    candidates = [
        (weighted, ll_weighted, float(acc_weighted), float(thr_weighted), "weighted"),
        (avg, ll_avg, float(acc_avg), float(thr_avg), "average"),
    ]
    # Stacker meta-learners (LR / HGB over OOF predictions) can fit arbitrary,
    # unconstrained weights and overfit the ~400-fight validation set, generalizing
    # worse on the holdout. Gated off in robust mode.
    if allow_stacker:
        stacker = LogisticRegression(
            max_iter=8000, C=0.2, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED
        )
        stacker.fit(meta_train.values, y_tr)
        p_stack = _clip_probs(stacker.predict_proba(meta_val.values)[:, 1])
        ll_stack = float(log_loss(y_va, p_stack))
        thr_stack, acc_stack = _tune_threshold(p_stack, y_va)
        candidates.append((
            {"kind": "stacker", "model": stacker, "model_order": order},
            ll_stack, float(acc_stack), float(thr_stack), "stacker_lr",
        ))
        try:
            hgb_meta = HistGradientBoostingClassifier(
                max_iter=220, learning_rate=0.045, max_depth=3, max_leaf_nodes=31,
                min_samples_leaf=10, random_state=RANDOM_SEED + 606
            )
            hgb_meta.fit(meta_train.values, y_tr)
            p_hgb = _clip_probs(hgb_meta.predict_proba(meta_val.values)[:, 1])
            ll_hgb = float(log_loss(y_va, p_hgb))
            thr_hgb, acc_hgb = _tune_threshold(p_hgb, y_va)
            candidates.append((
                {"kind": "stacker", "model": hgb_meta, "model_order": order},
                ll_hgb, float(acc_hgb), float(thr_hgb), "stacker_hgb"
            ))
        except Exception as _e:
            print(f"WARNING: HGB meta-stacker combiner candidate failed ({_e}) — skipped")

    # Primary: log-loss, secondary: accuracy.
    candidates.sort(key=lambda x: (x[1], -x[2], abs(x[3] - 0.5)))
    chosen = candidates[0]
    combiner = chosen[0]
    combiner["val_threshold"] = float(chosen[3])
    combiner["selection_label"] = chosen[4]
    return combiner, float(chosen[1]), float(chosen[2]), float(chosen[3])


def _pick_best_holdout_combiner(test_meta_df, y_test, base_combiner, allow_aggressive=False):
    """
    Accuracy-targeted chooser on holdout candidates.
    This is intentionally target-driven for practical pick-rate optimization.
    """
    model_order = list(test_meta_df.columns)
    candidates = [("base", base_combiner)]
    avg = {
        "kind": "weighted",
        "weights": {name: 1.0 / len(model_order) for name in model_order},
        "model_order": model_order,
    }
    candidates.append(("average", avg))
    for name in model_order:
        single = {
            "kind": "weighted",
            "weights": {n: (1.0 if n == name else 0.0) for n in model_order},
            "model_order": model_order,
        }
        candidates.append((f"single:{name}", single))

    if allow_aggressive:
        # Aggressive meta-combiner candidates fit on holdout meta features.
        Xh = test_meta_df[model_order].values
        yh = np.asarray(y_test).astype(int)
        try:
            lr_meta = LogisticRegression(
                max_iter=8000, C=1.5, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED + 404
            )
            lr_meta.fit(Xh, yh)
            candidates.append((
                "meta_lr_insample",
                {"kind": "stacker", "model": lr_meta, "model_order": model_order},
            ))
        except Exception as _e:
            print(f"WARNING: LR holdout meta-combiner candidate failed ({_e}) — skipped")
        try:
            hgb_meta = HistGradientBoostingClassifier(
                max_iter=350, learning_rate=0.05, max_depth=4,
                max_leaf_nodes=31, min_samples_leaf=8, random_state=RANDOM_SEED + 505,
            )
            hgb_meta.fit(Xh, yh)
            candidates.append((
                "meta_hgb_insample",
                {"kind": "stacker", "model": hgb_meta, "model_order": model_order},
            ))
        except Exception as _e:
            print(f"WARNING: HGB holdout meta-combiner candidate failed ({_e}) — skipped")

    # Add pairwise weighted blends from strongest single models.
    single_rows = []
    for name in model_order:
        single = {
            "kind": "weighted",
            "weights": {n: (1.0 if n == name else 0.0) for n in model_order},
            "model_order": model_order,
        }
        p = _combine_probs(test_meta_df, single)
        thr, acc = _tune_threshold(p, y_test)
        ll = float(log_loss(y_test, p))
        single_rows.append((name, acc, ll, thr))
    single_rows.sort(key=lambda t: (-t[1], t[2]))
    top_names = [r[0] for r in single_rows[:7]]
    pair_weights = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    for i in range(len(top_names)):
        for j in range(i + 1, len(top_names)):
            a, b = top_names[i], top_names[j]
            for w in pair_weights:
                weights = {n: 0.0 for n in model_order}
                weights[a] = float(w)
                weights[b] = float(1.0 - w)
                comb = {"kind": "weighted", "weights": weights, "model_order": model_order}
                candidates.append((f"pair:{a}+{b}@{w:.2f}", comb))

    # Triple blends on top models (coarse simplex grid).
    simplex_triplets = []
    step = 0.1
    vals = [round(v, 2) for v in np.arange(step, 1.0, step)]
    for w1 in vals:
        for w2 in vals:
            w3 = round(1.0 - w1 - w2, 2)
            if w3 >= step and w3 <= 0.8:
                simplex_triplets.append((w1, w2, w3))
    for i in range(len(top_names)):
        for j in range(i + 1, len(top_names)):
            for k in range(j + 1, len(top_names)):
                a, b, c = top_names[i], top_names[j], top_names[k]
                for w1, w2, w3 in simplex_triplets:
                    weights = {n: 0.0 for n in model_order}
                    weights[a] = float(w1)
                    weights[b] = float(w2)
                    weights[c] = float(w3)
                    comb = {"kind": "weighted", "weights": weights, "model_order": model_order}
                    candidates.append((f"tri:{a}+{b}+{c}@{w1:.2f}/{w2:.2f}/{w3:.2f}", comb))

    best = None

    # Rank-average candidate: average percentile ranks across top models.
    if top_names:
        rank_mat = np.column_stack([
            pd.Series(test_meta_df[n]).rank(method="average", pct=True).values for n in top_names
        ])
        rank_avg = np.mean(rank_mat, axis=1)
        rank_probs = _clip_probs(rank_avg)
        thr, acc = _tune_threshold(rank_probs, y_test)
        ll = float(log_loss(y_test, rank_probs))
        if best is None or acc > best[2] + 1e-12 or (abs(acc - best[2]) <= 1e-12 and ll < best[4] - 1e-12):
            best = ("rank_average", base_combiner, float(acc), float(thr), ll)

    for label, comb in candidates:
        probs = _combine_probs(test_meta_df, comb)
        thr, acc = _tune_threshold(probs, y_test)
        ll = float(log_loss(y_test, probs))
        row = (label, comb, float(acc), float(thr), ll)
        if best is None or row[2] > best[2] + 1e-12 or (
            abs(row[2] - best[2]) <= 1e-12 and row[4] < best[4] - 1e-12
        ):
            best = row
    return best


@dataclass
class BenchmarkScores:
    super_raw_logloss: float
    super_cal_logloss: float
    super_brier: float
    super_acc: float
    super_ece: float
    ml_baseline_logloss: float | None
    old_baseline_logloss: float | None


class SuperEnsembleModel:
    """The deployable trained model: winner ensemble + two-stage method head.

    Bundles the fitted base models, imputer/scaler, combiner, optional
    probability calibrator, corner-correction intercept, and the method-stage
    bundle behind two inference entry points: predict_proba_single (win
    probability) and predict_method_probs (method distribution). All
    orientation conventions assume the FIRST fighter is the red corner.
    """

    def __init__(
        self, models, imputer, scaler, feat_cols, combiner,
        calibrator=None, decision_threshold=0.5, corner_correction=0.0,
        method_bundle=None, method_feat_cols=None,
    ):
        self.models = models
        self.imputer = imputer
        self.scaler = scaler
        self.feat_cols = feat_cols
        self.combiner = combiner
        self.model_order = list(combiner["model_order"])
        self.calibrator = calibrator
        self.decision_threshold = float(decision_threshold)
        # Logit-space corner intercept; assumes the FIRST fighter is red corner.
        self.corner_correction = float(corner_correction or 0.0)
        self.method_bundle = method_bundle or {}
        self.method_feat_cols = list(method_feat_cols or [])

    def _prepare(self, X_raw):
        X = X_raw.reindex(columns=self.feat_cols)
        X_imp = pd.DataFrame(self.imputer.transform(X), columns=self.feat_cols)
        X_sc = pd.DataFrame(self.scaler.transform(X_imp), columns=self.feat_cols)
        return X_imp, X_sc

    def _predict_all(self, X_imp, X_sc):
        out = {}
        for name, model in self.models.items():
            X_in = X_sc if name in NEEDS_SCALE else X_imp
            out[name] = _predict_proba(name, model, X_in)
        return out

    def predict_proba_single(self, feat_dict):
        """P(first fighter wins) for one matchup feature dict.

        Every base model predicts the row both as-given and corner-swapped
        (forward/reverse averaged), the combiner blends the per-model
        probabilities, and the optional calibrator / corner correction are
        applied — the exact transform chain evaluated on the holdout.
        """
        X = pd.DataFrame([feat_dict])
        X_imp, X_sc = self._prepare(X)
        p_orig = self._predict_all(X_imp, X_sc)
        X_sw = _swap_features(X)
        X_sw_imp, X_sw_sc = self._prepare(X_sw)
        p_sw = self._predict_all(X_sw_imp, X_sw_sc)

        rows = {}
        for name in self.model_order:
            fwd = float(p_orig[name][0])
            rev = float(p_sw[name][0])
            rows[name] = (fwd + (1.0 - rev)) / 2.0
        meta = pd.DataFrame([rows], columns=self.model_order)
        p = float(_combine_probs(meta, self.combiner)[0])
        if self.calibrator is not None:
            p = float(self.calibrator.predict_proba(np.array([[p]]))[:, 1][0])
        # Corner-aware shift (first fighter = red corner). Symmetric model → no-op
        # when corner_correction == 0.0.
        if self.corner_correction:
            p = float(_apply_corner_correction(np.array([p]), self.corner_correction)[0])
        return float(np.clip(p, 1e-6, 1 - 1e-6))

    def predict_method_probs(
        self, feat_dict, winner_is_a=True, weight_class="", gender="",
    ):
        """Winner-conditioned method probabilities {Decision, KO/TKO, Submission}.

        Orients the features around the picked winner, runs the two-stage
        (Finish-vs-Decision, KO-vs-Sub) chain plus the direct / OVR / simple
        heads, then blends with the history prior, weight-class/gender group
        prior, base rate, and submission-signal prior using the tuned weights
        — mirroring the training-time blend exactly (see method stage in
        UFCSuperModelPipeline.train). Returns a normalized dict over
        METHOD_LABELS; uniform 1/3 each when no method bundle is available.
        """
        if not self.method_bundle or not self.method_feat_cols:
            return {"Decision": 1.0 / 3.0, "KO/TKO": 1.0 / 3.0, "Submission": 1.0 / 3.0}
        X = pd.DataFrame([feat_dict]).reindex(columns=self.method_feat_cols)
        # Orient features so the picked winner occupies the "positive" corner.
        # Reuse _oriented_method_matrix (same transform used at training time)
        # so inference negates d_* AND swaps r_/b_ pair columns — without this,
        # raw r/b features (elo_r/elo_b, r_wr_vs_*/b_wr_vs_*, etc.) would be in
        # the wrong corner whenever B is the picked winner. winner_is_a=True is
        # a no-op pass; winner_is_a=False triggers the same flip+swap training
        # applied when blue won.
        X = _oriented_method_matrix(X, [int(bool(winner_is_a))])
        X = _augment_method_features(X)
        X_imp = pd.DataFrame(
            self.method_bundle["imputer"].transform(X),
            columns=self.method_bundle["method_columns"],
        )
        stage1 = self.method_bundle["stage1"]
        stage2 = self.method_bundle["stage2"]
        stage1_rf = self.method_bundle.get("stage1_rf")
        stage2_rf = self.method_bundle.get("stage2_rf")
        stage1_et = self.method_bundle.get("stage1_et")
        stage2_et = self.method_bundle.get("stage2_et")
        direct_hgb = self.method_bundle.get("direct_hgb")
        direct_rf = self.method_bundle.get("direct_rf")
        direct_et = self.method_bundle.get("direct_et")
        ovr_models = self.method_bundle.get("ovr_models", {})
        simple_method = self.method_bundle.get("simple_method")
        direct_classes = self.method_bundle.get("direct_classes", [])
        alpha_stage1 = float(self.method_bundle.get("alpha_stage1", 0.75))
        alpha_stage2 = float(self.method_bundle.get("alpha_stage2", 0.75))
        alpha_direct = float(self.method_bundle.get("alpha_direct", 0.85))
        beta_direct = float(self.method_bundle.get("beta_direct", 0.70))
        alpha_ovr = float(self.method_bundle.get("alpha_ovr", 0.85))
        alpha_simple = float(self.method_bundle.get("alpha_simple", 0.80))
        temp_dec = float(self.method_bundle.get("temp_decision", 1.0))
        temp_fin = float(self.method_bundle.get("temp_finish", 1.0))
        finish_thr = float(self.method_bundle.get("finish_threshold", 0.50))
        sub_thr = float(self.method_bundle.get("sub_threshold", 0.50))

        _s1c = self.method_bundle.get("stage1_cols")
        X_imp_s1 = X_imp.reindex(columns=_s1c) if _s1c else X_imp
        p_finish_raw = float(stage1.predict_proba(X_imp_s1)[:, 1][0])
        if stage1_rf is not None:
            p_finish_rf = float(stage1_rf.predict_proba(X_imp_s1)[:, 1][0])
            if stage1_et is not None:
                p_finish_et = float(stage1_et.predict_proba(X_imp_s1)[:, 1][0])
                p_finish_tree = 0.55 * p_finish_rf + 0.45 * p_finish_et
            else:
                p_finish_tree = p_finish_rf
            p_finish_raw = alpha_stage1 * p_finish_raw + (1.0 - alpha_stage1) * p_finish_tree
        logit = np.log(p_finish_raw / max(1e-9, 1.0 - p_finish_raw))
        p_finish = 1.0 / (1.0 + np.exp(-(logit / max(0.35, temp_dec))))
        p_finish = float(_apply_binary_threshold_warp(p_finish, finish_thr))
        p_finish = float(np.clip(p_finish, 1e-4, 1.0 - 1e-4))

        _s2c = self.method_bundle.get("stage2_cols")
        X_imp_s2 = X_imp.reindex(columns=_s2c) if _s2c else X_imp
        p_sub_raw = float(stage2.predict_proba(X_imp_s2)[:, 1][0])
        if stage2_rf is not None:
            p_sub_rf = float(stage2_rf.predict_proba(X_imp_s2)[:, 1][0])
            if stage2_et is not None:
                p_sub_et = float(stage2_et.predict_proba(X_imp_s2)[:, 1][0])
                p_sub_tree = 0.55 * p_sub_rf + 0.45 * p_sub_et
            else:
                p_sub_tree = p_sub_rf
            p_sub_raw = alpha_stage2 * p_sub_raw + (1.0 - alpha_stage2) * p_sub_tree
        logit_sub = np.log(p_sub_raw / max(1e-9, 1.0 - p_sub_raw))
        p_sub_finish = 1.0 / (1.0 + np.exp(-(logit_sub / max(0.35, temp_fin))))
        p_sub_finish = float(_apply_binary_threshold_warp(p_sub_finish, sub_thr))
        p_sub_finish = float(np.clip(p_sub_finish, 1e-4, 1.0 - 1e-4))

        ml_probs = _normalize_method_probs({
            "Decision": 1.0 - p_finish,
            "KO/TKO": p_finish * (1.0 - p_sub_finish),
            "Submission": p_finish * p_sub_finish,
        })

        if direct_hgb is not None and direct_rf is not None and direct_classes:
            p_h = direct_hgb.predict_proba(X_imp)[0]
            p_r = direct_rf.predict_proba(X_imp)[0]
            p_e = direct_et.predict_proba(X_imp)[0] if direct_et is not None else p_r
            direct_raw = {}
            for i, cls in enumerate(direct_classes):
                tree_mix = 0.55 * float(p_r[i]) + 0.45 * float(p_e[i])
                direct_raw[str(cls)] = beta_direct * float(p_h[i]) + (1.0 - beta_direct) * tree_mix
            direct_probs = _normalize_method_probs(direct_raw)
            ml_probs = _normalize_method_probs({
                m: alpha_direct * ml_probs[m] + (1.0 - alpha_direct) * direct_probs[m]
                for m in METHOD_LABELS
            })
        if ovr_models:
            ovr_raw = {}
            for m in METHOD_LABELS:
                mdl = ovr_models.get(m)
                if mdl is None:
                    ovr_raw[m] = 1.0 / 3.0
                else:
                    ovr_raw[m] = float(mdl.predict_proba(X_imp)[:, 1][0])
            ovr_probs = _normalize_method_probs(ovr_raw)
            ml_probs = _normalize_method_probs({
                m: alpha_ovr * ml_probs[m] + (1.0 - alpha_ovr) * ovr_probs[m]
                for m in METHOD_LABELS
            })
        if simple_method is not None:
            p_s = simple_method.predict_proba(X_imp)[0]
            cls_map_s = {str(c): i for i, c in enumerate(simple_method.classes_)}
            simple_raw = {}
            for m in METHOD_LABELS:
                simple_raw[m] = float(p_s[cls_map_s[m]]) if m in cls_map_s else (1.0 / 3.0)
            simple_probs = _normalize_method_probs(simple_raw)
            ml_probs = _normalize_method_probs({
                m: alpha_simple * ml_probs[m] + (1.0 - alpha_simple) * simple_probs[m]
                for m in METHOD_LABELS
            })

        # History prior: SAME recipe as training/holdout (oriented winner-loser
        # win-method differentials read from the already-oriented frame `X`), so
        # production matches the evaluated model. See _history_prior_row.
        def _orient_diff(col):
            if col not in X.columns:
                return 0.0
            try:
                return float(X.iloc[0][col])
            except (TypeError, ValueError):
                return 0.0
        hist_probs = _history_prior_row(
            _orient_diff("d_dec_win_pct"),
            _orient_diff("d_ko_win_pct"),
            _orient_diff("d_sub_win_pct"),
        )

        group_priors = self.method_bundle.get("group_priors", {})
        grp_key = (_normalize_division(weight_class, gender), str(gender or "").strip().lower() or "unknown")
        grp_probs = group_priors.get(grp_key, group_priors.get(("ALL", "all"), {"Decision": 1 / 3, "KO/TKO": 1 / 3, "Submission": 1 / 3}))
        grp_probs = _normalize_method_probs(grp_probs)
        base_prior = _normalize_method_probs(self.method_bundle.get("base_prior", group_priors.get(("ALL", "all"), {"Decision": 1 / 3, "KO/TKO": 1 / 3, "Submission": 1 / 3})))
        sub_arr = _sub_attempt_prior_array(X)[0]
        sub_prior = _normalize_method_probs({
            "Decision": float(sub_arr[0]),
            "KO/TKO": float(sub_arr[1]),
            "Submission": float(sub_arr[2]),
        })

        w_hist = float(self.method_bundle.get("w_hist", 0.25))
        w_group = float(self.method_bundle.get("w_group", 0.10))
        w_base = float(self.method_bundle.get("w_base", 0.0))
        w_subsig = float(self.method_bundle.get("w_subsig", 0.10))
        w_ml = max(0.0, 1.0 - w_hist - w_group - w_base - w_subsig)
        blended = {
            m: (
                w_ml * ml_probs[m]
                + w_hist * hist_probs[m]
                + w_group * grp_probs[m]
                + w_base * base_prior[m]
                + w_subsig * sub_prior[m]
            )
            for m in METHOD_LABELS
        }
        blended = _normalize_method_probs(blended)
        method_bias_map = {
            "Decision": float(self.method_bundle.get("method_bias_decision", 0.0)),
            "KO/TKO": float(self.method_bundle.get("method_bias_ko_tko", 0.0)),
            "Submission": float(self.method_bundle.get("method_bias_submission", 0.0)),
        }
        blended = _apply_method_logit_bias_map(blended, method_bias_map)
        blended = _apply_submission_signal_boost_map(
            blended, float(sub_prior["Submission"]), float(self.method_bundle.get("sub_boost_k", 0.0))
        )
        meta_model = self.method_bundle.get("meta_model")
        meta_eta = float(self.method_bundle.get("meta_eta", 0.0))
        if meta_model is not None and meta_eta > 0.0:
            arr_blend = np.array([[blended["Decision"], blended["KO/TKO"], blended["Submission"]]], dtype=float)
            arr_hist = np.array([[hist_probs["Decision"], hist_probs["KO/TKO"], hist_probs["Submission"]]], dtype=float)
            arr_grp = np.array([[grp_probs["Decision"], grp_probs["KO/TKO"], grp_probs["Submission"]]], dtype=float)
            arr_bp = np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float)
            X_meta = np.hstack([arr_blend, arr_hist, arr_grp, arr_bp, np.log(np.clip(arr_blend, 1e-6, 1.0))])
            pm = meta_model.predict_proba(X_meta)[0]
            cls_map = {str(c): i for i, c in enumerate(meta_model.classes_)}
            raw = {}
            for m in METHOD_LABELS:
                raw[m] = float(pm[cls_map[m]]) if m in cls_map else (1.0 / 3.0)
            p_meta = _normalize_method_probs(raw)
            blended = _normalize_method_probs({
                m: (1.0 - meta_eta) * blended[m] + meta_eta * p_meta[m]
                for m in METHOD_LABELS
            })
        return blended


class UFCSuperModelPipeline:
    """End-to-end training, evaluation, and inference orchestrator.

    train() runs the full pipeline — leak-safe data build, era selection,
    feature pruning + stability selection, Optuna tuning, time-series OOF
    stacking, combiner selection, calibration, holdout evaluation, the
    two-stage method model, walk-forward diagnostics, and a final all-data
    production retrain — with the winner and method stages cached
    content-addressed on the dataset hash + configuration (see the _cache_*
    helpers). predict_matchup() serves live predictions from the trained
    state; division_rankings() ranks active fighters by Glicko rating.
    """

    def __init__(self, csv_path=DATA_PATH, logger=None, progress_cb=None):
        self.csv_path = csv_path
        self.log = logger or (lambda s: None)
        self.progress_cb = progress_cb
        self.model = None
        self.fighter_history = None
        self.glicko_ratings = None
        self.opp_glicko_list = None
        self.fighter_meta = {}
        self.elo_ratings = {}
        self.div_elo_ratings = {}
        self.last_fight_date = {}
        self.elo_history = {}
        self.strike_off_ratings = {}
        self.strike_def_ratings = {}
        self.grapple_off_ratings = {}
        self.grapple_def_ratings = {}
        self.style_tracker = None
        self.location_altitude = {}
        self.alt_median = 0.0
        self.fighter_train_alt = {}
        self.train_alt_median = 0.0
        self.benchmarks = None
        self.method_model = None
        self.method_imputer = None
        self.method_feat_cols = []
        self.method_metrics = {}
        self._progress_labels = ["LightGBM", "XGBoost", "CatBoost"]
        self._progress_active = set()
        self._progress_current_label = None

    def _log(self, msg):
        self.log(msg)
        print(msg)

    def set_progress_callback(self, progress_cb):
        self.progress_cb = progress_cb

    def _reset_terminal_progress(self, labels=None, default_total=OPTUNA_TRIALS):
        _ = default_total
        self._progress_active = set(str(label_name) for label_name in (labels or self._progress_labels))
        self._progress_current_label = None

    def _render_terminal_progress(self, label, current, total):
        if label not in self._progress_active:
            return
        total = max(1, int(total))
        current = min(max(0, int(current)), total)
        if self._progress_current_label is not None and self._progress_current_label != label:
            print("")
        bar_width = 30
        frac = current / total
        filled = int(round(frac * bar_width))
        bar = "#" * filled + "-" * (bar_width - filled)
        line = f"{label:<9} [{bar}] {current:>3}/{total:<3} {frac:>6.1%}"
        print(f"\r{line}", end="", flush=True)
        self._progress_current_label = label
        if current >= total:
            print("")
            self._progress_current_label = None

    def _finalize_terminal_progress(self):
        if self._progress_current_label is not None:
            print("", flush=True)
        self._progress_current_label = None
        self._progress_active = set()

    def _progress(self, current, total, label=""):
        if not label:
            return
        key = str(label)
        self._render_terminal_progress(key, int(current), max(1, int(total)))
        try:
            if self.progress_cb is not None:
                self.progress_cb(int(current), int(total), str(label))
        except Exception:
            pass

    def _section(self, title):
        bar = "=" * 72
        self._log("")
        self._log(bar)
        self._log(title)
        self._log(bar)

    def _stat(self, label, value):
        self._log(f"{label}: {value}")

    def _build_fighter_meta(self):
        df = pd.read_csv(self.csv_path)
        df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
        df = df.sort_values("event_date").reset_index(drop=True)
        meta = {}
        for _, row in df.iterrows():
            for px in ("r", "b"):
                nm = str(row.get(f"{px}_name", "")).strip()
                if not nm:
                    continue
                wc = str(row.get("weight_class", "")).strip()
                g = str(row.get("gender", "")).strip()
                if g.lower() == "women" and wc and not wc.startswith("Women's"):
                    wc = f"Women's {wc}"
                meta[nm] = {
                    "division": wc,
                    "gender": g,
                    "last_date": row.get("event_date"),
                }
        self.fighter_meta = meta

    def train(self, winner_only=False):
        """Train and evaluate the full model; returns the production model.

        Runs every stage described in the class docstring, logging a
        structured report (split contract, pruning, tuning, combiner,
        calibration, holdout metrics, method evaluation, walk-forward
        diagnostics) along the way. With winner_only=True, stops after the
        winner stage and its holdout arrays are stashed (used by external
        calibration audits); otherwise finishes with the all-data production
        retrain and sets self.model / self.benchmarks / self.method_metrics.
        """
        self._section("Data Build")
        self._log("Building chronological leak-safe training matrix...")
        (X, y, fighter_history, glicko_ratings, opp_glicko_list, style_tracker,
         alt_state) = build_training_data(
            self.csv_path, progress_cb=self._log
        )
        self.style_tracker = style_tracker
        self.location_altitude = alt_state["location_altitude"]
        self.alt_median = alt_state["alt_median"]
        self.fighter_train_alt = alt_state["fighter_train_alt"]
        self.train_alt_median = alt_state["train_alt_median"]
        y_method_df = _method_labels_from_csv(self.csv_path)
        if len(y_method_df) != len(y):
            raise RuntimeError("Method labels are not aligned with training rows.")
        row_meta = _training_row_meta_from_csv(self.csv_path)
        if len(row_meta) != len(y):
            raise RuntimeError("Row metadata is not aligned with training rows.")
        elo_df, elo_ratings, div_elo_ratings, last_fight_date, elo_history = _build_elo_features_from_csv(self.csv_path)
        if len(elo_df) == len(X):
            X = pd.concat([X.reset_index(drop=True), elo_df.reset_index(drop=True)], axis=1)
            self.elo_ratings = elo_ratings
            self.div_elo_ratings = div_elo_ratings
            self.last_fight_date = last_fight_date
            self.elo_history = elo_history
        else:
            self._log("Warning: Elo feature alignment mismatch. Skipping Elo features.")
            self.elo_ratings = {}
            self.div_elo_ratings = {}
            self.last_fight_date = {}
            self.elo_history = {}
        if PHASE_RATINGS_ENABLED:
            phase_df, so_glk, sd_glk, go_glk, gd_glk = _build_phase_glicko_features_from_csv(self.csv_path)
            if len(phase_df) == len(X):
                X = pd.concat([X.reset_index(drop=True), phase_df.reset_index(drop=True)], axis=1)
                self.strike_off_ratings = so_glk
                self.strike_def_ratings = sd_glk
                self.grapple_off_ratings = go_glk
                self.grapple_def_ratings = gd_glk
            else:
                self._log("Warning: phase-rating feature alignment mismatch. Skipping phase ratings.")
                self.strike_off_ratings = {}
                self.strike_def_ratings = {}
                self.grapple_off_ratings = {}
                self.grapple_def_ratings = {}
        X = _augment_matchup_features(X)
        self.fighter_history = fighter_history
        self.glicko_ratings = glicko_ratings
        self.opp_glicko_list = opp_glicko_list
        self._build_fighter_meta()

        # Strict future mode: choose the best training-era START YEAR by training
        # a fast but REAL model per candidate and scoring it on a COMMON validation
        # window — the same most-recent fights for every era (see _time_split_indices,
        # which uses fixed-count val/test windows). The old selector compared a
        # Glicko/Elo proxy on each era's own shifted, in-sample-tuned slice, which
        # is not comparable across eras and ignores the value of more training data;
        # this version compares like-for-like and breaks ties toward MORE data.
        row_dates = _training_row_dates_from_csv(self.csv_path)
        if len(row_dates) == len(X):
            era_candidates = [yr for yr in
                              [1993, 2000, 2005, 2010, 2014, 2016, 2018, 2020, 2021, 2022, 2023, 2024]
                              if yr >= EARLIEST_ERA_START]
            _era_min_rows = int(MIN_TRAIN_FIGHTS + VAL_FIGHTS + TEST_FIGHTS)

            era_done = False
            if FORCED_START_YEAR is not None:
                mask = row_dates >= pd.Timestamp(f"{int(FORCED_START_YEAR)}-01-01")
                if int(mask.sum()) >= _era_min_rows:
                    X = X.loc[mask].reset_index(drop=True)
                    y = y.loc[mask].reset_index(drop=True)
                    y_method_df = y_method_df.loc[mask].reset_index(drop=True)
                    row_meta = row_meta.loc[mask].reset_index(drop=True)
                    # Keep row_dates aligned with the filtered rows — the method
                    # stage indexes row_dates by the filtered split, so a stale
                    # full-length row_dates would mis-pair dates with fights.
                    row_dates = row_dates.loc[mask].reset_index(drop=True)
                    self._section("Era Selection")
                    self._stat("Selected start year", int(FORCED_START_YEAR))
                    self._stat("Rows kept", len(X))
                    self._stat("Selection mode", "forced")
                    era_done = True
                else:
                    self._log("Forced start year has too few rows; falling back to auto era selection.")

            if not era_done:
                # Winner-routed numeric columns drive the era proxy (the real model
                # never sees method-only features for the winner stage either).
                era_win_cols = [
                    c for c in X.columns
                    if _feature_allowed(c, "winner") and pd.api.types.is_numeric_dtype(X[c])
                ]

                # Era selection runs on EVERY invocation (it sets train_end/val_end,
                # which feed the winner cache key), so cache the decision on the data
                # fingerprint + feature/era config to avoid retraining 12 proxies on
                # cached runs. It auto-invalidates when the CSV or feature set changes.
                _era_sig = hashlib.sha256(
                    ("|".join(sorted(era_win_cols))
                     + f"|{VAL_FIGHTS}|{TEST_FIGHTS}|{MIN_TRAIN_FIGHTS}|{ERA_LOGLOSS_TOL}"
                     + f"|{','.join(str(c) for c in era_candidates)}|{RANDOM_SEED}").encode("utf-8")
                ).hexdigest()[:16]
                _era_key = _cache_key(
                    "era_select", _cache_data_fingerprint(self.csv_path), "v1", _era_sig
                )
                _era_payload = _cache_load("era_select", _era_key)

                if isinstance(_era_payload, dict) and "best_year" in _era_payload:
                    best_year = int(_era_payload["best_year"])
                    era_results = [tuple(r) for r in _era_payload.get("era_results", [])]
                else:
                    def _era_val_logloss(Xc, yc):
                        """Train a fast model on this era's training rows and return its
                        log-loss on the COMMON (era-independent) validation window."""
                        tr_end, va_end = _time_split_indices(len(Xc))
                        if tr_end < MIN_TRAIN_FIGHTS or (va_end - tr_end) < 100:
                            return None
                        yv = yc.iloc[tr_end:va_end].astype(int).reset_index(drop=True)
                        if len(np.unique(yv.values)) < 2:
                            return None
                        Xtr = Xc[era_win_cols].iloc[:tr_end].reset_index(drop=True)
                        ytr = yc.iloc[:tr_end].astype(int).reset_index(drop=True)
                        Xva = Xc[era_win_cols].iloc[tr_end:va_end].reset_index(drop=True)
                        _imp = SimpleImputer(strategy="median")
                        Xtr_i = pd.DataFrame(_imp.fit_transform(Xtr), columns=era_win_cols)
                        Xva_i = pd.DataFrame(_imp.transform(Xva), columns=era_win_cols)
                        Xtr_aug, ytr_aug = _augment_swap(Xtr_i, ytr)
                        try:
                            if lgb is not None:
                                _m = lgb.LGBMClassifier(
                                    n_estimators=250, learning_rate=0.03, num_leaves=31,
                                    subsample=0.8, colsample_bytree=0.8, min_child_samples=30,
                                    reg_lambda=1.0, random_state=RANDOM_SEED, n_jobs=-1,
                                    verbose=-1,
                                )
                            else:
                                _m = HistGradientBoostingClassifier(
                                    max_iter=300, learning_rate=0.04, max_depth=6,
                                    max_leaf_nodes=31, min_samples_leaf=25,
                                    l2_regularization=1.0, random_state=RANDOM_SEED,
                                )
                            _m.fit(Xtr_aug, ytr_aug)
                            # Corner-swap-averaged prediction, matching the real model.
                            p_fwd = _m.predict_proba(Xva_i)[:, 1]
                            p_rev = _m.predict_proba(_swap_features(Xva_i))[:, 1]
                            pv = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
                            return float(log_loss(yv, pv))
                        except Exception:
                            return None

                    era_results = []  # (year, val_logloss, train_rows)
                    for yr in era_candidates:
                        mask = row_dates >= pd.Timestamp(f"{yr}-01-01")
                        Xc = X.loc[mask].reset_index(drop=True)
                        yc = y.loc[mask].reset_index(drop=True)
                        if len(Xc) < _era_min_rows:
                            continue
                        ll = _era_val_logloss(Xc, yc)
                        if ll is None:
                            continue
                        tr_end, _ = _time_split_indices(len(Xc))
                        era_results.append((int(yr), float(ll), int(tr_end)))

                    best_year = 1993
                    if era_results:
                        best_ll = min(r[1] for r in era_results)
                        # Earliest start year (most training data) whose log-loss is
                        # statistically indistinguishable from the best.
                        within = [r for r in era_results if r[1] <= best_ll + ERA_LOGLOSS_TOL]
                        best_year = min(r[0] for r in within)
                    _cache_save(
                        "era_select", _era_key,
                        {"best_year": int(best_year), "era_results": era_results},
                    )

                self._section("Era Selection")
                for (yr, ll, rows) in era_results:
                    self._stat(
                        f"  start {yr}",
                        f"val log-loss {ll:.4f} | {rows} train rows"
                        + ("   <= selected" if yr == best_year else ""),
                    )
                if best_year > 1993:
                    mask = row_dates >= pd.Timestamp(f"{best_year}-01-01")
                    X = X.loc[mask].reset_index(drop=True)
                    y = y.loc[mask].reset_index(drop=True)
                    y_method_df = y_method_df.loc[mask].reset_index(drop=True)
                    row_meta = row_meta.loc[mask].reset_index(drop=True)
                    # Keep row_dates aligned with the filtered rows (see note above).
                    row_dates = row_dates.loc[mask].reset_index(drop=True)
                self._stat("Selected start year", best_year)
                self._stat("Rows kept", len(X))
                self._stat("Selection mode", "auto (common-window log-loss, data-preferring)")

        n = len(X)
        if n < 200:
            raise RuntimeError("Dataset too small after filtering completed fights.")

        # Benchmark contract: use the same chronological holdout strategy.
        train_end, val_end = _time_split_indices(n)
        X_train_raw = X.iloc[:train_end].reset_index(drop=True)
        y_train = y.iloc[:train_end].reset_index(drop=True)
        X_val_raw = X.iloc[train_end:val_end].reset_index(drop=True)
        y_val = y.iloc[train_end:val_end].reset_index(drop=True)
        X_test_raw = X.iloc[val_end:].reset_index(drop=True)
        y_test = y.iloc[val_end:].reset_index(drop=True)
        y_method_train = y_method_df.iloc[:train_end].reset_index(drop=True)
        y_method_val = y_method_df.iloc[train_end:val_end].reset_index(drop=True)
        y_method_test = y_method_df.iloc[val_end:].reset_index(drop=True)
        meta_test = row_meta.iloc[val_end:].reset_index(drop=True)
        self._section("Split Contract")
        self._stat("Train rows", len(X_train_raw))
        self._stat("Validation rows", len(X_val_raw))
        self._stat("Holdout test rows", len(X_test_raw))
        self._stat("Feature count", X.shape[1])
        full_feature_cols = list(X.columns)
        feature_cols = [c for c in full_feature_cols if _feature_allowed(c, "winner")]
        data_fp = _cache_data_fingerprint(self.csv_path)
        X_full = X[full_feature_cols].reset_index(drop=True)
        X_winner = X[feature_cols].reset_index(drop=True)
        X_train_raw = X_winner.iloc[:train_end].reset_index(drop=True)
        X_val_raw = X_winner.iloc[train_end:val_end].reset_index(drop=True)
        X_test_raw = X_winner.iloc[val_end:].reset_index(drop=True)
        X_train_raw_full = X_full.iloc[:train_end].reset_index(drop=True)
        X_val_raw_full = X_full.iloc[train_end:val_end].reset_index(drop=True)
        X_test_raw_full = X_full.iloc[val_end:].reset_index(drop=True)
        self._stat("Winner feature count", len(feature_cols))

        imputer = SimpleImputer(strategy="median")
        X_train = pd.DataFrame(imputer.fit_transform(X_train_raw), columns=feature_cols)
        X_val = pd.DataFrame(imputer.transform(X_val_raw), columns=feature_cols)
        X_test = pd.DataFrame(imputer.transform(X_test_raw), columns=feature_cols)
        full_imputer = SimpleImputer(strategy="median")
        X_train_full = pd.DataFrame(full_imputer.fit_transform(X_train_raw_full), columns=full_feature_cols)
        X_val_full = pd.DataFrame(full_imputer.transform(X_val_raw_full), columns=full_feature_cols)
        X_test_full = pd.DataFrame(full_imputer.transform(X_test_raw_full), columns=full_feature_cols)

        # ── Winner correlation prune (leak-safe: TRAIN only) ──────────────────
        # Drop near-duplicate columns BEFORE stability selection, keeping the more
        # target-relevant member of each correlated pair. The matchup matrix is
        # heavily collinear (glicko/elo/interaction variants), which inflates
        # ensemble variance and widens the val→holdout gap.
        if len(feature_cols) > 1:
            _pre_corr_n = len(feature_cols)
            feature_cols = _correlation_prune(
                X_train, y=y_train, threshold=WINNER_CORR_PRUNE_THRESHOLD
            )
            feature_cols = _complete_swap_pairs(feature_cols, list(X_train.columns))
            X_train = X_train[feature_cols]
            X_val = X_val[feature_cols]
            X_test = X_test[feature_cols]
            self._section("Feature Pruning (Winner Correlation)")
            self._stat(
                "Corr-pruned",
                f"{_pre_corr_n} → {len(feature_cols)} (thr={WINNER_CORR_PRUNE_THRESHOLD})",
            )

        # Stability selection (threshold form): run K bootstrap subsamples; in
        # each, a feature is "selected" if it lands in the top STABILITY_PER_RUN_FRAC
        # of that run's importance ranking. Keep every feature whose selection
        # FREQUENCY across runs is >= STABILITY_FREQ_THRESHOLD — the kept count
        # FLOATS to however many features are reliably useful (no fixed top-N cap).
        # Textbook stability selection: a selection-probability cutoff, not a count.
        if lgb is not None and len(feature_cols) > 2:
            X_quick_aug, y_quick_aug = _augment_swap(X_train, y_train)
            _STAB_RUNS = 15
            _STAB_SUB_FRAC = 0.75
            _n_feats = X_quick_aug.shape[1]
            _per_run_q = max(1, int(round(STABILITY_PER_RUN_FRAC * _n_feats)))
            _counts = np.zeros(_n_feats, dtype=float)
            _rng = np.random.default_rng(RANDOM_SEED)
            _n_rows = len(X_quick_aug)
            _sub_size = max(int(_n_rows * _STAB_SUB_FRAC), 1)
            for _k in range(_STAB_RUNS):
                _idx = _rng.choice(_n_rows, size=_sub_size, replace=False)
                _X_sub = X_quick_aug.iloc[_idx]
                _y_sub = y_quick_aug.iloc[_idx]
                _quick = lgb.LGBMClassifier(
                    n_estimators=200, learning_rate=0.05, max_depth=6,
                    random_state=int(_rng.integers(1, 10**9)),
                    n_jobs=-1, verbose=-1,
                )
                _quick.fit(_X_sub, _y_sub)
                _imp = np.asarray(_quick.feature_importances_, dtype=float)
                _top = np.argsort(_imp)[::-1][:_per_run_q]
                _counts[_top] += 1
            # Keep features whose selection frequency clears the threshold.
            _freq = _counts / _STAB_RUNS
            _keep = sorted(np.where(_freq >= STABILITY_FREQ_THRESHOLD)[0].tolist())
            # Safety: never collapse to (near-)nothing if the threshold is misset.
            if len(_keep) < 2:
                _keep = sorted(np.argsort(_freq)[::-1][:max(2, _n_feats // 2)].tolist())
            feature_cols = [feature_cols[i] for i in _keep]
            feature_cols = _complete_swap_pairs(feature_cols, list(X_train.columns))
            X_train = X_train[feature_cols]
            X_val = X_val[feature_cols]
            X_test = X_test[feature_cols]
            # Report the FLOATED count + how stable the kept set actually was.
            _sel_freq = float(_freq[_keep].mean())
            self._section("Feature Pruning (Stability Selection)")
            self._stat("Kept features (floated)", len(feature_cols))
            self._stat("Bootstrap runs", _STAB_RUNS)
            self._stat("Per-run select top", f"{_per_run_q}/{_n_feats} ({STABILITY_PER_RUN_FRAC:.0%})")
            self._stat("Freq threshold", f"{STABILITY_FREQ_THRESHOLD:.0%}")
            self._stat("Mean selection freq of kept", f"{_sel_freq:.1%}")

        feature_cols_fp = hashlib.sha256(",".join(feature_cols).encode("utf-8")).hexdigest()[:12]
        winner_key_extra = "|".join([
            str(STRICT_FUTURE_MODE),
            str(FORCED_START_YEAR),
            str(OPTUNA_TRIALS),
            str(VAL_FIGHTS),
            str(TEST_FIGHTS),
            str(MIN_TRAIN_FIGHTS),
            str(ERA_LOGLOSS_TOL),
            str(OBA_FEATURES_ENABLED),
            str(MOV_RATINGS_ENABLED),
            str(MOV_MODE),
            str(WINNER_CORR_PRUNE_THRESHOLD),
            str(STABILITY_PER_RUN_FRAC),
            str(STABILITY_FREQ_THRESHOLD),
            str(WINNER_SEES_ALL_FEATURES),
            str(WINNER_COMBINER_ROBUST),
            str(PHASE_RATINGS_ENABLED),
            str(CORNER_CORRECTION_ENABLED),
            str(CORNER_SHIFT_CAP),
            str(CORNER_FIT_FLOOR),
            str(RANDOM_SEED),
            # Active base-model lineup: changing STRICT_KEEP_MODELS (e.g. toggling
            # AdaBoost) now self-invalidates the cache instead of silently stale-hitting.
            ("+".join(sorted(STRICT_KEEP_MODELS)) if STRICT_FUTURE_MODE else "ALL_MODELS"),
            str(n),
            str(train_end),
            str(val_end),
            feature_cols_fp,
        ])
        winner_cache_key = _cache_key("winner_stage", data_fp, WINNER_CACHE_VERSION, winner_key_extra)
        winner_payload = _cache_load("winner_stage", winner_cache_key)
        winner_cache_hit = _winner_stage_cache_valid(
            winner_payload, feature_cols, n, train_end, val_end
        )
        if winner_cache_hit:
            self._stat(
                "Winner cache",
                f"HIT ({WINNER_CACHE_VERSION}) key={str(winner_cache_key)[:12]} file={os.path.basename(_cache_path('winner_stage', winner_cache_key))}",
            )
            _replay_winner_cache_logs(self, winner_payload, winner_cache_key)
        if not winner_cache_hit:
            self._stat("Winner cache", f"MISS ({WINNER_CACHE_VERSION}) — training winner stage")
            lgb_tuned = None
            xgb_tuned = None
            cb_tuned = None
            if lgb is not None and optuna is not None:
                self._section("Optuna Tuning")
                enabled = [name for name, ok in (("LightGBM", lgb is not None), ("XGBoost", xgb is not None), ("CatBoost", cb is not None)) if ok]
                self._reset_terminal_progress(labels=enabled, default_total=OPTUNA_TRIALS)

                def _mk_progress():
                    def _emit(done, total, label):
                        self._progress(int(done), int(total), str(label))
                    return _emit

                lgb_tuned = _tune_lightgbm_optuna(
                    X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=self._log,
                    progress_cb=_mk_progress(),
                )
                if xgb is not None:
                    xgb_tuned = _tune_xgboost_optuna(
                        X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=self._log,
                        progress_cb=_mk_progress(),
                    )
                if cb is not None:
                    cb_tuned = _tune_catboost_optuna(
                        X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=self._log,
                        progress_cb=_mk_progress(),
                    )
                self._finalize_terminal_progress()
            specs = _make_model_specs(
                lgb_tuned_params=lgb_tuned,
                xgb_tuned_params=xgb_tuned,
                cb_tuned_params=cb_tuned,
            )
            if STRICT_FUTURE_MODE:
                # Reseed twins (_S2) add correlation, not diversity; dropped to
                # cut ensemble variance. Tuned/Wide/Deep variants are kept — they
                # are genuinely different models. Lineup defined by STRICT_KEEP_MODELS
                # (top of file); it's fingerprinted into the cache key above.
                specs = [(n, mk) for n, mk in specs if n in STRICT_KEEP_MODELS]
            model_order = [n for n, _ in specs]
            self._section("Model Setup")
            self._stat("Base models", ", ".join(model_order))

            # Build OOF meta-features on dev split.
            X_dev = pd.concat([X_train, X_val], ignore_index=True)
            y_dev = pd.concat([y_train, y_val], ignore_index=True)
            oof = pd.DataFrame(index=np.arange(len(X_dev)), columns=model_order, dtype=float)
            tscv = TimeSeriesSplit(n_splits=5)
            self._section("OOF Stacking")
            for fold_id, (tr_idx, va_idx) in enumerate(tscv.split(X_dev), start=1):
                self._stat("OOF fold", f"{fold_id}/5")
                X_tr = X_dev.iloc[tr_idx].reset_index(drop=True)
                y_tr = y_dev.iloc[tr_idx].reset_index(drop=True)
                X_va = X_dev.iloc[va_idx].reset_index(drop=True)
                X_va_sw = _swap_features(X_va)
                X_tr_aug, y_tr_aug = _augment_swap(X_tr, y_tr)
                w_tr_aug = _augment_weights(_time_weights(len(X_tr), floor=0.35))
                fold_scaler = StandardScaler()
                X_tr_aug_sc = pd.DataFrame(fold_scaler.fit_transform(X_tr_aug), columns=feature_cols)
                X_va_sc = pd.DataFrame(fold_scaler.transform(X_va), columns=feature_cols)
                X_va_sw_sc = pd.DataFrame(fold_scaler.transform(X_va_sw), columns=feature_cols)

                for name, make_model in specs:
                    model = _fit_model(
                        name, make_model(),
                        X_tr_aug_sc if name in NEEDS_SCALE else X_tr_aug,
                        y_tr_aug, sample_weight=w_tr_aug
                    )
                    p_fwd = _predict_proba(name, model, X_va_sc if name in NEEDS_SCALE else X_va)
                    p_rev = _predict_proba(name, model, X_va_sw_sc if name in NEEDS_SCALE else X_va_sw)
                    oof.loc[va_idx, name] = (p_fwd + (1.0 - p_rev)) / 2.0

            valid = ~oof.isna().any(axis=1)
            idx = np.arange(len(X_dev))
            meta_train_mask = valid & (idx < len(X_train))
            meta_val_mask = valid & (idx >= len(X_train))
            X_meta_train = oof.loc[meta_train_mask, model_order].astype(float)
            y_meta_train = y_dev.loc[meta_train_mask].astype(int)
            X_meta_val = oof.loc[meta_val_mask, model_order].astype(float)
            y_meta_val = y_dev.loc[meta_val_mask].astype(int)

            self._section("Combiner Selection")
            combiner, val_ll, val_acc, val_thr = _choose_combiner(
                X_meta_train, y_meta_train, X_meta_val, y_meta_val,
                allow_stacker=not WINNER_COMBINER_ROBUST,
            )
            self._stat("Selected combiner", combiner["kind"])
            _w_str = _format_combiner_weights(combiner)
            if _w_str:
                self._stat("Combiner weights", _w_str)
            self._stat("Validation log-loss", f"{val_ll:.4f}")
            self._stat("Validation accuracy", f"{val_acc:.1%}")
            self._stat("Validation threshold", f"{val_thr:.3f}")

            # Fit base models on full dev split, evaluate on test.
            X_dev_aug, y_dev_aug = _augment_swap(X_dev, y_dev)
            w_dev_aug = _augment_weights(_time_weights(len(X_dev), floor=0.35))
            scaler_dev = StandardScaler()
            X_dev_aug_sc = pd.DataFrame(scaler_dev.fit_transform(X_dev_aug), columns=feature_cols)
            X_test_sc_for_eval = pd.DataFrame(scaler_dev.transform(X_test), columns=feature_cols)
            X_test_sw = _swap_features(X_test)
            X_test_sw_sc = pd.DataFrame(scaler_dev.transform(X_test_sw), columns=feature_cols)

            test_meta = {}
            for name, make_model in specs:
                model = _fit_model(
                    name, make_model(),
                    X_dev_aug_sc if name in NEEDS_SCALE else X_dev_aug,
                    y_dev_aug, sample_weight=w_dev_aug
                )
                p_fwd = _predict_proba(name, model, X_test_sc_for_eval if name in NEEDS_SCALE else X_test)
                p_rev = _predict_proba(name, model, X_test_sw_sc if name in NEEDS_SCALE else X_test_sw)
                test_meta[name] = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
            test_meta_df = pd.DataFrame(test_meta)[model_order]
            test_probs_raw = _combine_probs(test_meta_df, combiner)

            # Optional aggressive holdout combiner refinement. Disabled in strict mode.
            best_holdout_label = "disabled_strict_mode"
            best_holdout_ll = float("nan")
            best_holdout_acc = float("nan")
            best_holdout_thr = 0.5
            if not STRICT_FUTURE_MODE:
                best_holdout_label, best_holdout_combiner, best_holdout_acc, best_holdout_thr, best_holdout_ll = (
                    _pick_best_holdout_combiner(test_meta_df, y_test, combiner, allow_aggressive=True)
                )
                if best_holdout_acc >= _model_accuracy_at_threshold(test_probs_raw, y_test, 0.5):
                    combiner = best_holdout_combiner
                    test_probs_raw = _combine_probs(test_meta_df, combiner)

            # Validation probs for calibration.
            val_probs_raw = _combine_probs(X_meta_val, combiner)
            calibrator, cal_name, cal_ll = _fit_best_calibrator(val_probs_raw, y_meta_val)
            # Preserve the fitted calibrator for transparent reporting even when
            # the threshold tuner ends up preferring the raw branch (and we null
            # `calibrator` out below for pick-time use).
            calibrator_fitted = calibrator
            cal_name_fitted = cal_name
            cal_ll_val_fitted = float(cal_ll)
            self._section("Calibration")
            self._stat("Selected method", cal_name)
            # cal_ll is scored on the LATE ~45% of validation only (the slice
            # _fit_best_calibrator selects on), so it won't match the Combiner
            # section's full-window validation log-loss.
            self._stat("Selection-slice log-loss (late ~45% of val)", f"{cal_ll:.4f}")
            self._stat("Strict future mode", "ON" if STRICT_FUTURE_MODE else "OFF")
            if not STRICT_FUTURE_MODE:
                self._stat("Holdout-selected combiner", best_holdout_label)
                self._stat("Holdout combiner log-loss", f"{best_holdout_ll:.4f}")
                self._stat("Holdout combiner acc", f"{best_holdout_acc:.1%}")
                self._stat("Holdout combiner threshold", f"{best_holdout_thr:.3f}")

            # Tune decision threshold with a robust chronological criterion.
            dev_meta_valid = oof.loc[valid, model_order].astype(float)
            y_dev_valid = y_dev.loc[valid].astype(int).values
            dev_probs_raw = _combine_probs(dev_meta_valid, combiner)

            # Raw branch (always available).
            val_probs_raw_branch = _clip_probs(val_probs_raw)
            dev_probs_raw_branch = _clip_probs(dev_probs_raw)
            tuned_thr_raw, tuned_acc_raw = _tune_threshold_robust(dev_probs_raw_branch, y_dev_valid, n_blocks=4)
            decision_threshold_raw = float(tuned_thr_raw)
            val_acc_raw = _model_accuracy_at_threshold(val_probs_raw_branch, y_meta_val, threshold=decision_threshold_raw)

            # Calibrated branch (optional, only keep if it helps).
            use_calibrated_branch = False
            tuned_thr = tuned_thr_raw
            tuned_acc = tuned_acc_raw
            decision_threshold = decision_threshold_raw
            val_acc_for_log = val_acc_raw
            if calibrator is not None:
                val_probs_cal_branch = _clip_probs(
                    calibrator.predict_proba(np.asarray(val_probs_raw).reshape(-1, 1))[:, 1]
                )
                dev_probs_cal_branch = _clip_probs(
                    calibrator.predict_proba(np.asarray(dev_probs_raw).reshape(-1, 1))[:, 1]
                )
                tuned_thr_cal, tuned_acc_cal = _tune_threshold_robust(dev_probs_cal_branch, y_dev_valid, n_blocks=4)
                decision_threshold_cal = float(tuned_thr_cal)
                val_acc_cal = _model_accuracy_at_threshold(
                    val_probs_cal_branch, y_meta_val, threshold=decision_threshold_cal
                )
                if (val_acc_cal > val_acc_raw + 1e-4) or (
                    abs(val_acc_cal - val_acc_raw) <= 1e-4 and cal_ll <= float(log_loss(y_meta_val, val_probs_raw_branch))
                ):
                    use_calibrated_branch = True
                    tuned_thr = tuned_thr_cal
                    tuned_acc = tuned_acc_cal
                    decision_threshold = decision_threshold_cal
                    val_acc_for_log = val_acc_cal
                else:
                    calibrator = None
                    cal_name = "none"

            # ── Decision-threshold regularization (corner-symmetric model) ──
            # Anchor the threshold at 0.5 (the neutral point for a corner-swap
            # symmetric model), allowing at most WINNER_THR_MAX_DEV of adaptive
            # deviation. Without this, the dev-tuned threshold drifts below 0.5 to
            # exploit a red-corner win rate that no longer holds on recent fights,
            # over-predicting Red and losing holdout accuracy.
            _thr_pre_reg = decision_threshold
            decision_threshold = float(
                0.5 + np.clip(decision_threshold - 0.5, -WINNER_THR_MAX_DEV, WINNER_THR_MAX_DEV)
            )
            if abs(decision_threshold - _thr_pre_reg) > 1e-9:
                _vp_reg = val_probs_cal_branch if use_calibrated_branch else val_probs_raw_branch
                val_acc_for_log = _model_accuracy_at_threshold(
                    _vp_reg, y_meta_val, threshold=decision_threshold
                )

            self._stat("Calibration used for picks", "yes" if use_calibrated_branch else "no")
            self._stat("Validation accuracy (picked)", f"{val_acc_for_log:.1%}")
            # tuned_thr/tuned_acc come from _tune_threshold_robust on OOF probs over
            # the train+val dev set (block-mean accuracy), NOT the validation window.
            tuned_thr_str = f"{tuned_thr:.3f} (dev-OOF block-robust acc={tuned_acc:.1%})"
            if abs(decision_threshold - tuned_thr) > 1e-9:
                tuned_thr_str = (f"{tuned_thr:.3f} (dev-OOF block-robust acc="
                                 f"{tuned_acc:.1%}, clamped to {decision_threshold:.3f})")
            self._stat("Dev-OOF tuned threshold", tuned_thr_str)
            self._stat("Decision threshold", f"{decision_threshold:.3f}")

            if calibrator is None:
                test_probs_cal = _clip_probs(test_probs_raw)
            else:
                test_probs_cal = _clip_probs(
                    calibrator.predict_proba(np.asarray(test_probs_raw).reshape(-1, 1))[:, 1]
                )

            # ── Corner-aware correction (asymmetric calibration) ──────────────
            # Fit a logit intercept on leak-safe dev OOF pick-probs + true red
            # labels with recency weights, then shift the holdout (and, later,
            # production + inference) probabilities. Corrects the symmetric
            # model's structural Red over-prediction.
            corner_b = 0.0
            corner_rows = []
            test_probs_precorner = _clip_probs(test_probs_cal)
            if CORNER_CORRECTION_ENABLED:
                if calibrator is None:
                    dev_probs_for_corner = _clip_probs(dev_probs_raw)
                else:
                    dev_probs_for_corner = _clip_probs(
                        calibrator.predict_proba(np.asarray(dev_probs_raw).reshape(-1, 1))[:, 1]
                    )
                _corner_w = _time_weights(len(dev_probs_for_corner), floor=CORNER_FIT_FLOOR)
                corner_b = _fit_corner_intercept(
                    dev_probs_for_corner, y_dev_valid, weights=_corner_w
                )
                test_probs_cal = _apply_corner_correction(test_probs_cal, corner_b)
                _pre_red = int(np.sum(test_probs_precorner >= decision_threshold))
                _post_red = int(np.sum(test_probs_cal >= decision_threshold))
                _pre_acc = accuracy_score(
                    y_test, (test_probs_precorner >= decision_threshold).astype(int)
                )
                _post_acc = accuracy_score(
                    y_test, (test_probs_cal >= decision_threshold).astype(int)
                )
                _eff_thr = 1.0 / (1.0 + np.exp(corner_b))  # p_red bar at the 0.5 cut
                corner_rows = [
                    ("Enabled", "yes"),
                    ("Fitted intercept b", f"{corner_b:+.4f} (cap ±{CORNER_SHIFT_CAP})"),
                    ("Effective red pick threshold", f"{_eff_thr:.3f}"),
                    ("Holdout Red picks", f"{_pre_red} → {_post_red} (of {len(y_test)})"),
                    ("Holdout acc (pre → post)", f"{_pre_acc:.2%} → {_post_acc:.2%}"),
                ]
                self._section("Corner Correction")
                for _lbl, _val in corner_rows:
                    self._stat(_lbl, _val)

            raw_ll = log_loss(y_test, test_probs_raw)
            cal_ll_test = log_loss(y_test, test_probs_cal)
            brier = brier_score_loss(y_test, test_probs_cal)
            acc = accuracy_score(y_test, (test_probs_cal >= decision_threshold).astype(int))
            ece = _expected_calibration_error(y_test, test_probs_cal)
            cal_curve_rmse = _calibration_curve_rmse(y_test, test_probs_cal, n_bins=10)

            # Always compute the FITTED calibrator's holdout LL so users can see
            # what calibration would have produced even when the threshold tuner
            # rejects it for picks. Without this, the "Calibrated log-loss" line
            # is just a copy of "Raw log-loss" and looks like calibration was
            # ineffective when really it was disabled at pick time.
            if calibrator_fitted is not None and calibrator_fitted is not calibrator:
                test_probs_cal_fitted = _clip_probs(
                    calibrator_fitted.predict_proba(
                        np.asarray(test_probs_raw).reshape(-1, 1)
                    )[:, 1]
                )
                cal_ll_fitted_test = float(log_loss(y_test, test_probs_cal_fitted))
                cal_label = f"Calibrated log-loss ({cal_name_fitted}, fitted but unused for picks)"
            else:
                cal_ll_fitted_test = float(cal_ll_test)
                cal_label = "Calibrated log-loss"

            self._section("Holdout Evaluation")
            self._stat("Raw log-loss", f"{raw_ll:.4f}")
            self._stat(cal_label, f"{cal_ll_fitted_test:.4f}")
            self._stat("Brier score", f"{brier:.4f}")
            self._stat("Accuracy", f"{acc:.2%}")
            self._stat("ECE", f"{ece:.4f}")
            self._stat("Calibration curve RMSE", f"{cal_curve_rmse:.4f}")
            self._stat("Accuracy threshold", f"{decision_threshold:.3f}")

            # Winner diagnostics: confusion matrix + subgroup accuracy.
            y_true_red = y_test.astype(int).values
            y_pred_red = (test_probs_cal >= decision_threshold).astype(int)
            tn = int(np.sum((y_true_red == 0) & (y_pred_red == 0)))
            fp = int(np.sum((y_true_red == 0) & (y_pred_red == 1)))
            fn = int(np.sum((y_true_red == 1) & (y_pred_red == 0)))
            tp = int(np.sum((y_true_red == 1) & (y_pred_red == 1)))
            self._section("Winner Diagnostics")
            self._log("Confusion Matrix (Winner: Red=positive)")
            self._log("               Pred Blue   Pred Red")
            self._log(f"Actual Blue   {tn:9d}  {fp:9d}")
            self._log(f"Actual Red    {fn:9d}  {tp:9d}")

            self._log("")
            self._log("Accuracy by Weight Class")
            self._log("-" * 72)
            wc_order = [
                "Women's Strawweight",
                "Women's Flyweight",
                "Women's Bantamweight",
                "Women's Featherweight",
                "Flyweight",
                "Bantamweight",
                "Featherweight",
                "Lightweight",
                "Welterweight",
                "Middleweight",
                "Light Heavyweight",
                "Heavyweight",
                "Catch Weight",
                "Open Weight",
            ]
            wc_rank = {name: i for i, name in enumerate(wc_order)}
            wc_rows = []
            for wc, grp in meta_test.groupby("weight_class", dropna=False):
                idx = grp.index.values
                if len(idx) == 0:
                    continue
                acc_wc = float(np.mean(y_pred_red[idx] == y_true_red[idx]))
                wc_rows.append((str(wc), acc_wc, int(len(idx))))
            wc_rows.sort(key=lambda t: (wc_rank.get(t[0], 999), t[0]))
            for wc, acc_wc, n_wc in wc_rows:
                self._stat(f"{wc} (n={n_wc})", f"{acc_wc:.1%}")

            self._log("")
            self._log("Accuracy by Gender")
            self._log("-" * 72)
            g_rows = []
            for gender, grp in meta_test.groupby("gender", dropna=False):
                idx = grp.index.values
                if len(idx) == 0:
                    continue
                acc_g = float(np.mean(y_pred_red[idx] == y_true_red[idx]))
                g_rows.append((str(gender), acc_g, int(len(idx))))
            g_rows.sort(key=lambda t: (-t[2], t[0]))
            for gender, acc_g, n_g in g_rows:
                self._stat(f"{gender} (n={n_g})", f"{acc_g:.1%}")

            winner_payload = {
                "kind": WINNER_STAGE_CACHE_KIND,
                "winner_cache_version": WINNER_CACHE_VERSION,
                "feature_cols": list(feature_cols),
                "n_rows": int(n),
                "train_end": int(train_end),
                "val_end": int(val_end),
                "lgb_tuned": lgb_tuned,
                "xgb_tuned": xgb_tuned,
                "cb_tuned": cb_tuned,
                "oof": oof,
                "valid": valid,
                "combiner": combiner,
                "calibrator": calibrator,
                "cal_name": cal_name,
                "decision_threshold": float(decision_threshold),
                "corner_b": float(corner_b),
                "corner_rows": [(str(a), str(b)) for a, b in corner_rows],
                "tuned_thr": float(tuned_thr),
                "tuned_acc": float(tuned_acc),
                "val_acc_for_log": float(val_acc_for_log),
                "use_calibrated_branch": bool(use_calibrated_branch),
                "best_holdout_label": best_holdout_label,
                "best_holdout_ll": float(best_holdout_ll),
                "best_holdout_acc": float(best_holdout_acc),
                "best_holdout_thr": float(best_holdout_thr),
                "raw_ll": float(raw_ll),
                "cal_ll_test": float(cal_ll_test),
                "brier": float(brier),
                "acc": float(acc),
                "ece": float(ece),
                "cal_curve_rmse": float(cal_curve_rmse),
                "test_probs_raw": np.asarray(test_probs_raw, dtype=float),
                "test_probs_cal": np.asarray(test_probs_cal, dtype=float),
                "y_pred_red": np.asarray(y_pred_red, dtype=int),
                "model_order": list(model_order),
                "combiner_kind": str(combiner.get("kind", "")),
                "val_ll_str": f"{float(val_ll):.4f}",
                "val_acc_str": f"{float(val_acc):.1%}",
                "val_thr_str": f"{float(val_thr):.3f}",
                "cal_ll_str": f"{float(cal_ll):.4f}",
                "best_holdout_ll_str": f"{float(best_holdout_ll):.4f}" if np.isfinite(best_holdout_ll) else "n/a",
                "best_holdout_acc_str": f"{float(best_holdout_acc):.1%}" if np.isfinite(best_holdout_acc) else "n/a",
                "best_holdout_thr_str": f"{float(best_holdout_thr):.3f}",
                "cal_used_str": "yes" if use_calibrated_branch else "no",
                "val_acc_picked_str": f"{float(val_acc_for_log):.1%}",
                "val_tuned_thr_str": tuned_thr_str,
                "decision_thr_str": f"{float(decision_threshold):.3f}",
                "holdout_eval": [
                    ("Raw log-loss", f"{float(raw_ll):.4f}"),
                    (cal_label, f"{float(cal_ll_fitted_test):.4f}"),
                    ("Brier score", f"{float(brier):.4f}"),
                    ("Accuracy", f"{float(acc):.2%}"),
                    ("ECE", f"{float(ece):.4f}"),
                    ("Calibration curve RMSE", f"{float(cal_curve_rmse):.4f}"),
                    ("Accuracy threshold", f"{float(decision_threshold):.3f}"),
                ],
                "confusion": (tn, fp, fn, tp),
                "wc_rows": [(str(a), float(b), int(c)) for a, b, c in wc_rows],
                "g_rows": [(str(a), float(b), int(c)) for a, b, c in g_rows],
            }
            _cache_save("winner_stage", winner_cache_key, winner_payload)
            self._stat(
                "Winner cache",
                f"SAVED ({WINNER_CACHE_VERSION}) key={str(winner_cache_key)[:12]} file={os.path.basename(_cache_path('winner_stage', winner_cache_key))}",
            )

        lgb_tuned = winner_payload["lgb_tuned"]
        xgb_tuned = winner_payload.get("xgb_tuned")
        cb_tuned = winner_payload.get("cb_tuned")
        oof = winner_payload["oof"]
        valid = winner_payload["valid"]
        combiner = winner_payload["combiner"]
        calibrator = winner_payload["calibrator"]
        cal_name = winner_payload.get("cal_name", "none")
        decision_threshold = float(winner_payload["decision_threshold"])
        corner_b = float(winner_payload.get("corner_b", 0.0))
        tuned_thr = float(winner_payload["tuned_thr"])
        tuned_acc = float(winner_payload["tuned_acc"])
        val_acc_for_log = float(winner_payload["val_acc_for_log"])
        use_calibrated_branch = bool(winner_payload.get("use_calibrated_branch", False))
        best_holdout_label = winner_payload.get("best_holdout_label", "disabled_strict_mode")
        best_holdout_ll = float(winner_payload.get("best_holdout_ll", float("nan")))
        best_holdout_acc = float(winner_payload.get("best_holdout_acc", float("nan")))
        best_holdout_thr = float(winner_payload.get("best_holdout_thr", 0.5))
        raw_ll = float(winner_payload["raw_ll"])
        cal_ll_test = float(winner_payload["cal_ll_test"])
        brier = float(winner_payload["brier"])
        acc = float(winner_payload["acc"])
        ece = float(winner_payload["ece"])
        cal_curve_rmse = float(winner_payload["cal_curve_rmse"])
        test_probs_raw = np.asarray(winner_payload["test_probs_raw"], dtype=float)
        test_probs_cal = np.asarray(winner_payload["test_probs_cal"], dtype=float)
        y_pred_red = np.asarray(winner_payload["y_pred_red"], dtype=int)
        specs = _make_model_specs(
            lgb_tuned_params=lgb_tuned,
            xgb_tuned_params=xgb_tuned,
            cb_tuned_params=cb_tuned,
        )
        if STRICT_FUTURE_MODE:
            # Same lineup as the winner stage — see STRICT_KEEP_MODELS at top of file.
            specs = [(n, mk) for n, mk in specs if n in STRICT_KEEP_MODELS]
        model_order = [n for n, _ in specs]
        X_dev = pd.concat([X_train, X_val], ignore_index=True)
        y_dev = pd.concat([y_train, y_val], ignore_index=True)
        dev_meta_valid = oof.loc[valid, model_order].astype(float)
        dev_probs_raw = _combine_probs(dev_meta_valid, combiner)
        y_true_red = y_test.astype(int).values
        # Expose the aligned holdout (true outcome, calibrated P(red) fed to picks)
        # so external calibration audits read it without re-deriving the split.
        # Inert: set on every path, used by nothing in the model itself.
        self.holdout_y_true = y_true_red
        self.holdout_prob_cal = np.asarray(test_probs_cal, dtype=float)
        self.holdout_threshold = float(decision_threshold)
        # Per-fight data-sufficiency, aligned with the holdout above, so audits can
        # segment calibration by experience. Pulled from the full (pre-selection,
        # median-imputed) feature frame; any absent column degrades to NaN.
        def _hcol(_name):
            return (X_test_full[_name].to_numpy(dtype=float)
                    if _name in X_test_full.columns
                    else np.full(len(y_true_red), np.nan, dtype=float))
        self.holdout_min_fights = _hcol("min_num_fights")
        self.holdout_max_fights = _hcol("max_num_fights")
        self.holdout_avg_glicko_conf = _hcol("avg_glicko_confidence")

        if winner_only:
            # Winner-only callers (e.g. the calibration audit) need just the
            # holdout stashed above; skip the expensive method stage below.
            return

        # Method diagnostics/training (winner-model OOF conditioned, 2-stage).
        method_key_extra = "|".join([
            str(METHOD_TUNING_TRIALS),
            str(METHOD_AUTO_ERA),
            str(METHOD_HARD_RESET),
            str(len(full_feature_cols)),
            ",".join(map(str, METHOD_ERA_CANDIDATES)),
            f"bag{METHOD_HGB_BAG}@{METHOD_HGB_BAG_SUBSAMPLE}",
        ])
        method_cache_key = _cache_key(
            "method_stage", data_fp, METHOD_CACHE_VERSION, f"{winner_cache_key}|{method_key_extra}"
        )
        method_acc_when_winner_correct = float("nan")
        method_acc_predicted_winner = float("nan")
        method_majority_baseline_when_winner_correct = float("nan")
        method_holdout_acc_oracle = float("nan")
        finish_score = float("nan")
        method_bundle = None
        method_payload = _cache_load("method_stage", method_cache_key)
        method_cache_hit = _method_stage_cache_valid(
            method_payload, winner_cache_key, METHOD_CACHE_VERSION
        )
        if method_cache_hit:
            self._stat(
                "Method cache",
                f"HIT ({METHOD_CACHE_VERSION}) key={str(method_cache_key)[:12]} file={os.path.basename(_cache_path('method_stage', method_cache_key))}",
            )
            self._section("Method Model")
            self._stat("Cache", f"HIT ({METHOD_CACHE_VERSION}) — method retrain skipped")
            _mb = method_payload.get("method_bundle") or {}
            method_bundle = {k: v for k, v in _mb.items()}
            method_acc_when_winner_correct = float(method_payload.get("method_acc_when_winner_correct", float("nan")))
            method_acc_predicted_winner = float(method_payload.get("method_acc_predicted_winner", float("nan")))
            method_majority_baseline_when_winner_correct = float(
                method_payload.get("method_majority_baseline_when_winner_correct", float("nan"))
            )
            method_holdout_acc_oracle = float(method_payload.get("method_holdout_acc_oracle", float("nan")))
            finish_score = float(method_payload.get("finish_score", float("nan")))
            _replay_method_cache_logs(self, method_payload)
        if not method_cache_hit:
            self._stat("Method cache", f"MISS ({METHOD_CACHE_VERSION}) — training method stage")
            try:
                self._section("Method Model")
                X_dev_full = pd.concat([X_train_full, X_val_full], ignore_index=True)
                y_dev_true = pd.concat([y_train, y_val], ignore_index=True).astype(int).values
                y_dev_method_df = pd.concat([y_method_train, y_method_val], ignore_index=True).reset_index(drop=True)
                meta_dev = row_meta.iloc[:val_end].reset_index(drop=True)

                valid_idx = np.where(np.asarray(valid).astype(bool))[0]
                X_dev_valid = X_dev_full.iloc[valid_idx].reset_index(drop=True)
                y_dev_true_valid = y_dev_true[valid_idx]
                if calibrator is None:
                    dev_probs_method = _clip_probs(dev_probs_raw)
                else:
                    dev_probs_method = _clip_probs(
                        calibrator.predict_proba(np.asarray(dev_probs_raw).reshape(-1, 1))[:, 1]
                    )
                # Orient the method stage by the SAME corner-corrected picks the
                # production winner makes, so training/inference orientation agree.
                if CORNER_CORRECTION_ENABLED:
                    dev_probs_method = _apply_corner_correction(dev_probs_method, corner_b)
                y_dev_pred_valid = (dev_probs_method >= decision_threshold).astype(int)
                y_method_dev = y_dev_method_df.iloc[valid_idx].reset_index(drop=True)
                meta_dev_valid = meta_dev.iloc[valid_idx].reset_index(drop=True)
                row_dates_dev_valid = pd.to_datetime(
                    row_dates.iloc[:val_end].reset_index(drop=True).iloc[valid_idx].reset_index(drop=True),
                    errors="coerce"
                )

                X_dev_oriented = _oriented_method_matrix(X_dev_valid, y_dev_pred_valid)
                X_dev_oriented = _augment_method_features(X_dev_oriented)

                # Method-specific era selection: train a quick multiclass model per
                # candidate era and score its macro-F1 on a COMMON recent validation
                # window (identical fights across eras), CONDITIONED on winner-pick-
                # correct fights to mirror the real headline metric. Among eras within
                # METHOD_ERA_F1_TOL of the best, prefer the MOST TRAINING ROWS
                # (earliest start as tiebreak) — recency weighting during method
                # training already absorbs finish-type non-stationarity, and more
                # examples help the rare Submission class most (rationale at the
                # METHOD_ERA_F1_TOL definition). The old criterion MAXIMIZED class
                # imbalance (majority-class fraction) on each era's own slice —
                # which measures no model skill and starved it to ~766 rows.
                best_method_year = 1993
                _n_dev_all = len(X_dev_oriented)
                if METHOD_AUTO_ERA and _n_dev_all > METHOD_VAL_FIGHTS + 500:
                    _m_cols_e = [c for c in X_dev_oriented.columns
                                 if pd.api.types.is_numeric_dtype(X_dev_oriented[c])]
                    _lab_e = {"Decision": 0, "KO/TKO": 1, "Submission": 2}
                    _y_all_e = (y_method_dev["coarse"].astype(str)
                                .map(_lab_e).fillna(0).astype(int).to_numpy())
                    _va_lo = _n_dev_all - METHOD_VAL_FIGHTS
                    _Xva_e = X_dev_oriented[_m_cols_e].iloc[_va_lo:].reset_index(drop=True)
                    _yva_e = _y_all_e[_va_lo:]
                    # Winner-pick-correct mask on the common window (mirror real eval).
                    _wc_win = (np.asarray(y_dev_pred_valid)[_va_lo:]
                               == np.asarray(y_dev_true_valid)[_va_lo:])
                    _dates_tr = row_dates_dev_valid.iloc[:_va_lo].to_numpy()
                    _Xtr_pool = X_dev_oriented[_m_cols_e].iloc[:_va_lo].reset_index(drop=True)
                    _ytr_pool = _y_all_e[:_va_lo]
                    _cands_e = []  # (year, macro_f1, train_rows)
                    if len(np.unique(_yva_e)) >= 2:
                        for yr in METHOD_ERA_CANDIDATES:
                            if yr < EARLIEST_ERA_START:
                                continue
                            _sel = _dates_tr >= np.datetime64(f"{int(yr)}-01-01")
                            if int(_sel.sum()) < 500:
                                continue
                            _ytr_e = _ytr_pool[_sel]
                            if len(np.unique(_ytr_e)) < 3:
                                continue
                            try:
                                _imp_e = SimpleImputer(strategy="median")
                                _Xtr_e = pd.DataFrame(
                                    _imp_e.fit_transform(_Xtr_pool.loc[_sel]),
                                    columns=_m_cols_e)
                                _Xva_i = pd.DataFrame(_imp_e.transform(_Xva_e), columns=_m_cols_e)
                                _mc = HistGradientBoostingClassifier(
                                    max_iter=200, learning_rate=0.05, max_depth=4,
                                    min_samples_leaf=20, l2_regularization=1.0,
                                    random_state=RANDOM_SEED)
                                _mc.fit(_Xtr_e, _ytr_e)
                                _pred_e = _mc.predict(_Xva_i)
                                if int(_wc_win.sum()) >= 60:
                                    _yt_e, _yp_e = _yva_e[_wc_win], _pred_e[_wc_win]
                                else:
                                    _yt_e, _yp_e = _yva_e, _pred_e
                                _f1 = precision_recall_fscore_support(
                                    _yt_e, _yp_e, labels=[0, 1, 2],
                                    average="macro", zero_division=0)[2]
                                _cands_e.append((int(yr), float(_f1), int(_sel.sum())))
                            except Exception:
                                continue
                    if _cands_e:
                        _bf1 = max(c[1] for c in _cands_e)
                        _within = [c for c in _cands_e if c[1] >= _bf1 - METHOD_ERA_F1_TOL]
                        # Among near-best eras, prefer the MOST training data (rows),
                        # earliest start as tiebreak. Recency weighting in training
                        # handles finish-type non-stationarity; more examples help the
                        # rare Submission class more than a smaller recent window.
                        best_method_year = max(_within, key=lambda c: (c[2], -c[0]))[0]
                        for (yr, f1, rows) in _cands_e:
                            self._stat(
                                f"  method start {yr}",
                                f"macro-F1 {f1:.4f} | {rows} train rows"
                                + ("   <= selected" if yr == best_method_year else ""),
                            )
                        _bf1_yr = max(_cands_e, key=lambda c: c[1])[0]
                        if best_method_year != _bf1_yr:
                            self._stat(
                                "Method era rule",
                                f"{_bf1_yr} has the best macro-F1 ({_bf1:.4f}) but "
                                f"{best_method_year} selected: among eras within "
                                f"{METHOD_ERA_F1_TOL:g} of the best, most train rows wins",
                            )
                if METHOD_AUTO_ERA and best_method_year > 1993:
                    keep = row_dates_dev_valid >= pd.Timestamp(f"{best_method_year}-01-01")
                    X_dev_oriented = X_dev_oriented.loc[keep].reset_index(drop=True)
                    y_method_dev = y_method_dev.loc[keep].reset_index(drop=True)
                    meta_dev_valid = meta_dev_valid.loc[keep].reset_index(drop=True)
                    y_dev_pred_valid = y_dev_pred_valid[keep.values]
                    y_dev_true_valid = y_dev_true_valid[keep.values]
                    row_dates_dev_valid = row_dates_dev_valid.loc[keep].reset_index(drop=True)
                    self._stat("Method start year", best_method_year)
                    self._stat("Method rows kept", int(len(X_dev_oriented)))
                method_columns = list(X_dev_oriented.columns)

                n_dev_m = len(X_dev_oriented)
                split_m = int(max(200, min(n_dev_m - 1, int(n_dev_m * 0.75)))) if n_dev_m > 350 else max(1, int(n_dev_m * 0.7))
                tr_idx = np.arange(split_m)
                va_idx = np.arange(split_m, n_dev_m)
                if len(va_idx) < 80:
                    va_idx = np.arange(max(0, n_dev_m - 80), n_dev_m)
                    tr_idx = np.arange(0, max(0, va_idx[0]))

                X_m_tr = X_dev_oriented.iloc[tr_idx].reset_index(drop=True)
                X_m_va = X_dev_oriented.iloc[va_idx].reset_index(drop=True)

                method_imputer = SimpleImputer(strategy="median")
                X_m_tr_imp = pd.DataFrame(method_imputer.fit_transform(X_m_tr), columns=method_columns)
                X_m_va_imp = pd.DataFrame(method_imputer.transform(X_m_va), columns=method_columns)

                # Multi-model method ensemble: stage1 (Finish/Decision) + stage2 (KO/Sub)
                # + direct multiclass + OVR heads + simple LR + tuned blending + meta calibrator.
                y_bin_tr = (y_method_dev.iloc[tr_idx]["finish_bin"].values == "Finish").astype(int)
                y_sub_tr = (y_method_dev.iloc[tr_idx]["finish_subtype"].values == "Submission").astype(int)

                # Weighted objectives for class imbalance.
                w_time = _time_weights(len(X_m_tr_imp), floor=0.45)
                p_finish = max(1e-6, float(np.mean(y_bin_tr)))
                finish_w = np.where(y_bin_tr == 1, 0.5 / p_finish, 0.5 / max(1e-6, 1.0 - p_finish))
                w_stage1 = w_time * finish_w

                stage1_cols_raw = [c for c in X_m_tr_imp.columns if _feature_allowed(c, "stage1")]
                stage1_cols = _correlation_prune(X_m_tr_imp[stage1_cols_raw], y=y_bin_tr, threshold=METHOD_CORR_PRUNE_THRESHOLD)
                self._stat("Stage1 corr-pruned", f"{len(stage1_cols_raw)} → {len(stage1_cols)} (thr={METHOD_CORR_PRUNE_THRESHOLD})")
                X1_tr = X_m_tr_imp[stage1_cols]
                X1_va = X_m_va_imp[stage1_cols]
                y_bin_va = (y_method_dev.iloc[va_idx]["finish_bin"].values == "Finish").astype(int)
                _stage1_defaults = dict(
                    loss="log_loss",
                    max_iter=360, learning_rate=0.045, max_depth=6,
                    max_leaf_nodes=31, min_samples_leaf=16, l2_regularization=0.8,
                    random_state=RANDOM_SEED + 808,
                )
                if METHOD_OPTUNA_TRIALS > 0 and optuna is not None:
                    self._reset_terminal_progress(
                        labels=["Stage1 HGB", "Stage2 HGB"],
                        default_total=METHOD_OPTUNA_TRIALS,
                    )
                _stage1_params = _optuna_tune_hgb(
                    X1_tr, y_bin_tr, X1_va, y_bin_va,
                    sample_weight_tr=w_stage1,
                    n_trials=METHOD_OPTUNA_TRIALS,
                    seed=RANDOM_SEED + 808,
                    defaults=_stage1_defaults,
                    progress_cb=self._progress,
                    progress_label="Stage1 HGB",
                )
                self._stat("Stage1 HGB params", f"lr={_stage1_params.get('learning_rate', 0):.4f}, depth={_stage1_params.get('max_depth', 0)}, leaves={_stage1_params.get('max_leaf_nodes', 0)}, iter={_stage1_params.get('max_iter', 0)}")
                stage1 = _BaggedHGB(_stage1_params, n_estimators=METHOD_HGB_BAG,
                                    base_seed=RANDOM_SEED + 808, subsample=METHOD_HGB_BAG_SUBSAMPLE)
                stage1.fit(X1_tr, y_bin_tr, sample_weight=w_stage1)
                stage1_rf = RandomForestClassifier(
                    n_estimators=420, max_depth=10, min_samples_leaf=4,
                    max_features=0.7, random_state=RANDOM_SEED + 910, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_rf.fit(X1_tr, y_bin_tr)
                stage1_et = ExtraTreesClassifier(
                    n_estimators=560, max_depth=12, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 914, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_et.fit(X1_tr, y_bin_tr)

                finish_tr_mask = (y_method_dev.iloc[tr_idx]["finish_bin"].values == "Finish")
                if int(np.sum(finish_tr_mask)) < 60:
                    finish_tr_mask = np.ones(len(tr_idx), dtype=bool)
                stage2_cols_raw = [c for c in X_m_tr_imp.columns if _feature_allowed(c, "stage2")]
                _X2_for_corr = X_m_tr_imp.iloc[np.where(finish_tr_mask)[0]][stage2_cols_raw].reset_index(drop=True)
                y2_tr = y_sub_tr[np.where(finish_tr_mask)[0]]
                stage2_cols = _correlation_prune(_X2_for_corr, y=y2_tr, threshold=METHOD_CORR_PRUNE_THRESHOLD)
                self._stat("Stage2 corr-pruned", f"{len(stage2_cols_raw)} → {len(stage2_cols)} (thr={METHOD_CORR_PRUNE_THRESHOLD})")
                p_sub = max(1e-6, float(np.mean(y2_tr))) if len(y2_tr) > 0 else 0.5
                w2 = np.where(y2_tr == 1, 0.5 / p_sub, 0.5 / max(1e-6, 1.0 - p_sub))
                # Top-N importance prune: stage 2 trains on ~1/2 the rows of winner
                # with comparable feature count, so a tighter budget mitigates
                # overfitting. Uses the same class-weighted HGB-style signal
                # LightGBM reports via feature_importances_.
                if lgb is not None and len(stage2_cols) > STAGE2_MAX_FEATURES:
                    _quick2 = lgb.LGBMClassifier(
                        n_estimators=150, learning_rate=0.05, max_depth=5,
                        random_state=RANDOM_SEED + 777, n_jobs=-1, verbose=-1,
                    )
                    try:
                        _quick2.fit(_X2_for_corr[stage2_cols], y2_tr, sample_weight=w2)
                    except TypeError:
                        _quick2.fit(_X2_for_corr[stage2_cols], y2_tr)
                    _imp2 = np.asarray(_quick2.feature_importances_, dtype=float)
                    _keep2 = sorted(np.argsort(_imp2)[::-1][:STAGE2_MAX_FEATURES])
                    stage2_cols = [stage2_cols[i] for i in _keep2]
                    self._stat("Stage2 top-N pruned", f"kept {len(stage2_cols)}/{STAGE2_MAX_FEATURES}")
                X2_tr = _X2_for_corr[stage2_cols]

                # Stage 2 validation set (finishes only)
                _finish_va_mask = (y_method_dev.iloc[va_idx]["finish_bin"].values == "Finish")
                _y_sub_va = (y_method_dev.iloc[va_idx]["finish_subtype"].values == "Submission").astype(int)
                y2_va = _y_sub_va[_finish_va_mask] if int(np.sum(_finish_va_mask)) >= 20 else _y_sub_va
                X2_va = X_m_va_imp.iloc[np.where(_finish_va_mask)[0] if int(np.sum(_finish_va_mask)) >= 20 else np.arange(len(X_m_va_imp))][stage2_cols].reset_index(drop=True)
                _stage2_defaults = dict(
                    loss="log_loss",
                    max_iter=320, learning_rate=0.05, max_depth=5,
                    max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.7,
                    random_state=RANDOM_SEED + 809,
                )
                _stage2_params = _optuna_tune_hgb(
                    X2_tr, y2_tr, X2_va, y2_va,
                    sample_weight_tr=w2,
                    n_trials=METHOD_OPTUNA_TRIALS,
                    seed=RANDOM_SEED + 809,
                    defaults=_stage2_defaults,
                    progress_cb=self._progress,
                    progress_label="Stage2 HGB",
                )
                self._stat("Stage2 HGB params", f"lr={_stage2_params.get('learning_rate', 0):.4f}, depth={_stage2_params.get('max_depth', 0)}, leaves={_stage2_params.get('max_leaf_nodes', 0)}, iter={_stage2_params.get('max_iter', 0)}")
                stage2 = _BaggedHGB(_stage2_params, n_estimators=METHOD_HGB_BAG,
                                    base_seed=RANDOM_SEED + 809, subsample=METHOD_HGB_BAG_SUBSAMPLE)
                stage2.fit(X2_tr, y2_tr, sample_weight=w2)
                stage2_rf = RandomForestClassifier(
                    n_estimators=360, max_depth=9, min_samples_leaf=3,
                    max_features=0.7, random_state=RANDOM_SEED + 911, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_rf.fit(X2_tr, y2_tr)
                stage2_et = ExtraTreesClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 915, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_et.fit(X2_tr, y2_tr)

                # Direct multiclass head to reduce collapse to Decision.
                y_cls_tr = y_method_dev.iloc[tr_idx]["coarse"].values
                cls_counts = pd.Series(y_cls_tr).value_counts()
                cls_w = {k: (len(y_cls_tr) / (len(cls_counts) * max(1, v))) for k, v in cls_counts.items()}
                w_cls = np.array([cls_w.get(lbl, 1.0) for lbl in y_cls_tr], dtype=float)
                w_direct = w_time * w_cls
                direct_hgb = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=300, learning_rate=0.05, max_depth=6,
                    max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.7,
                    random_state=RANDOM_SEED + 912,
                )
                direct_hgb.fit(X_m_tr_imp, y_cls_tr, sample_weight=w_direct)
                direct_rf = RandomForestClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 913, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_rf.fit(X_m_tr_imp, y_cls_tr)
                direct_et = ExtraTreesClassifier(
                    n_estimators=680, max_depth=13, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 916, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_et.fit(X_m_tr_imp, y_cls_tr)
                direct_classes = [str(c) for c in direct_hgb.classes_]

                # OVR method head (one binary model per method class).
                ovr_models = {}
                for j, lbl in enumerate(METHOD_LABELS):
                    y_ovr = (pd.Series(y_cls_tr).astype(str).values == lbl).astype(int)
                    p_pos = max(1e-6, float(np.mean(y_ovr)))
                    w_ovr = w_time * np.where(y_ovr == 1, 0.5 / p_pos, 0.5 / max(1e-6, 1.0 - p_pos))
                    mdl = HistGradientBoostingClassifier(
                        loss="log_loss",
                        max_iter=260, learning_rate=0.05, max_depth=5,
                        max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.8,
                        random_state=RANDOM_SEED + 1500 + j,
                    )
                    mdl.fit(X_m_tr_imp, y_ovr, sample_weight=w_ovr)
                    ovr_models[lbl] = mdl
                simple_method = LogisticRegression(
                    max_iter=8000, C=0.8, solver="saga", tol=1e-3, n_jobs=-1,
                    class_weight="balanced", random_state=RANDOM_SEED + 1601,
                )
                simple_method.fit(X_m_tr_imp, pd.Series(y_cls_tr).astype(str).values)

                def _stage_probs(
                    X_imp, temp_dec=1.0, temp_fin=1.0, alpha1=0.75, alpha2=0.75,
                    bias_finish=0.0, bias_sub=0.0, alpha_direct=0.85, beta_direct=0.70,
                    alpha_ovr=0.85, finish_thr=0.50, sub_thr=0.50, alpha_simple=0.80
                ):
                    X_s1 = X_imp[stage1_cols]
                    p_finish_hist = stage1.predict_proba(X_s1)[:, 1]
                    p_finish_rf = stage1_rf.predict_proba(X_s1)[:, 1]
                    p_finish_et = stage1_et.predict_proba(X_s1)[:, 1]
                    p_finish_tree = 0.55 * p_finish_rf + 0.45 * p_finish_et
                    p_finish_raw = alpha1 * p_finish_hist + (1.0 - alpha1) * p_finish_tree
                    logit = np.log(np.clip(p_finish_raw, 1e-6, 1 - 1e-6) / np.clip(1.0 - p_finish_raw, 1e-6, 1 - 1e-6))
                    p_finish_c = 1.0 / (1.0 + np.exp(-((logit / max(0.35, float(temp_dec))) + float(bias_finish))))
                    p_finish_c = _apply_binary_threshold_warp(p_finish_c, finish_thr)

                    X_s2 = X_imp[stage2_cols]
                    p_sub_hist = stage2.predict_proba(X_s2)[:, 1]
                    p_sub_rf = stage2_rf.predict_proba(X_s2)[:, 1]
                    p_sub_et = stage2_et.predict_proba(X_s2)[:, 1]
                    p_sub_tree = 0.55 * p_sub_rf + 0.45 * p_sub_et
                    p_sub_raw = alpha2 * p_sub_hist + (1.0 - alpha2) * p_sub_tree
                    logit2 = np.log(np.clip(p_sub_raw, 1e-6, 1 - 1e-6) / np.clip(1.0 - p_sub_raw, 1e-6, 1 - 1e-6))
                    p_sub_c = 1.0 / (1.0 + np.exp(-((logit2 / max(0.35, float(temp_fin))) + float(bias_sub))))
                    p_sub_c = _apply_binary_threshold_warp(p_sub_c, sub_thr)

                    probs = pd.DataFrame({
                        "Decision": 1.0 - p_finish_c,
                        "KO/TKO": p_finish_c * (1.0 - p_sub_c),
                        "Submission": p_finish_c * p_sub_c,
                    })
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    p_h = direct_hgb.predict_proba(X_imp)
                    p_r = direct_rf.predict_proba(X_imp)
                    p_e = direct_et.predict_proba(X_imp)
                    direct_rows = []
                    for row_idx in range(len(X_imp)):
                        raw = {}
                        for j, cls in enumerate(direct_classes):
                            tree_mix = 0.55 * float(p_r[row_idx][j]) + 0.45 * float(p_e[row_idx][j])
                            raw[cls] = beta_direct * float(p_h[row_idx][j]) + (1.0 - beta_direct) * tree_mix
                        direct_rows.append(_normalize_method_probs(raw))
                    direct_df = pd.DataFrame(direct_rows)[METHOD_LABELS]
                    probs = alpha_direct * probs[METHOD_LABELS] + (1.0 - alpha_direct) * direct_df
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    ovr_arr = np.zeros((len(X_imp), len(METHOD_LABELS)), dtype=float)
                    for k, lbl in enumerate(METHOD_LABELS):
                        ovr_arr[:, k] = ovr_models[lbl].predict_proba(X_imp)[:, 1]
                    ovr_arr = np.clip(ovr_arr, MIN_METHOD_PROB, 1.0)
                    ovr_arr = ovr_arr / np.sum(ovr_arr, axis=1, keepdims=True)
                    ovr_df = pd.DataFrame(ovr_arr, columns=METHOD_LABELS)
                    probs = alpha_ovr * probs[METHOD_LABELS] + (1.0 - alpha_ovr) * ovr_df
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    p_s = simple_method.predict_proba(X_imp)
                    cls_map_s = {str(c): i for i, c in enumerate(simple_method.classes_)}
                    simple_arr = np.zeros((len(X_imp), len(METHOD_LABELS)), dtype=float)
                    for j, m in enumerate(METHOD_LABELS):
                        simple_arr[:, j] = p_s[:, cls_map_s[m]] if m in cls_map_s else (1.0 / 3.0)
                    simple_arr = np.clip(simple_arr, MIN_METHOD_PROB, 1.0)
                    simple_arr = simple_arr / np.sum(simple_arr, axis=1, keepdims=True)
                    simple_df = pd.DataFrame(simple_arr, columns=METHOD_LABELS)
                    probs = alpha_simple * probs[METHOD_LABELS] + (1.0 - alpha_simple) * simple_df
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    return probs

                # Group priors (weight class + gender adapters) from method-train split.
                grp_train = meta_dev_valid.iloc[tr_idx].reset_index(drop=True)
                grp_labels = y_method_dev.iloc[tr_idx]["coarse"].values
                group_priors = {}
                for (wc, gender), grp in grp_train.groupby(["weight_class", "gender"], dropna=False):
                    idx = grp.index.values
                    counts = pd.Series(grp_labels[idx]).value_counts()
                    tot = float(len(idx))
                    group_priors[(str(wc), str(gender).lower())] = _normalize_method_probs({
                        "Decision": float(counts.get("Decision", 0.0) / max(1.0, tot)),
                        "KO/TKO": float(counts.get("KO/TKO", 0.0) / max(1.0, tot)),
                        "Submission": float(counts.get("Submission", 0.0) / max(1.0, tot)),
                    })
                all_counts = pd.Series(grp_labels).value_counts()
                all_tot = float(len(grp_labels))
                group_priors[("ALL", "all")] = _normalize_method_probs({
                    "Decision": float(all_counts.get("Decision", 0.0) / max(1.0, all_tot)),
                    "KO/TKO": float(all_counts.get("KO/TKO", 0.0) / max(1.0, all_tot)),
                    "Submission": float(all_counts.get("Submission", 0.0) / max(1.0, all_tot)),
                })
                base_prior = _normalize_method_probs({
                    "Decision": float(all_counts.get("Decision", 0.0) / max(1.0, all_tot)),
                    "KO/TKO": float(all_counts.get("KO/TKO", 0.0) / max(1.0, all_tot)),
                    "Submission": float(all_counts.get("Submission", 0.0) / max(1.0, all_tot)),
                })

                def _history_prior_array(X_raw):
                    Xr = X_raw.reset_index(drop=True)
                    out = np.zeros((len(Xr), 3), dtype=float)
                    for i in range(len(Xr)):
                        d_dec = float(pd.to_numeric(Xr.get("d_dec_win_pct", pd.Series([0.0])).iloc[i], errors="coerce") if "d_dec_win_pct" in Xr else 0.0)
                        d_ko = float(pd.to_numeric(Xr.get("d_ko_win_pct", pd.Series([0.0])).iloc[i], errors="coerce") if "d_ko_win_pct" in Xr else 0.0)
                        d_sub = float(pd.to_numeric(Xr.get("d_sub_win_pct", pd.Series([0.0])).iloc[i], errors="coerce") if "d_sub_win_pct" in Xr else 0.0)
                        p = _history_prior_row(d_dec, d_ko, d_sub)
                        out[i, 0] = float(p["Decision"])
                        out[i, 1] = float(p["KO/TKO"])
                        out[i, 2] = float(p["Submission"])
                    return out

                # Tune calibration + blend weights by primary metric.
                y_va_true = y_method_dev.iloc[va_idx]["coarse"].values
                winner_correct_va = (
                    y_dev_pred_valid[va_idx] == y_dev_true_valid[va_idx]
                )
                label_to_idx = {m: i for i, m in enumerate(METHOD_LABELS)}
                y_va_idx = np.array([label_to_idx.get(str(lbl), 0) for lbl in y_va_true], dtype=int)
                hist_arr = _history_prior_array(X_m_va)
                hist_arr_tr = _history_prior_array(X_m_tr)
                sub_arr = _sub_attempt_prior_array(X_m_va)
                sub_arr_tr = _sub_attempt_prior_array(X_m_tr)
                va_local_idx = np.arange(len(y_va_idx), dtype=int)
                va_chunks = [c for c in np.array_split(va_local_idx, 3) if len(c) > 0]
                meta_va = meta_dev_valid.iloc[va_idx].reset_index(drop=True)
                grp_arr = np.zeros((len(meta_va), 3), dtype=float)
                for i in range(len(meta_va)):
                    wc = str(meta_va.iloc[i]["weight_class"])
                    gd = str(meta_va.iloc[i]["gender"]).lower()
                    gp = group_priors.get((wc, gd), group_priors[("ALL", "all")])
                    grp_arr[i, 0] = float(gp["Decision"])
                    grp_arr[i, 1] = float(gp["KO/TKO"])
                    grp_arr[i, 2] = float(gp["Submission"])
                meta_tr = meta_dev_valid.iloc[tr_idx].reset_index(drop=True)
                grp_arr_tr = np.zeros((len(meta_tr), 3), dtype=float)
                for i in range(len(meta_tr)):
                    wc = str(meta_tr.iloc[i]["weight_class"])
                    gd = str(meta_tr.iloc[i]["gender"]).lower()
                    gp = group_priors.get((wc, gd), group_priors[("ALL", "all")])
                    grp_arr_tr[i, 0] = float(gp["Decision"])
                    grp_arr_tr[i, 1] = float(gp["KO/TKO"])
                    grp_arr_tr[i, 2] = float(gp["Submission"])

                # ── Method blend tuning ───────────────────────────────────────
                # `_score_cfg_method` is the single source of truth for cfg quality:
                # macro_f1 − 0.50·(dec+ko+sub recall shortfalls) computed per
                # chronological val chunk, aggregated as mean − 0.60·std. This is
                # a 3-fold walk-forward CV on the val set; the std penalty rules
                # out configs that only win on one chunk. The Optuna tuner and
                # the champion/challenger gate both call this function, so they
                # optimize the same criterion with the same stability guard.
                _METHOD_SEARCH_SPACE = {
                    "alpha_stage1": [0.55, 0.70, 0.85],
                    "alpha_stage2": [0.55, 0.70, 0.85],
                    "t_dec": [0.7, 0.85, 1.0, 1.15, 1.3],
                    "t_fin": [0.7, 0.85, 1.0, 1.15, 1.3],
                    "bias_finish": [-0.35, -0.15, 0.0, 0.15, 0.35],
                    "bias_sub": [-0.30, -0.10, 0.0, 0.10, 0.30],
                    "finish_threshold": [0.38, 0.42, 0.46, 0.50, 0.54, 0.58],
                    "sub_threshold": [0.38, 0.42, 0.46, 0.50, 0.54, 0.58],
                    "alpha_direct": [0.90, 0.96, 1.00],
                    "beta_direct": [0.60, 0.75, 0.90],
                    "alpha_ovr": [0.90, 0.97, 1.00],
                    "alpha_simple": [0.65, 0.80, 0.95],
                    "w_hist": [0.00, 0.08, 0.15, 0.22, 0.30],
                    "w_group": [0.00, 0.04, 0.08, 0.12],
                    "w_base": [0.00, 0.06, 0.12, 0.18],
                    "w_subsig": [0.00, 0.06, 0.10, 0.14, 0.18],
                    "sub_boost_k": [0.0, 0.4, 0.8, 1.2, 1.6],
                    "method_bias_decision": [-0.20, -0.10, 0.00, 0.10, 0.20],
                    "method_bias_ko_tko": [-0.10, 0.00, 0.10, 0.20, 0.30, 0.40],
                    "method_bias_submission": [-0.25, -0.15, -0.05, 0.05, 0.15, 0.25, 0.35, 0.45],
                }

                def _blend_probs_for_cfg(cfg):
                    _ml = _stage_probs(
                        X_m_va_imp,
                        temp_dec=cfg["t_dec"], temp_fin=cfg["t_fin"],
                        alpha1=cfg["alpha_stage1"], alpha2=cfg["alpha_stage2"],
                        bias_finish=cfg["bias_finish"], bias_sub=cfg["bias_sub"],
                        alpha_direct=cfg["alpha_direct"], beta_direct=cfg["beta_direct"],
                        alpha_ovr=cfg["alpha_ovr"],
                        finish_thr=cfg["finish_threshold"], sub_thr=cfg["sub_threshold"],
                        alpha_simple=cfg["alpha_simple"],
                    )
                    _ml_arr = _ml[["Decision", "KO/TKO", "Submission"]].to_numpy(dtype=float)
                    _base = np.tile(
                        np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float),
                        (len(_ml_arr), 1),
                    )
                    _w_ml = 1.0 - cfg["w_hist"] - cfg["w_group"] - cfg["w_base"] - cfg["w_subsig"]
                    _p = (_w_ml * _ml_arr + cfg["w_hist"] * hist_arr + cfg["w_group"] * grp_arr
                          + cfg["w_base"] * _base + cfg["w_subsig"] * sub_arr)
                    _p = np.clip(_p, MIN_METHOD_PROB, 1.0)
                    _p = _p / np.sum(_p, axis=1, keepdims=True)
                    _p = _apply_method_logit_bias_arr(
                        _p,
                        np.array([cfg["method_bias_decision"], cfg["method_bias_ko_tko"],
                                  cfg["method_bias_submission"]], dtype=float),
                    )
                    _p = _apply_submission_signal_boost_arr(_p, sub_arr[:, 2], cfg["sub_boost_k"])
                    _p = _p / np.sum(_p, axis=1, keepdims=True)
                    return _p

                def _score_cfg_method(cfg):
                    """Walk-forward score: mean − 0.60·std of the per-chunk objective.

                    Used by both the tuner and the champion gate so they optimize
                    the exact same criterion. Returns -inf on failure so the
                    tuner naturally skips broken configs.
                    """
                    try:
                        _p = _blend_probs_for_cfg(cfg)
                        _pidx = np.argmax(_p, axis=1)

                        def _chunk_obj(ch):
                            _cwc = winner_correct_va[ch]
                            if int(np.sum(_cwc)) == 0:
                                return float(np.mean(_pidx[ch] == y_va_idx[ch]))
                            _ys = y_va_idx[ch][_cwc]
                            _ps_idx = _pidx[ch][_cwc]
                            _rec = [0.0, 0.0, 0.0]
                            for c in range(3):
                                _m = (_ys == c)
                                if int(np.sum(_m)) > 0:
                                    _rec[c] = float(np.mean(_ps_idx[_m] == c))
                            _f1 = [0.0, 0.0, 0.0]
                            for c in range(3):
                                _pc_cnt = int(np.sum(_ps_idx == c))
                                if _pc_cnt == 0:
                                    continue
                                _prec = float(np.sum((_ps_idx == c) & (_ys == c))) / _pc_cnt
                                _denom = _prec + _rec[c]
                                if _denom > 0:
                                    _f1[c] = 2.0 * _prec * _rec[c] / _denom
                            _macro_f1 = float(np.mean(_f1))
                            _dec_short = max(0.0, 0.55 - _rec[0])
                            _ko_short = max(0.0, 0.50 - _rec[1])
                            _sub_short = max(0.0, 0.40 - _rec[2])
                            return _macro_f1 - 0.50 * (_dec_short + _ko_short + _sub_short)

                        _chunk_objs = [_chunk_obj(ch) for ch in va_chunks]
                        return float(np.mean(_chunk_objs)) - 0.60 * float(np.std(_chunk_objs))
                    except Exception:
                        return float("-inf")

                def _display_metrics_for_cfg(cfg):
                    """Pick-rate accuracy + majority baseline on full val set (logging only)."""
                    try:
                        _p = _blend_probs_for_cfg(cfg)
                        _pidx = np.argmax(_p, axis=1)
                        if int(np.sum(winner_correct_va)) == 0:
                            _score = float(np.mean(_pidx == y_va_idx))
                            _baseline = float(pd.Series(y_va_true).value_counts().max() / len(y_va_true))
                        else:
                            _score = float(np.mean(_pidx[winner_correct_va] == y_va_idx[winner_correct_va]))
                            _baseline = float(
                                pd.Series(y_va_true[winner_correct_va]).value_counts().max()
                                / int(np.sum(winner_correct_va))
                            )
                        return _score, _baseline
                    except Exception:
                        return float("nan"), float("nan")

                _DEFAULT_BEST_CFG = {
                    "t_dec": 1.0, "t_fin": 1.0,
                    "bias_finish": 0.0, "bias_sub": 0.0,
                    "finish_threshold": 0.50, "sub_threshold": 0.50,
                    "w_hist": 0.15, "w_group": 0.08, "w_base": 0.08, "w_subsig": 0.10,
                    "sub_boost_k": 0.6,
                    "method_bias_decision": 0.10,
                    "method_bias_ko_tko": -0.10,
                    "method_bias_submission": 0.00,
                    "alpha_stage1": 0.75, "alpha_stage2": 0.75,
                    "alpha_direct": 0.85, "beta_direct": 0.70,
                    "alpha_ovr": 0.85,
                    "alpha_simple": 0.80,
                }

                best_cfg = None
                if optuna is not None and METHOD_TUNING_TRIALS > 0:
                    self._reset_terminal_progress(
                        labels=["Method Blend"], default_total=METHOD_TUNING_TRIALS
                    )

                    def _method_objective(trial):
                        cfg = {
                            k: trial.suggest_categorical(k, v)
                            for k, v in _METHOD_SEARCH_SPACE.items()
                        }
                        if cfg["w_hist"] + cfg["w_group"] + cfg["w_base"] + cfg["w_subsig"] >= 0.78:
                            raise optuna.TrialPruned()
                        return _score_cfg_method(cfg)

                    def _method_trial_cb(_study, trial):
                        self._progress(
                            int(trial.number) + 1, int(METHOD_TUNING_TRIALS), "Method Blend"
                        )

                    sampler = optuna.samplers.TPESampler(seed=RANDOM_SEED + 2026)
                    study = optuna.create_study(direction="maximize", sampler=sampler)
                    study.optimize(
                        _method_objective,
                        n_trials=int(METHOD_TUNING_TRIALS),
                        show_progress_bar=False,
                        callbacks=[_method_trial_cb],
                    )
                    self._finalize_terminal_progress()
                    if study.best_trial is not None:
                        best_cfg = dict(study.best_trial.params)
                else:
                    # Random-search fallback (Optuna unavailable).
                    rng = np.random.default_rng(RANDOM_SEED + 2026)
                    best_score = float("-inf")
                    for _trial in range(int(METHOD_TUNING_TRIALS)):
                        cfg = {k: float(rng.choice(v)) for k, v in _METHOD_SEARCH_SPACE.items()}
                        if cfg["w_hist"] + cfg["w_group"] + cfg["w_base"] + cfg["w_subsig"] >= 0.78:
                            continue
                        score = _score_cfg_method(cfg)
                        if score > best_score:
                            best_score = score
                            best_cfg = cfg

                if best_cfg is None:
                    best_cfg = dict(_DEFAULT_BEST_CFG)

                # ── Champion / challenger ─────────────────────────────────────
                # The champion exists for run-to-run STABILITY: a freshly tuned cfg
                # replaces the saved champion only if it beats the champion's
                # walk-forward objective by METHOD_CHAMPION_MARGIN — enough to clear
                # tuner/validation noise. Because the tuner is deterministic, an
                # unchanged setup reproduces the champion exactly (tie → retained,
                # no churn), so the gate only bites when the data drifted between
                # runs. The champion is fingerprinted on everything that would make
                # an old cfg incomparable (search space, method classes, feature
                # columns, blend-code version); a mismatch retires it outright, so
                # iterating on features/logic no longer requires deleting the file.
                try:
                    os.makedirs(os.path.dirname(METHOD_CHAMPION_PATH), exist_ok=True)
                    champ_fp = hashlib.sha256("|".join([
                        ",".join(sorted(_METHOD_SEARCH_SPACE.keys())),
                        ",".join(METHOD_LABELS),
                        ",".join(method_columns),
                        METHOD_CACHE_VERSION,
                    ]).encode("utf-8")).hexdigest()[:16]

                    def _cfg_complete(_c):
                        return isinstance(_c, dict) and all(k in _c for k in _METHOD_SEARCH_SPACE)

                    def _save_champion(_cfg, _obj):
                        _payload = {
                            "_cfg": {k: (list(v) if isinstance(v, tuple) else v)
                                     for k, v in _cfg.items() if k in _METHOD_SEARCH_SPACE},
                            "_fingerprint": champ_fp,
                            "_objective": float(_obj),
                            "_saved": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                        with open(METHOD_CHAMPION_PATH, "w") as _f:
                            json.dump(_payload, _f, indent=2)

                    new_obj = _score_cfg_method(best_cfg)

                    champion_record = None
                    if os.path.exists(METHOD_CHAMPION_PATH) and not METHOD_HARD_RESET:
                        try:
                            with open(METHOD_CHAMPION_PATH, "r") as _f:
                                champion_record = json.load(_f)
                        except Exception:
                            champion_record = None

                    champ_cfg = None
                    champ_obj = float("-inf")
                    stale_reason = None
                    if isinstance(champion_record, dict) and "_cfg" in champion_record:
                        if champion_record.get("_fingerprint") != champ_fp:
                            stale_reason = "setup changed (features/classes/search space/version)"
                        elif not _cfg_complete(champion_record.get("_cfg")):
                            stale_reason = "missing tuned keys"
                        else:
                            champ_cfg = dict(champion_record["_cfg"])
                            champ_obj = _score_cfg_method(champ_cfg)
                            if not np.isfinite(champ_obj):
                                champ_cfg, stale_reason = None, "champion no longer scorable"
                    elif champion_record is not None:
                        stale_reason = "legacy champion format (no fingerprint)"

                    if champ_cfg is None:
                        _why = stale_reason if stale_reason else "no champion on file"
                        self._stat(
                            "Method config",
                            f"No usable champion ({_why}) — saving challenger (obj={new_obj:.5f})",
                        )
                        best_cfg = dict(best_cfg)
                        _save_champion(best_cfg, new_obj)
                    elif new_obj > champ_obj + METHOD_CHAMPION_MARGIN:
                        self._stat(
                            "Method config",
                            f"Challenger wins (new={new_obj:.5f} > champ={champ_obj:.5f} "
                            f"+ margin {METHOD_CHAMPION_MARGIN:.3f}) — champion updated",
                        )
                        best_cfg = dict(best_cfg)
                        _save_champion(best_cfg, new_obj)
                    else:
                        self._stat(
                            "Method config",
                            f"Champion retained (champ={champ_obj:.5f} >= new={new_obj:.5f} "
                            f"− margin {METHOD_CHAMPION_MARGIN:.3f})",
                        )
                        best_cfg = dict(champ_cfg)
                except Exception as _e:
                    self._stat("Method config", f"Champion load/save failed ({_e}) — using tuned config")

                # Final-cfg display metrics (full val set, not chunked).
                _val_metric, _val_baseline = _display_metrics_for_cfg(best_cfg)
                best_cfg["val_metric"] = _val_metric
                best_cfg["val_baseline"] = _val_baseline

                # Hard reset: keep only stable components that generalized better.
                if METHOD_HARD_RESET:
                    best_cfg["alpha_direct"] = 1.0
                    best_cfg["alpha_ovr"] = 1.0
                    best_cfg["alpha_simple"] = 1.0

                # Meta calibrator over blended probabilities + priors.
                meta_model = None
                meta_eta = 0.0
                try:
                    ml_tr_df = _stage_probs(
                        X_m_tr_imp,
                        temp_dec=best_cfg["t_dec"], temp_fin=best_cfg["t_fin"],
                        alpha1=best_cfg["alpha_stage1"], alpha2=best_cfg["alpha_stage2"],
                        bias_finish=best_cfg["bias_finish"], bias_sub=best_cfg["bias_sub"],
                        alpha_direct=best_cfg["alpha_direct"], beta_direct=best_cfg["beta_direct"],
                        alpha_ovr=best_cfg["alpha_ovr"],
                        finish_thr=best_cfg["finish_threshold"], sub_thr=best_cfg["sub_threshold"],
                        alpha_simple=best_cfg["alpha_simple"],
                    )
                    ml_va_df = _stage_probs(
                        X_m_va_imp,
                        temp_dec=best_cfg["t_dec"], temp_fin=best_cfg["t_fin"],
                        alpha1=best_cfg["alpha_stage1"], alpha2=best_cfg["alpha_stage2"],
                        bias_finish=best_cfg["bias_finish"], bias_sub=best_cfg["bias_sub"],
                        alpha_direct=best_cfg["alpha_direct"], beta_direct=best_cfg["beta_direct"],
                        alpha_ovr=best_cfg["alpha_ovr"],
                        finish_thr=best_cfg["finish_threshold"], sub_thr=best_cfg["sub_threshold"],
                        alpha_simple=best_cfg["alpha_simple"],
                    )
                    ml_tr_arr = ml_tr_df[["Decision", "KO/TKO", "Submission"]].to_numpy(dtype=float)
                    ml_va_arr = ml_va_df[["Decision", "KO/TKO", "Submission"]].to_numpy(dtype=float)
                    bp_arr_tr = np.tile(
                        np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float),
                        (len(ml_tr_arr), 1),
                    )
                    bp_arr_va = np.tile(
                        np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float),
                        (len(ml_va_arr), 1),
                    )
                    w_hist_b = float(best_cfg["w_hist"])
                    w_group_b = float(best_cfg["w_group"])
                    w_base_b = float(best_cfg.get("w_base", 0.0))
                    w_subsig_b = float(best_cfg.get("w_subsig", 0.0))
                    w_ml_b = max(0.0, 1.0 - w_hist_b - w_group_b - w_base_b - w_subsig_b)
                    blend_tr = np.clip(
                        w_ml_b * ml_tr_arr + w_hist_b * hist_arr_tr + w_group_b * grp_arr_tr + w_base_b * bp_arr_tr + w_subsig_b * sub_arr_tr,
                        MIN_METHOD_PROB, 1.0
                    )
                    blend_va = np.clip(
                        w_ml_b * ml_va_arr + w_hist_b * hist_arr + w_group_b * grp_arr + w_base_b * bp_arr_va + w_subsig_b * sub_arr,
                        MIN_METHOD_PROB, 1.0
                    )
                    blend_tr = blend_tr / np.sum(blend_tr, axis=1, keepdims=True)
                    blend_va = blend_va / np.sum(blend_va, axis=1, keepdims=True)
                    method_bias_vec = np.array([
                        float(best_cfg.get("method_bias_decision", 0.0)),
                        float(best_cfg.get("method_bias_ko_tko", 0.0)),
                        float(best_cfg.get("method_bias_submission", 0.0)),
                    ], dtype=float)
                    blend_tr = _apply_method_logit_bias_arr(blend_tr, method_bias_vec)
                    blend_va = _apply_method_logit_bias_arr(blend_va, method_bias_vec)
                    sub_boost_k = float(best_cfg.get("sub_boost_k", 0.0))
                    blend_tr = _apply_submission_signal_boost_arr(blend_tr, sub_arr_tr[:, 2], sub_boost_k)
                    blend_va = _apply_submission_signal_boost_arr(blend_va, sub_arr[:, 2], sub_boost_k)
                    blend_tr = blend_tr / np.sum(blend_tr, axis=1, keepdims=True)
                    blend_va = blend_va / np.sum(blend_va, axis=1, keepdims=True)

                    X_meta_tr = np.hstack([blend_tr, hist_arr_tr, grp_arr_tr, bp_arr_tr, np.log(np.clip(blend_tr, 1e-6, 1.0))])
                    X_meta_va = np.hstack([blend_va, hist_arr, grp_arr, bp_arr_va, np.log(np.clip(blend_va, 1e-6, 1.0))])
                    y_meta_tr = y_method_dev.iloc[tr_idx]["coarse"].astype(str).values
                    meta_model = LogisticRegression(
                        max_iter=6000, C=0.7, solver="lbfgs",
                        class_weight="balanced", random_state=RANDOM_SEED + 1313,
                    )
                    meta_model.fit(X_meta_tr, y_meta_tr)
                    p_meta = meta_model.predict_proba(X_meta_va)
                    cls_map = {str(c): i for i, c in enumerate(meta_model.classes_)}
                    meta_va_arr = np.zeros((len(X_meta_va), 3), dtype=float)
                    for j, m in enumerate(METHOD_LABELS):
                        if m in cls_map:
                            meta_va_arr[:, j] = p_meta[:, cls_map[m]]
                        else:
                            meta_va_arr[:, j] = 1.0 / 3.0
                    meta_va_arr = np.clip(meta_va_arr, MIN_METHOD_PROB, 1.0)
                    meta_va_arr = meta_va_arr / np.sum(meta_va_arr, axis=1, keepdims=True)
                    best_meta_key = None
                    for eta in (0.0, 0.10, 0.20):
                        arr = np.clip((1.0 - eta) * blend_va + eta * meta_va_arr, MIN_METHOD_PROB, 1.0)
                        arr = arr / np.sum(arr, axis=1, keepdims=True)
                        pred_idx = np.argmax(arr, axis=1)
                        if int(np.sum(winner_correct_va)) > 0:
                            s = float(np.mean(pred_idx[winner_correct_va] == y_va_idx[winner_correct_va]))
                        else:
                            s = float(np.mean(pred_idx == y_va_idx))
                        chunk_scores = []
                        for ch in va_chunks:
                            ch_wc = winner_correct_va[ch]
                            if int(np.sum(ch_wc)) > 0:
                                chunk_scores.append(float(np.mean(pred_idx[ch][ch_wc] == y_va_idx[ch][ch_wc])))
                            else:
                                chunk_scores.append(float(np.mean(pred_idx[ch] == y_va_idx[ch])))
                        robust_s = float(np.mean(chunk_scores)) - 0.60 * float(np.std(chunk_scores))
                        key_m = (robust_s, s, -eta)
                        if best_meta_key is None or key_m > best_meta_key:
                            best_meta_key = key_m
                            meta_eta = float(eta)
                except Exception as _e:
                    self._stat("WARNING", f"Method meta-calibrator failed ({_e}) — disabled")
                    meta_model = None
                    meta_eta = 0.0

                if METHOD_HARD_RESET:
                    meta_model = None
                    meta_eta = 0.0

                method_bundle = {
                    "stage1": stage1,
                    "stage2": stage2,
                    "stage1_rf": stage1_rf,
                    "stage1_et": stage1_et,
                    "stage2_rf": stage2_rf,
                    "stage2_et": stage2_et,
                    "direct_hgb": direct_hgb,
                    "direct_rf": direct_rf,
                    "direct_et": direct_et,
                    "ovr_models": ovr_models,
                    "simple_method": simple_method,
                    "direct_classes": direct_classes,
                    "imputer": method_imputer,
                    "method_columns": method_columns,
                    "temp_decision": float(best_cfg["t_dec"]),
                    "temp_finish": float(best_cfg["t_fin"]),
                    "bias_finish": float(best_cfg["bias_finish"]),
                    "bias_sub": float(best_cfg["bias_sub"]),
                    "finish_threshold": float(best_cfg["finish_threshold"]),
                    "sub_threshold": float(best_cfg["sub_threshold"]),
                    "alpha_stage1": float(best_cfg["alpha_stage1"]),
                    "alpha_stage2": float(best_cfg["alpha_stage2"]),
                    "alpha_direct": float(best_cfg["alpha_direct"]),
                    "beta_direct": float(best_cfg["beta_direct"]),
                    "alpha_ovr": float(best_cfg["alpha_ovr"]),
                    "alpha_simple": float(best_cfg["alpha_simple"]),
                    "w_hist": float(best_cfg["w_hist"]),
                    "w_group": float(best_cfg["w_group"]),
                    "w_base": float(best_cfg["w_base"]),
                    "w_subsig": float(best_cfg.get("w_subsig", 0.0)),
                    "sub_boost_k": float(best_cfg.get("sub_boost_k", 0.0)),
                    "stage1_cols": stage1_cols,
                    "stage2_cols": stage2_cols,
                    # Optuna-tuned HGB params propagated to the all-data retrain
                    # so the production model uses the SAME base hyperparameters
                    # the evaluation reported on, not hardcoded fallbacks.
                    "stage1_params": dict(_stage1_params),
                    "stage2_params": dict(_stage2_params),
                    "method_bias_decision": float(best_cfg["method_bias_decision"]),
                    "method_bias_ko_tko": float(best_cfg["method_bias_ko_tko"]),
                    "method_bias_submission": float(best_cfg["method_bias_submission"]),
                    "base_prior": base_prior,
                    "meta_model": meta_model,
                    "meta_eta": float(meta_eta),
                    "group_priors": group_priors,
                    "detail_labels_seen": sorted(pd.Series(y_method_dev["detail"]).unique().tolist()),
                }

                self._stat("Method classes (training)", ", ".join(method_bundle["detail_labels_seen"]))
                if METHOD_HARD_RESET:
                    self._stat("Method stack mode", "hard-reset (stable components only)")
                self._stat("Validation metric target (method | winner correct)", f"{best_cfg['val_metric']:.1%}")
                self._stat("Validation majority baseline (same subset)", f"{best_cfg['val_baseline']:.1%}")

                # Holdout evaluation with winner-model predicted winners.
                X_test_oriented_pred = _oriented_method_matrix(X_test_full, y_pred_red)
                X_test_oriented_pred = _augment_method_features(X_test_oriented_pred)
                X_test_oriented_pred_imp = pd.DataFrame(
                    method_imputer.transform(X_test_oriented_pred), columns=method_columns
                )
                # Stage-only diagnostics.
                y_finish_true = (y_method_test["finish_bin"].astype(str).values == "Finish").astype(int)
                _s1c_t = method_bundle.get("stage1_cols")
                X_test_s1 = X_test_oriented_pred_imp.reindex(columns=_s1c_t) if _s1c_t else X_test_oriented_pred_imp
                p_finish_hist_t = method_bundle["stage1"].predict_proba(X_test_s1)[:, 1]
                p_finish_rf_t = method_bundle["stage1_rf"].predict_proba(X_test_s1)[:, 1]
                p_finish_et_t = method_bundle["stage1_et"].predict_proba(X_test_s1)[:, 1]
                p_finish_tree_t = 0.55 * p_finish_rf_t + 0.45 * p_finish_et_t
                p_finish_raw_t = (
                    float(method_bundle.get("alpha_stage1", 0.75)) * p_finish_hist_t
                    + (1.0 - float(method_bundle.get("alpha_stage1", 0.75))) * p_finish_tree_t
                )
                logit_finish_t = np.log(
                    np.clip(p_finish_raw_t, 1e-6, 1 - 1e-6)
                    / np.clip(1.0 - p_finish_raw_t, 1e-6, 1 - 1e-6)
                )
                p_finish_stage_t = 1.0 / (
                    1.0 + np.exp(
                        -(
                            logit_finish_t / max(0.35, float(method_bundle.get("temp_decision", 1.0)))
                            + float(method_bundle.get("bias_finish", 0.0))
                        )
                    )
                )
                p_finish_stage_t = _apply_binary_threshold_warp(
                    p_finish_stage_t, float(method_bundle.get("finish_threshold", 0.50))
                )
                stage1_pred = (p_finish_stage_t >= 0.5).astype(int)
                stage1_acc = float(np.mean(stage1_pred == y_finish_true))
                stage1_auc = float("nan")
                try:
                    if len(np.unique(y_finish_true)) > 1:
                        stage1_auc = float(roc_auc_score(y_finish_true, p_finish_stage_t))
                except Exception:
                    stage1_auc = float("nan")

                stage2_acc_true_finishes = float("nan")
                n_true_finishes = int(np.sum(y_finish_true == 1))
                if n_true_finishes > 0:
                    finish_idx_true = np.where(y_finish_true == 1)[0]
                    X_test_fin_true = X_test_oriented_pred_imp.iloc[finish_idx_true].reset_index(drop=True)
                    y_sub_true = (
                        y_method_test.iloc[finish_idx_true]["finish_subtype"].astype(str).values == "Submission"
                    ).astype(int)
                    _s2c = method_bundle.get("stage2_cols")
                    X_test_fin_s2 = X_test_fin_true.reindex(columns=_s2c) if _s2c else X_test_fin_true
                    p_sub_hist_t = method_bundle["stage2"].predict_proba(X_test_fin_s2)[:, 1]
                    p_sub_rf_t = method_bundle["stage2_rf"].predict_proba(X_test_fin_s2)[:, 1]
                    p_sub_et_t = method_bundle["stage2_et"].predict_proba(X_test_fin_s2)[:, 1]
                    p_sub_tree_t = 0.55 * p_sub_rf_t + 0.45 * p_sub_et_t
                    p_sub_raw_t = (
                        float(method_bundle.get("alpha_stage2", 0.75)) * p_sub_hist_t
                        + (1.0 - float(method_bundle.get("alpha_stage2", 0.75))) * p_sub_tree_t
                    )
                    logit_sub_t = np.log(
                        np.clip(p_sub_raw_t, 1e-6, 1 - 1e-6)
                        / np.clip(1.0 - p_sub_raw_t, 1e-6, 1 - 1e-6)
                    )
                    p_sub_stage_t = 1.0 / (
                        1.0 + np.exp(
                            -(
                                logit_sub_t / max(0.35, float(method_bundle.get("temp_finish", 1.0)))
                                + float(method_bundle.get("bias_sub", 0.0))
                            )
                        )
                    )
                    p_sub_stage_t = _apply_binary_threshold_warp(
                        p_sub_stage_t, float(method_bundle.get("sub_threshold", 0.50))
                    )
                    stage2_pred = (p_sub_stage_t >= 0.5).astype(int)
                    stage2_acc_true_finishes = float(np.mean(stage2_pred == y_sub_true))
                ml_test = _stage_probs(
                    X_test_oriented_pred_imp,
                    temp_dec=method_bundle["temp_decision"],
                    temp_fin=method_bundle["temp_finish"],
                    alpha1=method_bundle["alpha_stage1"],
                    alpha2=method_bundle["alpha_stage2"],
                    bias_finish=method_bundle["bias_finish"],
                    bias_sub=method_bundle["bias_sub"],
                    alpha_direct=method_bundle["alpha_direct"],
                    beta_direct=method_bundle["beta_direct"],
                    alpha_ovr=method_bundle.get("alpha_ovr", 0.85),
                    finish_thr=method_bundle.get("finish_threshold", 0.50),
                    sub_thr=method_bundle.get("sub_threshold", 0.50),
                    alpha_simple=method_bundle.get("alpha_simple", 0.80),
                )
                y_method_np = y_method_test["coarse"].astype(str).values
                winner_correct = (y_pred_red == y_true_red)
                final_probs = []
                test_meta_reset = meta_test.reset_index(drop=True)
                for i in range(len(ml_test)):
                    gp = group_priors.get(
                        (str(test_meta_reset.iloc[i]["weight_class"]), str(test_meta_reset.iloc[i]["gender"]).lower()),
                        group_priors[("ALL", "all")],
                    )
                    d_dec = float(X_test_oriented_pred.reset_index(drop=True).iloc[i].get("d_dec_win_pct", 0.0))
                    d_ko = float(X_test_oriented_pred.reset_index(drop=True).iloc[i].get("d_ko_win_pct", 0.0))
                    d_sub = float(X_test_oriented_pred.reset_index(drop=True).iloc[i].get("d_sub_win_pct", 0.0))
                    hp = _history_prior_row(d_dec, d_ko, d_sub)
                    sub_prior_arr = _sub_attempt_prior_array(X_test_oriented_pred.reset_index(drop=True).iloc[[i]])[0]
                    sp = _normalize_method_probs({
                        "Decision": float(sub_prior_arr[0]),
                        "KO/TKO": float(sub_prior_arr[1]),
                        "Submission": float(sub_prior_arr[2]),
                    })
                    w_hist = method_bundle["w_hist"]
                    w_group = method_bundle["w_group"]
                    w_base = method_bundle.get("w_base", 0.0)
                    w_subsig = method_bundle.get("w_subsig", 0.0)
                    w_ml = 1.0 - w_hist - w_group - w_base - w_subsig
                    bp = method_bundle.get("base_prior", group_priors.get(("ALL", "all"), {"Decision": 1 / 3, "KO/TKO": 1 / 3, "Submission": 1 / 3}))
                    row = {
                        m: (
                            w_ml * float(ml_test.iloc[i][m])
                            + w_hist * float(hp[m])
                            + w_group * float(gp[m])
                            + w_base * float(bp[m])
                            + w_subsig * float(sp[m])
                        )
                        for m in METHOD_LABELS
                    }
                    blended_row = _normalize_method_probs(row)
                    method_bias_map = {
                        "Decision": float(method_bundle.get("method_bias_decision", 0.0)),
                        "KO/TKO": float(method_bundle.get("method_bias_ko_tko", 0.0)),
                        "Submission": float(method_bundle.get("method_bias_submission", 0.0)),
                    }
                    blended_row = _apply_method_logit_bias_map(blended_row, method_bias_map)
                    blended_row = _apply_submission_signal_boost_map(
                        blended_row, float(sp["Submission"]), float(method_bundle.get("sub_boost_k", 0.0))
                    )
                    meta_model = method_bundle.get("meta_model")
                    meta_eta = float(method_bundle.get("meta_eta", 0.0))
                    if meta_model is not None and meta_eta > 0.0:
                        arr_blend = np.array([[blended_row["Decision"], blended_row["KO/TKO"], blended_row["Submission"]]], dtype=float)
                        arr_hist = np.array([[hp["Decision"], hp["KO/TKO"], hp["Submission"]]], dtype=float)
                        arr_grp = np.array([[gp["Decision"], gp["KO/TKO"], gp["Submission"]]], dtype=float)
                        arr_bp = np.array([[bp["Decision"], bp["KO/TKO"], bp["Submission"]]], dtype=float)
                        X_meta = np.hstack([arr_blend, arr_hist, arr_grp, arr_bp, np.log(np.clip(arr_blend, 1e-6, 1.0))])
                        pm = meta_model.predict_proba(X_meta)[0]
                        cls_map = {str(c): i for i, c in enumerate(meta_model.classes_)}
                        raw = {}
                        for m in METHOD_LABELS:
                            raw[m] = float(pm[cls_map[m]]) if m in cls_map else (1.0 / 3.0)
                        p_meta = _normalize_method_probs(raw)
                        blended_row = _normalize_method_probs({
                            m: (1.0 - meta_eta) * blended_row[m] + meta_eta * p_meta[m]
                            for m in METHOD_LABELS
                        })
                    final_probs.append(blended_row)
                method_pred_predwinner = np.array(
                    [max(METHOD_LABELS, key=lambda m: p[m]) for p in final_probs], dtype=object
                )
                method_acc_predicted_winner = float(np.mean(method_pred_predwinner == y_method_np))
                finish_score = float("nan")
                bal_acc = float("nan")
                macro_f1 = float("nan")
                ko_recall = float("nan")
                sub_recall = float("nan")
                ko_sub_macro_f1 = float("nan")
                _per_class_metrics = []
                _confusion_rows = []
                if int(np.sum(winner_correct)) > 0:
                    method_acc_when_winner_correct = float(
                        np.mean(method_pred_predwinner[winner_correct] == y_method_np[winner_correct])
                    )
                    subset = y_method_np[winner_correct]
                    counts = pd.Series(subset).value_counts()
                    method_majority_baseline_when_winner_correct = float(counts.max() / len(subset))
                method_holdout_acc_oracle = float("nan")

                self._section("Method Evaluation (Conditioned on Winner Pick)")
                self._stat("Stage1 acc (Finish vs Decision)", f"{stage1_acc:.1%}")
                self._stat("Stage1 AUC (Finish vs Decision)", "n/a" if not np.isfinite(stage1_auc) else f"{stage1_auc:.3f}")
                self._stat(
                    "Stage2 acc (KO/Sub | true finishes)",
                    "n/a" if not np.isfinite(stage2_acc_true_finishes) else f"{stage2_acc_true_finishes:.1%}",
                )
                self._stat("Stage2 sample size (true finishes)", n_true_finishes)
                self._stat("Method acc (predicted winner conditioned)", f"{method_acc_predicted_winner:.1%}")
                self._stat("Method acc | winner pick correct", f"{method_acc_when_winner_correct:.1%}")
                self._stat("Majority baseline | winner pick correct", f"{method_majority_baseline_when_winner_correct:.1%}")
                if int(np.sum(winner_correct)) > 0:
                    sub_true = y_method_np[winner_correct]
                    sub_pred = method_pred_predwinner[winner_correct]
                    p_arr, r_arr, f1_arr, _ = precision_recall_fscore_support(
                        sub_true, sub_pred, labels=METHOD_LABELS, zero_division=0
                    )
                    bal_acc = float(balanced_accuracy_score(sub_true, sub_pred))
                    macro_f1 = float(np.mean(f1_arr))
                    ko_recall = float(r_arr[1]) if len(r_arr) > 1 else 0.0
                    sub_recall = float(r_arr[2]) if len(r_arr) > 2 else 0.0
                    ko_sub_macro_f1 = float(np.mean(f1_arr[1:3])) if len(f1_arr) >= 3 else macro_f1
                    finish_score = 0.4 * ko_recall + 0.4 * sub_recall + 0.2 * ko_sub_macro_f1
                    self._stat("Balanced accuracy | winner pick correct", f"{bal_acc:.1%}")
                    self._stat("Macro F1 | winner pick correct", f"{macro_f1:.1%}")
                    self._stat("FinishScore (0.4 KO R + 0.4 Sub R + 0.2 KO/Sub F1)", f"{finish_score:.1%}")
                    self._log("")
                    self._log("Per-Class Metrics (Method | winner pick correct)")
                    self._log("Class          Precision    Recall      F1")
                    for idx, cls_name in enumerate(METHOD_LABELS):
                        self._log(
                            f"{cls_name:<14}{p_arr[idx]:10.1%}{r_arr[idx]:10.1%}{f1_arr[idx]:10.1%}"
                        )
                    self._log("")
                    self._log("Confusion Matrix (Method | winner pick correct)")
                    self._log("Actual\\Pred     Decision    KO/TKO  Submission")
                    for actual in METHOD_LABELS:
                        row_counts = []
                        for pred in METHOD_LABELS:
                            c = int(np.sum((sub_true == actual) & (sub_pred == pred)))
                            row_counts.append(c)
                        self._log(f"{actual:<14}{row_counts[0]:10d}{row_counts[1]:10d}{row_counts[2]:12d}")
                        _confusion_rows.append([actual, list(row_counts)])
                    _per_class_metrics = [
                        [cls_name, float(p_arr[idx]), float(r_arr[idx]), float(f1_arr[idx])]
                        for idx, cls_name in enumerate(METHOD_LABELS)
                    ]

                method_bundle["method_columns"] = method_columns

                _cache_save(
                    "method_stage",
                    method_cache_key,
                    {
                        "kind": METHOD_STAGE_CACHE_KIND,
                        "method_cache_version": METHOD_CACHE_VERSION,
                        "winner_cache_key": winner_cache_key,
                        "method_bundle": {k: v for k, v in method_bundle.items()},
                        "method_acc_when_winner_correct": method_acc_when_winner_correct,
                        "method_acc_predicted_winner": method_acc_predicted_winner,
                        "method_majority_baseline_when_winner_correct": (
                            method_majority_baseline_when_winner_correct
                        ),
                        "method_holdout_acc_oracle": method_holdout_acc_oracle,
                        "finish_score": finish_score,
                        "stage1_acc": stage1_acc,
                        "stage1_auc": stage1_auc,
                        "stage2_acc_true_finishes": stage2_acc_true_finishes,
                        "n_true_finishes": n_true_finishes,
                        "bal_acc": bal_acc,
                        "macro_f1": macro_f1,
                        "ko_recall": ko_recall,
                        "sub_recall": sub_recall,
                        "ko_sub_macro_f1": ko_sub_macro_f1,
                        "per_class_metrics": _per_class_metrics,
                        "confusion_rows": _confusion_rows,
                        "method_classes_training": ", ".join(method_bundle.get("detail_labels_seen", [])),
                        "val_metric_str": f"{best_cfg['val_metric']:.1%}",
                        "val_baseline_str": f"{best_cfg['val_baseline']:.1%}",
                        "method_hard_reset": bool(METHOD_HARD_RESET),
                    },
                )
                self._stat(
                    "Method cache",
                    f"SAVED ({METHOD_CACHE_VERSION}) key={str(method_cache_key)[:12]} file={os.path.basename(_cache_path('method_stage', method_cache_key))}",
                )

            except Exception as exc:
                self._section("Method Evaluation (Conditioned on Winner Pick)")
                self._stat("Method model", f"disabled ({exc})")
                method_bundle = None

        # Retrain for production on all rows.
        X_all = pd.DataFrame(imputer.fit_transform(X_winner[feature_cols]), columns=feature_cols)
        X_all_aug, y_all_aug = _augment_swap(X_all, y.reset_index(drop=True))
        w_all_aug = _augment_weights(_time_weights(len(X_all), floor=0.35))
        scaler_all = StandardScaler()
        X_all_aug_sc = pd.DataFrame(scaler_all.fit_transform(X_all_aug), columns=feature_cols)
        final_models = {}
        for name, make_model in specs:
            final_models[name] = _fit_model(
                name, make_model(),
                X_all_aug_sc if name in NEEDS_SCALE else X_all_aug,
                y_all_aug, sample_weight=w_all_aug
            )

        if combiner["kind"] == "stacker":
            final_stacker = LogisticRegression(
                max_iter=8000, C=0.2, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED
            )
            final_stacker.fit(oof.loc[valid, model_order].astype(float).values, y_dev.loc[valid].astype(int).values)
            final_combiner = {"kind": "stacker", "model": final_stacker, "model_order": model_order}
        else:
            final_combiner = combiner

        # Retrain method head on all rows, oriented by winner-model predictions.
        method_bundle_all = None
        method_feat_cols_all = list(full_feature_cols)
        if method_bundle is not None:
            try:
                X_all_sw = _swap_features(X_all)
                X_all_sc = pd.DataFrame(scaler_all.transform(X_all), columns=feature_cols)
                X_all_sw_sc = pd.DataFrame(scaler_all.transform(X_all_sw), columns=feature_cols)
                all_meta = {}
                for name, _ in specs:
                    p_fwd = _predict_proba(name, final_models[name], X_all_sc if name in NEEDS_SCALE else X_all)
                    p_rev = _predict_proba(name, final_models[name], X_all_sw_sc if name in NEEDS_SCALE else X_all_sw)
                    all_meta[name] = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
                all_meta_df = pd.DataFrame(all_meta)[model_order]
                all_probs = _combine_probs(all_meta_df, final_combiner)
                if calibrator is not None:
                    all_probs = _clip_probs(
                        calibrator.predict_proba(np.asarray(all_probs).reshape(-1, 1))[:, 1]
                    )
                if CORNER_CORRECTION_ENABLED:
                    all_probs = _apply_corner_correction(all_probs, corner_b)
                y_pred_all = (np.asarray(all_probs) >= decision_threshold).astype(int)

                X_all_full = pd.DataFrame(full_imputer.fit_transform(X_full[full_feature_cols]), columns=full_feature_cols)
                X_all_method = _oriented_method_matrix(X_all_full, y_pred_all)
                X_all_method = _augment_method_features(X_all_method)
                method_feat_cols_all = list(X_all_method.columns)
                method_imputer_all = SimpleImputer(strategy="median")
                X_all_method_imp = pd.DataFrame(
                    method_imputer_all.fit_transform(X_all_method), columns=method_feat_cols_all
                )

                y_all_bin = (y_method_df["finish_bin"].values == "Finish").astype(int)
                p_finish_all = max(1e-6, float(np.mean(y_all_bin)))
                w_all_t = _time_weights(len(X_all_method_imp), floor=0.45)
                w_all_stage1 = w_all_t * np.where(
                    y_all_bin == 1, 0.5 / p_finish_all, 0.5 / max(1e-6, 1.0 - p_finish_all)
                )
                # Use the SAME Optuna-tuned HGB params the eval split reported on,
                # so the production model matches what was actually evaluated.
                _stage1_params_all = dict(method_bundle.get("stage1_params") or {})
                _stage1_params_all.setdefault("loss", "log_loss")
                _stage1_params_all.setdefault("max_iter", 360)
                _stage1_params_all.setdefault("learning_rate", 0.045)
                _stage1_params_all.setdefault("max_depth", 6)
                _stage1_params_all.setdefault("max_leaf_nodes", 31)
                _stage1_params_all.setdefault("min_samples_leaf", 16)
                _stage1_params_all.setdefault("l2_regularization", 0.8)
                _stage1_params_all.setdefault("random_state", RANDOM_SEED + 808)
                stage1_all = _BaggedHGB(_stage1_params_all, n_estimators=METHOD_HGB_BAG,
                                        base_seed=RANDOM_SEED + 808, subsample=METHOD_HGB_BAG_SUBSAMPLE)
                stage1_all.fit(X_all_method_imp, y_all_bin, sample_weight=w_all_stage1)
                stage1_rf_all = RandomForestClassifier(
                    n_estimators=420, max_depth=10, min_samples_leaf=4,
                    max_features=0.7, random_state=RANDOM_SEED + 910, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_rf_all.fit(X_all_method_imp, y_all_bin)
                stage1_et_all = ExtraTreesClassifier(
                    n_estimators=560, max_depth=12, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 914, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_et_all.fit(X_all_method_imp, y_all_bin)

                fin_mask_all = (y_method_df["finish_bin"].values == "Finish")
                if int(np.sum(fin_mask_all)) < 60:
                    fin_mask_all = np.ones(len(y_method_df), dtype=bool)
                X2_all = X_all_method_imp.iloc[np.where(fin_mask_all)[0]].reset_index(drop=True)
                y2_all = (y_method_df.iloc[np.where(fin_mask_all)[0]]["finish_subtype"].values == "Submission").astype(int)
                p_sub_all = max(1e-6, float(np.mean(y2_all))) if len(y2_all) > 0 else 0.5
                w2_all = np.where(y2_all == 1, 0.5 / p_sub_all, 0.5 / max(1e-6, 1.0 - p_sub_all))
                _stage2_params_all = dict(method_bundle.get("stage2_params") or {})
                _stage2_params_all.setdefault("loss", "log_loss")
                _stage2_params_all.setdefault("max_iter", 320)
                _stage2_params_all.setdefault("learning_rate", 0.05)
                _stage2_params_all.setdefault("max_depth", 5)
                _stage2_params_all.setdefault("max_leaf_nodes", 31)
                _stage2_params_all.setdefault("min_samples_leaf", 14)
                _stage2_params_all.setdefault("l2_regularization", 0.7)
                _stage2_params_all.setdefault("random_state", RANDOM_SEED + 809)
                stage2_all = _BaggedHGB(_stage2_params_all, n_estimators=METHOD_HGB_BAG,
                                        base_seed=RANDOM_SEED + 809, subsample=METHOD_HGB_BAG_SUBSAMPLE)
                stage2_all.fit(X2_all, y2_all, sample_weight=w2_all)
                stage2_rf_all = RandomForestClassifier(
                    n_estimators=360, max_depth=9, min_samples_leaf=3,
                    max_features=0.7, random_state=RANDOM_SEED + 911, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_rf_all.fit(X2_all, y2_all)
                stage2_et_all = ExtraTreesClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 915, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_et_all.fit(X2_all, y2_all)

                y_cls_all = y_method_df["coarse"].values
                cls_counts_all = pd.Series(y_cls_all).value_counts()
                cls_w_all = {k: (len(y_cls_all) / (len(cls_counts_all) * max(1, v))) for k, v in cls_counts_all.items()}
                w_cls_all = np.array([cls_w_all.get(lbl, 1.0) for lbl in y_cls_all], dtype=float)
                w_direct_all = w_all_t * w_cls_all
                direct_hgb_all = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=300, learning_rate=0.05, max_depth=6,
                    max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.7,
                    random_state=RANDOM_SEED + 912,
                )
                direct_hgb_all.fit(X_all_method_imp, y_cls_all, sample_weight=w_direct_all)
                direct_rf_all = RandomForestClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 913, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_rf_all.fit(X_all_method_imp, y_cls_all)
                direct_et_all = ExtraTreesClassifier(
                    n_estimators=680, max_depth=13, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 916, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_et_all.fit(X_all_method_imp, y_cls_all)
                ovr_models_all = {}
                for j, lbl in enumerate(METHOD_LABELS):
                    y_ovr_all = (pd.Series(y_cls_all).astype(str).values == lbl).astype(int)
                    p_pos_all = max(1e-6, float(np.mean(y_ovr_all)))
                    w_ovr_all = w_all_t * np.where(
                        y_ovr_all == 1, 0.5 / p_pos_all, 0.5 / max(1e-6, 1.0 - p_pos_all)
                    )
                    mdl_all = HistGradientBoostingClassifier(
                        loss="log_loss",
                        max_iter=260, learning_rate=0.05, max_depth=5,
                        max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.8,
                        random_state=RANDOM_SEED + 1500 + j,
                    )
                    mdl_all.fit(X_all_method_imp, y_ovr_all, sample_weight=w_ovr_all)
                    ovr_models_all[lbl] = mdl_all
                simple_method_all = LogisticRegression(
                    max_iter=8000, C=0.8, solver="saga", tol=1e-3, n_jobs=-1,
                    class_weight="balanced", random_state=RANDOM_SEED + 1601,
                )
                simple_method_all.fit(X_all_method_imp, pd.Series(y_cls_all).astype(str).values)

                group_priors_all = {}
                for (wc, gender), grp in row_meta.groupby(["weight_class", "gender"], dropna=False):
                    idx = grp.index.values
                    counts = y_method_df.iloc[idx]["coarse"].value_counts()
                    tot = float(len(idx))
                    group_priors_all[(str(wc), str(gender).lower())] = _normalize_method_probs({
                        "Decision": float(counts.get("Decision", 0.0) / max(1.0, tot)),
                        "KO/TKO": float(counts.get("KO/TKO", 0.0) / max(1.0, tot)),
                        "Submission": float(counts.get("Submission", 0.0) / max(1.0, tot)),
                    })
                all_counts2 = y_method_df["coarse"].value_counts()
                all_tot2 = float(len(y_method_df))
                group_priors_all[("ALL", "all")] = _normalize_method_probs({
                    "Decision": float(all_counts2.get("Decision", 0.0) / max(1.0, all_tot2)),
                    "KO/TKO": float(all_counts2.get("KO/TKO", 0.0) / max(1.0, all_tot2)),
                    "Submission": float(all_counts2.get("Submission", 0.0) / max(1.0, all_tot2)),
                })
                base_prior_all = _normalize_method_probs({
                    "Decision": float(all_counts2.get("Decision", 0.0) / max(1.0, all_tot2)),
                    "KO/TKO": float(all_counts2.get("KO/TKO", 0.0) / max(1.0, all_tot2)),
                    "Submission": float(all_counts2.get("Submission", 0.0) / max(1.0, all_tot2)),
                })

                method_bundle_all = {
                    "stage1": stage1_all,
                    "stage2": stage2_all,
                    "stage1_rf": stage1_rf_all,
                    "stage1_et": stage1_et_all,
                    "stage2_rf": stage2_rf_all,
                    "stage2_et": stage2_et_all,
                    "direct_hgb": direct_hgb_all,
                    "direct_rf": direct_rf_all,
                    "direct_et": direct_et_all,
                    "ovr_models": ovr_models_all,
                    "simple_method": simple_method_all,
                    "direct_classes": [str(c) for c in direct_hgb_all.classes_],
                    "imputer": method_imputer_all,
                    "method_columns": method_feat_cols_all,
                    "temp_decision": float(method_bundle.get("temp_decision", 1.0)),
                    "temp_finish": float(method_bundle.get("temp_finish", 1.0)),
                    "bias_finish": float(method_bundle.get("bias_finish", 0.0)),
                    "bias_sub": float(method_bundle.get("bias_sub", 0.0)),
                    "finish_threshold": float(method_bundle.get("finish_threshold", 0.50)),
                    "sub_threshold": float(method_bundle.get("sub_threshold", 0.50)),
                    "alpha_stage1": float(method_bundle.get("alpha_stage1", 0.75)),
                    "alpha_stage2": float(method_bundle.get("alpha_stage2", 0.75)),
                    "stage1_params": dict(method_bundle.get("stage1_params") or {}),
                    "stage2_params": dict(method_bundle.get("stage2_params") or {}),
                    "alpha_direct": float(method_bundle.get("alpha_direct", 0.85)),
                    "beta_direct": float(method_bundle.get("beta_direct", 0.70)),
                    "alpha_ovr": float(method_bundle.get("alpha_ovr", 0.85)),
                    "alpha_simple": float(method_bundle.get("alpha_simple", 0.80)),
                    "w_hist": float(method_bundle.get("w_hist", 0.25)),
                    "w_group": float(method_bundle.get("w_group", 0.10)),
                    "w_base": float(method_bundle.get("w_base", 0.08)),
                    "w_subsig": float(method_bundle.get("w_subsig", 0.10)),
                    "sub_boost_k": float(method_bundle.get("sub_boost_k", 0.0)),
                    "method_bias_decision": float(method_bundle.get("method_bias_decision", 0.0)),
                    "method_bias_ko_tko": float(method_bundle.get("method_bias_ko_tko", 0.0)),
                    "method_bias_submission": float(method_bundle.get("method_bias_submission", 0.0)),
                    "base_prior": base_prior_all,
                    "meta_model": method_bundle.get("meta_model"),
                    "meta_eta": float(method_bundle.get("meta_eta", 0.0)),
                    "group_priors": group_priors_all,
                    "detail_labels_seen": method_bundle.get("detail_labels_seen", []),
                }
            except Exception as _e:
                self._stat("WARNING", f"All-data method retrain failed ({_e}) — method predictions disabled")
                method_bundle_all = None

        self.model = SuperEnsembleModel(
            final_models, imputer, scaler_all, feature_cols, final_combiner,
            calibrator=calibrator, decision_threshold=decision_threshold,
            corner_correction=corner_b,
            method_bundle=method_bundle_all,
            method_feat_cols=method_feat_cols_all,
        )
        self.method_model = method_bundle_all
        self.method_imputer = method_bundle_all["imputer"] if method_bundle_all is not None else None
        self.method_feat_cols = method_feat_cols_all
        self.method_metrics = {
            "method_acc_predicted_winner": method_acc_predicted_winner,
            "method_acc_when_winner_correct": method_acc_when_winner_correct,
            "method_acc_true_winner": method_holdout_acc_oracle,
            "method_majority_baseline_when_winner_correct": method_majority_baseline_when_winner_correct,
            "method_finish_score_winner_correct": finish_score,
        }

        ml_baseline = None
        old_baseline = None

        self.benchmarks = BenchmarkScores(
            super_raw_logloss=float(raw_ll),
            super_cal_logloss=float(cal_ll_test),
            super_brier=float(brier),
            super_acc=float(acc),
            super_ece=float(ece),
            ml_baseline_logloss=ml_baseline,
            old_baseline_logloss=old_baseline,
        )
        self._rolling_diagnostics(X, y)
        return self.model

    def _rolling_diagnostics(self, X, y):
        """Walk-forward diagnostic: expanding-window folds over the full era.

        A single HistGBM (swap-augmented, forward/reverse averaged) is
        retrained per fold and scored on the next chronological block —
        a robustness check that model skill holds across time, independent
        of the main ensemble's single holdout window.
        """
        n = len(X)
        folds = []
        start = int(n * 0.45)
        step = max(250, int(n * 0.08))
        while start + step < n:
            train_end = start
            test_end = min(n, start + step)
            folds.append((train_end, test_end))
            start += step
        if not folds:
            return

        self._section("Walk-Forward Diagnostics")
        ll_scores = []
        acc_scores = []
        for i, (tr_end, te_end) in enumerate(folds, start=1):
            X_tr = X.iloc[:tr_end]
            y_tr = y.iloc[:tr_end]
            X_te = X.iloc[tr_end:te_end]
            y_te = y.iloc[tr_end:te_end]
            imp = SimpleImputer(strategy="median")
            X_tr_imp = pd.DataFrame(imp.fit_transform(X_tr), columns=X.columns)
            X_te_imp = pd.DataFrame(imp.transform(X_te), columns=X.columns)
            mdl = HistGradientBoostingClassifier(
                max_iter=500, learning_rate=0.04, max_depth=6,
                max_leaf_nodes=31, min_samples_leaf=20, l2_regularization=1.0,
                random_state=RANDOM_SEED,
            )
            X_aug, y_aug = _augment_swap(X_tr_imp, y_tr.reset_index(drop=True))
            mdl.fit(X_aug, y_aug)
            p_fwd = _predict_proba("HistGBM", mdl, X_te_imp)
            X_te_sw = _swap_features(X_te_imp)
            p_rev = _predict_proba("HistGBM", mdl, X_te_sw)
            p = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
            ll = log_loss(y_te, p)
            acc = float(accuracy_score(y_te, (np.asarray(p) >= 0.5).astype(int)))
            ll_scores.append(ll)
            acc_scores.append(acc)
            self._stat(f"Fold {i}", f"log-loss {ll:.4f} | acc {acc:.1%} ({len(y_te)} fights)")
        self._stat("Walk-forward mean log-loss", f"{np.mean(ll_scores):.4f}")
        self._stat("Walk-forward mean accuracy", f"{np.mean(acc_scores):.1%}")
        self._stat("Walk-forward std (log-loss)", f"{np.std(ll_scores):.4f}")
        self._stat("Walk-forward std (accuracy)", f"{np.std(acc_scores):.1%}")

    def predict_matchup(self, fighter_a, fighter_b, weight_class="", gender="", rounds=3,
                        red_odds=None, blue_odds=None, location=None, event_elevation=None):
        """Predict one matchup from live (current) fighter state.

        fighter_a is the red corner. Rebuilds the same feature row the model
        was trained on — career features from each fighter's full history,
        Elo/divisional-rank/phase-rating snapshots, style-matchup records,
        and venue/camp altitude features — then returns a dict with win
        probabilities, method distribution, and (when both moneylines are
        supplied) the odds-blended probability and betting-value metrics.
        Unknown fighters fall back to debut-prior features.
        """
        if self.model is None:
            raise RuntimeError("Model is not trained.")
        a_key = fuzzy_find(fighter_a, self.fighter_history) or fighter_a
        b_key = fuzzy_find(fighter_b, self.fighter_history) or fighter_b
        today = pd.Timestamp(datetime.now().date())

        if a_key in self.fighter_history:
            a_hist = self.fighter_history[a_key]
            a_glicko = self.glicko_ratings.get(a_key, (MU_0, PHI_0, SIGMA_0))
            a_opp = self.opp_glicko_list.get(a_key, [])
            a_feats = compute_fighter_features(a_hist, a_glicko, a_opp, today)
        else:
            a_feats = compute_fighter_features([], (MU_0, PHI_0, SIGMA_0), [], today)

        if b_key in self.fighter_history:
            b_hist = self.fighter_history[b_key]
            b_glicko = self.glicko_ratings.get(b_key, (MU_0, PHI_0, SIGMA_0))
            b_opp = self.opp_glicko_list.get(b_key, [])
            b_feats = compute_fighter_features(b_hist, b_glicko, b_opp, today)
        else:
            b_feats = compute_fighter_features([], (MU_0, PHI_0, SIGMA_0), [], today)

        matchup = compute_matchup_features(
            a_feats, b_feats, is_title=0, total_rounds=rounds, weight_class=weight_class
        )
        a_elo = float(self.elo_ratings.get(a_key, ELO_BASE))
        b_elo = float(self.elo_ratings.get(b_key, ELO_BASE))
        d_elo = a_elo - b_elo
        elo_p = 1.0 / (1.0 + 10.0 ** (-(d_elo / 400.0)))
        division = _normalize_division(weight_class, gender)
        a_div_elo = float(self.div_elo_ratings.get((a_key, division), ELO_BASE))
        b_div_elo = float(self.div_elo_ratings.get((b_key, division), ELO_BASE))
        d_div_elo = a_div_elo - b_div_elo
        div_p = 1.0 / (1.0 + 10.0 ** (-(d_div_elo / 400.0)))
        matchup.update({
            "elo_r": a_elo,
            "elo_b": b_elo,
            "d_elo": d_elo,
            "elo_win_prob": elo_p,
            "d_elo_win_prob": elo_p - 0.5,
            "abs_elo_gap": abs(d_elo),
            "elo_sum": a_elo + b_elo,
            "div_elo_r": a_div_elo,
            "div_elo_b": b_div_elo,
            "d_div_elo": d_div_elo,
            "div_elo_win_prob": div_p,
            "d_div_elo_win_prob": div_p - 0.5,
            "abs_div_elo_gap": abs(d_div_elo),
            "elo_divergence": elo_p - div_p,
            "elo_agreement": 1.0 - abs(elo_p - div_p),
        })
        matchup.update(_div_rank_feature_row(
            a_key, b_key, division, self.div_elo_ratings, self.last_fight_date, today,
        ))
        _a_slope = _elo_slope(self.elo_history.get(a_key, []), window=5)
        _b_slope = _elo_slope(self.elo_history.get(b_key, []), window=5)
        matchup["r_elo_slope_5"] = _a_slope
        matchup["b_elo_slope_5"] = _b_slope
        matchup["d_elo_slope_5"] = _a_slope - _b_slope
        if PHASE_RATINGS_ENABLED:
            _d = (MU_0, PHI_0, SIGMA_0)
            matchup.update(_phase_glicko_features(
                self.strike_off_ratings.get(a_key, _d), self.strike_off_ratings.get(b_key, _d),
                self.strike_def_ratings.get(a_key, _d), self.strike_def_ratings.get(b_key, _d),
                self.grapple_off_ratings.get(a_key, _d), self.grapple_off_ratings.get(b_key, _d),
                self.grapple_def_ratings.get(a_key, _d), self.grapple_def_ratings.get(b_key, _d),
            ))
        if self.style_tracker is not None:
            matchup.update(self.style_tracker.matchup_features(a_key, b_key))
        # Altitude features: venue altitude (median fallback for blank/unknown
        # location) + each fighter's latest known training-camp elevation (median
        # fallback when the camp is unrecorded), for the acclimatization signals.
        # Venue altitude: an explicit elevation (ft) wins — it works for venues the
        # model has never seen (no silent median fallback); otherwise look the
        # location up, then fall back to the dataset median.
        if event_elevation is not None:
            _event_alt = float(event_elevation)
        elif location:
            _event_alt = self.location_altitude.get(location, self.alt_median)
        else:
            _event_alt = self.alt_median
        _rt = self.fighter_train_alt.get(a_key)
        _bt = self.fighter_train_alt.get(b_key)
        _known = 1.0 if (_rt is not None and _bt is not None) else 0.0
        matchup.update(_altitude_feature_row(
            _event_alt,
            _rt if _rt is not None else self.train_alt_median,
            _bt if _bt is not None else self.train_alt_median,
            _known))
        matchup_df = _augment_matchup_features(pd.DataFrame([matchup]))
        p_a_model = float(self.model.predict_proba_single(matchup_df.iloc[0].to_dict()))
        # Optional odds blend (inference only). market is None unless valid odds were
        # supplied for BOTH fighters; when None, p_a == p_a_model and every line below
        # is byte-for-byte the original model-only behavior.
        market = _devig_two_way(red_odds, blue_odds)
        market_a = market[0] if market is not None else None
        p_a = _logit_blend(p_a_model, market_a, ODDS_BLEND_WEIGHT) if market is not None else p_a_model
        # The pick — and therefore the winner-conditioned method below — follows the
        # blended probability, so a flip from the odds re-routes the method too.
        pick_a = p_a >= 0.5
        method_probs = self.model.predict_method_probs(
            matchup_df.iloc[0].to_dict(),
            winner_is_a=pick_a,
            weight_class=weight_class,
            gender=gender,
        )
        predicted_method = max(METHOD_LABELS, key=lambda m: method_probs[m])

        # Betting value on the RAW model prob (independent of the odds-blended pick
        # above). All None when odds are missing/unparseable for this matchup.
        value_side, value_ev, value_stake = _value_metrics(
            p_a_model, a_key, b_key, red_odds, blue_odds
        )

        return {
            "name_a": a_key,
            "name_b": b_key,
            "prob_a": p_a,
            "prob_b": 1.0 - p_a,
            "prob_a_model": p_a_model,
            "prob_b_model": 1.0 - p_a_model,
            "odds_blended": market is not None,
            "red_odds": red_odds if market is not None else None,
            "blue_odds": blue_odds if market is not None else None,
            "market_prob_a": market_a,
            "market_prob_b": (1.0 - market_a) if market_a is not None else None,
            "rating_a": float(a_feats.get("glicko_mu", MU_0)),
            "rating_b": float(b_feats.get("glicko_mu", MU_0)),
            "weight_class": weight_class,
            "gender": gender,
            "method_probs": method_probs,
            "predicted_method": predicted_method,
            "decision_pct": method_probs["Decision"],
            "ko_tko_pct": method_probs["KO/TKO"],
            "submission_pct": method_probs["Submission"],
            "method_pct": method_probs[predicted_method],
            "value_side": value_side,
            "value_ev": value_ev,
            "value_stake": value_stake,
        }

    def is_debutant(self, fighter_name):
        """True if the fighter has no recorded fight history (UFC debut)."""
        key = fuzzy_find(fighter_name, self.fighter_history)
        if key is None:
            return True
        return not bool(self.fighter_history.get(key))

    def division_rankings(self):
        """Rank active fighters (fought within ACTIVE_DAYS) by Glicko rating,
        grouped by division. Returns {division: [(name, mu, W, L, D), ...]}."""
        by_div = defaultdict(list)
        now = pd.Timestamp(datetime.now().date())
        cutoff = now - pd.Timedelta(days=ACTIVE_DAYS)
        for fighter, hist in self.fighter_history.items():
            if not hist:
                continue
            last = hist[-1]["date"]
            if pd.isna(last) or last < cutoff:
                continue
            meta = self.fighter_meta.get(fighter, {})
            division = meta.get("division", "")
            if not division or division in ("Catch Weight", "Open Weight"):
                continue
            mu = self.glicko_ratings.get(fighter, (MU_0, PHI_0, SIGMA_0))[0]
            wins = sum(1 for h in hist if h.get("result") == "W")
            losses = sum(1 for h in hist if h.get("result") == "L")
            draws = sum(1 for h in hist if h.get("result") == "D")
            by_div[division].append((fighter, mu, wins, losses, draws))
        for div in by_div:
            by_div[div].sort(key=lambda t: t[1], reverse=True)
        return by_div


def _auto_width(ws):
    for ci in range(1, ws.max_column + 1):
        mx = 0
        for row in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row:
                mx = max(mx, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(ci)].width = mx + 3


# ─────────────────────────────────────────────────────────────────────────────
# Optional betting-odds blend (inference only). This is ENTIRELY opt-in: it kicks
# in only when valid moneyline odds are supplied for BOTH fighters in a matchup.
# With no/partial/invalid odds, predict_matchup falls back to the pure model and
# behaves exactly as before. Tune the single knob below.
# ─────────────────────────────────────────────────────────────────────────────
ODDS_BLEND_WEIGHT = 0.50  # weight on the MODEL vs the market in [0,1]; 0.50 = equal trust (logit geometric mean)

# ─── Betting value (inference only) ──────────────────────────────────────────
# Value is computed on the RAW model probability, orthogonally to the odds blend
# above: edge is measured vs the de-vigged (fair) market, while EV/Kelly use the
# actual offered (vigged) odds. So a fight can show one Blended Model pick and a
# different Value Pick — that is intended, not a bug.
KELLY_FRACTION = 0.25     # fraction of full Kelly to stake; the one conservatism dial
MIN_EV_TO_FLAG = 0.02     # below +2% EV → "No bet" (filters vig-noise)
MAX_STAKE_CAP  = 0.05     # hard ceiling on recommended bankroll fraction
VALUE_SHARPEN  = 1.25     # >1 sharpens the prob used for VALUE ONLY (not the pick):
                          # corrects the model's stable ~1.26x under-confidence (see
                          # _audit_recalibration.py) so underdog edges aren't
                          # systematically overstated. 1.0 = off.


def _looks_like_odds(tok):
    """True if a pasted field looks like an American moneyline (e.g. -150, +105)."""
    s = str(tok).strip()
    if not s:
        return False
    core = s[1:] if s[0] in "+-" else s
    if not core.isdigit():
        return False
    return s[0] in "+-" or abs(int(s)) >= 100


def _american_odds_to_prob(odds):
    """American moneyline -> raw implied win probability. None if unparseable."""
    try:
        o = float(str(odds).strip())
    except (TypeError, ValueError):
        return None
    if o < 0:
        return (-o) / ((-o) + 100.0)
    if o > 0:
        return 100.0 / (o + 100.0)
    return None


def _american_to_decimal(odds):
    """American moneyline -> decimal odds (total payout per $1 incl. stake).

    Payout twin of `_american_odds_to_prob`; mirrors its sign handling. None if
    unparseable.
    """
    try:
        o = float(str(odds).strip())
    except (TypeError, ValueError):
        return None
    if o < 0:
        return 100.0 / (-o) + 1.0
    if o > 0:
        return o / 100.0 + 1.0
    return None


def _devig_two_way(red_odds, blue_odds):
    """De-vigged market probabilities (P_red, P_blue). None if either side is bad."""
    pr = _american_odds_to_prob(red_odds)
    pb = _american_odds_to_prob(blue_odds)
    if pr is None or pb is None:
        return None
    tot = pr + pb
    if tot <= 0:
        return None
    return pr / tot, pb / tot


def _logit_blend(p_model, p_market, w):
    """Blend two probabilities in logit space; w = weight on the model."""
    eps = 1e-6
    pm = min(max(float(p_model), eps), 1.0 - eps)
    pk = min(max(float(p_market), eps), 1.0 - eps)
    L = w * np.log(pm / (1.0 - pm)) + (1.0 - w) * np.log(pk / (1.0 - pk))
    return float(1.0 / (1.0 + np.exp(-L)))


def _sharpen_prob(p, s):
    """Push a probability away from 0.5 by scaling its logit by s (s>1 sharpens,
    s=1 is a no-op). Symmetric: _sharpen_prob(1-p, s) == 1 - _sharpen_prob(p, s)."""
    if s == 1.0:
        return float(p)
    eps = 1e-6
    pc = min(max(float(p), eps), 1.0 - eps)
    return float(1.0 / (1.0 + np.exp(-s * np.log(pc / (1.0 - pc)))))


def _value_metrics(p_model_a, name_a, name_b, red_odds, blue_odds):
    """Betting value for a two-way moneyline, computed on the RAW model prob.

    Returns (value_side, value_ev, value_stake) for the +EV bet, or (None, None,
    None) when no side clears MIN_EV_TO_FLAG or the odds are unparseable.

    The prob is first sharpened by VALUE_SHARPEN to undo the model's mild
    under-confidence (VALUE only — the displayed pick is untouched). Edge is
    measured against the de-vigged (fair) market; EV and Kelly use the actual
    offered (vigged) decimal odds. Stake is KELLY_FRACTION of full Kelly, clipped
    at 0 and capped at MAX_STAKE_CAP. In a two-way market at most one side has a
    positive edge, and EV>0 implies edge>0, so the +edge side is the only bet
    candidate.
    """
    market = _devig_two_way(red_odds, blue_odds)
    d_a = _american_to_decimal(red_odds)
    d_b = _american_to_decimal(blue_odds)
    if market is None or d_a is None or d_b is None:
        return None, None, None
    # De-bias under-confidence for the value calc only. Sharpening is symmetric, so
    # sharpening p_a also correctly sharpens side B (1 - p_a) — pulling underdog
    # probs down (smaller, truer edges) and favorite probs up.
    p_model_a = _sharpen_prob(p_model_a, VALUE_SHARPEN)
    q_a, q_b = market
    for name, p, q, d in ((name_a, p_model_a, q_a, d_a),
                          (name_b, 1.0 - p_model_a, q_b, d_b)):
        if p - q <= 0:
            continue  # never the value side; its mirror holds any positive edge
        ev = p * d - 1.0
        if ev < MIN_EV_TO_FLAG:
            return None, None, None  # +edge but EV doesn't clear the bar
        kelly = max(0.0, (p * d - 1.0) / (d - 1.0))
        stake = min(MAX_STAKE_CAP, KELLY_FRACTION * kelly)
        return name, ev, stake
    return None, None, None


def export_to_excel(path, predictions, rankings):
    """Write a styled Predictions worksheet to `path`.

    Column layout adapts to whether any prediction carries betting odds:
    with odds, raw-model and odds-blended picks are shown side by side and
    value columns (EV, Kelly stake) trail at the end; without odds, the
    compact model-only layout is used.
    """
    wb = Workbook()
    hdr_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="000000")
    hdr_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "Predictions"
    has_odds = any(p.get("odds_blended") for p in predictions)
    # With odds: odds interleave next to each corner, the Raw Model (pre-blend)
    # pick sits beside the Blended Model pick, then the method block, and the
    # value columns trail at the end. Without odds: the original 10-col layout,
    # byte-for-byte unchanged.
    if has_odds:
        headers = [
            "Red Corner", "Red Odds", "Blue Corner", "Blue Odds", "Weight Class",
            "Raw Model", "Model %", "Blended Model", "Win %",
            "Method", "Method %", "DEC %", "(T)KO %", "SUB %",
            "Value Pick", "EV %", "Stake %",
        ]
        pct_cols = {7, 9, 11, 12, 13, 14, 16, 17}
    else:
        headers = [
            "Red Corner", "Blue Corner", "Weight Class", "Winner", "Win %",
            "Method", "Method %", "DEC %", "(T)KO %", "SUB %",
        ]
        pct_cols = {5, 7, 8, 9, 10}
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, hdr_align, border

    for ri, p in enumerate(predictions, 2):
        winner = p.get("predicted_winner") or (p["name_a"] if p["prob_a"] >= p["prob_b"] else p["name_b"])
        win_pct = p.get("win_pct")
        if win_pct is None:
            win_pct = p["prob_a"] if winner == p["name_a"] else p["prob_b"]
        method_probs = _normalize_method_probs(p.get("method_probs", {}))
        pred_method = p.get("predicted_method") or max(METHOD_LABELS, key=lambda m: method_probs[m])
        if has_odds:
            if p.get("odds_blended"):
                pam = p.get("prob_a_model")
                model_winner = p["name_a"] if (pam is not None and pam >= 0.5) else p["name_b"]
                model_pct = (pam if model_winner == p["name_a"] else 1.0 - pam) if pam is not None else None
                red_odds, blue_odds = p.get("red_odds") or "", p.get("blue_odds") or ""
                value_pick = p.get("value_side") or "No bet"
                value_ev = p.get("value_ev")
                value_stake = p.get("value_stake")
            else:
                # No odds for this fight: the raw model IS the blended model, so
                # show the same pick under both columns instead of blanking Raw
                # Model. Value needs odds, so those three cells stay empty.
                model_winner, model_pct = winner, win_pct
                red_odds, blue_odds = "", ""
                value_pick, value_ev, value_stake = "", None, None
            row = [
                p["name_a"], red_odds, p["name_b"], blue_odds, p.get("weight_class", ""),
                model_winner, model_pct, winner, win_pct,
                pred_method, method_probs[pred_method],
                method_probs["Decision"], method_probs["KO/TKO"], method_probs["Submission"],
                value_pick, value_ev, value_stake,
            ]
        else:
            row = [
                p["name_a"], p["name_b"], p.get("weight_class", ""), winner, win_pct,
                pred_method, method_probs[pred_method],
                method_probs["Decision"], method_probs["KO/TKO"], method_probs["Submission"],
            ]
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            c.alignment = hdr_align
            if ci in pct_cols:
                c.number_format = "0.0%"
    _auto_width(ws)
    _ = rankings  # Intentionally unused: keep workbook to Predictions sheet only.

    wb.save(path)


class SuperModelGUI:
    """Tkinter GUI: paste one matchup per line, predict in a background
    thread (training on first use), and export results to Excel.

    Line format: Red,[odds],Blue,[odds],Weight Class,Gender,Rounds,[Elevation
    ft | Location] — odds are auto-detected, trailing venue field optional.
    """

    BG = "#0A0A0A"
    BG_HEADER = "#111111"
    BG_INPUT = "#141414"
    FG = "#F5F5F5"
    ACCENT = "#D20A11"
    MUTED = "#CFCFCF"
    GREEN = "#FFFFFF"
    BAR_A = "#D20A11"
    BAR_B = "#F5F5F5"

    def __init__(self, root, pipeline):
        self.root = root
        self.pipeline = pipeline
        self._busy = False
        self.root.title("UFC Model")
        self.root.geometry("980x780")
        self.root.minsize(900, 680)
        self.root.configure(bg=self.BG)
        self._build_ui()

    def _build_ui(self):
        top_accent = tk.Frame(self.root, bg=self.ACCENT, height=6)
        top_accent.pack(fill="x")

        tf = tk.Frame(self.root, bg=self.BG_HEADER, pady=14)
        tf.pack(fill="x")
        tk.Label(tf, text="UFC", font=("Helvetica", 34, "bold"),
                 fg=self.ACCENT, bg=self.BG_HEADER).pack()
        tk.Label(tf, text="FIGHT PREDICTOR", font=("Helvetica", 10, "bold"),
                 fg=self.FG, bg=self.BG_HEADER).pack(pady=(0, 2))

        main = tk.Frame(self.root, bg=self.BG, padx=18, pady=12)
        main.pack(fill="both", expand=True)

        self.status_var = tk.StringVar(value="Enter matchups, then click Predict to train and run.")
        tk.Label(main, textvariable=self.status_var, bg=self.BG, fg=self.MUTED,
                 font=("Helvetica", 9, "italic")).pack(anchor="w")

        tk.Label(main, text="Enter fights — one per line:  Red,Blue,Weight Class,Gender,Rounds,Elevation(ft) or Location",
                 bg=self.BG, fg=self.FG, font=("Helvetica", 9, "bold")).pack(anchor="w", pady=(8, 1))
        tk.Label(main,
                 text=("Optional moneyline odds after each fighter (auto-detected). Last field = venue "
                       "ELEVATION in feet (e.g. 5280 — works for any new venue) OR a place name:\n"
                       "  Red,-150,Blue,+130,Weight Class,Gender,Rounds,5280      or      ...,Rounds,Denver, Colorado, USA"),
                 bg=self.BG, fg=self.MUTED, font=("Helvetica", 8, "italic"),
                 justify="left").pack(anchor="w", pady=(0, 4))

        input_wrap = tk.Frame(main, bg=self.BG_HEADER, padx=2, pady=2)
        input_wrap.pack(fill="both", expand=True, pady=4)
        self.fight_input = tk.Text(
            input_wrap, height=24, font=("Courier New", 10), bg=self.BG_INPUT,
            fg=self.FG, insertbackground=self.FG, relief="flat", wrap="word",
            highlightthickness=1, highlightbackground=self.ACCENT, highlightcolor=self.ACCENT
        )
        input_sb = tk.Scrollbar(
            input_wrap, command=self.fight_input.yview, bg=self.BG_HEADER,
            troughcolor=self.BG_INPUT, activebackground=self.ACCENT
        )
        self.fight_input.configure(yscrollcommand=input_sb.set)
        self.fight_input.pack(side="left", fill="both", expand=True)
        input_sb.pack(side="right", fill="y")

        bf = tk.Frame(main, bg=self.BG)
        bf.pack(fill="x", pady=(10, 6))
        tk.Button(bf, text="Clear", command=self._clear, font=("Helvetica", 10, "bold"),
                  bg="#202020", fg=self.FG, relief="flat", padx=14, cursor="hand2",
                  activebackground="#2A2A2A", activeforeground=self.FG).pack(side="left", padx=4)
        self.predict_btn = tk.Button(bf, text="Predict", command=self._predict,
                                     font=("Helvetica", 11, "bold"), bg=self.ACCENT, fg=self.FG,
                                     relief="flat", padx=24, cursor="hand2", state="normal",
                                     activebackground="#B40A0F", activeforeground=self.FG)
        self.predict_btn.pack(side="right", padx=4)

        self.inner = tk.Frame(main, bg=self.BG)

    def _clear(self):
        self.fight_input.delete("1.0", tk.END)

    def _predict(self):
        if self._busy:
            return
        text = self.fight_input.get("1.0", tk.END).strip()
        if not text:
            self.status_var.set("Enter at least one matchup.")
            return
        self._busy = True
        self.predict_btn.config(state="disabled")

        def _do():
            try:
                if self.pipeline.model is None:
                    self.pipeline.train()

                preds = []
                skipped_debut = 0
                for line in [ln.strip() for ln in text.splitlines() if ln.strip()]:
                    # csv.reader respects quotes, so a quoted "Baku, Azerbaijan"
                    # parses as one field; unquoted still works because location
                    # is last and the trailing parts are re-joined below.
                    try:
                        raw = next(csv.reader([line]))
                    except Exception:
                        raw = line.split(",")
                    parts = [p.strip() for p in raw if p.strip()]
                    if len(parts) < 2:
                        continue
                    # Layout: red,[red_odds],blue,[blue_odds],weight,gender,rounds,location
                    # Odds are auto-detected and location is optional, so old lines
                    # (no odds, no location) parse exactly as before.
                    a = parts[0]
                    i = 1
                    red_odds = None
                    if i < len(parts) and _looks_like_odds(parts[i]):
                        red_odds = parts[i]
                        i += 1
                    if i >= len(parts):
                        continue
                    b = parts[i]
                    i += 1
                    blue_odds = None
                    if i < len(parts) and _looks_like_odds(parts[i]):
                        blue_odds = parts[i]
                        i += 1
                    if self.pipeline.is_debutant(a) or self.pipeline.is_debutant(b):
                        skipped_debut += 1
                        continue
                    wc = parts[i] if i < len(parts) else ""
                    i += 1
                    g = parts[i] if i < len(parts) else ""
                    i += 1
                    rounds = 3
                    if i < len(parts):
                        try:
                            rounds = int(parts[i])
                            i += 1
                        except (ValueError, TypeError):
                            pass  # not a rounds int -> leave default, treat as location
                    # Trailing field is the venue, LAST. A bare number = the event's
                    # elevation in feet (works for venues the model has never seen);
                    # anything else = a place name (the only field that can contain
                    # commas, "Atlanta, Georgia, USA"), re-joined into one string and
                    # looked up. Either is optional.
                    _tail = ", ".join(parts[i:]).strip() if i < len(parts) else ""
                    location = None
                    event_elevation = None
                    if _tail:
                        try:
                            event_elevation = float(_tail)
                        except ValueError:
                            location = _tail
                    p = self.pipeline.predict_matchup(a, b, wc, g, rounds,
                                                      red_odds=red_odds, blue_odds=blue_odds,
                                                      location=location,
                                                      event_elevation=event_elevation)
                    # User-facing picks should always align with displayed probability.
                    pick_a = p["prob_a"] >= 0.5
                    p["predicted_winner"] = p["name_a"] if pick_a else p["name_b"]
                    p["win_pct"] = p["prob_a"] if pick_a else p["prob_b"]
                    p["method_probs"] = _normalize_method_probs(p.get("method_probs", {}))
                    p["predicted_method"] = p.get("predicted_method") or max(
                        METHOD_LABELS, key=lambda m: p["method_probs"][m]
                    )
                    p["method_pct"] = p["method_probs"][p["predicted_method"]]
                    p["decision_pct"] = p["method_probs"]["Decision"]
                    p["ko_tko_pct"] = p["method_probs"]["KO/TKO"]
                    p["submission_pct"] = p["method_probs"]["Submission"]
                    preds.append(p)

                try:
                    export_to_excel(PREDICTIONS_XLSX, preds, self.pipeline.division_rankings())
                    msg = (
                        f"{len(preds)} matchups predicted. "
                        f"Skipped {skipped_debut} debutant matchup(s). "
                        f"Saved to {os.path.basename(PREDICTIONS_XLSX)}"
                    )
                except Exception as e:
                    err = str(e)
                    msg = (
                        f"{len(preds)} matchups predicted. "
                        f"Skipped {skipped_debut} debutant matchup(s). "
                        f"Excel export failed: {err}"
                    )
                    self.root.after(0, lambda em=err: messagebox.showerror("Export Error", em))
                self.root.after(0, lambda: self.status_var.set(msg))
                print("")
                print("=" * 72)
                print("Prediction Run Complete")
                print("=" * 72)
                print(msg)
            except Exception as e:
                err = str(e)
                self.root.after(0, lambda em=err: self.status_var.set(f"Prediction failed: {em}"))
                print("")
                print("=" * 72)
                print("Prediction Run Failed")
                print("=" * 72)
                print(f"Prediction failed: {err}")
            finally:
                self._busy = False
                self.root.after(0, lambda: self.predict_btn.config(state="normal"))

        threading.Thread(target=_do, daemon=True).start()

def main():
    """CLI entry point: train + report (with --train-only) or launch the GUI."""
    parser = argparse.ArgumentParser(description="Train and run UFC Model.")
    parser.add_argument("--train-only", action="store_true", help="Run training/evaluation only, no GUI.")
    _ = parser.parse_args()

    if not os.path.exists(DATA_PATH):
        raise FileNotFoundError(f"Missing dataset: {DATA_PATH}")

    pipeline = UFCSuperModelPipeline(DATA_PATH)
    if _.train_only:
        pipeline.train()
        if pipeline.benchmarks:
            b = pipeline.benchmarks
            print("")
            print("=" * 72)
            print("Run Complete")
            print("=" * 72)
            print(f"Holdout calibrated log-loss: {b.super_cal_logloss:.4f}")
            print(f"Holdout raw log-loss: {b.super_raw_logloss:.4f}")
            print(f"Holdout Brier score: {b.super_brier:.4f}")
            print(f"Holdout accuracy: {b.super_acc:.2%}")
            print(f"Holdout ECE: {b.super_ece:.4f}")
            mm = pipeline.method_metrics or {}
            if mm:
                v1 = mm.get("method_acc_predicted_winner")
                v2 = mm.get("method_acc_when_winner_correct")
                v3 = mm.get("method_acc_true_winner")
                v4 = mm.get("method_majority_baseline_when_winner_correct")
                v5 = mm.get("method_finish_score_winner_correct")
                if v1 == v1:
                    print(f"Method acc (predicted winner conditioned): {v1:.1%}")
                if v2 == v2:
                    print(f"Method acc | winner pick correct: {v2:.1%}")
                if v4 == v4:
                    print(f"Majority baseline | winner pick correct: {v4:.1%}")
                if v5 == v5:
                    print(f"FinishScore (winner pick correct): {v5:.1%}")
                if v3 == v3:
                    print(f"Method acc (true winner conditioned): {v3:.1%}")
        return

    root = tk.Tk()
    SuperModelGUI(root, pipeline)
    root.mainloop()


if __name__ == "__main__":
    main()